from __future__ import annotations
import ast
import math
from pathlib import Path
import re

from openpyxl.utils.cell import column_index_from_string, coordinate_to_tuple

from ..config import cfg
from .xls_reader import open_workbook_any
from .bom_quantity import calculate_effective_needed_qty, coerce_scrap_factor
from ..models import BomFile, BomComponent

_DASH_LIKE = {"-", "—", "－", "–", "x", "X", "n", "N", "?", "N/A", "n/a", "無"}
_SCRAP_HEADER_STRONG_TERMS = (
    "拋料率",
    "抛料率",
    "損耗率",
    "损耗率",
    "耗損率",
    "耗损率",
    "不良率",
    "報廢率",
    "报废率",
    "scraprate",
    "lossrate",
    "attritionrate",
    "wastagerate",
)
_SCRAP_HEADER_WEAK_TERMS = (
    "拋料",
    "抛料",
    "損耗",
    "损耗",
    "耗損",
    "耗损",
    "報廢",
    "报废",
    "scrap",
    "attrition",
    "wastage",
)
_SCRAP_HEADER_NEGATIVE_TERMS = (
    "用量",
    "需求",
    "數量",
    "数量",
    "生產",
    "生产",
    "qty",
    "quantity",
    "needed",
)


def _try_float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _is_dash(v) -> bool:
    return v is not None and str(v).strip() in _DASH_LIKE


def _is_formula(v) -> bool:
    return isinstance(v, str) and v.lstrip().startswith("=")


def _row_value(row_vals, col_idx: int):
    return row_vals[col_idx] if len(row_vals) > col_idx else None


def _looks_like_order_quantity_cell(row_idx: int, col_idx: int, value) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return False

    order_qty_col = cfg("excel.bom_order_qty_col", 10) + 1
    data_start = cfg("excel.bom_data_start_row", 5)
    if row_idx < data_start and col_idx == order_qty_col:
        return True
    if row_idx == 2 and col_idx == 7:
        return True

    normalized = re.sub(r"[\s_:\-：／/\\()\[\]{}（）【】「」『』]+", "", text)
    quantity_terms = (
        "生產數量",
        "生产数量",
        "訂單數量",
        "订单数量",
        "數量",
        "数量",
        "orderqty",
        "quantity",
        "qty",
    )
    return row_idx < data_start and any(term in normalized for term in quantity_terms)


def _is_blank(value) -> bool:
    return value is None or str(value).strip() == ""


def _normalize_header_text(value) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[\s_:\-：／/\\()\[\]{}（）【】「」『』%％]+", "", text)


def _looks_like_scrap_header_exact(value) -> bool:
    normalized = _normalize_header_text(value)
    return bool(normalized) and normalized in _SCRAP_HEADER_STRONG_TERMS


def _looks_like_scrap_header(value) -> bool:
    normalized = _normalize_header_text(value)
    if not normalized:
        return False
    # 標題句（>10 字）例如「領料單(含0.6%拋料率)」不當成欄位 header
    if len(normalized) > 10:
        return False
    if normalized in _SCRAP_HEADER_STRONG_TERMS:
        return True
    if any(term in normalized for term in _SCRAP_HEADER_STRONG_TERMS):
        if not any(term in normalized for term in _SCRAP_HEADER_NEGATIVE_TERMS):
            return True
        return False
    if any(term in normalized for term in _SCRAP_HEADER_WEAK_TERMS):
        return not any(term in normalized for term in _SCRAP_HEADER_NEGATIVE_TERMS)
    return False


def _detect_scrap_column(rows_list: list[tuple], data_start: int, default_col: int) -> int:
    header_row_count = max(0, data_start - 1)
    # 第一輪：完全等於 STRONG 詞（最乾淨的 header cell，例如 row 3 col E = "拋料率"）
    for row_vals in rows_list[:header_row_count]:
        for col_idx, value in enumerate(row_vals or ()):
            if _looks_like_scrap_header_exact(value):
                return col_idx
    # 第二輪：寬鬆 header 匹配，配合長度＋否定詞過濾
    for row_vals in rows_list[:header_row_count]:
        for col_idx, value in enumerate(row_vals or ()):
            if _looks_like_scrap_header(value):
                return col_idx
    return default_col


def _coerce_scrap_cell_value(data_value, formula_value) -> float:
    if not _is_blank(data_value):
        return coerce_scrap_factor(data_value)
    return coerce_scrap_factor(formula_value)


def _cell_ref_to_col_idx(cell_ref: str) -> int | None:
    letters = re.sub(r"[^A-Za-z]", "", cell_ref or "")
    if not letters:
        return None
    try:
        return column_index_from_string(letters.upper()) - 1
    except ValueError:
        return None


def _scrap_factor_from_needed_formula(
    row_vals,
    formula_row_vals,
    formula_needed_raw,
    row_number: int,
    excluded_cols: set[int],
) -> float:
    if not _is_formula(formula_needed_raw):
        return 0.0

    formula = str(formula_needed_raw).upper()
    row_pattern = rf"\$?([A-Z]{{1,3}})\$?{row_number}\b"
    patterns = (
        rf"1\s*\+\s*{row_pattern}",
        rf"{row_pattern}\s*\+\s*1",
    )
    seen_cols: set[int] = set()
    for pattern in patterns:
        for match in re.finditer(pattern, formula):
            col_idx = _cell_ref_to_col_idx(match.group(1))
            if col_idx is None or col_idx in excluded_cols or col_idx in seen_cols:
                continue
            seen_cols.add(col_idx)
            factor = _coerce_scrap_cell_value(
                _row_value(row_vals, col_idx),
                _row_value(formula_row_vals, col_idx),
            )
            if factor > 0:
                return factor
    return 0.0


def _excel_round(value, digits=0):
    factor = 10 ** int(digits or 0)
    scaled = abs(float(value)) * factor
    rounded = math.floor(scaled + 0.5) / factor
    return math.copysign(rounded, float(value))


def _excel_roundup(value, digits=0):
    factor = 10 ** int(digits or 0)
    return math.copysign(math.ceil(abs(float(value)) * factor) / factor, float(value))


def _excel_rounddown(value, digits=0):
    factor = 10 ** int(digits or 0)
    return math.copysign(math.floor(abs(float(value)) * factor) / factor, float(value))


def _excel_ceiling(value, significance=1):
    amount = float(value)
    step = abs(float(significance or 1))
    if step <= 0:
        return amount
    return math.ceil(amount / step) * step


def _excel_floor(value, significance=1):
    amount = float(value)
    step = abs(float(significance or 1))
    if step <= 0:
        return amount
    return math.floor(amount / step) * step


_FORMULA_FUNCTIONS = {
    "ABS": abs,
    "CEILING": _excel_ceiling,
    "CEILING_MATH": _excel_ceiling,
    "FLOOR": _excel_floor,
    "INT": lambda value: math.floor(float(value)),
    "MAX": max,
    "MIN": min,
    "ROUND": _excel_round,
    "ROUNDDOWN": _excel_rounddown,
    "ROUNDUP": _excel_roundup,
    "SUM": lambda *values: sum(float(value or 0) for value in values),
}
_ALLOWED_FORMULA_AST_NODES = (
    ast.Expression,
    ast.BinOp,
    ast.UnaryOp,
    ast.Call,
    ast.Name,
    ast.Load,
    ast.Constant,
    ast.Add,
    ast.Sub,
    ast.Mult,
    ast.Div,
    ast.Pow,
    ast.Mod,
    ast.UAdd,
    ast.USub,
)


def _cell_formula_ref_value(data_rows, formula_rows, cell_ref: str) -> float | None:
    try:
        row_idx, col_idx = coordinate_to_tuple(cell_ref.replace("$", ""))
    except ValueError:
        return None

    data_row = data_rows[row_idx - 1] if len(data_rows) >= row_idx else ()
    formula_row = formula_rows[row_idx - 1] if len(formula_rows) >= row_idx else ()
    data_value = _row_value(data_row, col_idx - 1)
    formula_value = _row_value(formula_row, col_idx - 1)

    if _is_blank(data_value):
        if _is_blank(formula_value):
            return 0.0
        factor = coerce_scrap_factor(formula_value)
        if factor > 0:
            return factor
        return _try_float(formula_value)

    if isinstance(data_value, str) and ("%" in data_value or "％" in data_value):
        return coerce_scrap_factor(data_value)
    number = _try_float(data_value)
    if number is not None:
        return number
    if isinstance(data_value, str) and _looks_like_order_quantity_cell(row_idx, col_idx, data_value):
        return _extract_number(data_value)
    return None


def _validate_formula_ast(node) -> bool:
    for child in ast.walk(node):
        if not isinstance(child, _ALLOWED_FORMULA_AST_NODES):
            return False
        if isinstance(child, ast.Constant) and not isinstance(child.value, (int, float)):
            return False
        if isinstance(child, ast.Name) and child.id not in _FORMULA_FUNCTIONS:
            return False
        if isinstance(child, ast.Call):
            if not isinstance(child.func, ast.Name) or child.func.id not in _FORMULA_FUNCTIONS:
                return False
    return True


def _evaluate_numeric_formula(formula, data_rows, formula_rows) -> float | None:
    if not _is_formula(formula):
        return None

    expr = str(formula).strip()[1:].strip()
    if not expr or ":" in expr or "!" in expr or '"' in expr or "'" in expr:
        return None

    expr = re.sub(r"\bCEILING\.MATH\s*\(", "CEILING_MATH(", expr, flags=re.IGNORECASE)
    expr = re.sub(r"(?<![\w.])([+-]?\d+(?:\.\d+)?)\s*%", lambda m: str(float(m.group(1)) / 100), expr)

    cell_pattern = re.compile(r"(?<![A-Za-z_])\$?[A-Za-z]{1,3}\$?\d+(?![A-Za-z_])")

    def replace_cell(match):
        cell_ref = match.group(0)
        value = _cell_formula_ref_value(data_rows, formula_rows, cell_ref)
        if value is None:
            raise ValueError(f"unsupported formula cell value: {cell_ref}")
        return str(float(value))

    try:
        expr = cell_pattern.sub(replace_cell, expr)
        expr = expr.replace("^", "**")
        parsed = ast.parse(expr, mode="eval")
    except Exception:
        return None

    if not _validate_formula_ast(parsed):
        return None

    try:
        value = eval(compile(parsed, "<bom_formula>", "eval"), {"__builtins__": {}}, _FORMULA_FUNCTIONS)
    except Exception:
        return None

    try:
        amount = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(amount):
        return None
    return amount


def _extract_number(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s:
        return None
    # 嘗試直接轉 float
    try:
        return float(s)
    except (ValueError, TypeError):
        pass
    # 嘗試從字串中擷取數字（例如 "訂單數量: 100" -> 100）
    import re
    match = re.search(r"(\d+(\.\d+)?)", s)
    if match:
        try:
            return float(match.group(1))
        except (ValueError, TypeError):
            pass
    return None


def read_formula_needed_qty_cache(path: str) -> dict[tuple[str, int, str], float]:
    part_col = cfg("excel.bom_part_col", 2)
    f_col = cfg("excel.bom_needed_col", 5)
    data_start = cfg("excel.bom_data_start_row", 5)

    wb = open_workbook_any(path, read_only=True, data_only=True)
    formula_wb = open_workbook_any(path, read_only=True, data_only=False)
    ws = wb.worksheets[0]
    formula_ws = formula_wb.worksheets[0]
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    formula_rows = list(formula_ws.iter_rows(min_row=1, values_only=True))
    title = ws.title
    wb.close()
    formula_wb.close()

    cache: dict[tuple[str, int, str], float] = {}
    for row_number, row_vals in enumerate(all_rows[data_start - 1:], start=data_start):
        if not row_vals:
            continue
        part = str(_row_value(row_vals, part_col) or "").strip().upper()
        if not part:
            continue
        formula_row_vals = formula_rows[row_number - 1] if len(formula_rows) >= row_number else row_vals
        formula_needed_raw = _row_value(formula_row_vals, f_col)
        if not _is_formula(formula_needed_raw):
            continue
        evaluated_needed_qty = _evaluate_numeric_formula(formula_needed_raw, all_rows, formula_rows)
        if evaluated_needed_qty is not None and evaluated_needed_qty > 0:
            cache[(title, row_number, part)] = evaluated_needed_qty
            continue
        needed_qty = _try_float(_row_value(row_vals, f_col))
        if needed_qty is not None and needed_qty > 0:
            cache[(title, row_number, part)] = needed_qty
    return cache


def parse_bom(path: str, bom_id: str, filename: str, uploaded_at: str) -> BomFile:
    """
    解析 BOM 副檔（領料單）：
      Row 1：PO#、訂單數量
      Row 2：機種、PCB型號
      Row 5+：components
    """
    part_col = cfg("excel.bom_part_col", 2)
    desc_col = cfg("excel.bom_desc_col", 3)
    scrap_col = cfg("excel.bom_scrap_col", 4)
    qty_col = cfg("excel.bom_qty_per_board", 1)
    f_col = cfg("excel.bom_needed_col", 5)
    g_col = cfg("excel.bom_g_col", 6)
    h_col = cfg("excel.bom_h_col", 7)
    po_col = cfg("excel.bom_po_col", 7)
    oq_col = cfg("excel.bom_order_qty_col", 10)
    model_col = cfg("excel.bom_model_col", 2)
    pcb_col = cfg("excel.bom_pcb_col", 3)
    data_start = cfg("excel.bom_data_start_row", 5)

    wb = open_workbook_any(path, read_only=True, data_only=True)
    formula_wb = open_workbook_any(path, read_only=True, data_only=False)
    ws = wb.worksheets[0]
    formula_ws = formula_wb.worksheets[0]
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    formula_rows = list(formula_ws.iter_rows(min_row=1, values_only=True))
    wb.close()
    formula_wb.close()

    if len(all_rows) < 2:
        raise ValueError("BOM 檔案格式錯誤：行數不足")

    scrap_col = _detect_scrap_column(formula_rows, data_start, scrap_col)

    # Row 1（index 0）
    row1 = all_rows[0]
    po_raw = row1[po_col] if len(row1) > po_col else None
    po_extracted = _extract_number(po_raw)
    po = int(po_extracted) if po_extracted is not None else 0
    
    order_qty_raw = row1[oq_col] if len(row1) > oq_col else None
    order_qty = _extract_number(order_qty_raw) or 0.0

    # Row 2（index 1）
    row2 = all_rows[1]
    model = str(row2[model_col] or "").strip() if len(row2) > model_col else ""
    pcb = str(row2[pcb_col] or "").strip() if len(row2) > pcb_col else ""
    if not order_qty:
        order_qty = _extract_number(row2[6]) or 0.0

    # Row data_start+（index data_start-1）
    components: list[BomComponent] = []
    for row_number, row_vals in enumerate(all_rows[data_start - 1:], start=data_start):
        if not row_vals or len(row_vals) <= f_col:
            continue

        part = str(row_vals[part_col] or "").strip() if len(row_vals) > part_col else ""
        if not part:
            continue

        formula_row_vals = formula_rows[row_number - 1] if len(formula_rows) >= row_number else row_vals
        g_raw = _row_value(row_vals, g_col)
        h_raw = _row_value(row_vals, h_col)
        is_dash_flag = _is_dash(g_raw) or _is_dash(h_raw)

        qty_per = _try_float(_row_value(row_vals, qty_col)) or 0.0
        scrap_factor = _coerce_scrap_cell_value(
            _row_value(row_vals, scrap_col),
            _row_value(formula_row_vals, scrap_col),
        )
        needed_raw = _row_value(row_vals, f_col)
        formula_needed_raw = _row_value(formula_row_vals, f_col)
        if scrap_factor <= 0:
            scrap_factor = _scrap_factor_from_needed_formula(
                row_vals,
                formula_row_vals,
                formula_needed_raw,
                row_number,
                {part_col, desc_col, qty_col, f_col},
            )
        needed_qty = _evaluate_numeric_formula(formula_needed_raw, all_rows, formula_rows)
        if needed_qty is None:
            needed_qty = _try_float(needed_raw)
        if needed_qty is None:
            needed_qty = calculate_effective_needed_qty(
                needed_qty=0,
                qty_per_board=qty_per,
                scrap_factor=scrap_factor,
                schedule_order_qty=order_qty,
            )

        # 校正：F 欄快取值（needed_qty）實際上決定了這顆料真實用量，
        # 若跟 E 欄存進來的 scrap_factor 不一致，以 F 反推的 implied 為準。
        # 常見情況：E 欄填 "8"/"10" 被當 8%/10%，但 F 欄公式引用另一格的實際 scrap。
        if needed_qty and qty_per > 0 and order_qty > 0:
            expected_no_scrap = qty_per * order_qty
            if expected_no_scrap > 0:
                implied_scrap = (needed_qty / expected_no_scrap) - 1
                if 0 <= implied_scrap <= 0.5 and abs(implied_scrap - scrap_factor) > 0.005:
                    scrap_factor = implied_scrap

        # 最後保險：100% 以上拋料率實務上不合理，直接歸 0
        if scrap_factor >= 1:
            scrap_factor = 0.0

        prev_cs = _try_float(h_raw) or 0.0
        desc = str(row_vals[desc_col] or "").strip() if len(row_vals) > desc_col else ""

        components.append(BomComponent(
            part_number=part,
            description=desc,
            qty_per_board=qty_per,
            scrap_factor=scrap_factor,
            needed_qty=needed_qty or 0.0,
            prev_qty_cs=prev_cs,
            is_dash=is_dash_flag,
            is_customer_supplied=False,
            source_row=row_number,
            source_sheet=ws.title,
        ))

    return BomFile(
        id=bom_id,
        filename=filename,
        path=path,
        po_number=po,
        model=model,
        pcb=pcb,
        order_qty=order_qty,
        components=components,
        uploaded_at=uploaded_at,
    )
