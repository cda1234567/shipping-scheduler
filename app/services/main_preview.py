from __future__ import annotations

import json
from datetime import date, datetime, time
from functools import lru_cache
from pathlib import Path

from openpyxl.styles.numbers import is_date_format
from openpyxl.utils import get_column_letter, range_boundaries

from .xls_reader import is_xls, open_workbook_any

_DEFAULT_COLUMN_WIDTH_PX = 92
_DEFAULT_ROW_HEIGHT_PX = 24


def read_live_main_preview(path: str, sheet_name: str | None = None) -> dict:
    path_obj = Path(path)
    return _read_live_main_preview_cached(
        str(path_obj),
        path_obj.stat().st_mtime_ns,
        str(sheet_name or "").strip(),
    )


@lru_cache(maxsize=16)
def _read_live_main_preview_cached(path: str, modified_ns: int, requested_sheet: str) -> dict:
    workbook = open_workbook_any(path, read_only=False, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
        if not sheet_names:
            raise ValueError("Workbook has no sheets")

        selected_sheet = requested_sheet if requested_sheet in sheet_names else workbook.active.title
        worksheet = workbook[selected_sheet]
        return {
            "sheet_names": sheet_names,
            "selected_sheet": selected_sheet,
            "style_preserved": not is_xls(path),
            "sheet": _serialize_sheet(worksheet),
        }
    finally:
        try:
            workbook.close()
        except Exception:
            pass


def _serialize_sheet(worksheet) -> dict:
    max_row, max_col = _detect_used_bounds(worksheet)
    merged_anchors, merged_skips = _build_merged_maps(worksheet)
    styles: list[dict] = []
    style_map: dict[str, int] = {}
    rows: list[dict] = []

    for row_idx in range(1, max_row + 1):
        row_cells: list[dict] = []
        for col_idx in range(1, max_col + 1):
            if (row_idx, col_idx) in merged_skips:
                continue

            cell = worksheet.cell(row=row_idx, column=col_idx)
            style_id = _register_style(_serialize_style(cell), styles, style_map)
            cell_payload = {
                "col": col_idx,
                "value": _format_cell_value(cell.value, cell.number_format),
                "style_id": style_id,
            }
            merge = merged_anchors.get((row_idx, col_idx))
            if merge:
                cell_payload["rowspan"] = merge["rowspan"]
                cell_payload["colspan"] = merge["colspan"]
            row_cells.append(cell_payload)

        rows.append({
            "index": row_idx,
            "height_px": _row_height_to_px(worksheet.row_dimensions[row_idx].height),
            "cells": row_cells,
        })

    columns = [
        {
            "index": col_idx,
            "letter": get_column_letter(col_idx),
            "width_px": _column_width_to_px(worksheet.column_dimensions[get_column_letter(col_idx)].width),
        }
        for col_idx in range(1, max_col + 1)
    ]

    return {
        "name": worksheet.title,
        "row_count": max_row,
        "col_count": max_col,
        "columns": columns,
        "rows": rows,
        "styles": styles,
    }


def _detect_used_bounds(worksheet) -> tuple[int, int]:
    try:
        _, _, max_col, max_row = range_boundaries(worksheet.calculate_dimension())
    except ValueError:
        max_row, max_col = worksheet.max_row or 1, worksheet.max_column or 1

    for merged_range in worksheet.merged_cells.ranges:
        max_row = max(max_row, merged_range.max_row)
        max_col = max(max_col, merged_range.max_col)

    return max(max_row, 1), max(max_col, 1)


def _build_merged_maps(worksheet) -> tuple[dict[tuple[int, int], dict], set[tuple[int, int]]]:
    anchors: dict[tuple[int, int], dict] = {}
    skips: set[tuple[int, int]] = set()

    for merged_range in worksheet.merged_cells.ranges:
        anchor = (merged_range.min_row, merged_range.min_col)
        anchors[anchor] = {
            "rowspan": merged_range.max_row - merged_range.min_row + 1,
            "colspan": merged_range.max_col - merged_range.min_col + 1,
        }
        for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
            for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                if (row_idx, col_idx) == anchor:
                    continue
                skips.add((row_idx, col_idx))

    return anchors, skips


def _register_style(style: dict, styles: list[dict], style_map: dict[str, int]) -> int:
    style_key = json.dumps(style, sort_keys=True, ensure_ascii=False, separators=(",", ":"))
    cached = style_map.get(style_key)
    if cached is not None:
        return cached

    style_id = len(styles)
    styles.append(style)
    style_map[style_key] = style_id
    return style_id


def _serialize_style(cell) -> dict:
    style: dict[str, object] = {}

    background = _extract_fill_color(cell.fill)
    if background:
        style["background"] = background

    font_color = _color_to_hex(cell.font.color)
    if font_color:
        style["color"] = font_color
    if cell.font.name:
        style["font_name"] = cell.font.name
    if cell.font.sz:
        style["font_size"] = round(float(cell.font.sz), 1)
    if cell.font.bold:
        style["bold"] = True
    if cell.font.italic:
        style["italic"] = True
    if cell.font.underline and cell.font.underline != "none":
        style["underline"] = True

    horizontal = cell.alignment.horizontal or ("right" if isinstance(cell.value, (int, float)) else "left")
    vertical = cell.alignment.vertical or "center"
    if horizontal:
        style["align"] = horizontal
    if vertical:
        style["valign"] = vertical
    if cell.alignment.wrap_text:
        style["wrap"] = True

    for side_name in ("top", "right", "bottom", "left"):
        border_css = _border_side_to_css(getattr(cell.border, side_name))
        if border_css:
            style[f"border_{side_name}"] = border_css

    return style


def _extract_fill_color(fill) -> str | None:
    if getattr(fill, "fill_type", None) != "solid":
        return None
    return _color_to_hex(getattr(fill, "fgColor", None))


def _color_to_hex(color) -> str | None:
    if not color:
        return None

    color_type = getattr(color, "type", None)
    if color_type != "rgb":
        return None

    rgb = str(getattr(color, "rgb", "") or "").strip()
    if len(rgb) == 8:
        rgb = rgb[2:]
    if len(rgb) == 6 and rgb.upper() != "000000":
        return f"#{rgb.upper()}"
    return None


def _border_side_to_css(side) -> str | None:
    border_style = getattr(side, "style", None)
    if not border_style:
        return None

    width = "1px"
    css_style = "solid"
    if border_style in {"medium", "mediumDashDot", "mediumDashDotDot", "mediumDashed"}:
        width = "2px"
    elif border_style == "thick":
        width = "3px"
    elif border_style == "double":
        width = "3px"
        css_style = "double"
    elif border_style in {"dashed", "mediumDashed"}:
        css_style = "dashed"
    elif border_style == "dotted":
        css_style = "dotted"

    color = _color_to_hex(getattr(side, "color", None)) or "#d0d7e2"
    return f"{width} {css_style} {color}"


def _format_cell_value(value, number_format: str | None = None) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y/%m/%d")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"

    if isinstance(value, (int, float)):
        if number_format and is_date_format(number_format):
            return str(value)
        return _format_numeric(value, number_format or "")

    return str(value)


def _format_numeric(value: int | float, number_format: str) -> str:
    format_code = (number_format or "").split(";")[0]
    decimals = _count_decimal_places(format_code)
    use_grouping = "," in format_code and "%" not in format_code

    if "%" in format_code:
        digits = 0 if decimals is None else decimals
        return f"{value * 100:.{digits}f}%"

    if decimals is not None:
        template = f"{{:{',' if use_grouping else ''}.{decimals}f}}"
        rendered = template.format(value)
        if decimals > 0:
            rendered = rendered.rstrip("0").rstrip(".")
        return rendered

    if float(value).is_integer():
        integer_value = int(round(float(value)))
        return f"{integer_value:,}" if use_grouping else str(integer_value)

    rendered = f"{value:,.4f}" if use_grouping else f"{value:.4f}"
    return rendered.rstrip("0").rstrip(".")


def _count_decimal_places(format_code: str) -> int | None:
    if "." not in format_code:
        return 0 if "0" in format_code else None

    decimals = 0
    started = False
    for char in format_code.split(".", 1)[1]:
        if char in {"0", "#"}:
            decimals += 1
            started = True
            continue
        if started:
            break
    return decimals


def _column_width_to_px(width: float | None) -> int:
    if width is None:
        return _DEFAULT_COLUMN_WIDTH_PX
    return max(48, int(round(width * 7 + 14)))


def _row_height_to_px(height: float | None) -> int:
    if height is None:
        return _DEFAULT_ROW_HEIGHT_PX
    return max(22, int(round(height * 1.34)))
