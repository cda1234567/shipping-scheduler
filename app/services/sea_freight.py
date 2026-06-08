"""海運出貨匯入、匹配與出貨單輸出。"""
from __future__ import annotations

import math
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from ..config import BASE_DIR, SEA_FREIGHT_DIR
from .local_time import local_now

PACKING_SPEC_TEMPLATE = BASE_DIR / "templates" / "opentext 出貨淨重.xlsx"
SEA_OUTPUT_TEMPLATE = BASE_DIR / "templates" / "OPENTEXT 260310海運出貨單.xlsx"


def _text(value: Any) -> str:
    return str(value or "").strip()


def _upper(value: Any) -> str:
    return _text(value).upper()


def _num(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(str(value).replace(",", "").replace("'", "").strip())
    except Exception:
        return 0.0


def _date_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d")
    return _text(value)


def _infer_harmonized_code(item_no: str, description: str) -> str:
    text = f"{item_no} {description}".upper()
    if "POWER" in text or item_no.startswith("OC-10937") or item_no.startswith("OC-10838"):
        return "8504.40.6001"
    if "BAG" in text or "MEMBRANE" in text or item_no.startswith(("PK-", "OC-103", "OC-107", "OC-1084", "OC-10960")):
        return "3923.29.0000"
    if "USB3" in text or "TC7" in text:
        return "3923.29.0000"
    if "CABLE" in text or item_no.startswith("AC-"):
        return "8544.42.2000"
    return ""


def _load_packing_specs(path: Path = PACKING_SPEC_TEMPLATE) -> list[dict]:
    if not path.exists():
        return []
    ws = load_workbook(path, data_only=True).active
    specs = []
    for row in range(1, ws.max_row + 1):
        name = _text(ws.cell(row, 1).value)
        per_box = _num(ws.cell(row, 5).value)
        if not name or not per_box:
            continue
        specs.append({
            "name": name,
            "key": name.upper(),
            "per_box_qty": per_box,
            "net_weight": _num(ws.cell(row, 6).value),
            "gross_weight": _num(ws.cell(row, 7).value),
            "volume": _num(ws.cell(row, 8).value),
            "vendor": _text(ws.cell(row, 10).value),
        })
    return specs


def _find_spec(item_no: str, specs: list[dict]) -> dict | None:
    key = _upper(item_no)
    if not key:
        return None
    return next((spec for spec in specs if key == _upper(spec.get("item_no"))), None) \
        or next((spec for spec in specs if key in _upper(spec.get("packing_name") or spec.get("name"))), None)


def _extract_item_no(name: str) -> str:
    text = _upper(name)
    patterns = [
        r"\b[A-Z]{2}-[A-Z0-9]+(?:/[A-Z])?\b",
        r"\bPB-\d+[A-Z]?\b",
        r"\bPK-[A-Z0-9]+\b",
    ]
    matches: list[str] = []
    for pattern in patterns:
        matches.extend(re.findall(pattern, text))
    return matches[-1] if matches else ""


def _calc_boxes(qty: float, per_box_qty: float) -> tuple[int, float]:
    if qty <= 0 or per_box_qty <= 0:
        return 0, 0
    box_count = int(math.ceil(qty / per_box_qty))
    tail = qty % per_box_qty
    return box_count, tail


def load_packing_specs_from_template(path: Path = PACKING_SPEC_TEMPLATE) -> list[dict]:
    specs = []
    for spec in _load_packing_specs(path):
        item_no = _extract_item_no(spec["name"])
        if not item_no:
            continue
        specs.append({
            "item_no": item_no,
            "packing_name": spec["name"],
            "per_box_qty": spec["per_box_qty"],
            "net_weight": spec["net_weight"],
            "gross_weight": spec["gross_weight"],
            "volume": spec["volume"],
            "vendor": spec.get("vendor", ""),
        })
    return specs


def parse_sea_order_file(
    path: str | Path,
    hs_codes: dict[str, str] | None = None,
    packing_specs: list[dict] | None = None,
) -> tuple[dict, list[dict]]:
    hs_codes = hs_codes or {}
    workbook = load_workbook(path, data_only=True)
    ws = workbook.active
    specs = packing_specs if packing_specs is not None else load_packing_specs_from_template()
    items: list[dict] = []

    for row in range(2, ws.max_row + 1):
        item_no = _text(ws.cell(row, 4).value)
        if not item_no:
            continue
        qty = _num(ws.cell(row, 6).value)
        spec = _find_spec(item_no, specs)
        per_box = float(spec["per_box_qty"]) if spec else 0
        box_count, tail_qty = _calc_boxes(qty, per_box)
        description = _text(ws.cell(row, 7).value)
        code = hs_codes.get(item_no) or _infer_harmonized_code(item_no, description)
        items.append({
            "order_date": _date_text(ws.cell(row, 1).value),
            "customer": _text(ws.cell(row, 2).value),
            "cust_po": _text(ws.cell(row, 3).value),
            "item_no": item_no,
            "price": _num(ws.cell(row, 5).value),
            "qty": qty,
            "description": description,
            "delivery_date": _date_text(ws.cell(row, 8).value),
            "packing_name": spec.get("packing_name") or spec.get("name") if spec else description,
            "per_box_qty": per_box,
            "net_weight": float(spec["net_weight"]) if spec else 0,
            "gross_weight": float(spec["gross_weight"]) if spec else 0,
            "volume": float(spec["volume"]) if spec else 0,
            "box_count": box_count,
            "tail_qty": tail_qty,
            "carton_no": "",
            "harmonized_code": code,
            "match_status": "matched" if spec else "missing",
        })

    first = items[0] if items else {}
    meta = {
        "customer": first.get("customer", ""),
        "cust_po": first.get("cust_po", ""),
        "shipment_date": local_now().strftime("%Y/%m/%d"),
        "delivery_date": first.get("delivery_date", ""),
        "maker": "Andy",
        "mark_text": "HILLIARD",
        "invoice_no": "",
    }
    return meta, items


def save_uploaded_sea_file(filename: str, content: bytes) -> Path:
    SEA_FREIGHT_DIR.mkdir(parents=True, exist_ok=True)
    safe_name = re.sub(r'[\\/:*?"<>|]+', "_", filename or "sea_freight.xlsx")
    stamp = local_now().strftime("%Y%m%d_%H%M%S")
    path = SEA_FREIGHT_DIR / f"{stamp}_{safe_name}"
    path.write_bytes(content)
    return path


def _set(ws, row: int, col: int, value: Any):
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def _clear_cells(ws, rows: list[int]):
    for row in rows:
        for col in range(1, 18):
            if not isinstance(ws.cell(row, col), MergedCell):
                ws.cell(row, col).value = None


def _line_amount(item: dict) -> float:
    return float(item.get("price") or 0) * float(item.get("qty") or 0)


def _weight_total(item: dict, key: str) -> float:
    boxes = int(item.get("box_count") or 0)
    tail = float(item.get("tail_qty") or 0)
    per_box = float(item.get("per_box_qty") or 0)
    unit = float(item.get(key) or 0)
    full_boxes = boxes - 1 if tail > 0 else boxes
    total = max(full_boxes, 0) * unit
    if tail > 0 and per_box > 0:
        total += unit * tail / per_box
    return total


def _volume_total(item: dict) -> float:
    boxes = int(item.get("box_count") or 0)
    tail = float(item.get("tail_qty") or 0)
    full_boxes = boxes - 1 if tail > 0 else boxes
    total = max(full_boxes, 0) * float(item.get("volume") or 0)
    if tail > 0:
        total += float(item.get("volume") or 0)
    return total


def _write_item(ws, item: dict, index: int, start_row: int):
    qty = float(item.get("qty") or 0)
    per_box = float(item.get("per_box_qty") or 0)
    boxes = int(item.get("box_count") or 0)
    tail = float(item.get("tail_qty") or 0)
    full_boxes = boxes - 1 if tail > 0 else boxes
    full_qty = full_boxes * per_box
    net = float(item.get("net_weight") or 0)
    gross = float(item.get("gross_weight") or 0)
    volume = float(item.get("volume") or 0)
    price = float(item.get("price") or 0)

    _set(ws, start_row, 1, index)
    _set(ws, start_row, 2, f"HARMONIZED CODE: {item.get('harmonized_code') or ''}")
    _set(ws, start_row + 1, 2, f"PO: {item.get('cust_po') or ''}")
    _set(ws, start_row + 1, 9, item.get("carton_no") or "")
    _set(ws, start_row + 1, 10, full_boxes or None)
    _set(ws, start_row + 1, 11, per_box or None)
    _set(ws, start_row + 1, 12, net or None)
    _set(ws, start_row + 1, 13, gross or None)
    _set(ws, start_row + 1, 14, volume or None)

    _set(ws, start_row + 2, 2, item.get("packing_name") or item.get("description") or item.get("item_no"))
    _set(ws, start_row + 2, 11, full_qty or None)
    _set(ws, start_row + 2, 12, full_boxes * net if full_boxes else None)
    _set(ws, start_row + 2, 13, full_boxes * gross if full_boxes else None)
    _set(ws, start_row + 2, 14, full_boxes * volume if full_boxes else None)
    _set(ws, start_row + 2, 15, qty)
    _set(ws, start_row + 2, 16, price)
    _set(ws, start_row + 2, 17, _line_amount(item))

    if tail > 0:
        _set(ws, start_row + 3, 10, 1)
        _set(ws, start_row + 3, 11, tail)
        _set(ws, start_row + 3, 12, net * tail / per_box if per_box else None)
        _set(ws, start_row + 3, 13, gross * tail / per_box if per_box else None)
        _set(ws, start_row + 3, 14, volume)
    else:
        for col in range(9, 18):
            _set(ws, start_row + 3, col, None)


def _write_summary(ws, page_items: list[dict]):
    total_boxes = sum(int(item.get("box_count") or 0) for item in page_items)
    total_volume = sum(_volume_total(item) for item in page_items)
    total_weight = sum(_weight_total(item, "gross_weight") for item in page_items)
    top_amount = sum(_line_amount(item) for item in page_items[:6])
    bottom_amount = sum(_line_amount(item) for item in page_items[6:12])
    grand_total = top_amount + bottom_amount
    for row in (30, 63):
        _set(ws, row, 2, "Andy")
        _set(ws, row, 5, "HILLIARD")
        _set(ws, row, 15, total_boxes)
    for row in (31, 64):
        _set(ws, row, 15, total_volume)
        _set(ws, row, 17, 0)
    for row in (32, 65):
        _set(ws, row, 15, total_weight)
    _set(ws, 30, 17, top_amount)
    _set(ws, 32, 17, top_amount)
    _set(ws, 33, 17, grand_total)
    _set(ws, 63, 17, bottom_amount)
    _set(ws, 65, 17, bottom_amount)
    _set(ws, 66, 17, grand_total)


def export_sea_shipment(shipment: dict) -> Path:
    if not SEA_OUTPUT_TEMPLATE.exists():
        raise FileNotFoundError(f"找不到海運輸出模板：{SEA_OUTPUT_TEMPLATE}")

    items = list(shipment.get("items") or [])
    if not items:
        raise ValueError("沒有可匯出的海運明細")

    workbook = load_workbook(SEA_OUTPUT_TEMPLATE)
    template = workbook.active
    chunks = [items[i:i + 12] for i in range(0, len(items), 12)]
    slot_rows = [6, 10, 14, 18, 22, 26, 39, 43, 47, 51, 55, 59]
    clear_rows = []
    for row in slot_rows:
        clear_rows.extend([row, row + 1, row + 2, row + 3])

    for page_index, page_items in enumerate(chunks):
        ws = template if page_index == 0 else workbook.copy_worksheet(template)
        ws.title = f"海運出貨{page_index + 1}" if len(chunks) > 1 else "海運出貨單"
        _clear_cells(ws, clear_rows)

        invoice_no = shipment.get("invoice_no") or ""
        customer = shipment.get("customer") or "OPEN TEXT"
        ship_date = shipment.get("shipment_date") or local_now().strftime("%Y/%m/%d")
        for row in (3, 36):
            _set(ws, row, 17, ship_date)
        for row in (2, 35):
            _set(ws, row, 17, invoice_no)
        for row in (4, 37):
            _set(ws, row, 4, customer)

        for offset, item in enumerate(page_items):
            _write_item(ws, item, page_index * 12 + offset + 1, slot_rows[offset])
        _write_summary(ws, page_items)

    if len(chunks) > 1 and "0106空運" in workbook.sheetnames:
        pass

    output_dir = SEA_FREIGHT_DIR / "exports"
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = local_now().strftime("%Y%m%d_%H%M%S")
    output = output_dir / f"海運出貨單_{shipment.get('id', '')}_{stamp}.xlsx"
    workbook.save(output)
    return output


def copy_template_for_review(target: Path):
    shutil.copy2(SEA_OUTPUT_TEMPLATE, target)
