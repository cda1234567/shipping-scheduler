"""
不良品 Excel 匯入 + 主檔扣帳邏輯。

副檔格式（BOM 領料單）：
  Row 1: PO# (col H)
  Row 2: 機種 (col C), PCB (col D)
  Row 5+: 料號 (col C), 說明 (col D), 不良數量 (col F)
"""
from __future__ import annotations

from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from ..config import cfg
from .local_time import local_now
from .xls_reader import open_workbook_any
from .merge_to_main import (
    PART_COL,
    RED_FILL,
    STOCK_SEARCH_START_COL,
    _build_part_row_map,
    _read_latest_stock,
    _round_away,
    backup_main_file,
)


def _try_float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def parse_defective_excel(path: str) -> list[dict]:
    """解析副檔格式的不良品 Excel，回傳 [{part_number, description, defective_qty}]。"""
    part_col = cfg("excel.bom_part_col", 2)
    desc_col = cfg("excel.bom_desc_col", 3)
    qty_col = cfg("excel.bom_needed_col", 5)
    data_start = cfg("excel.bom_data_start_row", 5)

    wb = open_workbook_any(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    items: list[dict] = []
    for offset, row_vals in enumerate(all_rows[data_start - 1:], start=data_start):
        if not row_vals or len(row_vals) <= qty_col:
            continue

        part = str(row_vals[part_col] or "").strip().upper()
        if not part:
            continue

        qty = _try_float(row_vals[qty_col])
        if not qty or qty <= 0:
            continue

        desc = str(row_vals[desc_col] or "").strip() if len(row_vals) > desc_col else ""
        items.append({
            "source_row": offset,
            "part_number": part,
            "description": desc,
            "defective_qty": qty,
        })

    return items


HEADER_FONT = Font(bold=True, size=9)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
REVERSE_HEADER_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


def _get_main_worksheet(workbook):
    if getattr(workbook, "worksheets", None):
        return workbook.worksheets[0]
    return workbook.active


def _copy_column_layout(ws, source_col: int, target_col: int):
    if source_col <= 0 or target_col <= 0 or source_col == target_col:
        return

    source_letter = openpyxl.utils.get_column_letter(source_col)
    target_letter = openpyxl.utils.get_column_letter(target_col)
    source_dimension = ws.column_dimensions[source_letter]
    target_dimension = ws.column_dimensions[target_letter]

    for attr in ("width", "hidden", "bestFit", "style", "outlineLevel", "collapsed"):
        try:
            setattr(target_dimension, attr, copy(getattr(source_dimension, attr)))
        except Exception:
            continue

    for row_idx in range(1, ws.max_row + 1):
        source_cell = ws.cell(row=row_idx, column=source_col)
        target_cell = ws.cell(row=row_idx, column=target_col)
        target_cell._style = copy(source_cell._style)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)


def _prepare_new_entry_columns(ws, max_col: int, col_delta: int, col_stock: int):
    source_stock_col = max_col
    source_delta_col = max_col - 1 if max_col - 1 >= STOCK_SEARCH_START_COL else source_stock_col
    _copy_column_layout(ws, source_delta_col, col_delta)
    _copy_column_layout(ws, source_stock_col, col_stock)


def deduct_defectives_from_main(
    main_path: str,
    items: list[dict],
    backup_dir: str | None = None,
    entry_header: str = "不良品扣帳",
) -> dict:
    """
    在主檔新增 2 欄扣帳：扣帳數量 + 扣帳後庫存。

    回傳 {"backup_path", "deducted_count", "skipped_parts", "results"}。
    """
    if not items:
        return {"backup_path": None, "deducted_count": 0, "skipped_parts": [], "results": []}

    backup_path = backup_main_file(main_path, backup_dir) if backup_dir else None

    is_xlsm = Path(main_path).suffix.lower() == ".xlsm"
    wb = openpyxl.load_workbook(main_path, keep_vba=is_xlsm)
    ws = _get_main_worksheet(wb)
    part_row_map = _build_part_row_map(ws)
    max_col = ws.max_column

    col_deduct = max_col + 1
    col_stock = max_col + 2
    _prepare_new_entry_columns(ws, max_col, col_deduct, col_stock)

    # 寫表頭
    ts_label = local_now().strftime("%m/%d %H:%M")
    deduct_header = ws.cell(row=1, column=col_deduct)
    deduct_header.value = str(entry_header or "不良品扣帳").strip() or "不良品扣帳"
    deduct_header.font = HEADER_FONT
    deduct_header.fill = HEADER_FILL
    deduct_header.alignment = CENTER_ALIGN

    stock_header = ws.cell(row=1, column=col_stock)
    stock_header.value = ts_label
    stock_header.font = HEADER_FONT
    stock_header.fill = HEADER_FILL
    stock_header.alignment = CENTER_ALIGN

    results: list[dict] = []
    skipped: list[str] = []

    for item in items:
        part_upper = item["part_number"].upper()
        row_idx = part_row_map.get(part_upper)
        if row_idx is None:
            skipped.append(part_upper)
            continue

        current_stock = _read_latest_stock(ws, row_idx, max_col)
        deduct_qty = float(item["defective_qty"])
        new_stock = _round_away(current_stock - deduct_qty)

        ws.cell(row=row_idx, column=col_deduct).value = _round_away(deduct_qty)
        stock_cell = ws.cell(row=row_idx, column=col_stock)
        stock_cell.value = new_stock
        if new_stock < 0:
            stock_cell.fill = RED_FILL

        results.append({
            "part_number": part_upper,
            "description": item.get("description", ""),
            "defective_qty": deduct_qty,
            "stock_before": current_stock,
            "stock_after": new_stock,
        })

    wb.save(main_path)
    wb.close()

    return {
        "backup_path": backup_path,
        "deducted_count": len(results),
        "skipped_parts": skipped,
        "results": results,
    }


def reverse_defectives_from_main(
    main_path: str,
    items: list[dict],
    backup_dir: str | None = None,
    entry_header: str = "不良品回復",
) -> dict:
    """
    將已扣帳的數量加回主檔（刪除批次時用）。

    items: [{part_number, defective_qty}]
    在主檔新增 2 欄：回復數量 + 回復後庫存。
    回傳 {"backup_path", "reversed_count", "skipped_parts", "results"}。
    """
    if not items:
        return {"backup_path": None, "reversed_count": 0, "skipped_parts": [], "results": []}

    backup_path = backup_main_file(main_path, backup_dir) if backup_dir else None

    is_xlsm = Path(main_path).suffix.lower() == ".xlsm"
    wb = openpyxl.load_workbook(main_path, keep_vba=is_xlsm)
    ws = _get_main_worksheet(wb)
    part_row_map = _build_part_row_map(ws)
    max_col = ws.max_column

    col_reverse = max_col + 1
    col_stock = max_col + 2
    _prepare_new_entry_columns(ws, max_col, col_reverse, col_stock)

    ts_label = local_now().strftime("%m/%d %H:%M")
    reverse_header = ws.cell(row=1, column=col_reverse)
    reverse_header.value = str(entry_header or "不良品回復").strip() or "不良品回復"
    reverse_header.font = HEADER_FONT
    reverse_header.fill = REVERSE_HEADER_FILL
    reverse_header.alignment = CENTER_ALIGN

    stock_header = ws.cell(row=1, column=col_stock)
    stock_header.value = ts_label
    stock_header.font = HEADER_FONT
    stock_header.fill = REVERSE_HEADER_FILL
    stock_header.alignment = CENTER_ALIGN

    results: list[dict] = []
    skipped: list[str] = []

    for item in items:
        part_upper = item["part_number"].upper()
        row_idx = part_row_map.get(part_upper)
        if row_idx is None:
            skipped.append(part_upper)
            continue

        current_stock = _read_latest_stock(ws, row_idx, max_col)
        reverse_qty = float(item["defective_qty"])
        new_stock = _round_away(current_stock + reverse_qty)

        ws.cell(row=row_idx, column=col_reverse).value = _round_away(reverse_qty)
        stock_cell = ws.cell(row=row_idx, column=col_stock)
        stock_cell.value = new_stock
        if new_stock < 0:
            stock_cell.fill = RED_FILL

        results.append({
            "part_number": part_upper,
            "reverse_qty": reverse_qty,
            "stock_before": current_stock,
            "stock_after": new_stock,
        })

    wb.save(main_path)
    wb.close()

    return {
        "backup_path": backup_path,
        "reversed_count": len(results),
        "skipped_parts": skipped,
        "results": results,
    }
