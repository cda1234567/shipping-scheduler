"""
Merge BOM demand into the live main workbook.

This module now exposes both:
- a preview path that simulates how selected orders will write into main
- the real write path used by dispatch
"""
from __future__ import annotations

from copy import copy
import shutil
from math import copysign, floor
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from ..config import cfg
from ..models import calc_suggested_qty
from .local_time import local_now
from .shortage_rules import (
    calculate_current_order_shortage_amount,
    calculate_shortage_amount,
    is_order_scoped_shortage_part,
    summarize_requested_supply,
    summarize_st_supply,
)

PART_COL = 1
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
CLEAR_FILL = PatternFill(fill_type=None)
STOCK_SEARCH_START_COL = cfg("excel.main_moq_col", 2) + 2


def _try_float(value) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _round_away(value: float) -> float:
    if value == 0:
        return 0.0
    return copysign(floor(abs(value) + 0.5), value)


def _normalize_decisions(decisions: dict[str, str] | None = None) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for part, decision in (decisions or {}).items():
        key = str(part or "").strip().upper()
        if not key or not decision:
            continue
        normalized[key] = str(decision)
    return normalized


def _normalize_supplements(supplements: dict[str, float] | None = None) -> dict[str, float]:
    normalized: dict[str, float] = {}
    for part, qty in (supplements or {}).items():
        key = str(part or "").strip().upper()
        amount = _try_float(qty) or 0.0
        if not key or amount <= 0:
            continue
        normalized[key] = float(amount)
    return normalized


def _resolve_decision(part_number: str, decisions: dict[str, str]) -> str:
    key = str(part_number or "").strip().upper()
    return decisions.get(key, "None")


def _build_part_row_map(ws) -> dict[str, int]:
    part_row_map: dict[str, int] = {}
    for row_idx in range(2, ws.max_row + 1):
        raw = ws.cell(row=row_idx, column=PART_COL).value
        part = str(raw or "").strip().upper()
        if part:
            part_row_map[part] = row_idx
    return part_row_map


def _read_latest_stock(ws, row_idx: int, max_col: int) -> float:
    current_stock = 0.0
    for col_idx in range(max_col, STOCK_SEARCH_START_COL - 1, -1):
        value = _try_float(ws.cell(row=row_idx, column=col_idx).value)
        if value is not None:
            current_stock = value
            break
    return current_stock


def clear_cell_fill(cell) -> None:
    cell.fill = CLEAR_FILL


def ensure_main_header_wrap(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        wrapped_alignment = copy(cell.alignment)
        wrapped_alignment.wrap_text = True
        if not wrapped_alignment.vertical:
            wrapped_alignment.vertical = "center"
        cell.alignment = wrapped_alignment


def backup_main_file(main_path: str, backup_dir: str) -> str:
    source = Path(main_path)
    destination_dir = Path(backup_dir)
    destination_dir.mkdir(parents=True, exist_ok=True)
    ts = local_now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{source.stem}_backup_{ts}{source.suffix}"
    destination = destination_dir / backup_name
    shutil.copy2(main_path, destination)
    return str(destination)


def _build_preview_for_batches(
    ws,
    batches: list[dict],
    decisions: dict[str, str],
    moq_map: dict[str, float] | None = None,
    st_inventory_stock: dict[str, float] | None = None,
) -> dict:
    part_row_map = _build_part_row_map(ws)
    running_stock: dict[str, float] = {}
    planned_batches: list[dict] = []
    shortages: list[dict] = []
    max_col = ws.max_column
    total_merged = 0
    effective_moq = {
        str(part or "").strip().upper(): float(qty or 0)
        for part, qty in (moq_map or {}).items()
        if str(part or "").strip()
    }

    for batch in batches:
        remaining_supplements = _normalize_supplements(batch.get("supplements") or {})
        batch_decisions = _normalize_decisions(batch.get("decisions") or decisions)
        planned_groups: list[dict] = []

        for group in batch.get("groups", []):
            components = group.get("components", []) or []
            if not components:
                continue

            col_h = max_col + 1
            col_f = max_col + 2
            col_j = max_col + 3
            group_rows: list[dict] = []
            group_shortages: list[dict] = []

            for comp in components:
                needed_qty = float(comp.get("needed_qty") or 0)
                if comp.get("is_dash") or needed_qty <= 0:
                    continue

                part_number = str(comp.get("part_number") or "").strip()
                part_upper = part_number.upper()
                row_idx = part_row_map.get(part_upper)
                if row_idx is None:
                    continue

                current_stock = running_stock.get(part_upper)
                if current_stock is None:
                    current_stock = _read_latest_stock(ws, row_idx, max_col)

                prev_qty_cs = float(comp.get("prev_qty_cs") or 0)
                decision = _resolve_decision(part_number, batch_decisions)

                available_before = current_stock + prev_qty_cs
                supplement_qty = 0.0
                ending_without_supplement = available_before - needed_qty
                shortage_before = calculate_current_order_shortage_amount(part_upper, available_before, needed_qty)
                if decision != "Shortage" and shortage_before > 0 and remaining_supplements.get(part_upper, 0) > 0:
                    supplement_qty = float(remaining_supplements.get(part_upper, 0))
                    remaining_supplements[part_upper] = 0.0

                effective_h = prev_qty_cs + supplement_qty
                available_after_supply = current_stock + effective_h
                # 永遠扣帳（即使缺料也照扣，讓庫存反映真實狀態）
                ending_stock = available_after_supply - needed_qty
                shortage_after = calculate_current_order_shortage_amount(part_upper, available_after_supply, needed_qty)
                f_value = _round_away(needed_qty)

                running_stock[part_upper] = ending_stock
                total_merged += 1

                row_plan = {
                    "row_idx": row_idx,
                    "part_number": part_upper,
                    "description": str(comp.get("description") or ""),
                    "decision": decision,
                    "current_stock": float(current_stock),
                    "prev_qty_cs": prev_qty_cs,
                    "supplement_qty": supplement_qty,
                    "effective_h": _round_away(effective_h) if effective_h else 0,
                    "needed_qty": needed_qty,
                    "f_value": f_value,
                    "j_value": _round_away(ending_stock),
                    "shortage_amount": shortage_after,
                    "col_h": col_h,
                    "col_f": col_f,
                    "col_j": col_j,
                }
                group_rows.append(row_plan)

                if shortage_after > 0:
                    item_moq = float(effective_moq.get(part_upper, 0) or 0)
                    st_stock_qty = (st_inventory_stock or {}).get(part_upper, 0.0)
                    if is_order_scoped_shortage_part(part_upper):
                        st_context = summarize_requested_supply(shortage_after, st_stock_qty)
                        st_available_qty = float(st_context["st_available_qty"] or 0.0)
                        purchase_needed_qty = float(st_context["purchase_needed_qty"] or 0.0)
                        purchase_suggested_qty = purchase_needed_qty
                        suggested_qty = shortage_after
                    else:
                        st_context = summarize_st_supply(shortage_after, st_stock_qty, item_moq)
                        st_available_qty = float(st_context["st_available_qty"] or 0.0)
                        purchase_needed_qty = float(st_context["purchase_needed_qty"] or 0.0)
                        purchase_suggested_qty = (
                            calc_suggested_qty(purchase_needed_qty, item_moq) if purchase_needed_qty > 0 else 0.0
                        )
                        suggested_qty = st_available_qty + purchase_suggested_qty
                    shortage = {
                        "order_id": batch.get("order_id"),
                        "batch_code": group.get("batch_code", ""),
                        "po_number": group.get("po_number", ""),
                        "model": batch.get("model", ""),
                        "bom_model": group.get("bom_model", ""),
                        "part_number": part_upper,
                        "description": str(comp.get("description") or ""),
                        "current_stock": available_after_supply,
                        "needed": needed_qty,
                        "shortage_amount": shortage_after,
                        "moq": item_moq,
                        "supplement_qty": supplement_qty,
                        "decision": decision,
                        "resulting_stock": ending_stock,
                        "suggested_qty": suggested_qty if shortage_after > 0 else 0.0,
                        "purchase_suggested_qty": purchase_suggested_qty,
                        **st_context,
                    }
                    group_shortages.append(shortage)
                    shortages.append(shortage)

            if group_rows:
                planned_groups.append({
                    "batch_code": group.get("batch_code", ""),
                    "po_number": group.get("po_number", ""),
                    "bom_model": group.get("bom_model", ""),
                    "col_h": col_h,
                    "col_f": col_f,
                    "col_j": col_j,
                    "rows": group_rows,
                    "shortages": group_shortages,
                })
                max_col = col_j

        planned_batches.append({
            "order_id": batch.get("order_id"),
            "model": batch.get("model", ""),
            "groups": planned_groups,
        })

    return {
        "batches": planned_batches,
        "shortages": shortages,
        "merged_parts": total_merged,
        "new_col_count": max_col,
    }


def preview_order_batches(
    main_path: str,
    batches: list[dict],
    decisions: dict[str, str] | None = None,
    moq_map: dict[str, float] | None = None,
    st_inventory_stock: dict[str, float] | None = None,
) -> dict:
    workbook = openpyxl.load_workbook(main_path, keep_vba=(Path(main_path).suffix.lower() == ".xlsm"))
    try:
        return _build_preview_for_batches(
            workbook.active,
            batches,
            _normalize_decisions(decisions),
            moq_map=moq_map,
            st_inventory_stock=st_inventory_stock,
        )
    finally:
        workbook.close()


def _write_group_headers(ws, group_plan: dict):
    header_font = Font(bold=True, size=9)
    for row_idx, column, value in (
        (1, group_plan["col_h"], group_plan.get("batch_code", "")),
        (1, group_plan["col_f"], group_plan.get("po_number", "")),
        (1, group_plan["col_j"], group_plan.get("bom_model", "")),
    ):
        cell = ws.cell(row=row_idx, column=column)
        cell.value = value
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_group_rows(ws, group_plan: dict):
    for row in group_plan.get("rows", []):
        h_cell = ws.cell(row=row["row_idx"], column=group_plan["col_h"])
        f_cell = ws.cell(row=row["row_idx"], column=group_plan["col_f"])
        j_cell = ws.cell(row=row["row_idx"], column=group_plan["col_j"])

        h_cell.value = row["effective_h"]

        f_cell.value = row["f_value"]
        if row["decision"] == "CreateRequirement":
            f_cell.fill = ORANGE_FILL
        elif row["decision"] == "Shortage":
            f_cell.fill = RED_FILL
        else:
            clear_cell_fill(f_cell)

        j_cell.value = row["j_value"]
        if row["j_value"] < 0:
            j_cell.fill = RED_FILL
        else:
            clear_cell_fill(j_cell)


def merge_row_to_main(
    main_path: str,
    groups: list[dict],
    decisions: dict[str, str],
    supplements: dict[str, float] | None = None,
    backup_dir: str | None = None,
) -> dict:
    backup_path = backup_main_file(main_path, backup_dir) if backup_dir else None

    workbook = openpyxl.load_workbook(main_path, keep_vba=(Path(main_path).suffix.lower() == ".xlsm"))
    try:
        plan = _build_preview_for_batches(
            workbook.active,
            [{
                "order_id": None,
                "model": groups[0].get("bom_model", "") if groups else "",
                "groups": groups,
                "supplements": supplements or {},
            }],
            _normalize_decisions(decisions),
        )

        for batch in plan["batches"]:
            for group_plan in batch.get("groups", []):
                _write_group_headers(workbook.active, group_plan)
                _write_group_rows(workbook.active, group_plan)

        ensure_main_header_wrap(workbook.active)
        workbook.save(main_path)
    finally:
        workbook.close()

    return {
        "backup_path": backup_path,
        "merged_parts": plan["merged_parts"],
        "new_col_count": plan["new_col_count"],
        "shortages": plan["shortages"],
    }


def _find_latest_stock_col(ws, row_idx: int, max_col: int) -> int | None:
    """找到該列最右邊有數值的欄位 index（從 max_col 往左掃）。"""
    for col_idx in range(max_col, STOCK_SEARCH_START_COL - 1, -1):
        value = _try_float(ws.cell(row=row_idx, column=col_idx).value)
        if value is not None:
            return col_idx
    return None


def supplement_part_in_main(
    main_path: str,
    part_number: str,
    supplement_qty: float,
    backup_dir: str | None = None,
) -> dict:
    """對主檔中指定料號補料：直接在缺料欄位加上補料量，並連帶更新右邊所有欄位。"""
    if supplement_qty <= 0:
        return {"ok": False, "message": "補料數量必須大於 0"}

    part_key = str(part_number or "").strip().upper()
    if not part_key:
        return {"ok": False, "message": "料號不能為空"}

    backup_path = backup_main_file(main_path, backup_dir) if backup_dir else None

    workbook = openpyxl.load_workbook(main_path, keep_vba=(Path(main_path).suffix.lower() == ".xlsm"))
    try:
        ws = workbook.active
        part_row_map = _build_part_row_map(ws)
        row_idx = part_row_map.get(part_key)
        if row_idx is None:
            workbook.close()
            return {"ok": False, "message": f"主檔中找不到料號 {part_key}"}

        max_col = ws.max_column
        stock_col = _find_latest_stock_col(ws, row_idx, max_col)
        if stock_col is None:
            workbook.close()
            return {"ok": False, "message": f"找不到 {part_key} 的庫存欄位"}

        current_stock = _try_float(ws.cell(row=row_idx, column=stock_col).value) or 0.0
        new_stock = current_stock + supplement_qty

        # 直接更新缺料欄位的庫存值
        ws.cell(row=row_idx, column=stock_col).value = new_stock

        # 右邊如果還有數值欄位（後續 dispatch 的結存），也一併加上補料量
        for col_idx in range(stock_col + 1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = _try_float(cell.value)
            if val is not None:
                cell.value = val + supplement_qty

        ensure_main_header_wrap(ws)
        workbook.save(main_path)
    finally:
        workbook.close()

    return {
        "ok": True,
        "part_number": part_key,
        "stock_before": current_stock,
        "supplement_qty": supplement_qty,
        "stock_after": new_stock,
        "backup_path": backup_path,
    }
