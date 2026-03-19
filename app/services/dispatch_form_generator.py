from __future__ import annotations

from copy import copy
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill
from ..runtime_paths import get_resource_base_dir

TEMPLATE_PATH = get_resource_base_dir() / "templates" / "dispatch_form_template.xlsx"
ITEM_STYLE_ROW = 3
HEADER_ROWS = (1, 2)
ORANGE_FILL = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")


def _roc_year(western_year: int) -> int:
    return western_year - 1911


def _load_template_sheet():
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Dispatch form template not found: {TEMPLATE_PATH}")
    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    return workbook, workbook.active


def _copy_sheet_settings(source_ws, target_ws):
    target_ws.title = source_ws.title
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    target_ws.sheet_format.defaultRowHeight = source_ws.sheet_format.defaultRowHeight
    target_ws.page_margins = copy(source_ws.page_margins)
    target_ws.page_setup = copy(source_ws.page_setup)
    target_ws.print_options = copy(source_ws.print_options)
    target_ws.sheet_properties = copy(source_ws.sheet_properties)

    for key, dimension in source_ws.column_dimensions.items():
        target_dimension = target_ws.column_dimensions[key]
        target_dimension.width = dimension.width
        target_dimension.hidden = dimension.hidden
        target_dimension.bestFit = dimension.bestFit
        target_dimension.outlineLevel = dimension.outlineLevel

    target_ws.column_dimensions["D"].width = 72


def _copy_cell_style(source_cell, target_cell):
    target_cell._style = copy(source_cell._style)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)


def _copy_row_template(source_ws, source_row: int, target_ws, target_row: int):
    for column in range(1, 6):
        _copy_cell_style(source_ws.cell(source_row, column), target_ws.cell(target_row, column))
    target_ws.row_dimensions[target_row].height = source_ws.row_dimensions[source_row].height


def _merge_section(target_ws, start_row: int):
    target_ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
    target_ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=3)
    target_ws.merge_cells(start_row=start_row, start_column=4, end_row=start_row + 1, end_column=4)


def _parse_display_date(date_str: str | None, now: datetime) -> tuple[int, int, int]:
    raw = str(date_str or "").strip()
    if raw:
        normalized = raw.replace("-", "/")
        parts = normalized.split("/")
        if len(parts) >= 3:
            try:
                return int(parts[0]), int(parts[1]), int(parts[2])
            except ValueError:
                pass
    return now.year, now.month, now.day


def _build_title(model: str, year: int, month: int) -> str:
    return f"辰尚-庚霖   {_roc_year(year)}年 {month}月份  {model}  之發料單\u3000\u3000\u3000\u3000"


def _coerce_po_number(po_number):
    text = str(po_number or "").strip()
    if text.isdigit():
        try:
            return int(text)
        except ValueError:
            return text
    return text


def _build_fill(fill_color: str | None, is_shortage: bool):
    if is_shortage:
        return copy(WHITE_FILL)

    color = str(fill_color or "").strip().lstrip("#").upper()
    if len(color) == 6:
        color = f"FF{color}"
    if len(color) == 8:
        return PatternFill(start_color=color, end_color=color, fill_type="solid")
    return copy(WHITE_FILL)


def _write_section_header(source_ws, target_ws, start_row: int, group: dict, now: datetime):
    for offset, source_row in enumerate(HEADER_ROWS):
        _copy_row_template(source_ws, source_row, target_ws, start_row + offset)

    _merge_section(target_ws, start_row)

    year, month, day = _parse_display_date(group.get("date"), now)
    target_ws.cell(start_row, 1).value = group.get("batch_code") or ""
    target_ws.cell(start_row + 1, 1).value = _coerce_po_number(group.get("po_number"))
    target_ws.cell(start_row, 4).value = _build_title(group.get("model", ""), year, month)
    target_ws.cell(start_row, 5).value = "日期"
    target_ws.cell(start_row + 1, 5).value = f"{year}/{month}/{day}"
    date_cell = target_ws.cell(start_row + 1, 5)
    date_font = copy(date_cell.font)
    date_font.sz = 9
    date_cell.font = date_font


def _write_item_row(source_ws, target_ws, row_idx: int, index: int, item: dict):
    _copy_row_template(source_ws, ITEM_STYLE_ROW, target_ws, row_idx)

    description = str(item.get("desc") or "")
    is_shortage = bool(item.get("is_shortage"))
    qty_value = "缺" if is_shortage else item.get("qty", "")

    target_ws.cell(row_idx, 1).value = index
    target_ws.cell(row_idx, 3).value = item.get("part", "")
    target_ws.cell(row_idx, 4).value = description
    target_ws.cell(row_idx, 5).value = qty_value
    target_ws.cell(row_idx, 5).fill = _build_fill(item.get("fill_color"), is_shortage)

    if len(description) > 90 or is_shortage:
        target_ws.row_dimensions[row_idx].height = 24.0


def generate_dispatch_form(groups: list[dict], output_path: str) -> str:
    template_wb, template_ws = _load_template_sheet()
    wb = openpyxl.Workbook()
    ws = wb.active
    _copy_sheet_settings(template_ws, ws)

    now = datetime.now()
    current_row = 1

    for group in groups:
        items = group.get("items", [])
        if not items:
            continue

        _write_section_header(template_ws, ws, current_row, group, now)
        current_row += 2

        for index, item in enumerate(items, start=1):
            _write_item_row(template_ws, ws, current_row, index, item)
            current_row += 1

    wb.save(output_path)
    wb.close()
    template_wb.close()
    return output_path
