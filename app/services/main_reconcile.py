from __future__ import annotations

from math import copysign, floor
from pathlib import Path

import openpyxl

from .main_reader import read_stock

_TOLERANCE = 1e-6


def _try_float(value) -> float | None:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _round_away(value: float) -> float:
    if abs(value) <= _TOLERANCE:
        return 0.0
    return copysign(floor(abs(value) + 0.5), value)


def _same_number(expected, actual) -> bool:
    expected_num = _try_float(expected)
    actual_num = _try_float(actual)
    if expected_num is None or actual_num is None:
        return expected == actual
    return abs(expected_num - actual_num) <= _TOLERANCE


def _delta(expected, actual) -> float | None:
    expected_num = _try_float(expected)
    actual_num = _try_float(actual)
    if expected_num is None or actual_num is None:
        return None
    return actual_num - expected_num


def _mismatch(part_number: str, kind: str, column: str, expected, actual) -> dict:
    return {
        "part_number": part_number,
        "kind": kind,
        "column": column,
        "expected": expected,
        "actual": actual,
        "delta": _delta(expected, actual),
    }


def verify_main_write(main_path: str, plan_rows: list[dict]) -> dict:
    mismatches: list[dict] = []
    expected_stock_by_part: dict[str, float] = {}

    suffix = Path(main_path).suffix.lower()
    wb = openpyxl.load_workbook(
        main_path,
        keep_vba=(suffix == ".xlsm"),
        data_only=True,
    )
    try:
        ws = wb.worksheets[0]
        for row in plan_rows or []:
            part = str(row.get("part_number") or "").strip().upper()
            if not part:
                continue

            row_idx = int(row.get("row_idx") or 0)
            col_h = int(row.get("col_h") or 0)
            col_f = int(row.get("col_f") or 0)
            col_j = int(row.get("col_j") or 0)
            if row_idx <= 0 or col_h <= 0 or col_f <= 0 or col_j <= 0:
                continue

            expected_h = _try_float(row.get("effective_h")) or 0.0
            expected_f = _try_float(row.get("f_value")) or 0.0
            expected_j = _try_float(row.get("j_value")) or 0.0
            current_stock = _try_float(row.get("current_stock")) or 0.0

            actual_h = _try_float(ws.cell(row=row_idx, column=col_h).value)
            actual_f = _try_float(ws.cell(row=row_idx, column=col_f).value)
            actual_j = _try_float(ws.cell(row=row_idx, column=col_j).value)

            for column, expected, actual in (
                ("H", expected_h, actual_h),
                ("F", expected_f, actual_f),
                ("J", expected_j, actual_j),
            ):
                if not _same_number(expected, actual):
                    mismatches.append(_mismatch(part, "cell", column, expected, actual))

            actual_h_num = actual_h or 0.0
            actual_f_num = actual_f or 0.0
            actual_j_num = actual_j or 0.0
            conserved_j = _round_away(current_stock + actual_h_num - actual_f_num)
            if not _same_number(conserved_j, actual_j_num):
                mismatches.append(_mismatch(part, "conservation", "J", conserved_j, actual_j_num))

            expected_stock_by_part[part] = expected_j
    finally:
        wb.close()

    stock = read_stock(main_path)
    for part, expected_stock in expected_stock_by_part.items():
        actual_stock = stock.get(part, 0.0)
        if not _same_number(expected_stock, actual_stock):
            mismatches.append(_mismatch(part, "cross_read", "stock", expected_stock, actual_stock))

    return {
        "ok": not mismatches,
        "checked_parts": len(expected_stock_by_part),
        "mismatches": mismatches,
    }
