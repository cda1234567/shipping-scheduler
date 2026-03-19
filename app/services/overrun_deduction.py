"""
加工多打扣帳邏輯。

情境：
- 單一機種在加工廠多打 X pcs
- 需要依對應 BOM，把每個料號額外消耗的數量扣回主檔
"""
from __future__ import annotations

from difflib import SequenceMatcher
from pathlib import Path

import openpyxl

from .. import database as db
from .merge_to_main import _build_part_row_map, _read_latest_stock, _round_away
from .main_reader import read_stock
from .xls_reader import open_workbook_any


def _try_float(value) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def normalize_model_key(value: str) -> str:
    return str(value or "").strip().upper()


def parse_overrun_detail_excel(path: str) -> dict:
    """
    解析加工廠提供的多打扣帳明細。

    範例格式：
    - Row 1: A=料號, B=標題/機種資訊
    - Row 2: A=M/O, B=單號/備註
    - Row 3+: A=料號, B=扣帳數量
    """
    workbook = open_workbook_any(path, read_only=True, data_only=True)
    ws = workbook.worksheets[0]
    try:
        title = str(ws.cell(row=1, column=2).value or "").strip()
        mo_info = str(ws.cell(row=2, column=2).value or "").strip()

        items: list[dict] = []
        for row_idx in range(3, ws.max_row + 1):
            part_number = str(ws.cell(row=row_idx, column=1).value or "").strip().upper()
            qty = _try_float(ws.cell(row=row_idx, column=2).value)
            if not part_number:
                continue
            if qty is None or qty <= 0:
                continue
            items.append({
                "source_row": row_idx,
                "part_number": part_number,
                "description": "",
                "defective_qty": float(qty),
            })

        if not items:
            raise ValueError("檔案中沒有有效的料號 / 數量資料")

        return {
            "title": title,
            "mo_info": mo_info,
            "items": items,
        }
    finally:
        workbook.close()


def build_model_overrun_plan(model: str, extra_pcs: float) -> dict:
    """依機種與多打 pcs 展開成要扣帳的料號清單。"""
    model_key = normalize_model_key(model)
    if not model_key:
        raise ValueError("請輸入機種")

    extra_qty = _try_float(extra_pcs)
    if extra_qty is None or extra_qty <= 0:
        raise ValueError("多打 pcs 必須大於 0")

    bom_files = db.get_bom_files_by_models([model_key])
    if not bom_files:
        raise LookupError(f"機種 {model_key} 找不到對應 BOM")

    aggregate: dict[str, dict] = {}
    matched_boms: list[dict] = []
    matched_models: set[str] = set()

    for bom in bom_files:
        bom_id = str(bom.get("id") or "")
        matched_boms.append({
            "id": bom_id,
            "filename": str(bom.get("filename") or ""),
            "model": str(bom.get("model") or ""),
            "group_model": str(bom.get("group_model") or ""),
        })

        raw_model = str(bom.get("group_model") or bom.get("model") or "").strip()
        if raw_model:
            matched_models.add(raw_model)

        for component in db.get_bom_components(bom_id):
            part_number = str(component.get("part_number") or "").strip().upper()
            if not part_number:
                continue
            if bool(component.get("is_dash")) or bool(component.get("is_customer_supplied")):
                continue

            qty_per_board = _try_float(component.get("qty_per_board")) or 0.0
            if qty_per_board <= 0:
                continue

            entry = aggregate.setdefault(part_number, {
                "part_number": part_number,
                "description": str(component.get("description") or ""),
                "qty_per_board_total": 0.0,
                "raw_deduction_qty": 0.0,
                "source_boms": set(),
            })

            if not entry["description"]:
                entry["description"] = str(component.get("description") or "")
            entry["qty_per_board_total"] += qty_per_board
            entry["raw_deduction_qty"] += qty_per_board * extra_qty
            if bom_id:
                entry["source_boms"].add(bom_id)

    items: list[dict] = []
    for row in aggregate.values():
        deduction_qty = _round_away(float(row.get("raw_deduction_qty") or 0))
        if deduction_qty <= 0:
            continue
        items.append({
            "part_number": row["part_number"],
            "description": row.get("description", ""),
            "defective_qty": deduction_qty,
            "qty_per_board_total": float(row.get("qty_per_board_total") or 0),
            "source_bom_ids": sorted(str(v) for v in row.get("source_boms", set())),
        })

    items.sort(key=lambda item: item["part_number"])

    return {
        "requested_model": str(model or "").strip(),
        "model": model_key,
        "extra_pcs": float(extra_qty),
        "matched_models": sorted(matched_models),
        "matched_boms": matched_boms,
        "items": items,
    }


def suggest_main_part_numbers(
    missing_part: str,
    main_stock_map: dict[str, float],
    limit: int = 5,
) -> list[dict]:
    key = str(missing_part or "").strip().upper()
    if not key:
        return []

    scored: list[tuple[float, str]] = []
    for part_number in (main_stock_map or {}).keys():
        part = str(part_number or "").strip().upper()
        if not part:
            continue
        if part == key:
            score = 999.0
        elif key in part or part in key:
            score = 200.0 + min(len(key), len(part))
        else:
            ratio = SequenceMatcher(None, key, part).ratio()
            if ratio < 0.45:
                continue
            score = ratio * 100.0
        scored.append((score, part))

    scored.sort(key=lambda item: (-item[0], item[1]))
    unique_parts: list[str] = []
    for _, part in scored:
        if part not in unique_parts:
            unique_parts.append(part)
        if len(unique_parts) >= limit:
            break

    return [
        {
            "part_number": part,
            "stock_qty": float((main_stock_map or {}).get(part, 0) or 0),
        }
        for part in unique_parts
    ]


def build_overrun_import_preview(main_path: str, parsed: dict) -> dict:
    main_stock_map = read_stock(main_path)
    preview = preview_deductions_against_main(main_path, parsed.get("items") or [])
    result_map = {
        str(item.get("part_number") or "").strip().upper(): item
        for item in (preview.get("results") or [])
        if str(item.get("part_number") or "").strip()
    }

    items: list[dict] = []
    missing_items: list[dict] = []
    for raw_item in parsed.get("items") or []:
        part_number = str(raw_item.get("part_number") or "").strip().upper()
        source_row = int(raw_item.get("source_row") or 0)
        item = {
            "source_row": source_row,
            "part_number": part_number,
            "description": str(raw_item.get("description") or ""),
            "defective_qty": float(raw_item.get("defective_qty") or 0),
            "found_in_main": part_number in main_stock_map,
            "suggestions": [],
        }
        matched = result_map.get(part_number)
        if matched:
            item["stock_before"] = float(matched.get("stock_before") or 0)
            item["stock_after"] = float(matched.get("stock_after") or 0)
            item["will_go_negative"] = bool(matched.get("will_go_negative"))
        else:
            item["suggestions"] = suggest_main_part_numbers(part_number, main_stock_map)
            missing_items.append(dict(item))
        items.append(item)

    return {
        "title": str(parsed.get("title") or ""),
        "mo_info": str(parsed.get("mo_info") or ""),
        "source_filename": str(parsed.get("source_filename") or ""),
        "item_count": len(items),
        "deducted_count": int(preview.get("deducted_count") or 0),
        "negative_count": int(preview.get("negative_count") or 0),
        "total_deduction_qty": float(preview.get("total_deduction_qty") or 0),
        "items": items,
        "results": preview.get("results") or [],
        "missing_items": missing_items,
        "missing_count": len(missing_items),
        "requires_confirmation": bool(missing_items),
    }


def apply_overrun_import_confirmations(main_path: str, items: list[dict]) -> dict:
    main_stock_map = read_stock(main_path)
    final_items: list[dict] = []
    skipped_items: list[dict] = []
    replaced_items: list[dict] = []
    unresolved_items: list[dict] = []

    for raw_item in items or []:
        source_row = int(raw_item.get("source_row") or 0)
        original_part = str(raw_item.get("part_number") or "").strip().upper()
        target_part = str(raw_item.get("target_part_number") or "").strip().upper()
        action = str(raw_item.get("action") or "").strip().lower() or "deduct"
        qty = float(raw_item.get("defective_qty") or 0)
        description = str(raw_item.get("description") or "")

        if not original_part or qty <= 0:
            continue

        if action == "skip":
            skipped_items.append({
                "source_row": source_row,
                "part_number": original_part,
                "defective_qty": qty,
            })
            continue

        if action == "replace":
            if not target_part:
                unresolved_items.append({
                    "source_row": source_row,
                    "part_number": original_part,
                    "reason": "請輸入要改成的正確料號",
                })
                continue
            if target_part not in main_stock_map:
                unresolved_items.append({
                    "source_row": source_row,
                    "part_number": original_part,
                    "reason": f"主檔找不到替代料號 {target_part}",
                })
                continue
            final_items.append({
                "source_row": source_row,
                "part_number": target_part,
                "description": description,
                "defective_qty": qty,
                "source_part_number": original_part,
            })
            replaced_items.append({
                "source_row": source_row,
                "source_part_number": original_part,
                "target_part_number": target_part,
                "defective_qty": qty,
            })
            continue

        if original_part not in main_stock_map:
            unresolved_items.append({
                "source_row": source_row,
                "part_number": original_part,
                "reason": "主檔找不到此料號，請改正或選不扣",
            })
            continue

        final_items.append({
            "source_row": source_row,
            "part_number": original_part,
            "description": description,
            "defective_qty": qty,
            "source_part_number": original_part,
        })

    return {
        "final_items": final_items,
        "skipped_items": skipped_items,
        "replaced_items": replaced_items,
        "unresolved_items": unresolved_items,
    }


def preview_deductions_against_main(main_path: str, items: list[dict]) -> dict:
    """用目前主檔庫存預覽扣帳結果，不實際寫檔。"""
    if not main_path or not Path(main_path).exists():
        raise FileNotFoundError("主檔尚未上傳，無法預覽")

    is_xlsm = Path(main_path).suffix.lower() == ".xlsm"
    workbook = openpyxl.load_workbook(main_path, data_only=True, keep_vba=is_xlsm)
    try:
        ws = workbook.active
        part_row_map = _build_part_row_map(ws)
        max_col = ws.max_column

        results: list[dict] = []
        skipped_parts: list[str] = []

        for item in items or []:
            part_number = str(item.get("part_number") or "").strip().upper()
            if not part_number:
                continue

            row_idx = part_row_map.get(part_number)
            if row_idx is None:
                skipped_parts.append(part_number)
                continue

            stock_before = _read_latest_stock(ws, row_idx, max_col)
            deduction_qty = float(item.get("defective_qty") or 0)
            stock_after = _round_away(stock_before - deduction_qty)

            results.append({
                "part_number": part_number,
                "description": str(item.get("description") or ""),
                "defective_qty": deduction_qty,
                "stock_before": stock_before,
                "stock_after": stock_after,
                "will_go_negative": stock_after < 0,
            })

        return {
            "deducted_count": len(results),
            "skipped_parts": skipped_parts,
            "results": results,
            "negative_count": sum(1 for item in results if item["will_go_negative"]),
            "total_deduction_qty": sum(float(item.get("defective_qty") or 0) for item in results),
        }
    finally:
        workbook.close()
