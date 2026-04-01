from __future__ import annotations

import shutil
import subprocess
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from ..config import BACKUP_DIR, BOM_DIR, cfg
from ..models import BomEditorSaveRequest, BomFile
from .bom_parser import parse_bom
from .xls_reader import open_workbook_any

_BOM_BACKUP_DIR = BACKUP_DIR / "bom"
_BOM_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
_DASH_MARKERS = {"-", "x", "X", "n", "N", "n/a", "N/A", "na", "NA", "?"}


def _ps_quote(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def _copy_xls_to_xlsx_values(src_path: Path, dest_path: Path):
    wb = open_workbook_any(str(src_path))
    wb.save(dest_path)
    wb.close()


def _find_libreoffice_executable() -> str | None:
    return shutil.which("soffice") or shutil.which("libreoffice")


def _convert_with_libreoffice(src_path: Path, dest_path: Path):
    soffice = _find_libreoffice_executable()
    if not soffice:
        raise FileNotFoundError("LibreOffice not found")

    expected_path = src_path.with_suffix(".xlsx")
    expected_path.unlink(missing_ok=True)
    if dest_path != expected_path:
        dest_path.unlink(missing_ok=True)

    subprocess.run(
        [
            soffice,
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(src_path.parent),
            str(src_path),
        ],
        check=True,
        capture_output=True,
        text=True,
    )

    if not expected_path.exists():
        raise FileNotFoundError(f"LibreOffice did not create converted file: {expected_path}")

    if expected_path != dest_path:
        expected_path.replace(dest_path)


def convert_xls_to_xlsx(src_path: str, dest_path: str):
    """
    將 .xls 轉成 .xlsx。

    優先使用 Excel COM 轉檔以保留格式；若環境無法使用，再退回值導向的相容轉檔。
    """
    src = Path(src_path)
    dest = Path(dest_path)
    dest.parent.mkdir(parents=True, exist_ok=True)

    script = f"""
$ErrorActionPreference = 'Stop'
$excel = $null
$workbook = $null
try {{
  $excel = New-Object -ComObject Excel.Application
  $excel.Visible = $false
  $excel.DisplayAlerts = $false
  $workbook = $excel.Workbooks.Open({_ps_quote(str(src.resolve()))})
  $workbook.SaveAs({_ps_quote(str(dest.resolve()))}, 51)
}} finally {{
  if ($workbook -ne $null) {{
    $workbook.Close($false)
    [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($workbook)
  }}
  if ($excel -ne $null) {{
    $excel.Quit()
    [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel)
  }}
  [GC]::Collect()
  [GC]::WaitForPendingFinalizers()
}}
"""

    try:
        subprocess.run(
            ["powershell", "-NoProfile", "-NonInteractive", "-Command", script],
            check=True,
            capture_output=True,
            text=True,
        )
        return
    except Exception:
        pass

    try:
        _convert_with_libreoffice(src, dest)
        return
    except Exception:
        _copy_xls_to_xlsx_values(src, dest)


def build_editable_filename(filename: str) -> str:
    name = filename or "BOM.xlsx"
    path = Path(name)
    if path.suffix.lower() == ".xls":
        return f"{path.stem}.xlsx"
    return path.name


def prepare_uploaded_bom_file(bom_id: str, upload_name: str, content: bytes) -> dict[str, object]:
    ext = Path(upload_name or "").suffix.lower()
    if ext == ".xls":
        source_path = BOM_DIR / f"{bom_id}.xls"
        target_path = BOM_DIR / f"{bom_id}.xlsx"
        source_path.write_bytes(content)
        try:
            convert_xls_to_xlsx(str(source_path), str(target_path))
        finally:
            source_path.unlink(missing_ok=True)
        return {
            "filepath": str(target_path),
            "filename": build_editable_filename(upload_name),
            "source_filename": upload_name or target_path.name,
            "source_format": ext,
            "is_converted": True,
        }

    target_path = BOM_DIR / f"{bom_id}{ext}"
    target_path.write_bytes(content)
    return {
        "filepath": str(target_path),
        "filename": (upload_name or target_path.name),
        "source_filename": upload_name or target_path.name,
        "source_format": ext,
        "is_converted": False,
    }


def build_bom_storage_payload(bom: BomFile) -> dict:
    return {
        "id": bom.id,
        "filename": bom.filename,
        "filepath": bom.path,
        "source_filename": bom.source_filename,
        "source_format": bom.source_format,
        "is_converted": bom.is_converted,
        "po_number": bom.po_number,
        "model": bom.model,
        "pcb": bom.pcb,
        "group_model": bom.group_model,
        "order_qty": bom.order_qty,
        "uploaded_at": bom.uploaded_at,
        "components": [c.dict() for c in bom.components],
    }


def parse_bom_for_storage(
    *,
    path: str,
    bom_id: str,
    filename: str,
    uploaded_at: str,
    group_model: str = "",
    source_filename: str = "",
    source_format: str = "",
    is_converted: bool = False,
) -> BomFile:
    parsed = parse_bom(
        path=path,
        bom_id=bom_id,
        filename=filename,
        uploaded_at=uploaded_at,
    )
    parsed.group_model = group_model or parsed.group_model
    parsed.source_filename = source_filename or filename
    parsed.source_format = source_format or Path(source_filename or filename).suffix.lower()
    parsed.is_converted = bool(is_converted)
    return parsed


def normalize_bom_record_to_editable(bom: dict) -> dict:
    """
    舊版 .xls BOM 會在第一次需要編輯時自動轉成 .xlsx。
    回傳的是已可編輯、可保存回檔案的正式 BOM metadata。
    """
    filepath = Path(str(bom.get("filepath", "")))
    if filepath.suffix.lower() != ".xls":
        return dict(bom)

    target_path = BOM_DIR / f"{bom['id']}.xlsx"
    convert_xls_to_xlsx(str(filepath), str(target_path))
    filepath.unlink(missing_ok=True)

    normalized = dict(bom)
    normalized["filepath"] = str(target_path)
    normalized["source_filename"] = bom.get("source_filename") or bom.get("filename") or filepath.name
    normalized["source_format"] = bom.get("source_format") or ".xls"
    normalized["filename"] = build_editable_filename(bom.get("filename") or filepath.name)
    normalized["is_converted"] = True
    return normalized


def backup_bom_file(path: str) -> str:
    src = Path(path)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = _BOM_BACKUP_DIR / f"{src.stem}_backup_{stamp}{src.suffix}"
    shutil.copy2(src, dest)
    return str(dest)


def _resolve_cell_for_write(ws: Worksheet, row_idx: int, col_idx: int):
    cell = ws.cell(row=row_idx, column=col_idx)
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell


def _get_cell_value(ws: Worksheet, row_idx: int, col_idx: int):
    return _resolve_cell_for_write(ws, row_idx, col_idx).value


def _set_cell_value(ws: Worksheet, row_idx: int, col_idx: int, value):
    _resolve_cell_for_write(ws, row_idx, col_idx).value = value


def _write_component_row(ws, row_idx: int, component):
    part_col = cfg("excel.bom_part_col", 2) + 1
    desc_col = cfg("excel.bom_desc_col", 3) + 1
    qty_col = cfg("excel.bom_qty_per_board", 1) + 1
    needed_col = cfg("excel.bom_needed_col", 5) + 1
    g_col = cfg("excel.bom_g_col", 6) + 1
    h_col = cfg("excel.bom_h_col", 7) + 1

    _set_cell_value(ws, row_idx, part_col, component.part_number.strip())
    _set_cell_value(ws, row_idx, desc_col, component.description)
    _set_cell_value(ws, row_idx, qty_col, component.qty_per_board)
    _set_cell_value(ws, row_idx, needed_col, component.needed_qty)

    if component.is_dash:
        _set_cell_value(ws, row_idx, g_col, "-")
        _set_cell_value(ws, row_idx, h_col, "-")
    else:
        if str(_get_cell_value(ws, row_idx, g_col) or "").strip() in _DASH_MARKERS:
            _set_cell_value(ws, row_idx, g_col, None)
        _set_cell_value(ws, row_idx, h_col, component.prev_qty_cs)


def apply_bom_editor_changes(path: str, req: BomEditorSaveRequest):
    workbook_path = Path(path)
    is_macro = workbook_path.suffix.lower() == ".xlsm"
    wb = openpyxl.load_workbook(str(workbook_path), keep_vba=is_macro)
    ws = wb.worksheets[0]

    po_col = cfg("excel.bom_po_col", 7) + 1
    order_qty_col = cfg("excel.bom_order_qty_col", 10) + 1
    model_col = cfg("excel.bom_model_col", 2) + 1
    pcb_col = cfg("excel.bom_pcb_col", 3) + 1
    data_start = cfg("excel.bom_data_start_row", 5)

    _set_cell_value(ws, 1, po_col, req.po_number)
    _set_cell_value(ws, 1, order_qty_col, req.order_qty)
    _set_cell_value(ws, 2, model_col, req.model)
    _set_cell_value(ws, 2, pcb_col, req.pcb)

    seen_rows: set[int] = set()
    for component in req.components:
        row_idx = int(component.source_row)
        if row_idx < data_start or row_idx > ws.max_row:
            wb.close()
            raise ValueError(f"找不到可更新的 BOM 列：{row_idx}")
        if row_idx in seen_rows:
            wb.close()
            raise ValueError(f"BOM 列號重複：{row_idx}")
        if not component.part_number.strip():
            wb.close()
            raise ValueError(f"BOM 第 {row_idx} 列料號不可為空白")
        seen_rows.add(row_idx)
        _write_component_row(ws, row_idx, component)

    wb.save(str(workbook_path))
    wb.close()
