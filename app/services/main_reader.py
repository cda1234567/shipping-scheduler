from __future__ import annotations

from pathlib import Path

import openpyxl

from ..config import cfg
from .xls_reader import open_workbook_any

_PART_COL = None
_VENDOR_COL = None
_MOQ_COL = None
_STOCK_SEARCH_START_COL = None


def _part_col():
    global _PART_COL
    if _PART_COL is None:
        _PART_COL = cfg("excel.main_part_col", 0)
    return _PART_COL


def _vendor_col():
    global _VENDOR_COL
    if _VENDOR_COL is None:
        _VENDOR_COL = cfg("excel.main_vendor_col", 1)
    return _VENDOR_COL


def _moq_col():
    global _MOQ_COL
    if _MOQ_COL is None:
        _MOQ_COL = cfg("excel.main_moq_col", 2)
    return _MOQ_COL


def _stock_search_start_col():
    global _STOCK_SEARCH_START_COL
    if _STOCK_SEARCH_START_COL is None:
        # 庫存只能從 MOQ 右側開始找，避免把 C 欄 MOQ 誤讀成目前庫存。
        _STOCK_SEARCH_START_COL = _moq_col() + 1
    return _STOCK_SEARCH_START_COL


def _try_float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def find_current_stock_cell_from_row_values(row_vals) -> float | None:
    """找出 MOQ 右側最後一個可用庫存值；若沒有任何數值則回傳 None。"""
    start_col = _stock_search_start_col()
    if not row_vals or len(row_vals) <= start_col:
        return None

    for value in reversed(row_vals[start_col:]):
        stock = _try_float(value)
        if stock is not None:
            return stock
    return None


def find_current_stock_from_row_values(row_vals) -> float:
    """找出 MOQ 右側最後一個可用庫存值；若沒有任何數值則視為 0。"""
    stock = find_current_stock_cell_from_row_values(row_vals)
    return stock if stock is not None else 0.0


def read_stock(path: str) -> dict[str, float]:
    """讀取主檔目前庫存，永遠忽略 MOQ 左右非庫存欄位。"""
    wb = open_workbook_any(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    result: dict[str, float] = {}
    pc = _part_col()

    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        if not row_vals:
            continue
        part = str(row_vals[pc] or "").strip()
        if not part:
            continue
        result[part.upper()] = find_current_stock_from_row_values(row_vals)

    wb.close()
    return result


def read_vendors(path: str) -> dict[str, str]:
    """讀取主檔 B 欄廠商，以料號為 key。"""
    wb = open_workbook_any(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    result: dict[str, str] = {}
    pc = _part_col()
    vc = _vendor_col()

    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        if not row_vals or len(row_vals) <= pc:
            continue
        part = str(row_vals[pc] or "").strip()
        if not part:
            continue
        vendor = str(row_vals[vc] if len(row_vals) > vc and row_vals[vc] is not None else "").strip()
        result[part.upper()] = vendor

    wb.close()
    return result


def update_vendor(path: str, part_number: str, vendor: str) -> dict:
    """更新主檔 B 欄廠商。"""
    part_key = str(part_number or "").strip().upper()
    if not part_key:
        raise ValueError("料號不可空白")

    suffix = Path(path).suffix.lower()
    if suffix == ".xls":
        raise ValueError("xls 主檔不支援直接修改廠商，請先轉成 xlsx 或 xlsm")

    wb = openpyxl.load_workbook(path, keep_vba=(suffix == ".xlsm"))
    try:
        ws = wb.worksheets[0]
        pc = _part_col() + 1
        vc = _vendor_col() + 1
        for row_idx in range(2, ws.max_row + 1):
            cell_part = str(ws.cell(row=row_idx, column=pc).value or "").strip().upper()
            if cell_part != part_key:
                continue
            cell = ws.cell(row=row_idx, column=vc)
            old_vendor = str(cell.value or "").strip()
            new_vendor = str(vendor or "").strip()
            cell.value = new_vendor
            wb.save(path)
            return {
                "part_number": part_key,
                "old_vendor": old_vendor,
                "vendor": new_vendor,
                "row": row_idx,
            }
    finally:
        wb.close()

    raise KeyError(f"主檔找不到料號 {part_key}")


def read_moq(path: str) -> dict[str, float]:
    """讀取主檔 MOQ。"""
    wb = open_workbook_any(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    result: dict[str, float] = {}
    pc = _part_col()
    mc = _moq_col()

    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        if not row_vals or len(row_vals) <= mc:
            continue
        part = str(row_vals[pc] or "").strip()
        if not part:
            continue
        moq = _try_float(row_vals[mc]) or 0.0
        result[part.upper()] = moq

    wb.close()
    return result


def find_legacy_snapshot_stock_fixes(path: str, snapshot: dict[str, dict]) -> dict[str, float]:
    """
    找出舊版快照把 MOQ 誤存成庫存的料號。

    只修正可以明確判斷的情況：
    1. snapshot.stock_qty == snapshot.moq 且 moq != 0
    2. 主檔在 MOQ 右側完全沒有任何庫存數字

    這種資料在舊邏輯下會被誤讀成「庫存 = MOQ」，正確值應為 0。
    """
    if not snapshot:
        return {}

    suspicious_parts = {
        str(part).strip().upper()
        for part, values in snapshot.items()
        if (
            str(part).strip()
            and float((values or {}).get("moq") or 0) != 0
            and float((values or {}).get("stock_qty") or 0) == float((values or {}).get("moq") or 0)
        )
    }
    if not suspicious_parts:
        return {}

    wb = open_workbook_any(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    fixes: dict[str, float] = {}
    pc = _part_col()

    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        if not row_vals:
            continue
        part = str(row_vals[pc] or "").strip().upper()
        if not part or part not in suspicious_parts:
            continue
        if find_current_stock_cell_from_row_values(row_vals) is None:
            fixes[part] = 0.0

    wb.close()
    return fixes
