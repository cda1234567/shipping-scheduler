from __future__ import annotations
from pathlib import Path
import openpyxl
import xlrd
import xlrd.xldate


def is_xls(path: str) -> bool:
    return Path(path).suffix.lower() == ".xls"


def open_workbook_any(path: str, read_only: bool = False, data_only: bool = True):
    """開啟 .xls 或 .xlsx，統一回傳 openpyxl workbook。"""
    if is_xls(path):
        return _xls_to_openpyxl(path)
    return openpyxl.load_workbook(path, read_only=read_only, data_only=data_only)


def _xls_to_openpyxl(path: str):
    xls_wb = xlrd.open_workbook(path)
    new_wb = openpyxl.Workbook()
    new_wb.remove(new_wb.active)

    for si in range(xls_wb.nsheets):
        xls_ws = xls_wb.sheets()[si]
        new_ws = new_wb.create_sheet(xls_ws.name)
        for row in range(xls_ws.nrows):
            for col in range(xls_ws.ncols):
                cell = xls_ws.cell(row, col)
                val = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    val = xlrd.xldate.xldate_as_datetime(val, xls_wb.datemode)
                new_ws.cell(row + 1, col + 1, val)

    return new_wb
