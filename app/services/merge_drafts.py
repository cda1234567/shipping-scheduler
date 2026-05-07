from __future__ import annotations

import io
import logging
import re
import shutil
import tempfile
import time
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import quote

log = logging.getLogger(__name__)

import openpyxl
from openpyxl.styles import PatternFill
from fastapi import HTTPException, Request
from fastapi.responses import FileResponse, StreamingResponse

from .. import database as db
from ..config import MERGE_DRAFT_DIR, cfg
from .server_downloads import maybe_server_save_bytes_response, maybe_server_save_response
from .download_names import build_bom_dispatch_filename, build_generated_filename
from .bom_editor import (
    build_bom_storage_payload,
    normalize_bom_record_to_editable,
    parse_bom_for_storage,
)
from .bom_quantity import (
    calculate_effective_needed_qty,
    coerce_qty,
    format_excel_qty,
    get_component_effective_needed_qty,
)
from .main_reader import read_moq, read_stock
from ..models import calc_suggested_qty
from .shortage_rules import (
    calculate_current_order_shortage_amount,
    calculate_shortage_amount,
    is_order_scoped_shortage_part,
    summarize_requested_supply,
    summarize_st_supply,
)
from .workbook_recalc import save_workbook_with_recalc
from .workbook_recalc import cell_has_formula

COMMITTED_DRAFT_RETENTION_DAYS = 365
ORANGE_FILL = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")


def normalize_part_key(value) -> str:
    return str(value or "").strip().upper()


def _sanitize_filename_piece(value: str, fallback: str = "draft") -> str:
    text = str(value or "").strip()
    if not text:
        return fallback
    text = re.sub(r'[\\/:*?"<>|]+', "-", text)
    text = re.sub(r"\s+", " ", text).strip(" .")
    return text or fallback


def _first_filename_model(*values) -> str:
    for value in values:
        for part in str(value or "").split(","):
            model = part.strip()
            if model:
                return model
    return ""


def _safe_optional_filename_piece(value) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    return _sanitize_filename_piece(text, "")


def _build_draft_display_filename(plan: dict, source_path: Path, now=None) -> str:
    """Build the user-visible draft BOM filename from the target order, not the source BOM name."""
    po_number = _safe_optional_filename_piece(plan.get("po_number"))
    model = _safe_optional_filename_piece(_first_filename_model(
        plan.get("model"),
        plan.get("group_model"),
        source_path.stem,
    ))
    suffix = source_path.suffix or ".xlsx"
    return build_bom_dispatch_filename(po_number, model, suffix, now=now)


def _normalize_decisions(decisions: dict[str, str] | None = None) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for part_number, decision in (decisions or {}).items():
        key = normalize_part_key(part_number)
        if key and decision:
            normalized[key] = str(decision)
    return normalized


def _normalize_supplements(supplements: dict[str, float] | None = None) -> dict[str, float]:
    normalized: dict[str, float] = {}
    for part_number, qty in (supplements or {}).items():
        key = normalize_part_key(part_number)
        try:
            amount = float(qty or 0)
        except (TypeError, ValueError):
            amount = 0.0
        if key and amount > 0:
            normalized[key] = amount
    return normalized


def _get_persisted_decision_map(order_ids: list[int]) -> dict[int, dict[str, str]]:
    if not order_ids:
        return {}
    return {
        int(order_id): _normalize_decisions(decisions)
        for order_id, decisions in db.get_order_decisions(order_ids).items()
    }


def _get_persisted_supplement_map(order_ids: list[int]) -> dict[int, dict[str, float]]:
    if not order_ids:
        return {}
    return {
        int(order_id): _normalize_supplements(supplements)
        for order_id, supplements in db.get_order_supplements(order_ids).items()
    }


def _get_main_signature(main_path: str) -> tuple[str, str]:
    path = Path(main_path)
    if not path.exists():
        raise HTTPException(400, "找不到主檔，無法建立副檔")
    return str(path), str(path.stat().st_mtime_ns)


def _load_effective_moq(main_path: str) -> dict[str, float]:
    live = {
        normalize_part_key(part): float(qty or 0)
        for part, qty in read_moq(main_path).items()
        if normalize_part_key(part)
    }
    snapshot = db.get_snapshot()
    for part, row in snapshot.items():
        key = normalize_part_key(part)
        if not key:
            continue
        live[key] = float((row or {}).get("moq") or 0)
    return live


def _build_running_stock(main_path: str) -> dict[str, float]:
    return {
        normalize_part_key(part): float(qty or 0)
        for part, qty in read_stock(main_path).items()
        if normalize_part_key(part)
    }


def _sanitize_resolved_shortage_decisions(
    part_totals: dict[str, dict[str, float]],
    running_stock: dict[str, float],
    decisions: dict[str, str] | None = None,
    supplements: dict[str, float] | None = None,
) -> dict[str, str]:
    sanitized = _normalize_decisions(decisions)
    normalized_supplements = _normalize_supplements(supplements)

    for part, summary in (part_totals or {}).items():
        if sanitized.get(part) != "Shortage":
            continue

        current_stock = float(running_stock.get(part, 0) or 0)
        prev_qty_cs = float(summary.get("prev_qty_cs") or 0)
        needed_qty = float(summary.get("needed_qty") or 0)
        available_before = current_stock + prev_qty_cs
        shortage_before = calculate_current_order_shortage_amount(part, available_before, needed_qty)
        planned_supplement = float(normalized_supplements.get(part, 0) or 0)
        shortage_with_planned_supply = calculate_current_order_shortage_amount(
            part,
            available_before + planned_supplement,
            needed_qty,
        )

        if shortage_before <= 0 or (planned_supplement > 0 and shortage_with_planned_supply <= 0):
            sanitized.pop(part, None)

    return sanitized


def _resolve_cell_for_write(ws, row_idx: int, col_idx: int):
    cell = ws.cell(row=row_idx, column=col_idx)
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell


def _set_cell_value(ws, row_idx: int, col_idx: int, value):
    _resolve_cell_for_write(ws, row_idx, col_idx).value = value


def _format_bom_po_value(existing_value, po_number):
    po_text = str(po_number or "").strip()
    if not po_text:
        return existing_value

    existing_text = str(existing_value or "").strip()
    if any(token in existing_text for token in ("製單號碼", "M/O", "PO")):
        if ":" in existing_text:
            prefix = existing_text.split(":", 1)[0]
            return f"{prefix}:{po_text}"
        return f"{existing_text}{po_text}"
    return po_text


def _read_ws_bom_order_qty(ws) -> float:
    order_qty_col = cfg("excel.bom_order_qty_col", 10) + 1
    order_qty_cell = _resolve_cell_for_write(ws, 1, order_qty_col)
    order_qty = coerce_qty(order_qty_cell.value)
    if order_qty > 0:
        return order_qty
    return coerce_qty(ws.cell(row=2, column=7).value)


def _apply_target_order_qty_to_ws(ws, target_order_qty: float | None, source_order_qty: float | None = None):
    effective_order_qty = coerce_qty(target_order_qty)
    if effective_order_qty <= 0:
        return

    base_order_qty = coerce_qty(source_order_qty) or _read_ws_bom_order_qty(ws)
    part_col = cfg("excel.bom_part_col", 2) + 1
    qty_col = cfg("excel.bom_qty_per_board", 1) + 1
    scrap_col = cfg("excel.bom_scrap_col", 4) + 1
    needed_col = cfg("excel.bom_needed_col", 5) + 1
    g_col = cfg("excel.bom_g_col", 6) + 1
    h_col = cfg("excel.bom_h_col", 7) + 1
    data_start = cfg("excel.bom_data_start_row", 5)
    dash_markers = {"-", "x", "X", "n", "N", "n/a", "N/A", "na", "NA", "?"}

    for row_idx in range(data_start, ws.max_row + 1):
        if not normalize_part_key(ws.cell(row=row_idx, column=part_col).value):
            continue
        g_text = str(ws.cell(row=row_idx, column=g_col).value or "").strip()
        h_text = str(ws.cell(row=row_idx, column=h_col).value or "").strip()
        if g_text in dash_markers or h_text in dash_markers:
            continue
        target_needed_cell = _resolve_cell_for_write(ws, row_idx, needed_col)
        if cell_has_formula(target_needed_cell):
            continue
        needed_qty = calculate_effective_needed_qty(
            needed_qty=target_needed_cell.value,
            qty_per_board=ws.cell(row=row_idx, column=qty_col).value,
            scrap_factor=ws.cell(row=row_idx, column=scrap_col).value,
            schedule_order_qty=effective_order_qty,
            bom_order_qty=base_order_qty,
        )
        _set_cell_value(ws, row_idx, needed_col, format_excel_qty(needed_qty))


def _write_bom_header_values(ws, po_number, order_qty: float | None = None):
    po_col = cfg("excel.bom_po_col", 7) + 1
    po_cell = _resolve_cell_for_write(ws, 1, po_col)
    po_cell.value = _format_bom_po_value(po_cell.value, po_number)
    if coerce_qty(order_qty) > 0:
        order_qty_col = cfg("excel.bom_order_qty_col", 10) + 1
        _resolve_cell_for_write(ws, 1, order_qty_col).value = format_excel_qty(order_qty)
        fallback_row2_qty = ws.cell(row=2, column=7).value
        if fallback_row2_qty in (None, "") or isinstance(fallback_row2_qty, (int, float)):
            ws.cell(row=2, column=7).value = format_excel_qty(order_qty)


def _get_bom_worksheet(workbook):
    if getattr(workbook, "worksheets", None):
        return workbook.worksheets[0]
    return workbook.active


def _select_bom_worksheet(workbook):
    ws = _get_bom_worksheet(workbook)
    try:
        workbook.active = workbook.index(ws)
    except Exception:
        pass
    return ws


def _write_dispatch_values_to_ws(
    ws,
    supplements: dict[str, float],
    carry_overs: dict[str, float],
    purchase_parts: set[str] | None = None,
    target_order_qty: float | None = None,
    source_order_qty: float | None = None,
):
    _apply_target_order_qty_to_ws(ws, target_order_qty, source_order_qty=source_order_qty)
    part_col = cfg("excel.bom_part_col", 2) + 1
    g_col = cfg("excel.bom_g_col", 6) + 1
    h_col = cfg("excel.bom_h_col", 7) + 1
    data_start = cfg("excel.bom_data_start_row", 5)
    dash_markers = {"-", "x", "X", "n", "N", "n/a", "N/A", "na", "NA", "?"}
    supplemented_parts: set[str] = set()

    for row_idx in range(data_start, ws.max_row + 1):
        part = normalize_part_key(ws.cell(row=row_idx, column=part_col).value)
        if not part:
            continue

        g_text = str(ws.cell(row=row_idx, column=g_col).value or "").strip()
        h_text = str(ws.cell(row=row_idx, column=h_col).value or "").strip()
        if g_text in dash_markers or h_text in dash_markers:
            continue

        if part in carry_overs:
            _set_cell_value(ws, row_idx, g_col, carry_overs[part])

        supplement_qty = 0
        if part not in supplemented_parts and part in supplements:
            supplement_qty = supplements[part]
            supplemented_parts.add(part)
        target_h_cell = _resolve_cell_for_write(ws, row_idx, h_col)
        target_h_cell.value = supplement_qty
        if supplement_qty and part in (purchase_parts or set()):
            target_h_cell.fill = ORANGE_FILL


def _ensure_editable_bom_for_draft(bom: dict, *, sync_components: bool = True) -> dict:
    normalized = normalize_bom_record_to_editable(bom or {})
    converted = normalized != (bom or {})
    if not converted and not sync_components:
        return normalized

    file_path = Path(str(normalized.get("filepath") or ""))
    if not file_path.exists():
        return normalized

    parsed = parse_bom_for_storage(
        path=normalized["filepath"],
        bom_id=normalized["id"],
        filename=normalized["filename"],
        uploaded_at=normalized["uploaded_at"],
        group_model=normalized.get("group_model", ""),
        source_filename=normalized.get("source_filename", ""),
        source_format=normalized.get("source_format", ""),
        is_converted=bool(normalized.get("is_converted")),
    )
    db.save_bom_file(build_bom_storage_payload(parsed))
    if converted:
        db.log_activity("bom_convert", f"{bom.get('filename') or bom['id']} 已在副檔生成前轉為可編輯 xlsx")
    return db.get_bom_file(normalized["id"]) or normalized


def _cleanup_draft_files(draft_id: int):
    for item in db.get_merge_draft_files(draft_id):
        Path(str(item.get("filepath") or "")).unlink(missing_ok=True)
    draft_dir = MERGE_DRAFT_DIR / f"draft_{draft_id}"
    if draft_dir.exists():
        shutil.rmtree(draft_dir, ignore_errors=True)


def _build_retention_cutoff_iso(retention_days: int = COMMITTED_DRAFT_RETENTION_DAYS) -> str:
    try:
        days = max(int(retention_days), 1)
    except (TypeError, ValueError):
        days = COMMITTED_DRAFT_RETENTION_DAYS
    return (datetime.now() - timedelta(days=days)).isoformat()


def cleanup_expired_committed_merge_drafts(retention_days: int = COMMITTED_DRAFT_RETENTION_DAYS) -> int:
    expired = db.get_expired_committed_merge_drafts(retention_days)
    cleaned = 0
    for draft in expired:
        draft_id = int(draft.get("id") or 0)
        if draft_id <= 0:
            continue
        _cleanup_draft_files(draft_id)
        cleaned += db.delete_merge_draft(draft_id)
    return cleaned


def restore_recent_committed_merge_drafts(
    order_ids: list[int],
    retention_days: int = COMMITTED_DRAFT_RETENTION_DAYS,
) -> list[int]:
    normalized_ids: list[int] = []
    for order_id in order_ids or []:
        try:
            normalized_ids.append(int(order_id))
        except (TypeError, ValueError):
            continue
    normalized_ids = list(dict.fromkeys(normalized_ids))
    if not normalized_ids:
        return []

    cutoff_iso = _build_retention_cutoff_iso(retention_days)
    restored_order_ids: list[int] = []
    for order_id in normalized_ids:
        if db.get_active_merge_draft_for_order(order_id):
            continue
        committed = db.get_latest_committed_merge_draft_for_order(order_id, committed_after=cutoff_iso)
        if not committed:
            continue
        if db.reactivate_merge_draft(int(committed["id"])):
            restored_order_ids.append(order_id)

    active_order_ids = [int(item["order_id"]) for item in db.get_active_merge_drafts()]
    if active_order_ids:
        rebuild_merge_drafts(active_order_ids)
    return restored_order_ids


def _build_download_response(file_entries: list[dict], archive_label: str = "副檔草稿", request: Request | None = None):
    if not file_entries:
        raise HTTPException(404, "副檔檔案不存在")

    if len(file_entries) == 1:
        entry = file_entries[0]
        if request is not None:
            return maybe_server_save_response(
                request,
                str(entry["path"]),
                entry["download_name"],
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        return FileResponse(
            path=str(entry["path"]),
            filename=entry["download_name"],
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    zip_buffer = io.BytesIO()
    used_names: set[str] = set()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for entry in file_entries:
            download_name = entry["download_name"]
            subdir = str(entry.get("subdir") or "").strip().strip("/\\")
            target_name = f"{subdir}/{download_name}" if subdir else download_name
            stem = Path(download_name).stem
            suffix = Path(download_name).suffix or Path(entry["path"]).suffix
            counter = 1
            while target_name in used_names:
                deduped_name = f"{stem}_{counter}{suffix}"
                target_name = f"{subdir}/{deduped_name}" if subdir else deduped_name
                counter += 1
            used_names.add(target_name)
            zf.write(str(entry["path"]), target_name)
    zip_buffer.seek(0)
    archive_name = build_generated_filename(archive_label, ".zip")
    if request is not None:
        server_response = maybe_server_save_bytes_response(request, zip_buffer.getvalue(), archive_name)
        if server_response is not None:
            return server_response
        zip_buffer.seek(0)
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(archive_name)}"},
    )


def _plan_order_draft(
    order: dict,
    draft: dict,
    bom_files: list[dict],
    running_stock: dict[str, float],
    moq_map: dict[str, float],
    st_inventory_stock: dict[str, float] | None = None,
    *,
    decisions: dict[str, str] | None = None,
    supplements: dict[str, float] | None = None,
) -> dict:
    st_inventory_stock = st_inventory_stock or {}
    decisions = _normalize_decisions(decisions if decisions is not None else draft.get("decisions"))
    remaining_supplements = _normalize_supplements(supplements if supplements is not None else draft.get("supplements"))
    file_plans: list[dict] = []
    shortages: list[dict] = []
    schedule_order_qty = coerce_qty(order.get("order_qty"))

    for bom in bom_files:
        components = db.get_bom_components(str(bom["id"]))
        if not components:
            continue
        target_order_qty = schedule_order_qty if schedule_order_qty > 0 else None

        carry_overs: dict[str, float] = {}
        supplement_allocations: dict[str, float] = {}
        purchase_parts: set[str] = set()
        part_totals: dict[str, dict[str, float]] = {}

        for component in components:
            needed_qty = get_component_effective_needed_qty(
                component,
                schedule_order_qty=schedule_order_qty,
                bom_order_qty=bom.get("order_qty"),
            )
            if component.get("is_dash") or needed_qty <= 0:
                continue

            part = normalize_part_key(component.get("part_number"))
            if not part:
                continue

            if part not in carry_overs:
                carry_overs[part] = float(running_stock.get(part, 0))

            summary = part_totals.setdefault(part, {
                "part_number": str(component.get("part_number") or ""),
                "description": str(component.get("description") or ""),
                "needed_qty": 0.0,
                "prev_qty_cs": 0.0,
            })
            summary["needed_qty"] += needed_qty
            summary["prev_qty_cs"] += float(component.get("prev_qty_cs") or 0)

        decisions = _sanitize_resolved_shortage_decisions(
            part_totals,
            running_stock,
            decisions,
            remaining_supplements,
        )

        for part, summary in part_totals.items():
            current_stock = float(running_stock.get(part, 0))
            prev_qty_cs = float(summary.get("prev_qty_cs") or 0)
            needed_qty = float(summary.get("needed_qty") or 0)
            decision = decisions.get(part, "None")
            st_stock_qty = float(st_inventory_stock.get(part, 0.0) or 0.0)

            available_before = current_stock + prev_qty_cs
            ending_without_supplement = available_before - needed_qty
            shortage_before = calculate_current_order_shortage_amount(part, available_before, needed_qty)
            supplement_qty = 0.0
            if decision != "Shortage" and remaining_supplements.get(part, 0) > 0:
                supplement_qty = float(remaining_supplements.get(part, 0))
                remaining_supplements[part] = 0.0
                supplement_allocations[part] = supplement_qty
                if bool(summarize_requested_supply(supplement_qty, st_stock_qty)["needs_purchase"]):
                    purchase_parts.add(part)

            available_after_supply = available_before + supplement_qty
            if decision == "Shortage":
                ending_stock = available_after_supply
                shortage_after = shortage_before
            else:
                ending_stock = available_after_supply - needed_qty
                shortage_after = calculate_current_order_shortage_amount(part, available_after_supply, needed_qty)

            running_stock[part] = ending_stock

            if shortage_after > 0:
                moq = float(moq_map.get(part, 0) or 0)
                if is_order_scoped_shortage_part(part):
                    st_context = summarize_requested_supply(shortage_after, st_stock_qty)
                    st_available_qty = float(st_context["st_available_qty"] or 0.0)
                    purchase_needed_qty = float(st_context["purchase_needed_qty"] or 0.0)
                    purchase_suggested_qty = purchase_needed_qty
                    suggested_qty = shortage_after
                else:
                    st_context = summarize_st_supply(shortage_after, st_stock_qty, moq)
                    st_available_qty = float(st_context["st_available_qty"] or 0.0)
                    purchase_needed_qty = float(st_context["purchase_needed_qty"] or 0.0)
                    purchase_suggested_qty = calc_suggested_qty(purchase_needed_qty, moq) if purchase_needed_qty > 0 else 0.0
                    suggested_qty = st_available_qty + purchase_suggested_qty
                shortages.append({
                    "part_number": summary.get("part_number") or part,
                    "description": summary.get("description") or "",
                    "current_stock": float(carry_overs.get(part, current_stock)),
                    "needed": needed_qty,
                    "prev_qty_cs": prev_qty_cs,
                    "shortage_amount": shortage_after,
                    "moq": moq,
                    "suggested_qty": suggested_qty if shortage_after > 0 else 0.0,
                    "purchase_suggested_qty": purchase_suggested_qty,
                    "decision": decision,
                    "supplement_qty": supplement_qty,
                    "resulting_stock": ending_stock,
                    "_row_code": order.get("code") or order.get("model") or "",
                    "_row_model": order.get("model") or "",
                    **st_context,
                })

        file_plans.append({
            "bom_file_id": str(bom["id"]),
            "source_filename": str(bom.get("filename") or ""),
            "source_format": str(bom.get("source_format") or Path(str(bom.get("filename") or "")).suffix.lower()),
            "model": str(bom.get("model") or ""),
            "group_model": str(bom.get("group_model") or ""),
            "po_number": str(order.get("po_number") or bom.get("po_number") or ""),
            "order_qty": target_order_qty,
            "source_order_qty": coerce_qty(bom.get("order_qty")),
            "carry_overs": carry_overs,
            "supplements": supplement_allocations,
            "purchase_parts": sorted(purchase_parts),
        })

    return {
        "running_stock": running_stock,
        "file_plans": file_plans,
        "shortages": shortages,
        "decisions": decisions,
    }


def _write_draft_files(draft_id: int, file_plans: list[dict], *, root_dir: Path | None = None) -> list[dict]:
    draft_dir = (root_dir or MERGE_DRAFT_DIR) / f"draft_{draft_id}"
    draft_dir.mkdir(parents=True, exist_ok=True)
    written: list[dict] = []

    for index, plan in enumerate(file_plans, start=1):
        bom = db.get_bom_file(plan["bom_file_id"])
        if not bom:
            continue
        bom = _ensure_editable_bom_for_draft(bom, sync_components=False)

        source_path = Path(str(bom.get("filepath") or ""))
        if not source_path.exists():
            continue

        display_name = _build_draft_display_filename(plan, source_path)
        internal_name = f"{index:02d}_{display_name}"
        output_path = draft_dir / internal_name

        ext = source_path.suffix.lower()
        workbook = openpyxl.load_workbook(str(source_path), keep_vba=(ext == ".xlsm"))
        try:
            sheet = _select_bom_worksheet(workbook)
            _write_dispatch_values_to_ws(
                sheet,
                plan.get("supplements") or {},
                plan.get("carry_overs") or {},
                purchase_parts={normalize_part_key(part) for part in (plan.get("purchase_parts") or [])},
                target_order_qty=plan.get("order_qty"),
                source_order_qty=plan.get("source_order_qty"),
            )
            _write_bom_header_values(sheet, plan.get("po_number", ""), plan.get("order_qty"))
            save_workbook_with_recalc(workbook, output_path)
        finally:
            workbook.close()

        written.append({
            "bom_file_id": plan["bom_file_id"],
            "filename": display_name,
            "filepath": str(output_path),
            "source_filename": plan.get("source_filename", ""),
            "source_format": plan.get("source_format", ""),
            "model": plan.get("model", ""),
            "group_model": plan.get("group_model", ""),
            "carry_overs": plan.get("carry_overs") or {},
            "supplements": plan.get("supplements") or {},
        })

    return written


def rebuild_merge_drafts(order_ids: list[int], overrides: dict[int, dict] | None = None) -> list[dict]:
    normalized_ids: list[int] = []
    for order_id in order_ids or []:
        try:
            normalized_ids.append(int(order_id))
        except (TypeError, ValueError):
            continue
    normalized_ids = list(dict.fromkeys(normalized_ids))
    if not normalized_ids:
        return []

    main_path = str(db.get_setting("main_file_path") or "").strip()
    if not main_path or not Path(main_path).exists():
        raise HTTPException(400, "請先載入主檔，才能建立副檔")

    main_loaded_at = str(db.get_setting("main_loaded_at") or "")
    main_path, main_mtime_ns = _get_main_signature(main_path)
    normalized_overrides: dict[int, dict] = {}
    for order_id, payload in (overrides or {}).items():
        if not (isinstance(order_id, int) or str(order_id).strip().isdigit()):
            continue
        normalized_overrides[int(order_id)] = {
            "decisions": _normalize_decisions((payload or {}).get("decisions")),
            "supplements": _normalize_supplements((payload or {}).get("supplements")),
        }

    persisted_decisions_by_order = _get_persisted_decision_map(normalized_ids)
    persisted_supplements_by_order = _get_persisted_supplement_map(normalized_ids)
    for raw_order_id in normalized_ids:
        order = db.get_order(raw_order_id)
        if not order:
            continue
        existing = db.get_active_merge_draft_for_order(raw_order_id) or {}
        payload = normalized_overrides.get(raw_order_id, {})
        decisions = payload.get("decisions", persisted_decisions_by_order.get(raw_order_id, {}))
        supplements = payload.get("supplements", persisted_supplements_by_order.get(raw_order_id, {}))
        db.replace_merge_draft(
            order_id=raw_order_id,
            main_file_path=main_path,
            main_file_mtime_ns=main_mtime_ns,
            main_loaded_at=main_loaded_at,
            decisions=decisions,
            supplements=supplements,
            shortages=existing.get("shortages", []),
        )

    t0 = time.monotonic()
    active_drafts = db.get_active_merge_drafts()
    log.info("[rebuild_merge_drafts] %d active drafts, reading main file...", len(active_drafts))

    t1 = time.monotonic()
    running_stock = _build_running_stock(main_path)
    t2 = time.monotonic()
    moq_map = _load_effective_moq(main_path)
    t3 = time.monotonic()
    log.info("[rebuild_merge_drafts] read_stock=%.1fs, read_moq=%.1fs", t2 - t1, t3 - t2)

    st_inventory_stock = db.get_st_inventory_stock()
    refreshed_ids: set[int] = set()
    active_order_ids = [int(draft["order_id"]) for draft in active_drafts]
    active_decisions_by_order = _get_persisted_decision_map(active_order_ids)
    active_supplements_by_order = _get_persisted_supplement_map(active_order_ids)

    for draft in active_drafts:
        draft_id = int(draft["id"])
        order_id = int(draft["order_id"])
        order = db.get_order(order_id)
        if not order or order.get("status") not in ("pending", "merged"):
            _cleanup_draft_files(draft_id)
            db.delete_merge_draft(draft_id)
            continue

        td0 = time.monotonic()
        bom_files = [
            _ensure_editable_bom_for_draft(bom)
            for bom in db.get_bom_files_by_models([str(order.get("model") or "")])
        ]
        effective_decisions = dict(active_decisions_by_order.get(order_id, {}))
        original_decisions = dict(effective_decisions)
        effective_supplements = active_supplements_by_order.get(order_id, {})
        plan = _plan_order_draft(
            order,
            draft,
            bom_files,
            running_stock,
            moq_map,
            st_inventory_stock,
            decisions=effective_decisions,
            supplements=effective_supplements,
        )
        sanitized_decisions = plan.get("decisions") if "decisions" in plan else effective_decisions
        effective_decisions = _normalize_decisions(sanitized_decisions)
        if effective_decisions != original_decisions:
            db.replace_order_decisions([order_id], {order_id: effective_decisions})
            active_decisions_by_order[order_id] = effective_decisions
        db.replace_merge_draft(
            order_id=int(order["id"]),
            main_file_path=main_path,
            main_file_mtime_ns=main_mtime_ns,
            main_loaded_at=main_loaded_at,
            decisions=effective_decisions,
            supplements=effective_supplements,
            shortages=plan["shortages"],
        )
        refreshed = db.get_active_merge_draft_for_order(int(order["id"]))
        if not refreshed:
            continue
        _cleanup_draft_files(int(refreshed["id"]))
        written = _write_draft_files(int(refreshed["id"]), plan["file_plans"])
        db.replace_merge_draft_files(int(refreshed["id"]), written)
        refreshed_ids.add(int(refreshed["id"]))
        log.info("[rebuild_merge_drafts] draft %d (order %d) done in %.1fs, %d files written",
                 draft_id, int(draft["order_id"]), time.monotonic() - td0, len(written))

    log.info("[rebuild_merge_drafts] total %.1fs for %d drafts", time.monotonic() - t0, len(refreshed_ids))
    return [draft for draft in db.get_active_merge_drafts() if int(draft["id"]) in refreshed_ids]


def delete_merge_draft_and_refresh(draft_id: int):
    draft = db.get_merge_draft(draft_id)
    if not draft or draft.get("status") != "active":
        raise HTTPException(404, "找不到副檔草稿")
    _cleanup_draft_files(draft_id)
    db.delete_merge_draft(draft_id)
    remaining = [item["order_id"] for item in db.get_active_merge_drafts()]
    if remaining:
        rebuild_merge_drafts(remaining)


def _build_draft_file_preview_rows(
    file_item: dict,
    decisions: dict[str, str],
    shortages_by_part: dict[str, dict],
    order: dict | None = None,
) -> list[dict]:
    bom_file_id = str(file_item.get("bom_file_id") or "").strip()
    if not bom_file_id:
        return []

    carry_overs = {
        normalize_part_key(part): float(qty or 0)
        for part, qty in (file_item.get("carry_overs") or {}).items()
        if normalize_part_key(part)
    }
    supplements = {
        normalize_part_key(part): float(qty or 0)
        for part, qty in (file_item.get("supplements") or {}).items()
        if normalize_part_key(part)
    }

    grouped: dict[str, dict] = {}
    schedule_order_qty = coerce_qty((order or {}).get("order_qty"))
    for component in db.get_bom_components(bom_file_id):
        needed_qty = get_component_effective_needed_qty(component, schedule_order_qty=schedule_order_qty)
        if component.get("is_dash") or needed_qty <= 0:
            continue

        part_key = normalize_part_key(component.get("part_number"))
        if not part_key:
            continue

        item = grouped.setdefault(part_key, {
            "part_number": str(component.get("part_number") or part_key),
            "description": str(component.get("description") or ""),
            "needed": 0.0,
            "prev_qty_cs": 0.0,
            "carry_over": float(carry_overs.get(part_key, 0)),
            "supplement_qty": float(supplements.get(part_key, 0)),
            "decision": decisions.get(part_key, "None"),
            "shortage_amount": 0.0,
            "resulting_stock": None,
            "suggested_qty": 0.0,
            "moq": 0.0,
            "st_stock_qty": 0.0,
            "st_available_qty": 0.0,
            "purchase_needed_qty": 0.0,
            "purchase_suggested_qty": 0.0,
            "needs_purchase": False,
            "is_customer_supplied": bool(component.get("is_customer_supplied")),
        })
        item["needed"] += needed_qty
        item["prev_qty_cs"] += float(component.get("prev_qty_cs") or 0)
        item["is_customer_supplied"] = item["is_customer_supplied"] or bool(component.get("is_customer_supplied"))

    for part_key, shortage in shortages_by_part.items():
        if part_key not in grouped:
            continue
        grouped[part_key]["shortage_amount"] = float(shortage.get("shortage_amount") or 0)
        grouped[part_key]["resulting_stock"] = shortage.get("resulting_stock")
        grouped[part_key]["suggested_qty"] = float(shortage.get("suggested_qty") or 0)
        grouped[part_key]["moq"] = float(shortage.get("moq") or 0)
        grouped[part_key]["st_stock_qty"] = float(shortage.get("st_stock_qty") or 0)
        grouped[part_key]["st_available_qty"] = float(shortage.get("st_available_qty") or 0)
        grouped[part_key]["purchase_needed_qty"] = float(shortage.get("purchase_needed_qty") or 0)
        grouped[part_key]["purchase_suggested_qty"] = float(shortage.get("purchase_suggested_qty") or 0)
        grouped[part_key]["needs_purchase"] = bool(shortage.get("needs_purchase"))

    return list(grouped.values())


def get_schedule_draft_map() -> dict[int, dict]:
    drafts = db.get_active_merge_drafts()
    order_ids = [int(draft["order_id"]) for draft in drafts]
    decisions_by_order = _get_persisted_decision_map(order_ids)
    supplements_by_order = _get_persisted_supplement_map(order_ids)
    return {
        int(draft["order_id"]): _serialize_draft_summary(
            draft,
            decisions=decisions_by_order.get(int(draft["order_id"]), {}),
            supplements=supplements_by_order.get(int(draft["order_id"]), {}),
        )
        for draft in drafts
    }


def _serialize_draft_summary(
    draft: dict,
    *,
    decisions: dict[str, str] | None = None,
    supplements: dict[str, float] | None = None,
) -> dict:
    order_id = int(draft["order_id"])
    return {
        "id": int(draft["id"]),
        "order_id": order_id,
        "status": draft.get("status", "active"),
        "model": draft.get("model", ""),
        "po_number": draft.get("po_number", ""),
        "main_loaded_at": draft.get("main_loaded_at", ""),
        "updated_at": draft.get("updated_at", ""),
        "committed_at": draft.get("committed_at", ""),
        "supplements": _normalize_supplements(supplements or {}),
        "decisions": _normalize_decisions(decisions or {}),
        "shortages": draft.get("shortages", []),
        "files": [
            {
                "id": int(file_item["id"]),
                "bom_file_id": file_item.get("bom_file_id", ""),
                "filename": file_item.get("filename", ""),
                "filepath": file_item.get("filepath", ""),
                "source_filename": file_item.get("source_filename", ""),
            }
            for file_item in draft.get("files", [])
        ],
    }


def get_committed_schedule_draft_map(order_ids: list[int]) -> dict[int, dict]:
    normalized_ids: list[int] = []
    for order_id in order_ids or []:
        try:
            normalized_ids.append(int(order_id))
        except (TypeError, ValueError):
            continue
    normalized_ids = list(dict.fromkeys(normalized_ids))
    if not normalized_ids:
        return {}

    result: dict[int, dict] = {}
    for order_id in normalized_ids:
        draft = db.get_latest_committed_merge_draft_for_order(order_id)
        if not draft:
            continue
        draft["files"] = db.get_merge_draft_files(int(draft["id"]))
        result[order_id] = _serialize_draft_summary(
            draft,
            decisions=draft.get("decisions", {}),
            supplements=draft.get("supplements", {}),
        )
    return result


def get_draft_detail(draft_id: int) -> dict:
    draft = db.get_merge_draft(draft_id)
    if not draft or draft.get("status") not in ("active", "committed"):
        raise HTTPException(404, "找不到副檔草稿")
    order_id = int(draft["order_id"])
    order = db.get_order(order_id)
    if draft.get("status") == "committed":
        decisions = _normalize_decisions(draft.get("decisions") or {})
        supplements = _normalize_supplements(draft.get("supplements") or {})
    else:
        decisions = _get_persisted_decision_map([order_id]).get(order_id, {})
        supplements = _get_persisted_supplement_map([order_id]).get(order_id, {})
    shortages_by_part = {
        normalize_part_key(item.get("part_number")): item
        for item in (draft.get("shortages") or [])
        if normalize_part_key(item.get("part_number"))
    }
    files = []
    for file_item in db.get_merge_draft_files(draft_id):
        enriched = dict(file_item)
        enriched["preview_rows"] = _build_draft_file_preview_rows(
            enriched,
            decisions,
            shortages_by_part,
            order=order or {},
        )
        files.append(enriched)
    return {
        "draft": {
            "id": int(draft["id"]),
            "order_id": order_id,
            "status": draft.get("status", "active"),
            "main_loaded_at": draft.get("main_loaded_at", ""),
            "updated_at": draft.get("updated_at", ""),
            "committed_at": draft.get("committed_at", ""),
            "supplements": supplements,
            "decisions": decisions,
            "shortages": draft.get("shortages", []),
            "files": files,
        },
        "order": order or {},
    }


def _replace_po_in_filename(filename: str, po_number: str) -> str:
    """將檔名中的 PO#xxxx 換成訂單實際的 PO 號。"""
    if not po_number:
        return filename
    safe_po = _sanitize_filename_piece(po_number, "PO")
    path = Path(filename or "副檔.xlsx")
    updated_stem = re.sub(r"PO#\s*[^_\s]+", safe_po, path.stem, count=1)
    return f"{updated_stem}{path.suffix}" if updated_stem != path.stem else filename


def _sanitize_committed_archive_piece(value) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    text = re.sub(r'[\\/:*?"<>|]+', " ", text)
    text = re.sub(r"\s+", " ", text).strip(" .")
    return text


def _format_committed_archive_date(value) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        parsed = datetime.strptime(text[:10], "%Y-%m-%d")
    except ValueError:
        return ""
    return f"{parsed.year}.{parsed.month}.{parsed.day}"


def _format_committed_archive_subdir(order: dict) -> str:
    code = _sanitize_committed_archive_piece(order.get("code"))
    po_number = _sanitize_committed_archive_piece(order.get("po_number"))
    model = _sanitize_committed_archive_piece(order.get("model"))
    ship_date = _format_committed_archive_date(order.get("delivery_date"))

    prefix_parts: list[str] = []
    if code:
        prefix_parts.append(code)
    if po_number:
        prefix_parts.append(f"PO#{po_number}")
    if model:
        prefix_parts.append(model)

    prefix = " ".join(prefix_parts)
    date_part = f"{ship_date} 出貨" if ship_date else ""
    if prefix and date_part:
        return f"{prefix}  {date_part}"
    if prefix:
        return prefix
    if date_part:
        return date_part
    return _sanitize_committed_archive_piece(f"order_{order.get('id') or ''}") or "order"


def _rebuild_committed_merge_draft_files(draft: dict, *, file_id: int | None = None) -> list[dict]:
    draft_id = int(draft["id"])
    order_id = int(draft["order_id"])
    order = db.get_order(order_id)
    if not order:
        raise HTTPException(404, "找不到已發料訂單")

    main_path = str(db.get_setting("main_file_path") or "").strip()
    if not main_path or not Path(main_path).exists():
        raise HTTPException(400, "請先載入主檔，才能下載已發料副檔")

    requested_bom_file_ids: set[str] | None = None
    if file_id is not None:
        requested_files = [
            item for item in db.get_merge_draft_files(draft_id)
            if int(item.get("id", -1)) == file_id
        ]
        if not requested_files:
            raise HTTPException(404, "找不到副檔檔案")
        requested_bom_file_ids = {
            str(item.get("bom_file_id") or "").strip()
            for item in requested_files
            if str(item.get("bom_file_id") or "").strip()
        }

    bom_files = [
        _ensure_editable_bom_for_draft(bom)
        for bom in db.get_bom_files_by_models([str(order.get("model") or "")])
    ]
    if requested_bom_file_ids is not None:
        bom_files = [
            bom for bom in bom_files
            if str(bom.get("id") or "").strip() in requested_bom_file_ids
        ]
    if not bom_files:
        raise HTTPException(404, "已發料訂單沒有可下載的副檔")

    decisions = _get_persisted_decision_map([order_id]).get(order_id, {})
    supplements = _get_persisted_supplement_map([order_id]).get(order_id, {})
    plan = _plan_order_draft(
        order,
        draft,
        bom_files,
        _build_running_stock(main_path),
        _load_effective_moq(main_path),
        db.get_st_inventory_stock(),
        decisions=decisions,
        supplements=supplements,
    )

    temp_root = Path(tempfile.mkdtemp(prefix=f"committed_merge_draft_{draft_id}_"))
    written = _write_draft_files(draft_id, plan["file_plans"], root_dir=temp_root)
    po = order.get("po_number", "")
    return [
        {
            "path": Path(str(item.get("filepath") or "")),
            "download_name": _replace_po_in_filename(item.get("filename") or "", po),
        }
        for item in written
        if Path(str(item.get("filepath") or "")).exists()
    ]


def download_merge_draft(draft_id: int, *, file_id: int | None = None, request: Request | None = None):
    draft = db.get_merge_draft(draft_id)
    if not draft or draft.get("status") not in ("active", "committed"):
        raise HTTPException(404, "找不到副檔草稿")
    if draft.get("status") == "committed":
        valid_files = _rebuild_committed_merge_draft_files(draft, file_id=file_id)
        return _build_download_response(valid_files, archive_label="已發料副檔", request=request)

    order = db.get_order(int(draft["order_id"])) or {}
    po = order.get("po_number", "")
    files = db.get_merge_draft_files(draft_id)
    if file_id is not None:
        files = [f for f in files if int(f.get("id", -1)) == file_id]
    valid_files = [
        {
            "path": file_path,
            "download_name": _replace_po_in_filename(item.get("filename") or file_path.name, po),
        }
        for item in files
        for file_path in [Path(str(item.get("filepath") or ""))]
        if file_path.exists()
    ]
    return _build_download_response(valid_files, request=request)


def download_selected_merge_drafts(order_ids: list[int], request: Request | None = None):
    normalized_ids: list[int] = []
    for order_id in order_ids or []:
        try:
            normalized_ids.append(int(order_id))
        except (TypeError, ValueError):
            continue
    normalized_ids = list(dict.fromkeys(normalized_ids))
    if not normalized_ids:
        raise HTTPException(400, "請先選擇要下載副檔的訂單")

    draft_id_map = db.get_active_merge_draft_ids_by_order_ids(normalized_ids)
    missing_orders = [order_id for order_id in normalized_ids if order_id not in draft_id_map]
    if missing_orders:
        raise HTTPException(404, "部分訂單尚未建立副檔")

    file_entries: list[dict] = []
    for order_id in normalized_ids:
        order = db.get_order(order_id) or {}
        po = order.get("po_number", "")
        draft_files = db.get_merge_draft_files(int(draft_id_map[order_id]))
        for item in draft_files:
            file_path = Path(str(item.get("filepath") or ""))
            if not file_path.exists():
                continue
            file_entries.append({
                "path": file_path,
                "download_name": _replace_po_in_filename(item.get("filename") or file_path.name, po),
            })

    return _build_download_response(file_entries, request=request)


def download_selected_committed_merge_drafts(order_ids: list[int], request: Request | None = None):
    normalized_ids: list[int] = []
    for order_id in order_ids or []:
        try:
            normalized_ids.append(int(order_id))
        except (TypeError, ValueError):
            continue
    normalized_ids = list(dict.fromkeys(normalized_ids))
    if not normalized_ids:
        raise HTTPException(400, "請先選擇要下載副檔的已發料訂單")

    file_entries: list[dict] = []
    missing_orders: list[int] = []
    for order_id in normalized_ids:
        draft = db.get_latest_committed_merge_draft_for_order(order_id)
        if not draft:
            missing_orders.append(order_id)
            continue

        rebuilt_entries = _rebuild_committed_merge_draft_files(draft)
        if not rebuilt_entries:
            missing_orders.append(order_id)
            continue

        order = db.get_order(order_id) or {"id": order_id}
        subdir = _format_committed_archive_subdir(order)
        for entry in rebuilt_entries:
            entry["subdir"] = subdir
        file_entries.extend(rebuilt_entries)

    if missing_orders:
        raise HTTPException(404, "部分已發料訂單沒有可下載的副檔")

    return _build_download_response(file_entries, archive_label="已發料副檔", request=request)
