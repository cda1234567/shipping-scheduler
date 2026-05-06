from __future__ import annotations

import io
from pathlib import Path
from urllib.parse import quote

import hashlib

import openpyxl
from fastapi import APIRouter, File, HTTPException, Request, Response, UploadFile
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from .. import database as db
from ..config import BACKUP_DIR, MAIN_FILE_DIR
from ..services.server_downloads import maybe_server_save_bytes_response, maybe_server_save_response
from ..models import UpdateMoqRequest
from ..services.main_preview import read_live_main_preview
from ..services.main_reader import (
    find_legacy_snapshot_stock_fixes,
    read_moq,
    read_stock,
    read_vendors,
    update_vendor,
)
from ..services.local_time import local_now
from ..services.main_file_recalc import recalc_batch_balances_for_cell
from ..services.merge_to_main import backup_main_file
from ..snapshot_sync import refresh_snapshot_from_main

router = APIRouter()


# ── In-memory cache for main-file/data (avoids re-reading xlsx every request) ─
_main_data_cache: dict | None = None
_main_data_cache_mtime: float = 0.0


def invalidate_main_data_cache():
    global _main_data_cache, _main_data_cache_mtime
    _main_data_cache = None
    _main_data_cache_mtime = 0.0


def _repair_legacy_snapshot_if_needed(main_path: str, snapshot: dict[str, dict]) -> dict[str, dict]:
    if not snapshot:
        return snapshot

    fixes = find_legacy_snapshot_stock_fixes(main_path, snapshot)
    repaired = db.update_snapshot_stock(fixes)
    if repaired:
        db.log_activity("snapshot_repaired", f"修正舊快照庫存 {repaired} 筆")
        for part, qty in fixes.items():
            if part in snapshot:
                snapshot[part]["stock_qty"] = qty
    return snapshot


@router.post("/main-file/upload")
async def upload_main_file(file: UploadFile = File(...)):
    ext = Path(file.filename or "").suffix.lower()
    if ext not in {".xlsx", ".xls", ".xlsm"}:
        raise HTTPException(400, "只支援 xlsx / xls / xlsm")

    dest = MAIN_FILE_DIR / f"main{ext}"
    dest.write_bytes(await file.read())

    stock = read_stock(str(dest))
    moq = read_moq(str(dest))

    db.set_setting("main_file_path", str(dest))
    db.set_setting("main_filename", file.filename or dest.name)
    db.set_setting("main_loaded_at", local_now().isoformat(timespec="seconds"))
    db.set_setting("main_part_count", str(len(stock)))

    # 上傳主檔 = 新基準，永遠更新快照
    refresh_snapshot_from_main(str(dest))
    invalidate_main_data_cache()
    db.log_activity("main_file_upload", f"{file.filename}, {len(stock)} 筆")
    return {"ok": True, "part_count": len(stock), "filename": file.filename}


@router.post("/main-file/snapshot")
async def set_snapshot():
    """把目前主檔重新設成缺料計算的快照基準。"""
    main_path = db.get_setting("main_file_path")
    if not main_path or not Path(main_path).exists():
        raise HTTPException(400, "請先上傳主檔")

    manual_moq = db.get_manual_snapshot_moq()
    stock = read_stock(main_path)
    moq = read_moq(main_path)
    moq.update(manual_moq)
    db.save_snapshot(stock, moq, manual_moq_parts=set(manual_moq))
    invalidate_main_data_cache()
    db.log_activity("snapshot_set", f"重設主檔快照，共 {len(stock)} 筆")
    return {"ok": True, "part_count": len(stock)}


@router.get("/main-file/data")
async def get_main_data():
    """回傳主檔庫存與 MOQ（帶 mtime 快取，主檔未變時直接回傳）。"""
    global _main_data_cache, _main_data_cache_mtime

    main_path = db.get_setting("main_file_path")
    if not main_path or not Path(main_path).exists():
        raise HTTPException(404, "找不到主檔")

    try:
        current_mtime = Path(main_path).stat().st_mtime
    except OSError:
        current_mtime = 0.0

    if _main_data_cache is not None and current_mtime == _main_data_cache_mtime:
        return _main_data_cache

    snapshot = db.get_snapshot()
    if snapshot:
        snapshot = _repair_legacy_snapshot_if_needed(main_path, snapshot)
        stock = {k: v["stock_qty"] for k, v in snapshot.items()}
        snapshot_moq = {k: v["moq"] for k, v in snapshot.items()}

        live_moq = read_moq(main_path)
        live_moq.update(snapshot_moq)
        moq = live_moq
    else:
        stock = read_stock(main_path)
        moq = read_moq(main_path)

    try:
        live_stock = read_stock(main_path)
    except Exception:
        live_stock = dict(stock)
    try:
        vendors = read_vendors(main_path)
    except Exception:
        vendors = {}

    result = {
        "stock": stock,
        "moq": moq,
        "live_stock": live_stock,
        "vendors": vendors,
        "purchase_reminder_statuses": db.get_purchase_reminder_statuses(),
        "part_count": int(db.get_setting("main_part_count", "0")),
        "loaded_at": db.get_setting("main_loaded_at"),
        "filename": db.get_setting("main_filename") or Path(main_path).name,
        "has_snapshot": bool(snapshot),
    }
    _main_data_cache = result
    _main_data_cache_mtime = current_mtime
    return result


@router.patch("/main-file/moq")
async def update_snapshot_moq(req: UpdateMoqRequest):
    part_number = str(req.part_number or "").strip().upper()
    if not part_number:
        raise HTTPException(400, "料號不可空白")

    saved_part = db.upsert_snapshot_moq(part_number, req.moq)
    invalidate_main_data_cache()
    db.log_activity("snapshot_moq_updated", f"{saved_part} MOQ -> {req.moq}")
    return {"ok": True, "part_number": saved_part, "moq": req.moq}


@router.get("/main-file/download")
async def download_main_file(request: Request):
    main_path = db.get_setting("main_file_path")
    if not main_path or not Path(main_path).exists():
        raise HTTPException(404, "找不到主檔")
    filename = db.get_setting("main_filename") or Path(main_path).name
    return maybe_server_save_response(
        request,
        main_path,
        filename,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/main-file/preview")
async def get_main_preview(request: Request, sheet: str | None = None):
    main_path = db.get_setting("main_file_path")
    if not main_path or not Path(main_path).exists():
        raise HTTPException(404, "找不到主檔")

    # ETag based on file mtime + sheet → 檔案沒變時回 304，瀏覽器跳過整個 download
    try:
        mtime_ns = Path(main_path).stat().st_mtime_ns
    except OSError:
        mtime_ns = 0
    etag_raw = f"{main_path}|{mtime_ns}|{sheet or ''}"
    etag = 'W/"' + hashlib.md5(etag_raw.encode("utf-8")).hexdigest() + '"'

    if request.headers.get("if-none-match") == etag:
        return Response(status_code=304, headers={"ETag": etag, "Cache-Control": "no-cache"})

    preview = read_live_main_preview(main_path, sheet_name=sheet)
    preview.update({
        "filename": db.get_setting("main_filename") or Path(main_path).name,
        "loaded_at": db.get_setting("main_loaded_at"),
    })
    return JSONResponse(content=preview, headers={"ETag": etag, "Cache-Control": "no-cache"})


@router.get("/main-file/info")
async def get_main_info():
    main_path = db.get_setting("main_file_path")
    snapshot = db.get_snapshot()
    exists = bool(main_path and Path(main_path).exists())
    return {
        "loaded": exists,
        "filename": db.get_setting("main_filename") or (Path(main_path).name if main_path else ""),
        "part_count": int(db.get_setting("main_part_count", "0")),
        "loaded_at": db.get_setting("main_loaded_at"),
        "has_snapshot": bool(snapshot),
    }


class EditCellRequest(BaseModel):
    sheet: str = ""
    row: int = Field(..., ge=1)
    col: int = Field(..., ge=1)
    value: str = ""


class UpdateVendorRequest(BaseModel):
    part_number: str = ""
    vendor: str = ""


class PurchaseReminderStatusRequest(BaseModel):
    part_number: str = ""
    notified: bool | None = None
    ignored: bool | None = None
    note: str = ""


class PurchaseReminderExportItem(BaseModel):
    vendor: str = ""
    part_number: str = ""
    description: str = ""
    current_stock: float = 0
    threshold: float = 0
    moq: float = 0
    suggested_qty: float = 0
    notified: bool = False
    notified_at: str = ""
    note: str = ""


class PurchaseReminderExportRequest(BaseModel):
    items: list[PurchaseReminderExportItem] = Field(default_factory=list)


def _clean_vendor(value: str) -> str:
    text = str(value or "").strip()
    return text or "未分類廠商"


def _build_purchase_reminder_export(items: list[PurchaseReminderExportItem]) -> bytes:
    normalized_items = sorted(
        [
            {
                "vendor": _clean_vendor(item.vendor),
                "part_number": str(item.part_number or "").strip().upper(),
                "description": str(item.description or "").strip(),
                "current_stock": float(item.current_stock or 0),
                "threshold": float(item.threshold or 0),
                "moq": float(item.moq or 0),
                "suggested_qty": float(item.suggested_qty or 0),
                "notified": bool(item.notified),
                "notified_at": str(item.notified_at or "").replace("T", " ")[:16],
                "note": str(item.note or "").strip(),
            }
            for item in (items or [])
            if str(item.part_number or "").strip()
        ],
        key=lambda row: (row["vendor"], row["notified"], row["part_number"]),
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "買料提醒"
    summary_ws = wb.create_sheet("廠商彙總")

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    notified_fill = PatternFill("solid", fgColor="E7F5E9")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:J1")
    title = ws["A1"]
    title.value = "IC / OC / UC ST 買料提醒"
    title.font = Font(bold=True, color="FFFFFF", size=14)
    title.fill = title_fill
    title.alignment = Alignment(horizontal="center")

    ws["A2"] = f"匯出時間：{local_now().isoformat(timespec='minutes').replace('T', ' ')}"
    ws["A2"].font = Font(color="666666")

    headers = ["廠商", "通知狀態", "料號", "說明", "ST 庫存", "安全線", "MOQ", "建議購買量", "通知時間", "備註"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if normalized_items:
        for row_idx, item in enumerate(normalized_items, start=5):
            values = [
                item["vendor"],
                "已通知採購" if item["notified"] else "待通知",
                item["part_number"],
                item["description"],
                item["current_stock"],
                item["threshold"],
                item["moq"],
                item["suggested_qty"],
                item["notified_at"],
                item["note"],
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(col in {4, 10}))
                if item["notified"]:
                    cell.fill = notified_fill
                elif col in {2, 5, 8}:
                    cell.fill = warning_fill
        ws.auto_filter.ref = f"A4:J{len(normalized_items) + 4}"
    else:
        ws.merge_cells("A5:J5")
        ws["A5"] = "目前沒有買料提醒。"
        ws["A5"].alignment = Alignment(horizontal="center")

    widths = [18, 13, 24, 32, 12, 12, 10, 14, 18, 28]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A5"

    summary_headers = ["廠商", "待通知", "已通知", "合計", "建議購買量合計"]
    for col, header in enumerate(summary_headers, start=1):
        cell = summary_ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    vendor_summary: dict[str, dict[str, float]] = {}
    for item in normalized_items:
        summary = vendor_summary.setdefault(
            item["vendor"],
            {"pending": 0, "notified": 0, "suggested": 0.0},
        )
        if item["notified"]:
            summary["notified"] += 1
        else:
            summary["pending"] += 1
        summary["suggested"] += item["suggested_qty"]

    for row_idx, (vendor, summary) in enumerate(sorted(vendor_summary.items()), start=2):
        values = [
            vendor,
            int(summary["pending"]),
            int(summary["notified"]),
            int(summary["pending"] + summary["notified"]),
            float(summary["suggested"]),
        ]
        for col, value in enumerate(values, start=1):
            cell = summary_ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            if col in {2, 5} and summary["pending"] > 0:
                cell.fill = warning_fill

    for idx, width in enumerate([18, 12, 12, 12, 18], start=1):
        summary_ws.column_dimensions[get_column_letter(idx)].width = width
    summary_ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    return buffer.getvalue()


@router.patch("/main-file/vendor")
async def update_main_vendor(req: UpdateVendorRequest):
    main_path = db.get_setting("main_file_path")
    if not main_path or not Path(main_path).exists():
        raise HTTPException(404, "找不到主檔")

    part_number = str(req.part_number or "").strip().upper()
    if not part_number:
        raise HTTPException(400, "料號不可空白")

    backup_main_file(main_path, str(BACKUP_DIR))
    try:
        result = update_vendor(main_path, part_number, req.vendor)
    except KeyError:
        raise HTTPException(404, f"主檔找不到料號 {part_number}") from None
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc

    invalidate_main_data_cache()
    db.log_activity("main_vendor_updated", f"{part_number}: {result['old_vendor']} -> {result['vendor']}")
    return {"ok": True, **result}


@router.patch("/main-file/purchase-reminder-status")
async def update_purchase_reminder_status(req: PurchaseReminderStatusRequest):
    part_number = str(req.part_number or "").strip().upper()
    if not part_number:
        raise HTTPException(400, "料號不可空白")

    if req.notified is True:
        status = db.set_purchase_reminder_status(part_number, True, req.note)
        log_detail = f"{part_number}: 已通知採購"
    elif req.ignored is True:
        status = db.set_purchase_reminder_ignored(part_number, True)
        log_detail = f"{part_number}: 已忽略買料提醒"
    elif req.notified is False and req.ignored is False:
        status = db.set_purchase_reminder_status(part_number, False, req.note)
        log_detail = f"{part_number}: 清除買料提醒狀態"
    elif req.ignored is False:
        status = db.set_purchase_reminder_ignored(part_number, False)
        log_detail = f"{part_number}: 取消忽略買料提醒"
    else:
        status = db.set_purchase_reminder_status(part_number, False, req.note)
        log_detail = f"{part_number}: 取消通知"

    invalidate_main_data_cache()
    db.log_activity("purchase_reminder_status", log_detail)
    return {"ok": True, "status": status}


@router.post("/main-file/purchase-reminders/export")
async def export_purchase_reminders(request: Request, req: PurchaseReminderExportRequest):
    content = _build_purchase_reminder_export(req.items)
    filename = f"ST買料提醒_{local_now().strftime('%Y%m%d_%H%M')}.xlsx"
    server_response = maybe_server_save_bytes_response(request, content, filename)
    if server_response:
        return server_response
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.patch("/main-file/cell")
async def edit_main_cell(req: EditCellRequest):
    """修改主檔指定 cell 的值（自動備份）。"""
    main_path = db.get_setting("main_file_path")
    if not main_path or not Path(main_path).exists():
        raise HTTPException(404, "找不到主檔")

    backup_main_file(main_path, str(BACKUP_DIR))

    is_xlsm = Path(main_path).suffix.lower() == ".xlsm"
    wb = openpyxl.load_workbook(main_path, keep_vba=is_xlsm)

    if req.sheet:
        ws = wb[req.sheet] if req.sheet in wb.sheetnames else wb.active
    else:
        ws = wb.active

    cell = ws.cell(row=req.row, column=req.col)
    old_value = cell.value

    # 嘗試轉數字，否則存字串
    new_value = req.value.strip()
    try:
        new_value = float(new_value)
        if new_value == int(new_value):
            new_value = int(new_value)
    except (ValueError, TypeError):
        pass

    snapshot = db.get_snapshot()
    snapshot_stock = {
        str(part).strip().upper(): float(values.get("stock_qty") or 0)
        for part, values in (snapshot or {}).items()
        if str(part).strip()
    }

    cell.value = new_value if new_value != "" else None
    recalc_result = recalc_batch_balances_for_cell(
        ws,
        row=req.row,
        col=req.col,
        snapshot_stock=snapshot_stock,
    )
    wb.save(main_path)
    wb.close()

    part_number = str(recalc_result.get("part_number") or "").strip().upper()
    current_stock = recalc_result.get("current_stock")
    if recalc_result.get("recalculated") and part_number and current_stock is not None:
        updated_snapshot_rows = db.update_snapshot_stock({part_number: float(current_stock)})
        if not updated_snapshot_rows:
            refresh_snapshot_from_main(main_path)
    else:
        refresh_snapshot_from_main(main_path)
    invalidate_main_data_cache()

    db.log_activity("主檔編輯", f"[{req.sheet or 'Sheet1'}] R{req.row}C{req.col}: {old_value} → {new_value}")
    return {
        "ok": True,
        "old_value": str(old_value or ""),
        "new_value": str(new_value),
        "affected_cells": recalc_result.get("affected_cells") or [],
    }
