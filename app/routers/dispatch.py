from __future__ import annotations

import os
import re
import tempfile
from pathlib import Path

import openpyxl
from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from starlette.concurrency import run_in_threadpool

from .. import database as db
from ..services.calculator import run as calc_run
from ..services.dispatch_form_generator import generate_dispatch_form
from ..services.download_names import build_generated_filename
from ..services.local_time import local_now
from ..services.main_reader import find_legacy_snapshot_stock_fixes, read_moq, read_stock
from ..services.shortage_rules import summarize_requested_supply
from ..services.server_downloads import maybe_server_save_response

router = APIRouter()


class DispatchRequest(BaseModel):
    order_ids: list[int] = Field(default_factory=list)
    decisions: dict[str, str] = Field(default_factory=dict)


def _normalize_part_key(value) -> str:
    return str(value or "").strip().upper()


def _normalize_decision_overrides(decisions: dict[str, str] | None) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for part_number, decision in (decisions or {}).items():
        key = _normalize_part_key(part_number)
        if key and decision:
            normalized[key] = decision
    return normalized


def _get_selected_orders(order_ids: list[int]) -> list[dict]:
    selected_orders: list[dict] = []
    seen_ids: set[int] = set()
    for order_id in order_ids:
        try:
            normalized_id = int(order_id)
        except (TypeError, ValueError):
            continue
        if normalized_id in seen_ids:
            continue
        seen_ids.add(normalized_id)

        order = db.get_order(normalized_id)
        if not order:
            continue
        if order.get("status") not in ("pending", "merged", "dispatched", "completed"):
            continue
        selected_orders.append(order)
    return selected_orders


def _load_shortage_inputs() -> tuple[dict[str, float], dict[str, float], dict[str, float]]:
    main_path = str(db.get_setting("main_file_path") or "").strip()
    if not main_path or not Path(main_path).exists():
        raise HTTPException(400, "請先上傳主檔")

    snapshot = db.get_snapshot()
    if snapshot:
        fixes = find_legacy_snapshot_stock_fixes(main_path, snapshot)
        if fixes:
            db.update_snapshot_stock(fixes)
            for part, qty in fixes.items():
                if part in snapshot:
                    snapshot[part]["stock_qty"] = qty

        stock = {
            _normalize_part_key(part): float((values or {}).get("stock_qty") or 0)
            for part, values in snapshot.items()
            if _normalize_part_key(part)
        }
        moq = {
            _normalize_part_key(part): float((values or {}).get("moq") or 0)
            for part, values in snapshot.items()
            if _normalize_part_key(part)
        }
        live_moq = {
            _normalize_part_key(part): float(qty or 0)
            for part, qty in read_moq(main_path).items()
            if _normalize_part_key(part)
        }
        live_moq.update(moq)
        moq = live_moq
    else:
        stock = {
            _normalize_part_key(part): float(qty or 0)
            for part, qty in read_stock(main_path).items()
            if _normalize_part_key(part)
        }
        moq = {
            _normalize_part_key(part): float(qty or 0)
            for part, qty in read_moq(main_path).items()
            if _normalize_part_key(part)
        }

    dispatched_consumption = db.get_all_dispatched_consumption(db.get_snapshot_taken_at())
    return stock, moq, dispatched_consumption


def _build_component_description_map(components: list[dict]) -> dict[str, str]:
    descriptions: dict[str, str] = {}
    for component in components:
        part = _normalize_part_key(component.get("part_number"))
        if not part:
            continue
        descriptions.setdefault(part, component.get("description", ""))
    return descriptions


def _get_active_reviewed_drafts_by_order(order_ids: list[int]) -> dict[int, dict]:
    drafts_by_order: dict[int, dict] = {}
    for draft in db.get_active_merge_drafts(order_ids):
        try:
            order_id = int(draft.get("order_id"))
        except (TypeError, ValueError):
            continue
        if order_id in drafts_by_order:
            continue
        drafts_by_order[order_id] = draft
    return drafts_by_order


def _is_reviewed_active_draft(draft: dict | None) -> bool:
    if not draft:
        return False
    return bool((draft.get("decisions") or {}) or (draft.get("supplements") or {}))


def _build_order_dispatch_context(
    order: dict,
    result_by_order: dict[int, dict],
    saved_supplements: dict[int, dict[str, float]],
    decision_overrides: dict[str, str],
    bom_map: dict[str, list[dict]],
    committed_main_supplements: dict[int, dict[str, float]] | None = None,
    active_draft: dict | None = None,
) -> dict | None:
    order_id = int(order["id"])
    model_key = _normalize_part_key(order.get("model"))
    components = bom_map.get(model_key, [])
    if not components:
        return None

    descriptions = _build_component_description_map(components)
    bom_parts = set(descriptions)
    saved_decisions = db.get_decisions_for_order(order_id)
    use_main_supplements = committed_main_supplements is not None and order_id in committed_main_supplements
    stored_order_supplements = (
        committed_main_supplements.get(order_id, {})
        if use_main_supplements
        else saved_supplements.get(order_id, {})
    )
    draft_supplements = {}
    if active_draft:
        for raw_part, raw_qty in (active_draft.get("supplements") or {}).items():
            part = _normalize_part_key(raw_part)
            if not part:
                continue
            try:
                draft_supplements[part] = float(raw_qty or 0)
            except (TypeError, ValueError):
                draft_supplements[part] = 0.0

    reviewed_draft = _is_reviewed_active_draft(active_draft)
    if reviewed_draft:
        # 已審閱的 draft 以 per-order DB decisions 為準，不被全域 override 覆蓋
        decisions = dict(saved_decisions)
        shortage_items = list(active_draft.get("shortages") or [])
    else:
        decisions = {**saved_decisions, **decision_overrides}
        result = result_by_order.get(order_id) or {}
        shortage_items = [
            *(result.get("shortages") or []),
            *(result.get("customer_material_shortages") or []),
        ]
    shortages_by_part = {
        _normalize_part_key(item.get("part_number")): item
        for item in shortage_items
        if _normalize_part_key(item.get("part_number"))
    }

    candidate_parts = set(shortages_by_part) | set(stored_order_supplements) | set(draft_supplements)
    candidate_parts.update(
        part_number
        for part_number, decision in decisions.items()
        if decision in {"CreateRequirement", "Shortage"}
    )
    candidate_parts &= bom_parts

    return {
        "order": order,
        "order_id": order_id,
        "descriptions": descriptions,
        "decisions": decisions,
        "stored_supplements": stored_order_supplements,
        "draft_supplements": draft_supplements,
        "shortages_by_part": shortages_by_part,
        "candidate_parts": sorted(candidate_parts),
        "reviewed_draft": reviewed_draft,
        "use_main_supplements": use_main_supplements,
    }


def _should_render_dispatch_item(
    decision: str,
    supplement_qty: float,
    shortage_item: dict | None,
    *,
    reviewed_draft: bool = False,
    has_explicit_supplement: bool = False,
) -> bool:
    if decision in {"MarkHasPO", "IgnoreOnce"}:
        return False
    if decision == "Shortage":
        return True
    if decision == "CreateRequirement" and has_explicit_supplement:
        return True
    if supplement_qty > 0:
        return True

    suggested_qty = float((shortage_item or {}).get("suggested_qty") or (shortage_item or {}).get("shortage_amount") or 0)
    if decision == "CreateRequirement":
        return bool(shortage_item) and suggested_qty > 0

    if reviewed_draft:
        # 已審閱的 draft，supplement=0 且非 CreateRequirement，代表使用者明確不補，不再 fallback 到 suggested_qty
        return False
    return bool(shortage_item) and suggested_qty > 0


def _should_highlight_dispatch_qty(
    part: str,
    qty: float,
    shortage_item: dict | None,
    st_inventory_stock: dict[str, float],
) -> bool:
    part_key = _normalize_part_key(part)
    shortage = shortage_item or {}
    st_stock_qty = max(
        float(shortage.get("st_stock_qty") or 0),
        float(shortage.get("st_available_qty") or 0),
        float(st_inventory_stock.get(part_key, 0) or 0),
    )
    return bool(summarize_requested_supply(qty, st_stock_qty)["needs_purchase"])


def _is_committed_status(order: dict) -> bool:
    return str(order.get("status") or "").strip().lower() in {"dispatched", "completed"}


def _to_number(value) -> float:
    if value is None or str(value).strip() == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _load_committed_main_supplements(
    orders: list[dict],
    bom_map: dict[str, list[dict]],
) -> dict[int, dict[str, float]]:
    """Read committed order supplements from the live main file batch columns.

    For dispatched/completed orders the main preview is the source of truth. If
    a batch code exists in main, stale order_supplements must not reappear on
    generated dispatch forms.
    """
    committed_orders = [order for order in orders if _is_committed_status(order)]
    if not committed_orders:
        return {}

    main_path = str(db.get_setting("main_file_path") or "").strip()
    if not main_path or not Path(main_path).exists():
        return {}

    try:
        wb = openpyxl.load_workbook(main_path, read_only=True, data_only=True)
    except Exception:
        return {}

    try:
        ws = wb.worksheets[0]
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        batch_cols_by_code: dict[str, list[int]] = {}
        for col_idx, value in enumerate(first_row, start=1):
            code = str(value or "").strip()
            if re.match(r"^\d+-\d+$", code):
                batch_cols_by_code.setdefault(code, []).append(col_idx)

        row_map: dict[str, int] = {}
        for row_idx, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row_values:
                continue
            part = _normalize_part_key(row_values[0] if len(row_values) >= 1 else "")
            if part and part not in row_map:
                row_map[part] = row_idx

        result: dict[int, dict[str, float]] = {}
        for order in committed_orders:
            try:
                order_id = int(order["id"])
            except (KeyError, TypeError, ValueError):
                continue
            code = str(order.get("code") or "").strip()
            batch_cols = batch_cols_by_code.get(code)
            if not batch_cols:
                continue

            model_key = _normalize_part_key(order.get("model"))
            parts = [
                _normalize_part_key(component.get("part_number"))
                for component in bom_map.get(model_key, [])
            ]
            supplements: dict[str, float] = {}
            for part in dict.fromkeys(part for part in parts if part):
                row_idx = row_map.get(part)
                if row_idx is None:
                    continue
                qty = 0.0
                for col_idx in batch_cols:
                    qty += _to_number(ws.cell(row=row_idx, column=col_idx).value)
                if qty > 0:
                    supplements[part] = qty
            result[order_id] = supplements
        return result
    finally:
        wb.close()


def _subtract_selected_committed_dispatch_records(
    orders: list[dict],
    dispatched_consumption: dict[str, float],
) -> dict[str, float]:
    adjusted = {
        _normalize_part_key(part): float(qty or 0)
        for part, qty in (dispatched_consumption or {}).items()
        if _normalize_part_key(part)
    }

    for order in orders:
        if not _is_committed_status(order):
            continue
        try:
            order_id = int(order["id"])
        except (KeyError, TypeError, ValueError):
            continue

        for record in db.get_dispatch_records(order_id):
            if str(record.get("decision") or "None") == "Shortage":
                continue
            part = _normalize_part_key(record.get("part_number"))
            if not part:
                continue
            adjusted[part] = adjusted.get(part, 0.0) - float(record.get("needed_qty") or 0)

    return adjusted


def _addback_committed_orders_st_consumption(
    orders: list[dict],
    st_inventory_stock: dict[str, float],
) -> dict[str, float]:
    """Mirror 主檔加回，避免重產已發料單時 ST 被自己扣到 0 而誤判缺料。"""
    adjusted = {
        _normalize_part_key(part): float(qty or 0)
        for part, qty in (st_inventory_stock or {}).items()
        if _normalize_part_key(part)
    }

    committed_ids: list[int] = []
    for order in orders:
        if not _is_committed_status(order):
            continue
        try:
            committed_ids.append(int(order["id"]))
        except (KeyError, TypeError, ValueError):
            continue

    if not committed_ids:
        return adjusted

    order_supplements = db.get_order_supplements(committed_ids)
    for supplements in order_supplements.values():
        for raw_part, raw_qty in (supplements or {}).items():
            part = _normalize_part_key(raw_part)
            if not part:
                continue
            adjusted[part] = adjusted.get(part, 0.0) + float(raw_qty or 0)

    return adjusted


def _build_dispatch_result_by_order(
    orders: list[dict],
    bom_map: dict[str, list[dict]],
) -> dict[int, dict]:
    result_by_order: dict[int, dict] = {}
    if orders:
        stock, moq, dispatched_consumption = _load_shortage_inputs()
        adjusted_dispatched_consumption = _subtract_selected_committed_dispatch_records(
            orders,
            dispatched_consumption,
        )
        adjusted_st_inventory = _addback_committed_orders_st_consumption(
            orders,
            db.get_st_inventory_stock(),
        )
        calc_results = calc_run(orders, bom_map, stock, moq, adjusted_dispatched_consumption, adjusted_st_inventory)
        for result in calc_results:
            if result.get("order_id") is not None:
                result_by_order[int(result["order_id"])] = result

    return result_by_order


def _generate_dispatch_response(req: DispatchRequest, request: Request):
    bom_map = db.get_all_bom_components_by_model()
    if not bom_map:
        raise HTTPException(400, "請先上傳 BOM 檔案")

    requested_ids = list(dict.fromkeys(req.order_ids))
    if not requested_ids:
        raise HTTPException(400, "請先勾選要生成發料單的訂單")

    orders = _get_selected_orders(requested_ids)
    if not orders:
        raise HTTPException(400, "勾選的訂單沒有可生成的待處理內容")

    result_by_order = _build_dispatch_result_by_order(orders, bom_map)
    saved_supplements = db.get_order_supplements([int(order["id"]) for order in orders])
    decision_overrides = _normalize_decision_overrides(req.decisions)
    active_drafts_by_order = _get_active_reviewed_drafts_by_order([int(order["id"]) for order in orders])
    committed_main_supplements = _load_committed_main_supplements(orders, bom_map)
    st_inventory_stock = _addback_committed_orders_st_consumption(orders, db.get_st_inventory_stock())

    today = local_now().strftime("%Y/%m/%d")
    groups = []
    order_contexts = [
        context
        for order in orders
        for context in [_build_order_dispatch_context(
            order,
            result_by_order,
            saved_supplements,
            decision_overrides,
            bom_map,
            committed_main_supplements=committed_main_supplements,
            active_draft=active_drafts_by_order.get(int(order["id"])),
        )]
        if context
    ]

    for context in order_contexts:
        order = context["order"]
        items = []

        for part in context["candidate_parts"]:
            decision = context["decisions"].get(part, "None")
            has_explicit_supplement = part in context["stored_supplements"] or part in context["draft_supplements"]
            supplement_qty = float(
                context["stored_supplements"].get(part, context["draft_supplements"].get(part, 0)) or 0
            )
            shortage_item = context["shortages_by_part"].get(part)
            final_shortage = shortage_item or {}
            if not _should_render_dispatch_item(
                decision,
                supplement_qty,
                shortage_item,
                reviewed_draft=bool(context.get("reviewed_draft")),
                has_explicit_supplement=has_explicit_supplement,
            ):
                continue
            if context.get("use_main_supplements") and supplement_qty <= 0 and decision != "Shortage":
                continue

            description = (
                context["descriptions"].get(part)
                or final_shortage.get("description")
                or (shortage_item or {}).get("description", "")
            )
            display_part = final_shortage.get("part_number") or (shortage_item or {}).get("part_number") or part
            effective_supplement_qty = supplement_qty

            if decision == "Shortage":
                items.append({
                    "part": display_part,
                    "desc": description,
                    "qty": "缺",
                    "fill_color": None,
                    "is_shortage": True,
                })
                continue

            if effective_supplement_qty > 0 or has_explicit_supplement:
                fill_color = (
                    "FFFFC000"
                    if _should_highlight_dispatch_qty(display_part, effective_supplement_qty, final_shortage or shortage_item, st_inventory_stock)
                    else None
                )
                items.append({
                    "part": display_part,
                    "desc": description,
                    "qty": round(effective_supplement_qty),
                    "fill_color": fill_color,
                    "is_shortage": False,
                })
                continue

            suggested_qty = float(final_shortage.get("suggested_qty") or final_shortage.get("shortage_amount") or 0)
            if suggested_qty <= 0:
                continue

            fill_color = (
                "FFFFC000"
                if _should_highlight_dispatch_qty(display_part, suggested_qty, final_shortage or shortage_item, st_inventory_stock)
                else None
            )
            items.append({
                "part": display_part,
                "desc": description,
                "qty": round(suggested_qty),
                "fill_color": fill_color,
                "is_shortage": False,
            })

        if not items:
            continue

        ship_date = order.get("delivery_date") or order.get("ship_date") or ""
        groups.append({
            "batch_code": order.get("code") or str(order.get("id", "")),
            "po_number": str(order.get("po_number", "")),
            "model": order.get("model", ""),
            "date": ship_date.replace("-", "/") if ship_date else today,
            "items": items,
        })

    if not groups:
        raise HTTPException(400, "勾選的訂單目前沒有可生成的發料內容")

    tmp_dir = tempfile.mkdtemp()
    filename = build_generated_filename("發料單", ".xlsx")
    out_path = os.path.join(tmp_dir, filename)
    generate_dispatch_form(groups, out_path)
    db.log_activity("dispatch_generated", f"生成發料單，{len(groups)} 筆訂單")

    response = maybe_server_save_response(
        request,
        out_path,
        filename,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return response


@router.post("/dispatch/generate")
async def generate(req: DispatchRequest, request: Request):
    return await run_in_threadpool(lambda: _generate_dispatch_response(req, request))
