from __future__ import annotations

import io
import shutil
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Dict, List
from urllib.parse import quote
from uuid import uuid4

import openpyxl
from openpyxl.styles import PatternFill
from fastapi import APIRouter, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field

from .. import database as db
from ..config import BOM_DIR, cfg
from ..services.server_downloads import maybe_server_save_response
from ..models import BomEditorSaveRequest
from ..services.bom_editor import (
    apply_bom_editor_changes,
    backup_bom_file,
    build_bom_storage_payload,
    normalize_uploaded_bom_layout,
    normalize_bom_record_to_editable,
    parse_bom_for_storage,
    prepare_uploaded_bom_file,
    validate_uploaded_bom_layout,
)
from ..services.bom_parser import read_formula_needed_qty_cache
from ..services.bom_quantity import (
    calculate_effective_needed_qty,
    coerce_qty,
    format_excel_qty,
    get_component_effective_needed_qty,
)
from ..services.bom_revision import (
    delete_bom_revision_files,
    ensure_bom_revision_history,
    snapshot_bom_revision,
)
from ..services.download_names import append_minute_timestamp, build_bom_dispatch_filename, build_generated_filename
from ..services.main_reader import find_legacy_snapshot_stock_fixes, read_stock
from ..services.order_supplements import build_order_supplement_allocations, merge_order_supplement_allocations
from ..services.shortage_rules import (
    calculate_current_order_shortage_amount,
    calculate_shortage_amount,
    is_order_scoped_shortage_part,
    summarize_requested_supply,
)
from ..services.workbook_recalc import save_workbook_bytes_with_recalc
from ..services.workbook_recalc import cell_has_formula
router = APIRouter()

ORANGE_FILL = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")
_BOM_STORAGE_SUFFIXES = (".xlsx", ".xls", ".xlsm")


def _get_required_bom(bom_id: str) -> dict:
    bom = db.get_bom_file(bom_id)
    if not bom:
        raise HTTPException(404, "找不到 BOM")
    return bom


def _build_revision_download_name(revision: dict) -> str:
    filename = str(revision.get("filename") or "bom.xlsx")
    stem = Path(filename).stem or "bom"
    suffix = Path(filename).suffix or ".xlsx"
    revision_number = int(revision.get("revision_number") or 0)
    return append_minute_timestamp(f"{stem}_v{revision_number:03d}{suffix}")


def _normalize_bom_identity_value(value: object) -> str:
    return " ".join(str(value or "").strip().upper().split())


def _build_bom_identity(source) -> dict[str, str]:
    getter = source.get if isinstance(source, dict) else lambda key, default="": getattr(source, key, default)
    filename = str(getter("filename", "") or "").strip()
    source_filename = str(getter("source_filename", "") or filename).strip()
    model = str(getter("model", "") or "").strip()
    group_model = str(getter("group_model", "") or model).strip()
    pcb = str(getter("pcb", "") or "").strip()
    return {
        "group_model": _normalize_bom_identity_value(group_model),
        "model": _normalize_bom_identity_value(model),
        "pcb": _normalize_bom_identity_value(pcb),
        "source_filename": _normalize_bom_identity_value(Path(source_filename).name),
        "filename": _normalize_bom_identity_value(Path(filename).name),
    }


def _is_same_bom_upload_target(uploaded, existing: dict) -> bool:
    uploaded_identity = _build_bom_identity(uploaded)
    existing_identity = _build_bom_identity(existing)

    if uploaded_identity["group_model"] and existing_identity["group_model"]:
        if uploaded_identity["group_model"] != existing_identity["group_model"]:
            return False

    shared_specific_identity = False
    for key in ("model", "pcb"):
        uploaded_value = uploaded_identity[key]
        existing_value = existing_identity[key]
        if uploaded_value and existing_value:
            shared_specific_identity = True
            if uploaded_value != existing_value:
                return False

    if shared_specific_identity:
        return True

    for key in ("source_filename", "filename"):
        uploaded_value = uploaded_identity[key]
        existing_value = existing_identity[key]
        if uploaded_value and existing_value and uploaded_value == existing_value:
            return True

    return False


def _find_matching_uploaded_boms(parsed) -> list[dict]:
    return [bom for bom in db.get_bom_files() if _is_same_bom_upload_target(parsed, bom)]


def _pick_overwrite_keeper(matches: list[dict], uploaded) -> tuple[dict | None, list[dict]]:
    if not matches:
        return None, []

    uploaded_identity = _build_bom_identity(uploaded)

    def _score(bom: dict) -> tuple:
        identity = _build_bom_identity(bom)
        return (
            0 if uploaded_identity["source_filename"] and identity["source_filename"] == uploaded_identity["source_filename"] else 1,
            0 if uploaded_identity["filename"] and identity["filename"] == uploaded_identity["filename"] else 1,
            0 if uploaded_identity["pcb"] and identity["pcb"] == uploaded_identity["pcb"] else 1,
            0 if uploaded_identity["model"] and identity["model"] == uploaded_identity["model"] else 1,
            int(bom.get("sort_order") or 0),
            str(bom.get("uploaded_at") or ""),
            str(bom.get("filename") or ""),
        )

    ordered = sorted(matches, key=_score)
    return ordered[0], ordered[1:]


def _cleanup_bom_storage_variants(bom_id: str, keep_path: str = ""):
    keep = str(Path(keep_path)) if keep_path else ""
    for suffix in _BOM_STORAGE_SUFFIXES:
        candidate = BOM_DIR / f"{bom_id}{suffix}"
        if keep and str(candidate) == keep:
            continue
        candidate.unlink(missing_ok=True)


def _copy_uploaded_bom_to_target_id(stored: dict, target_bom_id: str) -> dict[str, object]:
    source_path = Path(str(stored.get("filepath") or ""))
    if not source_path.exists():
        raise FileNotFoundError("上傳後的 BOM 暫存檔不存在")

    target_path = BOM_DIR / f"{target_bom_id}{source_path.suffix.lower()}"
    target_path.parent.mkdir(parents=True, exist_ok=True)
    _cleanup_bom_storage_variants(target_bom_id, keep_path=str(target_path))
    shutil.copy2(source_path, target_path)
    return {
        **stored,
        "filepath": str(target_path),
    }


def _component_cache_key(component) -> tuple[str, int, str] | None:
    row = getattr(component, "source_row", None)
    if row is None:
        return None
    part = str(getattr(component, "part_number", "") or "").strip().upper()
    if not part:
        return None
    return (str(getattr(component, "source_sheet", "") or ""), int(row), part)


def _restore_pre_normalize_needed_quantities(parsed, pre_normalize_parsed):
    if not pre_normalize_parsed:
        return parsed

    cached_needed: dict[tuple[str, int, str], float] = {}
    if isinstance(pre_normalize_parsed, dict):
        cached_needed = {
            key: coerce_qty(value)
            for key, value in pre_normalize_parsed.items()
            if coerce_qty(value) > 0
        }
    else:
        for component in pre_normalize_parsed.components:
            key = _component_cache_key(component)
            if key is None:
                continue
            value = coerce_qty(component.needed_qty)
            if value > 0:
                cached_needed[key] = value

    if not cached_needed:
        return parsed

    for component in parsed.components:
        key = _component_cache_key(component)
        if key in cached_needed:
            component.needed_qty = cached_needed[key]
    return parsed


def _delete_bom_record_with_files(bom: dict):
    bom_id = str(bom.get("id") or "").strip()
    if not bom_id:
        return

    filepath = Path(str(bom.get("filepath") or ""))
    filepath.unlink(missing_ok=True)
    _cleanup_bom_storage_variants(bom_id)
    delete_bom_revision_files(bom_id)
    db.delete_bom_file(bom_id)


def _parse_bom_record_for_storage(bom: dict):
    return parse_bom_for_storage(
        path=bom["filepath"],
        bom_id=bom["id"],
        filename=bom["filename"],
        uploaded_at=bom["uploaded_at"],
        group_model=bom.get("group_model", ""),
        source_filename=bom.get("source_filename", ""),
        source_format=bom.get("source_format", ""),
        is_converted=bool(bom.get("is_converted")),
    )


def _sync_bom_components_from_file(bom: dict) -> dict:
    parsed = _parse_bom_record_for_storage(bom)
    payload = build_bom_storage_payload(parsed)
    db.save_bom_file(payload)
    return payload


def _ensure_editable_bom_record(bom: dict) -> dict:
    normalized = normalize_bom_record_to_editable(bom)
    if normalized == bom:
        return normalized

    _sync_bom_components_from_file(normalized)
    db.log_activity("bom_convert", f"{bom.get('filename') or bom['id']} 已轉為可編輯 xlsx")
    return db.get_bom_file(bom["id"]) or normalized


@router.post("/bom/upload")
async def upload_bom_files(files: List[UploadFile] = File(...), group_model: str = Form("")):
    group_model = group_model.strip()
    saved = []
    errors = []

    for uf in files:
        ext = Path(uf.filename or "").suffix.lower()
        if ext not in {".xlsx", ".xls", ".xlsm"}:
            errors.append(f"{uf.filename}: 僅支援 xlsx / xls / xlsm")
            continue

        bom_id = uuid4().hex
        stored = None
        keep_uploaded_file_on_error = False
        try:
            stored = prepare_uploaded_bom_file(
                bom_id=bom_id,
                upload_name=uf.filename or f"{bom_id}{ext}",
                content=await uf.read(),
            )
            uploaded_at = datetime.now().isoformat()
            pre_normalize_needed_cache = {}
            try:
                pre_normalize_needed_cache = read_formula_needed_qty_cache(str(stored["filepath"]))
            except Exception:
                pre_normalize_needed_cache = {}
            auto_fixes = normalize_uploaded_bom_layout(str(stored["filepath"]))
            layout_errors = validate_uploaded_bom_layout(str(stored["filepath"]))
            if layout_errors:
                detail = "；".join(layout_errors[:6])
                if len(layout_errors) > 6:
                    detail += "；其餘列請打開原檔確認"
                raise ValueError(f"副檔欄位檢查失敗（G/H 應空白，I/J 應為公式）：{detail}")
            parsed = parse_bom_for_storage(
                path=str(stored["filepath"]),
                bom_id=bom_id,
                filename=str(stored["filename"]),
                uploaded_at=uploaded_at,
                group_model=group_model,
                source_filename=str(stored["source_filename"]),
                source_format=str(stored["source_format"]),
                is_converted=bool(stored["is_converted"]),
            )
            parsed = _restore_pre_normalize_needed_quantities(parsed, pre_normalize_needed_cache)
            overwrite_matches = _find_matching_uploaded_boms(parsed)
            overwrite_target, duplicate_targets = _pick_overwrite_keeper(overwrite_matches, parsed)
            replaced_existing = overwrite_target is not None
            removed_duplicates = len(duplicate_targets)

            payload = build_bom_storage_payload(parsed)
            snapshot_action = "upload"
            snapshot_note = "上傳 BOM"

            if overwrite_target:
                ensure_bom_revision_history(overwrite_target)

                payload = {
                    **payload,
                    "id": overwrite_target["id"],
                }
                db.save_bom_file(payload)
                keep_uploaded_file_on_error = True

                stored = _copy_uploaded_bom_to_target_id(stored, str(overwrite_target["id"]))
                payload = {
                    **payload,
                    "filepath": str(stored["filepath"]),
                }
                db.save_bom_file(payload)

                old_path = Path(str(overwrite_target.get("filepath") or ""))
                new_path = Path(str(stored["filepath"] or ""))
                if old_path != new_path:
                    old_path.unlink(missing_ok=True)

                temp_path = Path(str(parsed.path or ""))
                if temp_path != new_path:
                    temp_path.unlink(missing_ok=True)

                snapshot_action = "overwrite"
                snapshot_note = "重新上傳覆蓋 BOM"

                for duplicate_bom in duplicate_targets:
                    _delete_bom_record_with_files(duplicate_bom)
            else:
                db.save_bom_file(payload)

            snapshot_bom_revision(payload, snapshot_action, snapshot_note)
            saved.append({
                "id": payload["id"],
                "filename": payload["filename"],
                "source_filename": payload.get("source_filename") or payload["filename"],
                "source_format": payload.get("source_format") or Path(str(payload["filename"])).suffix.lower(),
                "is_converted": bool(payload.get("is_converted")),
                "po_number": payload.get("po_number", 0),
                "model": payload.get("model", ""),
                "pcb": payload.get("pcb", ""),
                "order_qty": payload.get("order_qty", 0),
                "components": len(payload.get("components", [])),
                "auto_fixes": auto_fixes,
                "replaced_existing": replaced_existing,
                "removed_duplicates": removed_duplicates,
            })
        except Exception as exc:
            if stored and stored.get("filepath") and not keep_uploaded_file_on_error:
                Path(str(stored["filepath"])).unlink(missing_ok=True)
            if not keep_uploaded_file_on_error:
                for suffix in _BOM_STORAGE_SUFFIXES:
                    (BOM_DIR / f"{bom_id}{suffix}").unlink(missing_ok=True)
            errors.append(f"{uf.filename}: {exc}")
            continue

    if saved:
        converted_count = sum(1 for item in saved if item["is_converted"])
        replaced_count = sum(1 for item in saved if item.get("replaced_existing"))
        deduped_count = sum(int(item.get("removed_duplicates") or 0) for item in saved)
        detail = f"上傳 {len(saved)} 份 BOM"
        if group_model:
            detail += f"（group_model: {group_model}）"
        if converted_count:
            detail += f"，其中 {converted_count} 份 xls 已轉為 xlsx"
        if replaced_count:
            detail += f"，覆蓋 {replaced_count} 份舊版"
        if deduped_count:
            detail += f"，清理 {deduped_count} 份重複舊檔"
        db.log_activity("bom_upload", detail)

    return {"saved": saved, "errors": errors}


@router.get("/bom/list")
async def list_bom_files():
    bom_files = db.get_bom_files()

    groups: dict[str, list] = {}
    for bom in bom_files:
        model_key = bom["group_model"] or bom["model"] or "未指定機種"
        groups.setdefault(model_key, [])
        components = db.get_bom_components(bom["id"])
        groups[model_key].append({
            "id": bom["id"],
            "filename": bom["filename"],
            "source_filename": bom.get("source_filename") or bom["filename"],
            "source_format": bom.get("source_format") or Path(bom["filename"]).suffix.lower(),
            "is_converted": bool(bom.get("is_converted")),
            "po_number": bom["po_number"],
            "model": bom["model"],
            "pcb": bom["pcb"],
            "order_qty": bom["order_qty"],
            "components": len(components),
            "uploaded_at": bom["uploaded_at"],
        })

    return {"groups": [{"model": model, "items": items} for model, items in groups.items()]}


class BomOrderGroupRequest(BaseModel):
    model: str = ""
    item_ids: List[str] = Field(default_factory=list)


class BomReorderRequest(BaseModel):
    groups: List[BomOrderGroupRequest] = Field(default_factory=list)


@router.post("/bom/reorder")
async def reorder_bom_files(req: BomReorderRequest):
    if not req.groups:
        raise HTTPException(400, "請提供排序資料")

    updated = db.save_bom_order([group.dict() for group in req.groups])
    db.log_activity("bom_reorder", f"BOM 排序已更新，{updated} 筆")
    return {"ok": True, "updated": updated}


@router.delete("/bom/{bom_id}")
async def delete_bom(bom_id: str):
    bom = _get_required_bom(bom_id)
    Path(bom["filepath"]).unlink(missing_ok=True)
    delete_bom_revision_files(bom_id)
    db.delete_bom_file(bom_id)
    db.log_activity("bom_delete", f"刪除 BOM {bom['filename']}")
    return {"ok": True}


@router.get("/bom/data")
async def get_bom_data():
    bom_map = db.get_all_bom_components_by_model()
    return {model: {"model": model, "components": components} for model, components in bom_map.items()}


@router.get("/bom/{bom_id}/file")
async def get_bom_file(bom_id: str, request: Request):
    bom = _ensure_editable_bom_record(_get_required_bom(bom_id))
    file_path = Path(bom["filepath"])
    if not file_path.exists():
        raise HTTPException(404, "BOM 檔案不存在")
    return maybe_server_save_response(
        request,
        str(file_path),
        append_minute_timestamp(bom["filename"] or file_path.name),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


class BomLookupRequest(BaseModel):
    models: List[str]


@router.post("/bom/lookup")
async def lookup_bom_files(req: BomLookupRequest):
    bom_files = [_ensure_editable_bom_record(bom) for bom in db.get_bom_files_by_models(req.models)]
    return {
        "files": [
            {
                "id": bom["id"],
                "filename": bom["filename"],
                "model": bom["model"],
                "group_model": bom["group_model"],
            }
            for bom in bom_files
            if Path(bom["filepath"]).exists()
        ]
    }


class BomDownloadRequest(BaseModel):
    models: List[str]


@router.post("/bom/download")
async def download_bom_files(req: BomDownloadRequest, request: Request):
    if not req.models:
        raise HTTPException(400, "請提供要下載的機種")

    bom_files = db.get_bom_files_by_models(req.models)
    if not bom_files:
        raise HTTPException(404, "找不到對應的 BOM")

    valid = [(bom, Path(bom["filepath"])) for bom in bom_files if Path(bom["filepath"]).exists()]
    if not valid:
        raise HTTPException(404, "BOM 檔案不存在")

    if len(valid) == 1:
        bom, file_path = valid[0]
        return maybe_server_save_response(
            request,
            str(file_path),
            append_minute_timestamp(bom["filename"] or file_path.name),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        seen_names: dict[str, int] = {}
        for bom, file_path in valid:
            name = append_minute_timestamp(bom["filename"] or file_path.name)
            seen_names[name] = seen_names.get(name, -1) + 1
            if seen_names[name] > 0:
                stem, suffix = Path(name).stem, Path(name).suffix
                name = f"{stem}_{seen_names[name]}{suffix}"
            zf.write(str(file_path), name)
    zip_buffer.seek(0)

    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(build_generated_filename('BOM', '.zip'))}"},
    )


@router.get("/bom/{bom_id}/editor")
async def get_bom_editor(bom_id: str):
    bom = _ensure_editable_bom_record(_get_required_bom(bom_id))
    payload = _sync_bom_components_from_file(bom)
    payload["component_count"] = len(payload.get("components", []))
    return payload


@router.get("/bom/{bom_id}/revisions")
async def list_bom_revisions(bom_id: str):
    bom = _ensure_editable_bom_record(_get_required_bom(bom_id))
    revisions = ensure_bom_revision_history(bom)
    return {
        "bom": {
            "id": bom["id"],
            "filename": bom["filename"],
            "source_filename": bom.get("source_filename") or bom["filename"],
        },
        "revisions": revisions,
    }


@router.get("/bom/{bom_id}/revisions/{revision_id}/file")
async def download_bom_revision(bom_id: str, revision_id: int, request: Request):
    revision = db.get_bom_revision(int(revision_id))
    if not revision or str(revision.get("bom_file_id")) != str(bom_id):
        raise HTTPException(404, "找不到 BOM 歷史版本")

    file_path = Path(str(revision.get("filepath") or ""))
    if not file_path.exists():
        raise HTTPException(404, "BOM 歷史檔案不存在")

    return maybe_server_save_response(
        request,
        str(file_path),
        _build_revision_download_name(revision),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.put("/bom/{bom_id}/editor")
async def save_bom_editor(bom_id: str, req: BomEditorSaveRequest):
    bom = _ensure_editable_bom_record(_get_required_bom(bom_id))
    file_path = Path(bom["filepath"])
    if not file_path.exists():
        raise HTTPException(404, "BOM 檔案不存在")

    ensure_bom_revision_history(bom)
    backup_path = backup_bom_file(str(file_path))
    try:
        apply_bom_editor_changes(str(file_path), req)
        parsed = parse_bom_for_storage(
            path=str(file_path),
            bom_id=bom["id"],
            filename=bom["filename"],
            uploaded_at=bom["uploaded_at"],
            group_model=req.group_model.strip(),
            source_filename=bom.get("source_filename", ""),
            source_format=bom.get("source_format", ""),
            is_converted=bool(bom.get("is_converted")),
        )
        payload = build_bom_storage_payload(parsed)
        db.save_bom_file(payload)
        snapshot_bom_revision(payload, "edit", "編輯後儲存")
    except Exception as exc:
        shutil.copy2(backup_path, file_path)
        raise HTTPException(400, f"儲存 BOM 失敗：{exc}")

    db.log_activity(
        "bom_edit",
        f"更新 BOM {parsed.filename}，{len(parsed.components)} 列元件已同步回正式檔",
    )
    return {
        "ok": True,
        "filename": parsed.filename,
        "components": len(parsed.components),
        "backup_path": backup_path,
    }


class BomDispatchDownloadRequest(BaseModel):
    bom_ids: List[str]
    order_ids: List[int] = Field(default_factory=list)
    supplements: Dict[str, float]
    order_supplements: Dict[str, Dict[str, float]] = Field(default_factory=dict)
    header_overrides: Dict[str, Dict[str, str]] = Field(default_factory=dict)
    carry_overs: Dict[str, Dict[str, float]] = Field(default_factory=dict)


def _normalize_lookup_key(value) -> str:
    return str(value or "").strip().upper()


def _get_bom_match_keys(bom: dict) -> set[str]:
    raw = str(bom.get("group_model") or bom.get("model") or "")
    return {_normalize_lookup_key(item) for item in raw.split(",") if _normalize_lookup_key(item)}


def _build_dispatch_running_stock() -> dict[str, float]:
    main_path = str(db.get_setting("main_file_path") or "").strip()
    if main_path and Path(main_path).exists():
        running = {
            _normalize_lookup_key(part): float(qty or 0)
            for part, qty in read_stock(main_path).items()
            if _normalize_lookup_key(part)
        }
    else:
        snapshot = db.get_snapshot()
        if snapshot and main_path and Path(main_path).exists():
            fixes = find_legacy_snapshot_stock_fixes(main_path, snapshot)
            if fixes:
                db.update_snapshot_stock(fixes)
                for part, qty in fixes.items():
                    if part in snapshot:
                        snapshot[part]["stock_qty"] = qty
        running = {
            _normalize_lookup_key(part): float((values or {}).get("stock_qty") or 0)
            for part, values in snapshot.items()
            if _normalize_lookup_key(part)
        }

    return running


def _build_order_based_export_values(
    target_boms: list[dict],
    order_ids: list[int],
    supplements: dict[str, float],
    order_supplements: dict[int, dict[str, float]] | None = None,
) -> tuple[dict[str, dict[str, float]], dict[str, dict[str, float]], dict[str, set[str]], dict[str, float], dict[str, dict]]:
    if not order_ids or not target_boms:
        return {}, {}, {}, {}, {}

    running = _build_dispatch_running_stock()
    st_inventory_stock = db.get_st_inventory_stock()
    components_by_bom = {str(bom["id"]): db.get_bom_components(str(bom["id"])) for bom in target_boms}
    carry_overs: dict[str, dict[str, float]] = {}
    supplement_allocations: dict[str, dict[str, float]] = {}
    purchase_highlights: dict[str, set[str]] = {}
    order_qty_by_bom: dict[str, float] = {}
    order_info_by_bom: dict[str, dict] = {}
    remaining_supplements = {
        _normalize_lookup_key(part): float(qty or 0)
        for part, qty in (supplements or {}).items()
        if _normalize_lookup_key(part) and float(qty or 0) > 0
    }
    normalized_order_supplements = {
        int(order_id): {
            _normalize_lookup_key(part): float(qty or 0)
            for part, qty in (part_map or {}).items()
            if _normalize_lookup_key(part) and float(qty or 0) > 0
        }
        for order_id, part_map in (order_supplements or {}).items()
    }

    for order_id in order_ids:
        order = db.get_order(int(order_id))
        if not order:
            continue

        order_model = _normalize_lookup_key(order.get("model"))
        if not order_model:
            continue
        remaining_order_supplements = dict(normalized_order_supplements.get(int(order_id), {}))

        matched_boms = [bom for bom in target_boms if order_model in _get_bom_match_keys(bom)]
        for bom in matched_boms:
            bom_id = str(bom["id"])
            if bom_id in carry_overs:
                continue
            target_order_qty = coerce_qty(order.get("order_qty"))

            part_map: dict[str, float] = {}
            supplement_map: dict[str, float] = {}
            purchase_parts: set[str] = set()
            part_totals: dict[str, dict[str, float]] = {}
            for component in components_by_bom.get(bom_id, []):
                needed_qty = get_component_effective_needed_qty(
                    component,
                    schedule_order_qty=order.get("order_qty"),
                    bom_order_qty=bom.get("order_qty"),
                )
                if component.get("is_dash") or needed_qty <= 0:
                    continue

                part = _normalize_lookup_key(component.get("part_number"))
                if not part:
                    continue

                if part not in part_map:
                    part_map[part] = float(running.get(part, 0))

                summary = part_totals.setdefault(part, {"needed_qty": 0.0, "prev_qty_cs": 0.0})
                summary["needed_qty"] += needed_qty
                summary["prev_qty_cs"] += float(component.get("prev_qty_cs") or 0)

            for part, totals in part_totals.items():
                current_stock = float(running.get(part, 0))
                st_stock_qty = float(st_inventory_stock.get(part, 0.0) or 0.0)
                available_before = (
                    current_stock
                    + float(totals.get("prev_qty_cs") or 0)
                )
                ending_without_supplement = (
                    available_before
                    - float(totals.get("needed_qty") or 0)
                )
                shortage_without_supplement = calculate_shortage_amount(part, ending_without_supplement)
                current_order_shortage = calculate_current_order_shortage_amount(
                    part,
                    available_before,
                    float(totals.get("needed_qty") or 0),
                )

                supplement_qty = 0.0
                available_supplement_qty = float(remaining_order_supplements.get(part, 0))
                if available_supplement_qty <= 0:
                    available_supplement_qty = float(remaining_supplements.get(part, 0))
                if shortage_without_supplement > 0 and available_supplement_qty > 0:
                    supplement_qty = available_supplement_qty
                    if is_order_scoped_shortage_part(part):
                        supplement_qty = min(supplement_qty, current_order_shortage)
                    supplement_map[part] = supplement_qty
                    if part in remaining_order_supplements:
                        remaining_order_supplements[part] = max(0.0, remaining_order_supplements.get(part, 0) - supplement_qty)
                    if part in remaining_supplements:
                        remaining_supplements[part] = max(0.0, remaining_supplements.get(part, 0) - supplement_qty)
                    if bool(summarize_requested_supply(supplement_qty, st_stock_qty)["needs_purchase"]):
                        purchase_parts.add(part)

                running[part] = ending_without_supplement + supplement_qty

            carry_overs[bom_id] = part_map
            supplement_allocations[bom_id] = supplement_map
            purchase_highlights[bom_id] = purchase_parts
            order_qty_by_bom[bom_id] = target_order_qty
            order_info_by_bom[bom_id] = {"po_number": order.get("po_number"), "model": order.get("model")}

    return carry_overs, supplement_allocations, purchase_highlights, order_qty_by_bom, order_info_by_bom


def _build_direct_purchase_highlights(supplements: dict[str, float]) -> set[str]:
    st_inventory_stock = {
        _normalize_lookup_key(part): float(qty or 0)
        for part, qty in db.get_st_inventory_stock().items()
        if _normalize_lookup_key(part)
    }
    return {
        part
        for part, qty in supplements.items()
        if bool(summarize_requested_supply(qty, st_inventory_stock.get(part, 0.0))["needs_purchase"])
    }


def _normalize_order_supplement_map(payload: dict | None) -> dict[int, dict[str, float]]:
    normalized: dict[int, dict[str, float]] = {}
    for raw_order_id, part_map in (payload or {}).items():
        try:
            order_id = int(raw_order_id)
        except (TypeError, ValueError):
            continue
        normalized_parts: dict[str, float] = {}
        for part, qty in (part_map or {}).items():
            key = _normalize_lookup_key(part)
            try:
                amount = float(qty or 0)
            except (TypeError, ValueError):
                amount = 0.0
            if key and amount > 0:
                normalized_parts[key] = amount
        normalized[order_id] = normalized_parts
    return normalized


def _resolve_cell_for_write(ws, row_idx: int, col_idx: int):
    cell = ws.cell(row=row_idx, column=col_idx)
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell


def _set_cell_value(ws, row_idx: int, col_idx: int, value):
    _resolve_cell_for_write(ws, row_idx, col_idx).value = value


def _format_bom_po_value(existing_value, po_number):
    po_text = str(po_number or "").strip()
    if not po_text:
        return existing_value

    existing_text = str(existing_value or "").strip()
    if any(token in existing_text for token in ("製單號碼", "M/O", "PO")):
        if ":" in existing_text:
            prefix = existing_text.split(":", 1)[0]
            return f"{prefix}:{po_text}"
        return f"{existing_text}{po_text}"
    return po_number


def _read_ws_bom_order_qty(ws) -> float:
    order_qty_col = cfg("excel.bom_order_qty_col", 10) + 1
    order_qty = coerce_qty(_resolve_cell_for_write(ws, 1, order_qty_col).value)
    if order_qty > 0:
        return order_qty
    return coerce_qty(ws.cell(row=2, column=7).value)


def _apply_target_order_qty_to_ws(ws, target_order_qty: float | None, source_order_qty: float | None = None):
    effective_order_qty = coerce_qty(target_order_qty)
    if effective_order_qty <= 0:
        return

    base_order_qty = coerce_qty(source_order_qty) or _read_ws_bom_order_qty(ws)
    part_col = cfg("excel.bom_part_col", 2) + 1
    qty_col = cfg("excel.bom_qty_per_board", 1) + 1
    needed_col = cfg("excel.bom_needed_col", 5) + 1
    g_col = cfg("excel.bom_g_col", 6) + 1
    h_col = cfg("excel.bom_h_col", 7) + 1
    data_start = cfg("excel.bom_data_start_row", 5)
    dash_markers = {"-", "x", "X", "n", "N", "n/a", "N/A", "na", "NA", "?"}

    for row_idx in range(data_start, ws.max_row + 1):
        if not _normalize_lookup_key(ws.cell(row=row_idx, column=part_col).value):
            continue
        g_text = str(ws.cell(row=row_idx, column=g_col).value or "").strip()
        h_text = str(ws.cell(row=row_idx, column=h_col).value or "").strip()
        if g_text in dash_markers or h_text in dash_markers:
            continue
        target_needed_cell = _resolve_cell_for_write(ws, row_idx, needed_col)
        if cell_has_formula(target_needed_cell):
            continue
        needed_qty = calculate_effective_needed_qty(
            needed_qty=target_needed_cell.value,
            qty_per_board=ws.cell(row=row_idx, column=qty_col).value,
            schedule_order_qty=effective_order_qty,
            bom_order_qty=base_order_qty,
        )
        _set_cell_value(ws, row_idx, needed_col, format_excel_qty(needed_qty))


def _write_bom_header_values(ws, po_number, order_qty: float | None = None):
    po_col = cfg("excel.bom_po_col", 7) + 1
    po_cell = _resolve_cell_for_write(ws, 1, po_col)
    po_cell.value = _format_bom_po_value(po_cell.value, po_number)
    if coerce_qty(order_qty) > 0:
        order_qty_col = cfg("excel.bom_order_qty_col", 10) + 1
        _resolve_cell_for_write(ws, 1, order_qty_col).value = format_excel_qty(order_qty)
        fallback_row2_qty = ws.cell(row=2, column=7).value
        if fallback_row2_qty in (None, "") or isinstance(fallback_row2_qty, (int, float)):
            ws.cell(row=2, column=7).value = format_excel_qty(order_qty)


def _write_dispatch_values_to_ws(
    ws,
    supplements: dict[str, float],
    carry_overs: dict[str, float],
    purchase_parts: set[str] | None = None,
    target_order_qty: float | None = None,
    source_order_qty: float | None = None,
):
    _apply_target_order_qty_to_ws(ws, target_order_qty, source_order_qty=source_order_qty)
    part_col = cfg("excel.bom_part_col", 2) + 1
    g_col = cfg("excel.bom_g_col", 6) + 1
    h_col = cfg("excel.bom_h_col", 7) + 1
    data_start = cfg("excel.bom_data_start_row", 5)

    written = 0
    dash_markers = {"-", "x", "X", "n", "N", "n/a", "N/A", "na", "NA", "?"}
    supplemented_parts: set[str] = set()
    for row_idx in range(data_start, ws.max_row + 1):
        part = str(ws.cell(row=row_idx, column=part_col).value or "").strip().upper()
        if not part:
            continue

        supplement_qty = 0
        if part not in supplemented_parts and part in supplements:
            supplement_qty = supplements[part]
            supplemented_parts.add(part)

        g_text = str(ws.cell(row=row_idx, column=g_col).value or "").strip()
        h_text = str(ws.cell(row=row_idx, column=h_col).value or "").strip()
        if g_text in dash_markers or h_text in dash_markers:
            continue

        carry_over_qty = carry_overs.get(part)
        if carry_over_qty is not None:
            _set_cell_value(ws, row_idx, g_col, carry_over_qty)
        elif g_text == "":
            _set_cell_value(ws, row_idx, g_col, 0)
        h_cell = _resolve_cell_for_write(ws, row_idx, h_col)
        h_cell.value = supplement_qty
        if supplement_qty and part in (purchase_parts or set()):
            h_cell.fill = ORANGE_FILL
        written += 1

    return written


@router.post("/bom/dispatch-download")
async def dispatch_download_bom(req: BomDispatchDownloadRequest):
    if not req.bom_ids:
        raise HTTPException(400, "請提供 BOM ID")

    target_ids = set(req.bom_ids)
    target_boms = [_ensure_editable_bom_record(bom) for bom in db.get_bom_files() if bom["id"] in target_ids]
    if not target_boms:
        raise HTTPException(404, "找不到指定的 BOM")

    supplements = {part.strip().upper(): qty for part, qty in req.supplements.items()}
    order_supplements = _normalize_order_supplement_map(req.order_supplements)
    effective_order_supplements = merge_order_supplement_allocations(
        req.order_ids,
        supplements,
        order_supplements,
        allocator=build_order_supplement_allocations,
    )
    direct_purchase_highlights = _build_direct_purchase_highlights(supplements)
    computed_carry_overs, computed_supplements, computed_purchase_highlights, computed_order_qtys, computed_order_info = _build_order_based_export_values(
        target_boms,
        req.order_ids,
        supplements,
        order_supplements=effective_order_supplements,
    )
    if req.order_ids:
        db.replace_order_supplements(
            req.order_ids,
            effective_order_supplements,
        )
    output_files: list[tuple[str, io.BytesIO]] = []

    for bom in target_boms:
        src = Path(bom["filepath"])
        if not src.exists():
            continue

        ext = src.suffix.lower()
        _order_info = computed_order_info.get(bom["id"], {})
        _po = _order_info.get("po_number") or bom.get("po_number")
        _raw_model = str(bom.get("model") or bom.get("group_model") or "")
        _model = _raw_model.split(",")[0].strip()
        output_name = build_bom_dispatch_filename(_po, _model, ext)
        wb = openpyxl.load_workbook(str(src), keep_vba=(ext == ".xlsm"))

        override = req.header_overrides.get(bom["id"], {})
        override_po = str(override.get("po_number", "") or "").strip()
        po_number = override_po or str(bom.get("po_number") or "").strip()
        if req.order_ids:
            carry_over_source = computed_carry_overs.get(bom["id"], {})
            supplement_source = computed_supplements.get(bom["id"], {})
            purchase_parts = computed_purchase_highlights.get(bom["id"], set())
            order_qty_source = computed_order_qtys.get(bom["id"], 0.0)
        else:
            carry_over_source = req.carry_overs.get(bom["id"], {})
            supplement_source = supplements
            purchase_parts = direct_purchase_highlights
            order_qty_source = 0.0
        carry_overs = {
            str(part).strip().upper(): qty
            for part, qty in carry_over_source.items()
            if str(part).strip()
        }
        per_bom_supplements = {
            str(part).strip().upper(): qty
            for part, qty in supplement_source.items()
            if str(part).strip()
        }
        _write_dispatch_values_to_ws(
            wb.active,
            per_bom_supplements,
            carry_overs,
            purchase_parts={str(part).strip().upper() for part in purchase_parts if str(part).strip()},
            target_order_qty=order_qty_source,
            source_order_qty=bom.get("order_qty"),
        )
        _write_bom_header_values(wb.active, po_number, order_qty_source)
        buffer = save_workbook_bytes_with_recalc(wb, output_name)
        wb.close()
        buffer.seek(0)
        output_files.append((output_name, buffer))

    if not output_files:
        raise HTTPException(404, "找不到可下載的 BOM 檔案")

    if len(output_files) == 1:
        name, buffer = output_files[0]
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(name)}"},
        )

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, buffer in output_files:
            zf.writestr(name, buffer.read())
    zip_buffer.seek(0)

    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(build_generated_filename('補料BOM', '.zip'))}"},
    )
