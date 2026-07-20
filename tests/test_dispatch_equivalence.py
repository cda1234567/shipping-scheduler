from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from app.services import merge_to_main as merge_service
from app.services.merge_to_main import (
    merge_order_batches_to_main,
    merge_row_to_main,
    preview_order_batches,
    restore_main_from_backup_reference,
    validate_dispatch_backup_reference,
)


class DispatchEquivalenceTests(unittest.TestCase):
    def _build_main_workbook(self, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "2026"
        ws.append(["料號", "廠商", "MOQ", "期初", "盤點", "", "M/O", "結存"])
        ws.append(["PART-A", "Vendor", 10, None, None, None, 0, 20])
        ws.append(["PART-B", "Vendor", 10, None, None, None, 0, 50])
        wb.save(path)
        wb.close()

    def _single_group(self, *, part_number: str, needed_qty: float, prev_qty_cs: float) -> list[dict]:
        return [{
            "batch_code": "1-1",
            "po_number": "4500000001",
            "bom_model": "MODEL-A",
            "components": [{
                "part_number": part_number,
                "description": f"{part_number} desc",
                "is_dash": False,
                "needed_qty": needed_qty,
                "prev_qty_cs": prev_qty_cs,
            }],
        }]

    def test_preview_and_commit_keep_same_remaining_shortage(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "main.xlsx"
            backup_dir = Path(temp_dir) / "backups"
            self._build_main_workbook(path)

            groups = self._single_group(part_number="PART-A", needed_qty=30, prev_qty_cs=5)
            preview = preview_order_batches(
                str(path),
                batches=[{
                    "order_id": 1,
                    "model": "MODEL-A",
                    "groups": groups,
                    "supplements": {},
                }],
                decisions={},
                moq_map={"PART-A": 10},
            )
            result = merge_row_to_main(
                main_path=str(path),
                groups=groups,
                decisions={},
                supplements={},
                backup_dir=str(backup_dir),
            )

            self.assertEqual(preview["merged_parts"], result["merged_parts"])
            self.assertEqual(len(preview["shortages"]), 1)
            self.assertEqual(len(result["shortages"]), 1)
            self.assertEqual(preview["shortages"][0]["part_number"], "PART-A")
            self.assertEqual(preview["shortages"][0]["shortage_amount"], 5.0)
            self.assertEqual(preview["shortages"][0]["resulting_stock"], -5.0)
            self.assertEqual(result["shortages"][0]["shortage_amount"], 5.0)
            self.assertEqual(result["shortages"][0]["resulting_stock"], -5.0)

            wb = load_workbook(path, data_only=True)
            ws = wb.active
            self.assertEqual(ws.cell(row=2, column=9).value, 5)
            self.assertEqual(ws.cell(row=2, column=10).value, 30)
            self.assertEqual(ws.cell(row=2, column=11).value, -5)
            wb.close()

    def test_preview_and_commit_apply_manual_supplement_consistently(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "main.xlsx"
            backup_dir = Path(temp_dir) / "backups"
            self._build_main_workbook(path)

            groups = self._single_group(part_number="PART-A", needed_qty=30, prev_qty_cs=5)
            preview = preview_order_batches(
                str(path),
                batches=[{
                    "order_id": 1,
                    "model": "MODEL-A",
                    "groups": groups,
                    "supplements": {"PART-A": 15},
                }],
                decisions={},
                moq_map={"PART-A": 10},
            )
            result = merge_row_to_main(
                main_path=str(path),
                groups=groups,
                decisions={},
                supplements={"PART-A": 15},
                backup_dir=str(backup_dir),
            )

            self.assertEqual(preview["shortages"], [])
            self.assertEqual(result["shortages"], [])
            self.assertEqual(preview["batches"][0]["groups"][0]["rows"][0]["supplement_qty"], 15.0)
            self.assertEqual(preview["batches"][0]["groups"][0]["rows"][0]["effective_h"], 20.0)

            wb = load_workbook(path, data_only=True)
            ws = wb.active
            self.assertEqual(ws.cell(row=2, column=9).value, 20)
            self.assertEqual(ws.cell(row=2, column=10).value, 30)
            self.assertEqual(ws.cell(row=2, column=11).value, 10)
            wb.close()

    def test_preview_and_commit_ignore_supplement_when_decision_is_shortage(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "main.xlsx"
            backup_dir = Path(temp_dir) / "backups"
            self._build_main_workbook(path)

            groups = self._single_group(part_number="PART-A", needed_qty=30, prev_qty_cs=5)
            preview = preview_order_batches(
                str(path),
                batches=[{
                    "order_id": 1,
                    "model": "MODEL-A",
                    "groups": groups,
                    "supplements": {"PART-A": 15},
                }],
                decisions={"PART-A": "Shortage"},
                moq_map={"PART-A": 10},
            )
            result = merge_row_to_main(
                main_path=str(path),
                groups=groups,
                decisions={"PART-A": "Shortage"},
                supplements={"PART-A": 15},
                backup_dir=str(backup_dir),
            )

            self.assertEqual(len(preview["shortages"]), 1)
            self.assertEqual(preview["shortages"][0]["decision"], "Shortage")
            self.assertEqual(preview["shortages"][0]["supplement_qty"], 0.0)
            self.assertEqual(result["shortages"][0]["decision"], "Shortage")
            self.assertEqual(result["shortages"][0]["supplement_qty"], 0.0)

            wb = load_workbook(path, data_only=True)
            ws = wb.active
            self.assertEqual(ws.cell(row=2, column=9).value, 5)
            self.assertEqual(ws.cell(row=2, column=10).value, 30)
            self.assertEqual(ws.cell(row=2, column=11).value, -5)
            wb.close()

    def test_multi_batch_preview_matches_sequential_commits_for_shared_part_running_balance(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "main.xlsx"
            backup_dir = Path(temp_dir) / "backups"
            self._build_main_workbook(path)

            first_groups = self._single_group(part_number="PART-A", needed_qty=15, prev_qty_cs=0)
            second_groups = [{
                "batch_code": "1-2",
                "po_number": "4500000002",
                "bom_model": "MODEL-B",
                "components": [{
                    "part_number": "PART-A",
                    "description": "PART-A desc",
                    "is_dash": False,
                    "needed_qty": 10,
                    "prev_qty_cs": 0,
                }],
            }]

            preview = preview_order_batches(
                str(path),
                batches=[
                    {
                        "order_id": 1,
                        "model": "MODEL-A",
                        "groups": first_groups,
                        "supplements": {},
                    },
                    {
                        "order_id": 2,
                        "model": "MODEL-B",
                        "groups": second_groups,
                        "supplements": {},
                    },
                ],
                decisions={},
                moq_map={"PART-A": 10},
            )
            first_result = merge_row_to_main(
                main_path=str(path),
                groups=first_groups,
                decisions={},
                supplements={},
                backup_dir=str(backup_dir),
            )
            second_result = merge_row_to_main(
                main_path=str(path),
                groups=second_groups,
                decisions={},
                supplements={},
                backup_dir=str(backup_dir),
            )

            self.assertEqual(first_result["shortages"], [])
            self.assertEqual(len(preview["shortages"]), 1)
            self.assertEqual(preview["shortages"][0]["batch_code"], "1-2")
            self.assertEqual(preview["shortages"][0]["shortage_amount"], 5.0)
            self.assertEqual(preview["shortages"][0]["resulting_stock"], -5.0)
            self.assertEqual(second_result["shortages"][0]["shortage_amount"], 5.0)
            self.assertEqual(second_result["shortages"][0]["resulting_stock"], -5.0)
            self.assertEqual(preview["batches"][0]["groups"][0]["rows"][0]["j_value"], 5.0)
            self.assertEqual(preview["batches"][1]["groups"][0]["rows"][0]["j_value"], -5.0)

            wb = load_workbook(path, data_only=True)
            ws = wb.active
            self.assertIsNone(ws.cell(row=2, column=9).value)
            self.assertEqual(ws.cell(row=2, column=10).value, 15)
            self.assertEqual(ws.cell(row=2, column=11).value, 5)
            self.assertIsNone(ws.cell(row=2, column=12).value)
            self.assertEqual(ws.cell(row=2, column=13).value, 10)
            self.assertEqual(ws.cell(row=2, column=14).value, -5)
            wb.close()

    def test_multi_batch_commit_saves_once_and_can_restore_from_middle_order(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "main.xlsx"
            backup_dir = Path(temp_dir) / "backups"
            self._build_main_workbook(path)

            first_groups = self._single_group(part_number="PART-A", needed_qty=15, prev_qty_cs=0)
            second_groups = [{
                "batch_code": "1-2",
                "po_number": "4500000002",
                "bom_model": "MODEL-B",
                "components": [{
                    "part_number": "PART-A",
                    "description": "PART-A desc",
                    "is_dash": False,
                    "needed_qty": 10,
                    "prev_qty_cs": 0,
                }],
            }]
            batches = [
                {"order_id": 1, "model": "MODEL-A", "groups": first_groups, "supplements": {}, "decisions": {}},
                {"order_id": 2, "model": "MODEL-B", "groups": second_groups, "supplements": {}, "decisions": {}},
            ]

            with patch.object(
                merge_service,
                "_save_workbook_atomically",
                wraps=merge_service._save_workbook_atomically,
            ) as mock_save:
                result = merge_order_batches_to_main(
                    str(path),
                    batches,
                    backup_dir=str(backup_dir),
                )

            self.assertEqual(mock_save.call_count, 1)
            self.assertEqual([item["merged_parts"] for item in result["order_results"]], [1, 1])
            self.assertTrue(validate_dispatch_backup_reference(result["backup_path"])["ok"])

            wb = load_workbook(path, data_only=True)
            ws = wb.active
            self.assertEqual(ws.cell(row=2, column=11).value, 5)
            self.assertEqual(ws.cell(row=2, column=14).value, -5)
            wb.close()

            current_bytes = path.read_bytes()
            with self.assertRaisesRegex(ValueError, "找不到要反悔的訂單"):
                restore_main_from_backup_reference(
                    result["backup_path"],
                    str(path),
                    before_order_id=999,
                )
            self.assertEqual(path.read_bytes(), current_bytes)

            restore_result = restore_main_from_backup_reference(
                result["backup_path"],
                str(path),
                before_order_id=2,
            )
            self.assertEqual(restore_result["replayed_order_count"], 1)

            wb = load_workbook(path, data_only=True)
            ws = wb.active
            self.assertEqual(ws.cell(row=2, column=11).value, 5)
            self.assertIsNone(ws.cell(row=2, column=14).value)
            wb.close()

            restore_result = restore_main_from_backup_reference(
                result["backup_path"],
                str(path),
                before_order_id=1,
            )
            self.assertEqual(restore_result["replayed_order_count"], 0)
            wb = load_workbook(path, data_only=True)
            ws = wb.active
            self.assertEqual(ws.max_column, 8)
            self.assertEqual(ws.cell(row=2, column=8).value, 20)
            wb.close()


if __name__ == "__main__":
    unittest.main()
