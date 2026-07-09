from __future__ import annotations

import tempfile
import unittest
from unittest.mock import patch
from pathlib import Path

import openpyxl

from app.services.dispatch_pipeline import DispatchContext, execute_dispatch_context
from app.services.main_reconcile import verify_main_write
from app.services.merge_to_main import merge_row_to_main


def _create_main(path: Path, stock: float = 100) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["料號", "廠商", "MOQ", "說明", "庫存"])
    ws.append(["PART-A", "", 1, "Part A", stock])
    wb.save(path)
    wb.close()


def _groups(needed_qty: float = 30, prev_qty_cs: float = 5) -> list[dict]:
    return [{
        "batch_code": "1-1",
        "po_number": "PO-1",
        "bom_model": "MODEL-A",
        "components": [{
            "part_number": "PART-A",
            "description": "Part A",
            "needed_qty": needed_qty,
            "prev_qty_cs": prev_qty_cs,
            "is_dash": 0,
        }],
    }]


class MainReconcileTests(unittest.TestCase):
    def test_verify_main_write_matches_merge_plan(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            _create_main(main_path)

            result = merge_row_to_main(str(main_path), _groups(), {}, backup_dir=None)
            reconcile = verify_main_write(str(main_path), result["plan_rows"])

        self.assertTrue(reconcile["ok"])
        self.assertEqual(reconcile["checked_parts"], 1)
        self.assertEqual(reconcile["mismatches"], [])

    def test_verify_main_write_detects_bad_written_cell(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            _create_main(main_path)
            result = merge_row_to_main(str(main_path), _groups(), {}, backup_dir=None)
            first_row = result["plan_rows"][0]

            wb = openpyxl.load_workbook(main_path)
            ws = wb.active
            ws.cell(row=first_row["row_idx"], column=first_row["col_j"]).value = 999
            wb.save(main_path)
            wb.close()

            reconcile = verify_main_write(str(main_path), result["plan_rows"])

        self.assertFalse(reconcile["ok"])
        self.assertTrue(any(item["kind"] == "cell" for item in reconcile["mismatches"]))

    def test_verify_main_write_uses_raw_values_for_decimal_conservation(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            _create_main(main_path, stock=0)

            result = merge_row_to_main(str(main_path), _groups(needed_qty=1.4, prev_qty_cs=0.5), {}, backup_dir=None)
            reconcile = verify_main_write(str(main_path), result["plan_rows"])

        self.assertTrue(reconcile["ok"])
        self.assertEqual(reconcile["mismatches"], [])

    def test_verify_main_write_treats_blank_h_as_zero(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            _create_main(main_path)

            result = merge_row_to_main(str(main_path), _groups(needed_qty=30, prev_qty_cs=0), {}, backup_dir=None)
            first_row = result["plan_rows"][0]

            wb = openpyxl.load_workbook(main_path)
            ws = wb.active
            self.assertIsNone(ws.cell(row=first_row["row_idx"], column=first_row["col_h"]).value)
            wb.close()

            reconcile = verify_main_write(str(main_path), result["plan_rows"])

        self.assertTrue(reconcile["ok"])
        self.assertEqual(reconcile["mismatches"], [])

    def test_verify_main_write_cross_read_detects_wrong_rightmost_stock(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            _create_main(main_path)
            result = merge_row_to_main(str(main_path), _groups(), {}, backup_dir=None)
            first_row = result["plan_rows"][0]

            wb = openpyxl.load_workbook(main_path)
            ws = wb.active
            ws.cell(row=first_row["row_idx"], column=first_row["col_j"] + 1).value = 999
            wb.save(main_path)
            wb.close()

            reconcile = verify_main_write(str(main_path), result["plan_rows"])

        self.assertFalse(reconcile["ok"])
        self.assertFalse(any(item["kind"] == "cell" for item in reconcile["mismatches"]))
        self.assertTrue(any(item["kind"] == "cross_read" for item in reconcile["mismatches"]))

    def test_dispatch_context_wraps_verify_error(self):
        context = DispatchContext(
            order={"id": 123, "status": "merged", "po_number": "PO-1", "model": "MODEL-A", "folder": ""},
            groups=[],
            all_components=[],
        )

        def fake_merge_executor(**_kwargs):
            return {
                "merged_parts": 1,
                "backup_path": "",
                "plan_rows": [{
                    "part_number": "PART-A",
                    "row_idx": 2,
                    "col_h": 6,
                    "col_f": 7,
                    "col_j": 8,
                }],
            }

        with (
            patch("app.services.dispatch_pipeline.verify_main_write", side_effect=RuntimeError("boom")),
            patch("app.services.dispatch_pipeline.db.save_dispatch_session", return_value={"id": 1}),
            patch("app.services.dispatch_pipeline.db.save_dispatch_records"),
            patch("app.services.dispatch_pipeline.db.update_order"),
            patch("app.services.dispatch_pipeline.db.log_activity"),
        ):
            result = execute_dispatch_context(
                context,
                "main.xlsx",
                merge_executor=fake_merge_executor,
                backup_dir="",
            )

        self.assertTrue(result["ok"])
        self.assertFalse(result["reconcile"]["ok"])
        self.assertEqual(result["reconcile"]["mismatches"][0]["kind"], "verify_error")


if __name__ == "__main__":
    unittest.main()
