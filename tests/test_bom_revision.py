from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

from app.services.bom_revision import snapshot_bom_revision


class BomRevisionServiceTests(unittest.TestCase):
    def test_snapshot_bom_revision_copies_current_bom_file(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "formal.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["A1"] = "current"
            wb.save(source_path)
            wb.close()

            history_dir = Path(temp_dir) / "history"

            saved_payloads = []

            def fake_save_revision(revision):
                saved_payloads.append(dict(revision))
                return {
                    "id": 1,
                    "bom_file_id": revision["bom_file_id"],
                    "revision_number": 1,
                    "filename": revision["filename"],
                    "filepath": revision["filepath"],
                    "source_action": revision["source_action"],
                    "note": revision["note"],
                    "created_at": "2026-03-12T18:00:00",
                }

            with patch("app.services.bom_revision.BOM_HISTORY_DIR", history_dir), \
                 patch("app.services.bom_revision.db.save_bom_revision", side_effect=fake_save_revision):
                result = snapshot_bom_revision({
                    "id": "bom-1",
                    "filename": "formal.xlsx",
                    "filepath": str(source_path),
                }, "upload", "上傳 BOM")

            self.assertEqual(result["revision_number"], 1)
            self.assertEqual(result["source_action"], "upload")
            self.assertEqual(len(saved_payloads), 1)
            copied_path = Path(saved_payloads[0]["filepath"])
            self.assertTrue(copied_path.exists())
            copied_wb = openpyxl.load_workbook(copied_path)
            self.assertEqual(copied_wb.active["A1"].value, "current")
            copied_wb.close()
