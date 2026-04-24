from __future__ import annotations

import tempfile
import unittest
import zipfile
from xml.etree import ElementTree as ET
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.services.main_reader import find_legacy_snapshot_stock_fixes, read_stock, read_vendors, update_vendor
from app.services.bom_parser import parse_bom, read_formula_needed_qty_cache
from app.services.bom_quantity import coerce_scrap_factor
from app.services.merge_to_main import merge_row_to_main, preview_order_batches


class ExcelLogicTests(unittest.TestCase):
    def _inject_cached_formula_value(self, path: Path, cell_ref: str, formula: str, cached_value: float) -> None:
        worksheet_path = "xl/worksheets/sheet1.xml"
        namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        ET.register_namespace("", namespace)

        replacement_path = path.with_suffix(".patched.xlsx")
        with zipfile.ZipFile(path, "r") as source_zip, zipfile.ZipFile(replacement_path, "w", zipfile.ZIP_DEFLATED) as target_zip:
            for item in source_zip.infolist():
                data = source_zip.read(item.filename)
                if item.filename == worksheet_path:
                    root = ET.fromstring(data)
                    cell = root.find(f".//{{{namespace}}}c[@r='{cell_ref}']")
                    if cell is None:
                        raise AssertionError(f"cell {cell_ref} not found")
                    for child in list(cell):
                        if child.tag in {f"{{{namespace}}}f", f"{{{namespace}}}v"}:
                            cell.remove(child)
                    formula_el = ET.SubElement(cell, f"{{{namespace}}}f")
                    formula_el.text = formula.lstrip("=")
                    value_el = ET.SubElement(cell, f"{{{namespace}}}v")
                    value_el.text = str(cached_value)
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                target_zip.writestr(item, data)
        replacement_path.replace(path)

    def _build_main_workbook(self, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "2026"
        ws.append(["料號", "廠商", "MOQ", "期初", "盤點", "", "M/O", "結存"])
        ws.append(["PART-A", "Vendor", 1000, None, None, None, None, None])
        ws.append(["PART-B", "Vendor", 500, 10, 20, None, 0, 20])
        wb.save(path)
        wb.close()

    def test_read_stock_does_not_treat_moq_as_inventory(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "main.xlsx"
            self._build_main_workbook(path)

            stock = read_stock(str(path))

        self.assertEqual(stock["PART-A"], 0.0)
        self.assertEqual(stock["PART-B"], 20.0)

    def test_read_vendors_reads_main_file_b_column(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "main.xlsx"
            self._build_main_workbook(path)

            vendors = read_vendors(str(path))

        self.assertEqual(vendors["PART-A"], "Vendor")
        self.assertEqual(vendors["PART-B"], "Vendor")

    def test_update_vendor_writes_main_file_b_column(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "main.xlsx"
            self._build_main_workbook(path)

            result = update_vendor(str(path), "part-b", "新廠商")
            wb = load_workbook(path)
            ws = wb.active
            saved_vendor = ws.cell(row=3, column=2).value
            wb.close()

        self.assertEqual(result["part_number"], "PART-B")
        self.assertEqual(result["old_vendor"], "Vendor")
        self.assertEqual(result["vendor"], "新廠商")
        self.assertEqual(result["row"], 3)
        self.assertEqual(saved_vendor, "新廠商")

    def test_find_legacy_snapshot_stock_fixes_detects_moq_only_rows(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "main.xlsx"
            self._build_main_workbook(path)

            fixes = find_legacy_snapshot_stock_fixes(
                str(path),
                {
                    "PART-A": {"stock_qty": 1000, "moq": 1000},
                    "PART-B": {"stock_qty": 20, "moq": 500},
                },
            )

        self.assertEqual(fixes, {"PART-A": 0.0})

    def test_parse_bom_calculates_formula_needed_qty_from_scrap_factor_without_cached_value(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "bom.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=8, value="PO:4500059234")
            ws.cell(row=1, column=11, value=10)
            ws.cell(row=2, column=3, value="MODEL-S")
            ws.cell(row=2, column=4, value="PCB-S")
            ws.cell(row=5, column=2, value=2)
            ws.cell(row=5, column=3, value="PART-S")
            ws.cell(row=5, column=4, value="Cap")
            ws.cell(row=5, column=5, value=0.06)
            ws.cell(row=5, column=6, value="=B5*$K$1*(1+E5)")
            wb.save(path)
            wb.close()

            parsed = parse_bom(str(path), "bom-s", "bom.xlsx", "2026-04-22T10:00:00")

        self.assertEqual(parsed.order_qty, 10)
        self.assertEqual(len(parsed.components), 1)
        self.assertAlmostEqual(parsed.components[0].scrap_factor, 0.06)
        self.assertAlmostEqual(parsed.components[0].needed_qty, 21.2)

    def test_parse_bom_evaluates_simple_formula_needed_qty_instead_of_stale_cache(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "bom.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=8, value="PO:4500059234")
            ws.cell(row=1, column=11, value=300)
            ws.cell(row=2, column=3, value="MODEL-S")
            ws.cell(row=2, column=4, value="PCB-S")
            ws.cell(row=5, column=2, value=1)
            ws.cell(row=5, column=3, value="PB-20111A-TAB")
            ws.cell(row=5, column=4, value="TX1-S1 bare PCB")
            ws.cell(row=5, column=5, value=0)
            ws.cell(row=5, column=6, value="=B5*$K$1*(1+E5)")
            wb.save(path)
            wb.close()
            self._inject_cached_formula_value(path, "F5", "=B5*$K$1*(1+E5)", 600)

            parsed = parse_bom(str(path), "bom-s", "bom.xlsx", "2026-04-22T10:00:00")

        self.assertEqual(len(parsed.components), 1)
        self.assertAlmostEqual(parsed.components[0].scrap_factor, 0)
        self.assertAlmostEqual(parsed.components[0].needed_qty, 300)

    def test_parse_bom_formula_reads_text_order_quantity_cell(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "bom.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=8, value="PO:4500059234")
            ws.cell(row=1, column=11, value="生產數量: 300")
            ws.cell(row=2, column=3, value="MODEL-M24C")
            ws.cell(row=2, column=4, value="PCB-M24C")
            ws.cell(row=5, column=2, value=1)
            ws.cell(row=5, column=3, value="IC-M24C")
            ws.cell(row=5, column=4, value="EEPROM")
            ws.cell(row=5, column=5, value=0)
            ws.cell(row=5, column=6, value="=B5*$K$1*(1+E5)")
            wb.save(path)
            wb.close()
            self._inject_cached_formula_value(path, "F5", "=B5*$K$1*(1+E5)", 600)

            parsed = parse_bom(str(path), "bom-m24c", "bom.xlsx", "2026-04-22T10:00:00")

        self.assertEqual(parsed.order_qty, 300)
        self.assertEqual(len(parsed.components), 1)
        self.assertAlmostEqual(parsed.components[0].qty_per_board, 1)
        self.assertAlmostEqual(parsed.components[0].needed_qty, 300)

    def test_formula_needed_cache_prefers_evaluated_formula_over_stale_cache(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "bom.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=8, value="PO:4500059234")
            ws.cell(row=1, column=11, value="生產數量: 300")
            ws.cell(row=2, column=3, value="MODEL-M24C")
            ws.cell(row=2, column=4, value="PCB-M24C")
            ws.cell(row=5, column=2, value=1)
            ws.cell(row=5, column=3, value="IC-M24C")
            ws.cell(row=5, column=5, value=0)
            ws.cell(row=5, column=6, value="=B5*$K$1*(1+E5)")
            wb.save(path)
            wb.close()
            self._inject_cached_formula_value(path, "F5", "=B5*$K$1*(1+E5)", 600)

            cache = read_formula_needed_qty_cache(str(path))

        self.assertEqual(cache[("Sheet", 5, "IC-M24C")], 300)

    def test_parse_bom_uses_cached_formula_value_when_formula_is_not_supported(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "bom.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=8, value="PO:4500059234")
            ws.cell(row=1, column=11, value=300)
            ws.cell(row=2, column=3, value="MODEL-S")
            ws.cell(row=2, column=4, value="PCB-S")
            ws.cell(row=5, column=2, value=1)
            ws.cell(row=5, column=3, value="OC-10849B")
            ws.cell(row=5, column=4, value="TX1-S1 spring")
            ws.cell(row=5, column=5, value=0.04)
            ws.cell(row=5, column=6, value='=IF("X"="X",306,999)')
            wb.save(path)
            wb.close()
            self._inject_cached_formula_value(path, "F5", '=IF("X"="X",306,999)', 306)

            parsed = parse_bom(str(path), "bom-s", "bom.xlsx", "2026-04-22T10:00:00")

        self.assertEqual(len(parsed.components), 1)
        self.assertAlmostEqual(parsed.components[0].scrap_factor, 0.04)
        self.assertAlmostEqual(parsed.components[0].needed_qty, 306)

    def test_parse_bom_detects_scrap_factor_column_from_header(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "bom.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=8, value="PO:4500059234")
            ws.cell(row=1, column=11, value=10)
            ws.cell(row=2, column=3, value="MODEL-S")
            ws.cell(row=2, column=4, value="PCB-S")
            ws.cell(row=4, column=12, value="拋料率")
            ws.cell(row=5, column=2, value=2)
            ws.cell(row=5, column=3, value="PART-S")
            ws.cell(row=5, column=4, value="Cap")
            ws.cell(row=5, column=12, value="6%")
            ws.cell(row=5, column=6, value="=B5*$K$1*(1+L5)")
            wb.save(path)
            wb.close()

            parsed = parse_bom(str(path), "bom-s", "bom.xlsx", "2026-04-22T10:00:00")

        self.assertEqual(len(parsed.components), 1)
        self.assertAlmostEqual(parsed.components[0].scrap_factor, 0.06)
        self.assertAlmostEqual(parsed.components[0].needed_qty, 21.2)

    def test_parse_bom_reads_literal_formula_scrap_factor(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "bom.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=8, value="PO:4500059234")
            ws.cell(row=1, column=11, value=10)
            ws.cell(row=2, column=3, value="MODEL-S")
            ws.cell(row=2, column=4, value="PCB-S")
            ws.cell(row=5, column=2, value=2)
            ws.cell(row=5, column=3, value="PART-S")
            ws.cell(row=5, column=4, value="Cap")
            ws.cell(row=5, column=5, value="=6%")
            ws.cell(row=5, column=6, value="=B5*$K$1*(1+E5)")
            wb.save(path)
            wb.close()

            parsed = parse_bom(str(path), "bom-s", "bom.xlsx", "2026-04-22T10:00:00")

        self.assertEqual(len(parsed.components), 1)
        self.assertAlmostEqual(parsed.components[0].scrap_factor, 0.06)
        self.assertAlmostEqual(parsed.components[0].needed_qty, 21.2)

    def test_parse_bom_reconciles_bad_scrap_cell_against_formula_needed(self):
        # 若 E 欄 scrap cell 被誤存整數 1（coerce 會吐 1.0 = 100%），
        # 但 F 欄公式實際評估出的 needed 隱含 0.6% 拋料，
        # parser 應該用反推值覆蓋存錯的 1.0。
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "bom.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=8, value="PO:4500059234")
            ws.cell(row=1, column=11, value=192)
            ws.cell(row=2, column=3, value="MODEL-RECONCILE")
            ws.cell(row=2, column=4, value="PCB-R")
            ws.cell(row=5, column=2, value=4)
            ws.cell(row=5, column=3, value="PART-R")
            ws.cell(row=5, column=4, value="Res")
            # E 欄誤存整數 1（不該 100% 拋料）
            ws.cell(row=5, column=5, value=1)
            # F 欄公式實際含 0.006 scrap：4 × 192 × 1.006 = 772.608
            ws.cell(row=5, column=6, value=772.608)
            wb.save(path)
            wb.close()

            parsed = parse_bom(str(path), "bom-reconcile", "bom.xlsx", "2026-04-24T10:00:00")

        self.assertEqual(len(parsed.components), 1)
        self.assertAlmostEqual(parsed.components[0].scrap_factor, 0.006)
        self.assertAlmostEqual(parsed.components[0].needed_qty, 772.608)

    def test_coerce_scrap_factor_extracts_labeled_percentage_but_ignores_cell_refs(self):
        self.assertAlmostEqual(coerce_scrap_factor("拋料率 6%"), 0.06)
        self.assertAlmostEqual(coerce_scrap_factor("E5: 6%"), 0.06)
        self.assertAlmostEqual(coerce_scrap_factor("=6%"), 0.06)
        self.assertEqual(coerce_scrap_factor("=E5"), 0.0)
        self.assertEqual(coerce_scrap_factor("E5"), 0.0)

    def test_merge_to_main_uses_zero_stock_when_only_moq_exists(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "main.xlsx"
            backup_dir = Path(temp_dir) / "backups"
            self._build_main_workbook(path)

            result = merge_row_to_main(
                main_path=str(path),
                groups=[{
                    "batch_code": "1-1",
                    "po_number": "12345",
                    "bom_model": "MODEL-A",
                    "components": [{
                        "part_number": "PART-A",
                        "is_dash": False,
                        "needed_qty": 50,
                        "prev_qty_cs": 0,
                    }],
                }],
                decisions={},
                backup_dir=str(backup_dir),
            )

            wb = load_workbook(path, data_only=True)
            ws = wb.active
            self.assertEqual(result["merged_parts"], 1)
            self.assertEqual(ws.cell(row=2, column=9).value, 0)
            self.assertEqual(ws.cell(row=2, column=10).value, 50)
            self.assertEqual(ws.cell(row=2, column=11).value, -50)
            wb.close()

    def test_merge_to_main_writes_supplement_into_live_main(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "main.xlsx"
            backup_dir = Path(temp_dir) / "backups"
            self._build_main_workbook(path)

            result = merge_row_to_main(
                main_path=str(path),
                groups=[{
                    "batch_code": "1-3",
                    "po_number": "4500059234",
                    "bom_model": "MODEL-A",
                    "components": [{
                        "part_number": "PART-A",
                        "description": "CAP",
                        "is_dash": False,
                        "needed_qty": 50,
                        "prev_qty_cs": 0,
                    }],
                }],
                decisions={},
                supplements={"PART-A": 70},
                backup_dir=str(backup_dir),
            )

            wb = load_workbook(path, data_only=True)
            ws = wb.active
            self.assertEqual(result["merged_parts"], 1)
            self.assertEqual(ws.cell(row=2, column=9).value, 70)
            self.assertEqual(ws.cell(row=2, column=10).value, 50)
            self.assertEqual(ws.cell(row=2, column=11).value, 20)
            wb.close()

    def test_merge_to_main_sets_header_row_to_wrap_text(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "main.xlsx"
            backup_dir = Path(temp_dir) / "backups"
            self._build_main_workbook(path)

            result = merge_row_to_main(
                main_path=str(path),
                groups=[{
                    "batch_code": "1-3",
                    "po_number": "4500059234",
                    "bom_model": "MODEL-A-LONG-NAME",
                    "components": [{
                        "part_number": "PART-A",
                        "description": "CAP",
                        "is_dash": False,
                        "needed_qty": 50,
                        "prev_qty_cs": 0,
                    }],
                }],
                decisions={},
                supplements={"PART-A": 70},
                backup_dir=str(backup_dir),
            )

            wb = load_workbook(path, data_only=False)
            ws = wb.active
            self.assertEqual(result["merged_parts"], 1)
            self.assertTrue(ws.cell(row=1, column=1).alignment.wrap_text)
            self.assertTrue(ws.cell(row=1, column=9).alignment.wrap_text)
            self.assertTrue(ws.cell(row=1, column=10).alignment.wrap_text)
            self.assertTrue(ws.cell(row=1, column=11).alignment.wrap_text)
            wb.close()

    def test_preview_order_batches_includes_moq_and_rounded_suggestion(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "main.xlsx"
            self._build_main_workbook(path)

            preview = preview_order_batches(
                str(path),
                batches=[{
                    "order_id": 1,
                    "model": "MODEL-A",
                    "groups": [{
                        "batch_code": "1-3",
                        "po_number": "4500059234",
                        "bom_model": "MODEL-A",
                        "components": [{
                            "part_number": "PART-A",
                            "description": "CAP",
                            "is_dash": False,
                            "needed_qty": 1200,
                            "prev_qty_cs": 0,
                        }],
                    }],
                    "supplements": {},
                }],
                decisions={},
                moq_map={"PART-A": 1000},
            )

        self.assertEqual(len(preview["shortages"]), 1)
        self.assertEqual(preview["shortages"][0]["moq"], 1000)
        self.assertEqual(preview["shortages"][0]["suggested_qty"], 2000)
