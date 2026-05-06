from __future__ import annotations

import io
import os
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest.mock import ANY, call, patch

import openpyxl
from fastapi import HTTPException
from fastapi.responses import Response
from fastapi.testclient import TestClient
from openpyxl.styles import Font, PatternFill

import app.routers.schedule as schedule_router
import app.routers.bom as bom_router
import app.database as database
from app.models import BomComponent, BomFile
from main import app


class ApiTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        os.environ.setdefault("PYTEST_CURRENT_TEST", "1")
        cls.client = TestClient(app)

    def _build_main_workbook(self, path: Path, rows: list[tuple[str, float, float]]) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "2026"
        ws.append(["料號", "廠商", "MOQ", "期初", "盤點", "", "M/O", "結存"])
        for part_number, moq, stock_qty in rows:
            ws.append([part_number, "Vendor", moq, None, None, None, 0, stock_qty])
        wb.save(path)
        wb.close()

    def test_restore_pre_normalize_needed_quantities_keeps_cached_formula_values(self):
        pre_normalize = BomFile(
            id="bom-1",
            filename="bom.xlsx",
            path="bom.xlsx",
            po_number=4500059234,
            components=[
                BomComponent(
                    part_number="OC-10849B",
                    needed_qty=306,
                    source_row=5,
                    source_sheet="Sheet",
                )
            ],
        )
        parsed = BomFile(
            id="bom-1",
            filename="bom.xlsx",
            path="bom.xlsx",
            po_number=4500059234,
            components=[
                BomComponent(
                    part_number="OC-10849B",
                    needed_qty=312,
                    source_row=5,
                    source_sheet="Sheet",
                )
            ],
        )

        restored = bom_router._restore_pre_normalize_needed_quantities(parsed, pre_normalize)

        self.assertEqual(restored.components[0].needed_qty, 306)

    def test_schedule_rows_include_decisions_and_dispatched_consumption(self):
        pending_rows = [{"id": 1, "model": "MODEL-A"}]
        completed_rows = [{"id": 2, "model": "MODEL-B"}]

        def fake_get_orders(statuses=None):
            if statuses == ["pending", "merged"]:
                return pending_rows
            if statuses == ["dispatched", "completed"]:
                return completed_rows
            return []

        with patch("app.routers.schedule.db.get_orders", side_effect=fake_get_orders), \
             patch("app.routers.schedule.db.get_setting", side_effect=lambda key, default="": {
                 "schedule_loaded_at": "2026-03-12T08:00:00",
                 "schedule_filename": "schedule.xlsx",
             }.get(key, default)), \
             patch("app.routers.schedule.db.get_all_dispatched_consumption", return_value={"PART-1": 12}), \
             patch("app.routers.schedule.db.get_all_decisions", return_value={"PART-1": "CreateRequirement"}), \
             patch("app.routers.schedule.db.get_order_supplements", return_value={1: {"PART-1": 1000}}), \
             patch("app.routers.schedule.db.get_order_supplement_details", return_value={1: {"PART-1": {"supplement_qty": 1000, "note": "補急單", "updated_at": "2026-04-02T10:00:00"}}}), \
             patch("app.routers.schedule.get_schedule_draft_map", return_value={1: {"id": 9, "files": [], "shortages": []}}):
            response = self.client.get("/api/schedule/rows")

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["rows"], pending_rows)
        self.assertEqual(data["completed_count"], 1)
        self.assertEqual(data["dispatched_consumption"], {"PART-1": 12})
        self.assertEqual(data["decisions"], {"PART-1": "CreateRequirement"})
        self.assertEqual(data["merge_drafts"], {"1": {"id": 9, "files": [], "shortages": []}})
        self.assertEqual(data["order_supplements"], {"1": {"PART-1": 1000}})
        self.assertEqual(data["order_supplement_details"], {"1": {"PART-1": {"supplement_qty": 1000, "note": "補急單", "updated_at": "2026-04-02T10:00:00"}}})

    def test_edit_auth_blocks_mutating_api_until_login(self):
        with patch.dict(os.environ, {"PYTEST_CURRENT_TEST": ""}, clear=False):
            blocked = self.client.post("/api/schedule/batch-merge", json={"order_ids": [1]})
            self.assertEqual(blocked.status_code, 403)
            self.assertEqual(blocked.json()["code"], "edit_auth_required")

            status = self.client.get("/api/system/edit-auth/status")
            self.assertEqual(status.status_code, 200)
            self.assertFalse(status.json()["authenticated"])

            login = self.client.post("/api/system/edit-auth/login", json={"password": "123"})
            self.assertEqual(login.status_code, 200)
            self.assertTrue(login.json()["authenticated"])

            with patch("app.routers.schedule.db.batch_merge_orders") as mock_batch_merge, \
                 patch("app.routers.schedule.rebuild_merge_drafts", return_value=[]) as mock_rebuild, \
                 patch("app.routers.schedule.db.log_activity"), \
                 patch("app.routers.schedule.db.create_alert"):
                allowed = self.client.post("/api/schedule/batch-merge", json={"order_ids": [1]})

            self.assertEqual(allowed.status_code, 200)
            mock_batch_merge.assert_called_once_with([1])
            mock_rebuild.assert_called_once_with([1])

            logout = self.client.post("/api/system/edit-auth/logout")
            self.assertEqual(logout.status_code, 200)

    def test_schedule_rows_use_snapshot_cutoff_for_dispatched_consumption(self):
        with patch("app.routers.schedule.db.get_orders", side_effect=[[], []]), \
             patch("app.routers.schedule.db.get_setting", side_effect=lambda key, default="": default), \
             patch("app.routers.schedule.db.get_snapshot_taken_at", return_value="2026-03-12T11:05:45.000000"), \
             patch("app.routers.schedule.db.get_all_dispatched_consumption", return_value={"PART-2": 15}) as mock_consumption, \
             patch("app.routers.schedule.db.get_all_decisions", return_value={}), \
             patch("app.routers.schedule.db.get_order_supplements", return_value={}), \
             patch("app.routers.schedule.db.get_order_supplement_details", return_value={}):
            response = self.client.get("/api/schedule/rows")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["dispatched_consumption"], {"PART-2": 15})
        mock_consumption.assert_called_once_with("2026-03-12T11:05:45.000000")

    def test_update_schedule_shortage_settings_rebuilds_active_drafts_and_persists_supplements(self):
        with patch("app.routers.schedule.build_order_decision_allocations", return_value={1: {}}), \
             patch("app.routers.schedule._merge_decision_updates", return_value={"PART-OLD": "IgnoreOnce"}), \
             patch("app.routers.schedule._merge_supplement_updates", return_value={"PART-1": 1200}) as mock_merge_supplements, \
             patch("app.routers.schedule.db.replace_order_decisions") as mock_replace_decisions, \
             patch("app.routers.schedule.db.replace_order_supplements") as mock_replace_supplements, \
             patch("app.routers.schedule.db.get_order_supplement_details", return_value={1: {"PART-1": {"supplement_qty": 1200, "note": "急件", "updated_at": "2026-04-02T11:00:00"}}}), \
             patch("app.routers.schedule.db.get_active_merge_drafts", return_value=[{"order_id": 1}, {"order_id": 2}]), \
             patch("app.routers.schedule.rebuild_merge_drafts") as mock_rebuild, \
             patch("app.routers.schedule.db.log_activity"):
            response = self.client.put("/api/schedule/shortage-settings", json={
                "order_ids": [1],
                "order_supplements": {
                    "1": {"part-1": 1200},
                },
                "order_supplement_notes": {
                    "1": {"part-1": "急件"},
                },
            })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {
            "ok": True,
            "count": 1,
            "order_supplement_details": {"1": {"PART-1": {"supplement_qty": 1200, "note": "急件", "updated_at": "2026-04-02T11:00:00"}}},
        })
        mock_merge_supplements.assert_called_once_with(1, {"PART-1": 1200.0})
        mock_replace_decisions.assert_called_once_with([1], {1: {"PART-OLD": "IgnoreOnce"}})
        mock_replace_supplements.assert_called_once_with([1], {1: {"PART-1": 1200}}, {1: {"PART-1": "急件"}})
        mock_rebuild.assert_called_once_with([1, 2])

    def test_bom_upload_rejects_invalid_ghij_layout(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            stored_path = Path(temp_dir) / "upload.xlsx"
            stored_path.write_bytes(b"dummy")

            with patch("app.routers.bom.prepare_uploaded_bom_file", return_value={
                "filepath": str(stored_path),
                "filename": "upload.xlsx",
                "source_filename": "upload.xlsx",
                "source_format": ".xlsx",
                "is_converted": False,
            }), \
                 patch("app.routers.bom.normalize_uploaded_bom_layout", return_value=[]), \
                 patch("app.routers.bom.validate_uploaded_bom_layout", return_value=[
                     "第 5 列 PART-A: G 欄需為空白、H 欄需為空白、I 欄需為公式、J 欄需為公式"
                 ]), \
                 patch("app.routers.bom.parse_bom_for_storage") as mock_parse:
                response = self.client.post(
                    "/api/bom/upload",
                    files=[("files", ("upload.xlsx", b"dummy", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
                )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["saved"], [])
        self.assertEqual(len(body["errors"]), 1)
        self.assertIn("副檔欄位檢查失敗", body["errors"][0])
        mock_parse.assert_not_called()

    def test_bom_upload_overwrites_matching_existing_bom(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_root = Path(temp_dir)
            incoming_path = temp_root / "temp-upload.xlsx"
            existing_path = temp_root / "bom-existing.xlsx"
            duplicate_path = temp_root / "bom-duplicate.xlsx"
            incoming_path.write_bytes(b"new")
            existing_path.write_bytes(b"old")
            duplicate_path.write_bytes(b"dup")

            parsed = BomFile(
                id="temp-upload",
                filename="upload.xlsx",
                path=str(incoming_path),
                po_number=123,
                model="MODEL-A-MAIN",
                pcb="PCB-A",
                group_model="MODEL-A",
                order_qty=10,
                uploaded_at="2026-04-13T10:00:00",
                source_filename="upload.xlsx",
                source_format=".xlsx",
                is_converted=False,
                components=[BomComponent(part_number="PART-1", description="Cap")],
            )
            existing_bom = {
                "id": "bom-existing",
                "filename": "old.xlsx",
                "filepath": str(existing_path),
                "source_filename": "upload.xlsx",
                "source_format": ".xlsx",
                "is_converted": 0,
                "group_model": "MODEL-A",
                "model": "MODEL-A-MAIN",
                "pcb": "PCB-A",
                "uploaded_at": "2026-04-12T08:00:00",
                "sort_order": 2,
            }
            duplicate_bom = {
                "id": "bom-duplicate",
                "filename": "old-dup.xlsx",
                "filepath": str(duplicate_path),
                "source_filename": "upload.xlsx",
                "source_format": ".xlsx",
                "is_converted": 0,
                "group_model": "MODEL-A",
                "model": "MODEL-A-MAIN",
                "pcb": "PCB-A",
                "uploaded_at": "2026-04-12T09:00:00",
                "sort_order": 3,
            }

            with patch("app.routers.bom.BOM_DIR", temp_root), \
                 patch("app.routers.bom.prepare_uploaded_bom_file", return_value={
                     "filepath": str(incoming_path),
                     "filename": "upload.xlsx",
                     "source_filename": "upload.xlsx",
                     "source_format": ".xlsx",
                     "is_converted": False,
                 }), \
                 patch("app.routers.bom.normalize_uploaded_bom_layout", return_value=[]), \
                 patch("app.routers.bom.validate_uploaded_bom_layout", return_value=[]), \
                 patch("app.routers.bom.parse_bom_for_storage", return_value=parsed), \
                 patch("app.routers.bom.db.get_bom_files", return_value=[existing_bom, duplicate_bom]), \
                 patch("app.routers.bom.ensure_bom_revision_history") as mock_ensure_history, \
                 patch("app.routers.bom.db.save_bom_file") as mock_save_bom, \
                 patch("app.routers.bom.snapshot_bom_revision") as mock_snapshot_revision, \
                 patch("app.routers.bom.delete_bom_revision_files") as mock_delete_revision_files, \
                 patch("app.routers.bom.db.delete_bom_file") as mock_delete_bom_file, \
                 patch("app.routers.bom.db.log_activity"):
                response = self.client.post(
                    "/api/bom/upload",
                    data={"group_model": "MODEL-A"},
                    files=[("files", ("upload.xlsx", b"dummy", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
                )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["errors"], [])
        self.assertEqual(len(body["saved"]), 1)
        saved = body["saved"][0]
        self.assertEqual(saved["id"], "bom-existing")
        self.assertTrue(saved["replaced_existing"])
        self.assertEqual(saved["removed_duplicates"], 1)
        self.assertEqual(saved["components"], 1)

        self.assertEqual(mock_save_bom.call_count, 2)
        first_payload = mock_save_bom.call_args_list[0].args[0]
        second_payload = mock_save_bom.call_args_list[1].args[0]
        self.assertEqual(first_payload["id"], "bom-existing")
        self.assertEqual(first_payload["filepath"], str(incoming_path))
        self.assertEqual(second_payload["id"], "bom-existing")
        self.assertEqual(Path(second_payload["filepath"]).name, "bom-existing.xlsx")
        mock_ensure_history.assert_called_once_with(existing_bom)
        mock_snapshot_revision.assert_called_once()
        mock_delete_revision_files.assert_called_once_with("bom-duplicate")
        mock_delete_bom_file.assert_called_once_with("bom-duplicate")

    def test_batch_merge_creates_merge_drafts(self):
        with patch("app.routers.schedule.db.batch_merge_orders") as mock_batch_merge, \
             patch("app.routers.schedule.rebuild_merge_drafts", return_value=[{"id": 7}, {"id": 8}]) as mock_rebuild, \
             patch("app.routers.schedule.db.log_activity"), \
             patch("app.routers.schedule.db.create_alert"):
            response = self.client.post("/api/schedule/batch-merge", json={"order_ids": [1, 2]})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {"ok": True, "count": 2, "draft_count": 2})
        mock_batch_merge.assert_called_once_with([1, 2])
        mock_rebuild.assert_called_once_with([1, 2])

    def test_update_selected_schedule_drafts_allocates_supplements_per_order(self):
        with patch("app.routers.schedule.db.get_active_merge_draft_ids_by_order_ids", return_value={1: 11, 2: 12}), \
             patch("app.routers.schedule.build_order_supplement_allocations", return_value={
                 1: {"PART-1": 3000},
                 2: {},
             }) as mock_allocations, \
             patch("app.routers.schedule.rebuild_merge_drafts", return_value=[{"id": 11}, {"id": 12}]) as mock_rebuild, \
             patch("app.routers.schedule.get_schedule_draft_map", return_value={
                 1: {"id": 11, "order_id": 1},
                 2: {"id": 12, "order_id": 2},
             }), \
             patch("app.routers.schedule.db.log_activity"):
            response = self.client.put("/api/schedule/drafts", json={
                "order_ids": [1, 2],
                "decisions": {"part-1": "Shortage"},
                "supplements": {"part-1": 3000},
            })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["count"], 2)
        self.assertEqual(response.json()["draft_count"], 2)
        mock_allocations.assert_called_once_with([1, 2], {"part-1": 3000})
        mock_rebuild.assert_called_once_with([1, 2])

    def test_update_selected_schedule_drafts_prefers_order_scoped_supplements(self):
        with patch("app.routers.schedule.db.get_active_merge_draft_ids_by_order_ids", return_value={1: 11, 2: 12}), \
             patch("app.routers.schedule.build_order_supplement_allocations", return_value={
                 1: {"IC-STM32F": 150},
                 2: {},
             }), \
             patch("app.routers.schedule.rebuild_merge_drafts", return_value=[{"id": 11}, {"id": 12}]) as mock_rebuild, \
             patch("app.routers.schedule.get_schedule_draft_map", return_value={
                 1: {"id": 11, "order_id": 1},
                 2: {"id": 12, "order_id": 2},
             }), \
             patch("app.routers.schedule.db.log_activity"):
            response = self.client.put("/api/schedule/drafts", json={
                "order_ids": [1, 2],
                "decisions": {},
                "supplements": {"IC-STM32F": 150},
                "order_supplements": {
                    "1": {"IC-STM32F": 100},
                    "2": {"IC-STM32F": 50},
                },
            })

        self.assertEqual(response.status_code, 200)
        mock_rebuild.assert_called_once_with([1, 2])

    def test_update_selected_schedule_drafts_prefers_explicit_repeated_normal_part_supplements(self):
        with patch("app.routers.schedule.db.get_active_merge_draft_ids_by_order_ids", return_value={1: 11, 2: 12}), \
             patch("app.routers.schedule.build_order_supplement_allocations", return_value={
                 1: {"EC-30059A": 750},
                 2: {"EC-30059A": 750},
             }), \
             patch("app.routers.schedule.db.replace_order_decisions") as mock_replace_decisions, \
             patch("app.routers.schedule.db.replace_order_supplements") as mock_replace_supplements, \
             patch("app.routers.schedule.rebuild_merge_drafts", return_value=[{"id": 11}, {"id": 12}]) as mock_rebuild, \
             patch("app.routers.schedule.get_schedule_draft_map", return_value={
                 1: {"id": 11, "order_id": 1},
                 2: {"id": 12, "order_id": 2},
             }), \
             patch("app.routers.schedule.db.log_activity"):
            response = self.client.put("/api/schedule/drafts", json={
                "order_ids": [1, 2],
                "decisions": {},
                "supplements": {"EC-30059A": 1500},
                "order_supplements": {
                    "1": {"EC-30059A": 1500},
                    "2": {},
                },
            })

        self.assertEqual(response.status_code, 200)
        mock_replace_decisions.assert_called_once()
        mock_replace_supplements.assert_called_once_with([1, 2], {
            1: {"EC-30059A": 1500.0},
            2: {},
        })
        mock_rebuild.assert_called_once_with([1, 2])

    def test_download_selected_schedule_drafts_proxies_to_bundle_service(self):
        with patch(
            "app.routers.schedule.download_selected_merge_drafts",
            return_value=Response(content=b"ok", media_type="application/zip"),
        ) as mock_download:
            response = self.client.post("/api/schedule/drafts/download", json={"order_ids": [1, 2]})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content, b"ok")
        mock_download.assert_called_once_with([1, 2], request=ANY)

    def test_dispatch_order_saves_dispatch_session(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.write_bytes(b"main")

            order = {
                "id": 1,
                "po_number": "4500059234",
                "model": "MODEL-A",
                "code": "1-3",
                "status": "merged",
            }
            bom_file = {"id": "bom-1", "model": "MODEL-A"}
            components = [{"part_number": "PART-1", "needed_qty": 5, "prev_qty_cs": 0, "is_dash": 0}]

            with patch("app.routers.schedule.db.get_order", return_value=order), \
                 patch("app.routers.schedule.db.get_setting", return_value=str(main_path)), \
                 patch("app.routers.schedule._get_effective_moq", return_value={}), \
                 patch("app.routers.schedule.db.get_bom_files_by_models", return_value=[bom_file]), \
                 patch("app.routers.schedule.db.get_bom_components", return_value=components), \
                 patch("app.routers.schedule.preview_order_batches", return_value={
                     "merged_parts": 1,
                     "shortages": [],
                 }), \
                 patch("app.routers.schedule.merge_row_to_main", return_value={"merged_parts": 1, "backup_path": "C:/backup.xlsx"}), \
                 patch("app.routers.schedule.db.save_dispatch_records") as mock_records, \
                 patch("app.routers.schedule.db.save_dispatch_session") as mock_session, \
                 patch("app.routers.schedule.db.update_order"), \
                 patch("app.routers.schedule.db.log_activity"):
                response = self.client.post("/api/schedule/orders/1/dispatch", json={"decisions": {}})

        self.assertEqual(response.status_code, 200)
        mock_records.assert_called_once()
        mock_session.assert_called_once_with(
            order_id=1,
            previous_status="merged",
            backup_path="C:/backup.xlsx",
            main_file_path=str(main_path),
        )

    def test_batch_dispatch_processes_checked_orders_in_order(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.write_bytes(b"main")

            context_a = ({"id": 1, "po_number": "4500059234", "model": "MODEL-A", "status": "merged"}, [{"components": []}], [])
            context_b = ({"id": 2, "po_number": "4500059235", "model": "MODEL-B", "status": "pending"}, [{"components": []}], [])

            with patch("app.routers.schedule.db.get_setting", return_value=str(main_path)), \
                 patch("app.routers.schedule._get_effective_moq", return_value={}), \
                 patch("app.routers.schedule._prepare_dispatch_context", side_effect=[context_a, context_b]) as mock_prepare, \
                 patch("app.routers.schedule.preview_order_batches", return_value={
                     "merged_parts": 7,
                     "shortages": [],
                 }), \
                 patch("app.routers.schedule._execute_dispatch", side_effect=[
                     {"order_id": 1, "merged_parts": 3, "backup_path": "C:/b1.xlsx", "session": {"id": 11}},
                     {"order_id": 2, "merged_parts": 4, "backup_path": "C:/b2.xlsx", "session": {"id": 12}},
                 ]) as mock_execute, \
                 patch("app.routers.schedule.db.log_activity"):
                response = self.client.post("/api/schedule/batch-dispatch", json={
                    "order_ids": [1, 2],
                    "decisions": {"part-1": "CreateRequirement"},
                })

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["count"], 2)
        self.assertEqual(data["merged_parts"], 7)
        self.assertEqual(data["order_ids"], [1, 2])
        mock_prepare.assert_has_calls([call(1, str(main_path)), call(2, str(main_path))])
        self.assertEqual(mock_execute.call_args_list[0].args[4], {"PART-1": "CreateRequirement"})
        self.assertEqual(mock_execute.call_args_list[1].args[4], {})

    def test_preview_defective_import_returns_missing_items(self):
        with patch("app.routers.defectives._require_main_path", return_value="C:/main.xlsx"), \
             patch("app.routers.defectives.parse_defective_excel", return_value=[
                 {"source_row": 6, "part_number": "PART-X", "description": "", "defective_qty": 15},
             ]), \
             patch("app.routers.defectives.build_overrun_import_preview", return_value={
                 "source_filename": "defective.xlsx",
                 "item_count": 2,
                 "deducted_count": 1,
                 "negative_count": 0,
                 "total_deduction_qty": 15,
                 "items": [],
                 "results": [],
                 "missing_items": [{"source_row": 6, "part_number": "PART-X", "suggestions": []}],
                 "missing_count": 1,
                 "requires_confirmation": True,
             }) as mock_preview:
            response = self.client.post(
                "/api/defectives/import-preview",
                files={"file": ("defective.xlsx", io.BytesIO(b"detail"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["missing_count"], 1)
        mock_preview.assert_called_once_with("C:/main.xlsx", {
            "source_filename": "defective.xlsx",
            "items": [{"source_row": 6, "part_number": "PART-X", "description": "", "defective_qty": 15}],
        })

    def test_confirm_defective_import_creates_batch_and_records(self):
        with patch("app.routers.defectives._require_main_path", return_value="C:/main.xlsx"), \
             patch("app.routers.defectives._get_main_file_mtime", return_value=321.0), \
             patch("app.routers.defectives.apply_overrun_import_confirmations", return_value={
                 "final_items": [
                     {"source_row": 5, "part_number": "PART-1", "description": "Cap", "defective_qty": 12, "source_part_number": "PART-1"},
                 ],
                 "replaced_items": [],
                 "skipped_items": [
                     {"source_row": 6, "part_number": "PART-X", "defective_qty": 3},
                 ],
                 "unresolved_items": [],
             }), \
             patch("app.routers.defectives.deduct_defectives_from_main", return_value={
                 "deducted_count": 1,
                 "skipped_parts": [],
                 "results": [
                     {"part_number": "PART-1", "stock_before": 50, "stock_after": 38},
                 ],
             }) as mock_deduct, \
             patch("app.routers.defectives.refresh_snapshot_from_main"), \
             patch("app.routers.defectives.db.create_defective_batch", return_value=66) as mock_batch, \
             patch("app.routers.defectives.db.create_defective_record", return_value=701) as mock_record, \
             patch("app.routers.defectives.db.log_activity") as mock_log:
            response = self.client.post("/api/defectives/import-confirm", json={
                "source_filename": "defective.xlsx",
                "items": [
                    {"source_row": 5, "part_number": "PART-1", "defective_qty": 12, "action": "deduct", "target_part_number": ""},
                    {"source_row": 6, "part_number": "PART-X", "defective_qty": 3, "action": "skip", "target_part_number": ""},
                ],
            })

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["batch_id"], 66)
        self.assertEqual(data["deducted_count"], 1)
        self.assertEqual(data["skipped_count"], 1)
        mock_batch.assert_called_once()
        mock_record.assert_called_once()
        mock_deduct.assert_called_once()
        mock_log.assert_called_once()

    def test_confirm_defective_import_can_append_existing_batch(self):
        with patch("app.routers.defectives._require_main_path", return_value="C:/main.xlsx"), \
             patch("app.routers.defectives.apply_overrun_import_confirmations", return_value={
                 "final_items": [
                     {"source_row": 5, "part_number": "PART-1", "description": "Cap", "defective_qty": 12, "source_part_number": "PART-1"},
                 ],
                 "replaced_items": [],
                 "skipped_items": [],
                 "unresolved_items": [],
             }), \
             patch("app.routers.defectives.deduct_defectives_from_main", return_value={
                 "deducted_count": 1,
                 "skipped_parts": [],
                 "results": [
                     {"part_number": "PART-1", "stock_before": 50, "stock_after": 38},
                 ],
             }), \
             patch("app.routers.defectives.refresh_snapshot_from_main"), \
             patch("app.routers.defectives.db.get_defective_batches", return_value=[{
                 "id": 90,
                 "filename": "old.xlsx",
                 "items": [],
             }]), \
             patch("app.routers.defectives.db.create_defective_batch") as mock_create_batch, \
             patch("app.routers.defectives.db.create_defective_record", return_value=702) as mock_record, \
             patch("app.routers.defectives.db.log_activity") as mock_log:
            response = self.client.post("/api/defectives/import-confirm", json={
                "batch_id": 90,
                "source_filename": "append.xlsx",
                "items": [
                    {"source_row": 5, "part_number": "PART-1", "defective_qty": 12, "action": "deduct", "target_part_number": ""},
                ],
            })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["batch_id"], 90)
        mock_create_batch.assert_not_called()
        mock_record.assert_called_once()
        mock_log.assert_called_once()

    def test_overrun_preview_returns_plan_and_stock_preview(self):
        with patch("app.routers.defectives._require_main_path", return_value="C:/main.xlsx"), \
             patch("app.routers.defectives.build_model_overrun_plan", return_value={
                 "model": "MODEL-A",
                 "requested_model": "model-a",
                 "extra_pcs": 25,
                 "matched_models": ["MODEL-A"],
                 "matched_boms": [{"id": "bom-1", "filename": "model-a.xlsx", "model": "MODEL-A", "group_model": "MODEL-A"}],
                 "items": [{"part_number": "PART-1", "description": "IC", "defective_qty": 50}],
             }) as mock_plan, \
             patch("app.routers.defectives.preview_deductions_against_main", return_value={
                 "deducted_count": 1,
                 "skipped_parts": [],
                 "results": [{"part_number": "PART-1", "stock_before": 120, "stock_after": 70}],
                 "negative_count": 0,
                 "total_deduction_qty": 50,
             }) as mock_preview:
            response = self.client.post("/api/defectives/overrun/preview", json={
                "model": "model-a",
                "extra_pcs": 25,
                "reason": "加工廠多打",
                "note": "",
                "reported_by": "Andy",
            })

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["model"], "MODEL-A")
        self.assertEqual(data["deducted_count"], 1)
        mock_plan.assert_called_once_with("model-a", 25.0)
        mock_preview.assert_called_once()

    def test_create_model_overrun_creates_batch_and_records(self):
        with patch("app.routers.defectives._require_main_path", return_value="C:/main.xlsx"), \
             patch("app.routers.defectives._get_main_file_mtime", return_value=123.0), \
             patch("app.routers.defectives.build_model_overrun_plan", return_value={
                 "model": "MODEL-A",
                 "requested_model": "MODEL-A",
                 "extra_pcs": 10,
                 "matched_models": ["MODEL-A"],
                 "matched_boms": [{"id": "bom-1", "filename": "model-a.xlsx", "model": "MODEL-A", "group_model": "MODEL-A"}],
                 "items": [
                     {"part_number": "PART-1", "description": "IC-1", "defective_qty": 30},
                     {"part_number": "PART-2", "description": "IC-2", "defective_qty": 20},
                 ],
             }), \
             patch("app.routers.defectives.deduct_defectives_from_main", return_value={
                 "deducted_count": 2,
                 "skipped_parts": [],
                 "results": [
                     {"part_number": "PART-1", "stock_before": 100, "stock_after": 70},
                     {"part_number": "PART-2", "stock_before": 80, "stock_after": 60},
                 ],
             }) as mock_deduct, \
             patch("app.routers.defectives.refresh_snapshot_from_main"), \
             patch("app.routers.defectives.db.create_defective_batch", return_value=55) as mock_batch, \
             patch("app.routers.defectives.db.create_defective_record", side_effect=[101, 102]) as mock_record, \
             patch("app.routers.defectives.db.log_activity") as mock_log:
            response = self.client.post("/api/defectives/overrun", json={
                "model": "MODEL-A",
                "extra_pcs": 10,
                "reason": "加工廠多打",
                "note": "晚班補單",
                "reported_by": "Andy",
            })

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["batch_id"], 55)
        self.assertEqual(data["batch_type"], "overrun")
        self.assertEqual(mock_record.call_count, 2)
        self.assertTrue(mock_batch.call_args.args[0].startswith("加工多打｜MODEL-A｜+10"))
        self.assertIn("加工廠多打", mock_batch.call_args.kwargs["note"])
        mock_deduct.assert_called_once()
        self.assertEqual(mock_deduct.call_args.kwargs["entry_header"], "加工多打扣帳")
        mock_log.assert_called_once()

    def test_create_model_overrun_rebuilds_active_merge_drafts_after_main_change(self):
        with patch("app.routers.defectives._require_main_path", return_value="C:/main.xlsx"), \
             patch("app.routers.defectives._get_main_file_mtime", return_value=123.0), \
             patch("app.routers.defectives.build_model_overrun_plan", return_value={
                 "model": "MODEL-A",
                 "requested_model": "MODEL-A",
                 "extra_pcs": 10,
                 "matched_models": ["MODEL-A"],
                 "matched_boms": [{"id": "bom-1", "filename": "model-a.xlsx", "model": "MODEL-A", "group_model": "MODEL-A"}],
                 "items": [{"part_number": "PART-1", "description": "IC-1", "defective_qty": 30}],
             }), \
             patch("app.routers.defectives.deduct_defectives_from_main", return_value={
                 "deducted_count": 1,
                 "skipped_parts": [],
                 "results": [{"part_number": "PART-1", "stock_before": 100, "stock_after": 70}],
             }), \
             patch("app.routers.defectives.refresh_snapshot_from_main"), \
             patch("app.routers.defectives.db.get_active_merge_drafts", return_value=[
                 {"id": 9, "order_id": 101},
                 {"id": 10, "order_id": 102},
             ]), \
             patch("app.routers.defectives.rebuild_merge_drafts") as mock_rebuild, \
             patch("app.routers.defectives.db.create_defective_batch", return_value=55), \
             patch("app.routers.defectives.db.create_defective_record", return_value=101), \
             patch("app.routers.defectives.db.log_activity"):
            response = self.client.post("/api/defectives/overrun", json={
                "model": "MODEL-A",
                "extra_pcs": 10,
                "reason": "加工廠多打",
                "note": "",
                "reported_by": "",
            })

        self.assertEqual(response.status_code, 200)
        mock_rebuild.assert_called_once_with([101, 102])

    def test_import_overrun_detail_file_creates_overrun_batch(self):
        with patch("app.routers.defectives._require_main_path", return_value="C:/main.xlsx"), \
             patch("app.routers.defectives._get_main_file_mtime", return_value=456.0), \
             patch("app.routers.defectives.parse_overrun_detail_excel", return_value={
                 "title": "12-7 02/02 T356789IU-A (REV3.4) * 4",
                 "mo_info": "4500059105 多打 S+D",
                 "items": [
                     {"part_number": "EC-10025A", "description": "", "defective_qty": 56},
                     {"part_number": "EC-10028A", "description": "", "defective_qty": 24},
                 ],
             }), \
             patch("app.routers.defectives.deduct_defectives_from_main", return_value={
                 "deducted_count": 2,
                 "skipped_parts": [],
                 "results": [
                     {"part_number": "EC-10025A", "stock_before": 100, "stock_after": 44},
                     {"part_number": "EC-10028A", "stock_before": 50, "stock_after": 26},
                 ],
             }) as mock_deduct, \
             patch("app.routers.defectives.refresh_snapshot_from_main"), \
             patch("app.routers.defectives.db.create_defective_batch", return_value=77) as mock_batch, \
             patch("app.routers.defectives.db.create_defective_record", side_effect=[201, 202]) as mock_record, \
             patch("app.routers.defectives.db.log_activity") as mock_log:
            response = self.client.post(
                "/api/defectives/overrun/import",
                files={"file": ("20260202 T356789IU多打扣帳明細.xlsx", io.BytesIO(b"detail"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["batch_id"], 77)
        self.assertEqual(data["batch_type"], "overrun")
        self.assertEqual(mock_record.call_count, 2)
        self.assertTrue(mock_batch.call_args.args[0].startswith("加工多打明細｜20260202 T356789IU多打扣帳明細.xlsx"))
        self.assertIn("4500059105 多打 S+D", mock_batch.call_args.kwargs["note"])
        mock_deduct.assert_called_once()
        self.assertEqual(mock_deduct.call_args.kwargs["entry_header"], "加工多打扣帳")
        mock_log.assert_called_once()

    def test_preview_overrun_detail_import_returns_missing_items(self):
        with patch("app.routers.defectives._require_main_path", return_value="C:/main.xlsx"), \
             patch("app.routers.defectives.parse_overrun_detail_excel", return_value={
                 "title": "12-7 02/02 T356789IU-A (REV3.4) * 4",
                 "mo_info": "4500059105 多打 S+D",
                 "items": [{"source_row": 3, "part_number": "EC-10029A", "description": "", "defective_qty": 24}],
             }), \
             patch("app.routers.defectives.build_overrun_import_preview", return_value={
                 "title": "12-7 02/02 T356789IU-A (REV3.4) * 4",
                 "mo_info": "4500059105 多打 S+D",
                 "source_filename": "20260202 T356789IU多打扣帳明細.xlsx",
                 "item_count": 1,
                 "deducted_count": 0,
                 "negative_count": 0,
                 "total_deduction_qty": 0,
                 "items": [],
                 "results": [],
                 "missing_items": [{"source_row": 3, "part_number": "EC-10029A", "suggestions": [{"part_number": "EC-10028A", "stock_qty": 50}]}],
                 "missing_count": 1,
                 "requires_confirmation": True,
             }) as mock_preview:
            response = self.client.post(
                "/api/defectives/overrun/import-preview",
                files={"file": ("20260202 T356789IU多打扣帳明細.xlsx", io.BytesIO(b"detail"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["missing_count"], 1)
        mock_preview.assert_called_once()

    def test_confirm_overrun_detail_import_applies_replacements(self):
        with patch("app.routers.defectives._require_main_path", return_value="C:/main.xlsx"), \
             patch("app.routers.defectives._get_main_file_mtime", return_value=789.0), \
             patch("app.routers.defectives.apply_overrun_import_confirmations", return_value={
                 "final_items": [
                     {"source_row": 3, "part_number": "EC-10025A", "description": "", "defective_qty": 56, "source_part_number": "EC-10025A"},
                     {"source_row": 4, "part_number": "EC-10028A", "description": "", "defective_qty": 24, "source_part_number": "EC-10029A"},
                 ],
                 "replaced_items": [
                     {"source_row": 4, "source_part_number": "EC-10029A", "target_part_number": "EC-10028A", "defective_qty": 24},
                 ],
                 "skipped_items": [
                     {"source_row": 5, "part_number": "EC-10099A", "defective_qty": 10},
                 ],
                 "unresolved_items": [],
             }), \
             patch("app.routers.defectives.deduct_defectives_from_main", return_value={
                 "deducted_count": 2,
                 "skipped_parts": [],
                 "results": [
                     {"part_number": "EC-10025A", "stock_before": 100, "stock_after": 44},
                     {"part_number": "EC-10028A", "stock_before": 50, "stock_after": 26},
                 ],
             }) as mock_deduct, \
             patch("app.routers.defectives.refresh_snapshot_from_main"), \
             patch("app.routers.defectives.db.create_defective_batch", return_value=88) as mock_batch, \
             patch("app.routers.defectives.db.create_defective_record", side_effect=[301, 302]) as mock_record, \
             patch("app.routers.defectives.db.log_activity") as mock_log:
            response = self.client.post("/api/defectives/overrun/import-confirm", json={
                "source_filename": "20260202 T356789IU多打扣帳明細.xlsx",
                "title": "12-7 02/02 T356789IU-A (REV3.4) * 4",
                "mo_info": "4500059105 多打 S+D",
                "items": [
                    {"source_row": 3, "part_number": "EC-10025A", "defective_qty": 56, "action": "deduct", "target_part_number": ""},
                    {"source_row": 4, "part_number": "EC-10029A", "defective_qty": 24, "action": "replace", "target_part_number": "EC-10028A"},
                    {"source_row": 5, "part_number": "EC-10099A", "defective_qty": 10, "action": "skip", "target_part_number": ""},
                ],
            })

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["batch_id"], 88)
        self.assertEqual(data["replaced_count"], 1)
        self.assertEqual(data["skipped_count"], 1)
        self.assertEqual(mock_record.call_count, 2)
        self.assertIn("改正料號", mock_batch.call_args.kwargs["note"])
        mock_deduct.assert_called_once()
        self.assertEqual(mock_deduct.call_args.kwargs["entry_header"], "加工多打扣帳")
        mock_log.assert_called_once()

    def test_delete_overrun_batch_uses_overrun_reverse_header(self):
        with patch("app.routers.defectives.db.get_defective_batches", return_value=[{
            "id": 88,
            "filename": "加工多打｜MODEL-A｜+10 pcs",
            "imported_at": "2026-03-20T08:00:00",
            "main_file_mtime": 123.0,
            "items": [
                {"part_number": "PART-1", "defective_qty": 30, "action_taken": "加工多打扣帳"},
            ],
        }]), \
             patch("app.services.inventory_restore_guard.db.get_defective_batch_summaries_after_id", return_value=[]), \
             patch("app.services.inventory_restore_guard.db.get_active_dispatch_sessions_after", return_value=[]), \
             patch("app.services.inventory_restore_guard.db.get_activity_logs_after", return_value=[]), \
             patch("app.routers.defectives.db.get_setting", return_value="C:/main.xlsx"), \
             patch("app.routers.defectives._get_main_file_mtime", return_value=123.0), \
             patch("app.routers.defectives.Path.exists", return_value=True), \
             patch("app.routers.defectives.reverse_defectives_from_main", return_value={
                 "reversed_count": 1,
                 "skipped_parts": [],
                 "results": [],
             }) as mock_reverse, \
             patch("app.routers.defectives.refresh_snapshot_from_main"), \
             patch("app.routers.defectives.db.delete_defective_batch", return_value=True), \
             patch("app.routers.defectives.db.log_activity"):
            response = self.client.delete("/api/defectives/batches/88")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["reversed_count"], 1)
        self.assertEqual(mock_reverse.call_args.kwargs["entry_header"], "加工多打回復")

    def test_delete_batch_rebuilds_active_merge_drafts_after_reverse(self):
        with patch("app.routers.defectives.db.get_defective_batches", return_value=[{
            "id": 88,
            "filename": "不良品批次.xlsx",
            "imported_at": "2026-03-20T08:00:00",
            "main_file_mtime": 123.0,
            "items": [{"part_number": "PART-1", "defective_qty": 30}],
        }]), \
             patch("app.services.inventory_restore_guard.db.get_defective_batch_summaries_after_id", return_value=[]), \
             patch("app.services.inventory_restore_guard.db.get_active_dispatch_sessions_after", return_value=[]), \
             patch("app.services.inventory_restore_guard.db.get_activity_logs_after", return_value=[]), \
             patch("app.routers.defectives.db.get_setting", return_value="C:/main.xlsx"), \
             patch("app.routers.defectives._get_main_file_mtime", return_value=123.0), \
             patch("app.routers.defectives.Path.exists", return_value=True), \
             patch("app.routers.defectives.reverse_defectives_from_main", return_value={
                 "reversed_count": 1,
                 "skipped_parts": [],
                 "results": [],
             }), \
             patch("app.routers.defectives.refresh_snapshot_from_main"), \
             patch("app.routers.defectives.db.get_active_merge_drafts", return_value=[
                 {"id": 9, "order_id": 101},
             ]), \
             patch("app.routers.defectives.rebuild_merge_drafts") as mock_rebuild, \
             patch("app.routers.defectives.db.delete_defective_batch", return_value=True), \
             patch("app.routers.defectives.db.log_activity"):
            response = self.client.delete("/api/defectives/batches/88")

        self.assertEqual(response.status_code, 200)
        mock_rebuild.assert_called_once_with([101])

    def test_delete_defective_batch_blocks_when_later_inventory_mutation_exists(self):
        with patch("app.routers.defectives.db.get_defective_batches", return_value=[{
            "id": 88,
            "filename": "不良品批次.xlsx",
            "imported_at": "2026-03-20T08:00:00",
            "main_file_mtime": 123.0,
            "items": [
                {"part_number": "PART-1", "defective_qty": 30},
            ],
        }]), \
             patch("app.services.inventory_restore_guard.db.get_defective_batch_summaries_after_id", return_value=[{"id": 99}]), \
             patch("app.services.inventory_restore_guard.db.get_active_dispatch_sessions_after", return_value=[]), \
             patch("app.services.inventory_restore_guard.db.get_activity_logs_after", return_value=[]), \
             patch("app.routers.defectives.reverse_defectives_from_main") as mock_reverse, \
             patch("app.routers.defectives.db.delete_defective_batch") as mock_delete_batch:
            response = self.client.delete("/api/defectives/batches/88")

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.json()["detail"],
            "後面已有其他庫存異動，不能直接回復。請先下載目前主檔，手動修正後重新上傳主檔。重新上傳後一定要重設快照。",
        )
        mock_reverse.assert_not_called()
        mock_delete_batch.assert_not_called()

    def test_load_active_merge_draft_context_accepts_repaired_main_path(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.write_bytes(b"main")
            signature = str(main_path.stat().st_mtime_ns)

            draft = {
                "id": 9,
                "order_id": 1,
                "status": "active",
                "main_file_path": "Z:/legacy/main.xlsx",
                "main_file_mtime_ns": signature,
                "decisions": {"PART-1": "Shortage"},
                "supplements": {"PART-1": 3000},
            }
            order = {"id": 1, "po_number": "4500059234", "model": "MODEL-A"}

            def fake_resolve(path_value: str, setting_key: str = ""):
                if str(path_value or "").strip():
                    return str(main_path)
                return ""

            with patch("app.routers.schedule.db.get_merge_draft", return_value=draft), \
                 patch("app.routers.schedule.db.get_order_decisions", return_value={1: {"PART-1": "Shortage"}}), \
                 patch("app.routers.schedule.db.get_order_supplements", return_value={1: {"PART-1": 3000}}), \
                 patch("app.routers.schedule.db.resolve_managed_path", side_effect=fake_resolve), \
                 patch("app.routers.schedule._prepare_dispatch_context", return_value=(order, [{"components": []}], [])):
                context = schedule_router._load_active_merge_draft_context(9, str(main_path))

        self.assertEqual(context.draft["id"], 9)
        self.assertEqual(context.order["id"], 1)
        self.assertEqual(context.decisions, {"PART-1": "Shortage"})
        self.assertEqual(context.supplements, {"PART-1": 3000.0})

    def test_main_write_preview_uses_allocated_supplements(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.write_bytes(b"main")

            context_a = ({"id": 1, "po_number": "4500059234", "model": "MODEL-A", "status": "merged"}, [{"components": []}], [])
            context_b = ({"id": 2, "po_number": "4500059235", "model": "MODEL-B", "status": "pending"}, [{"components": []}], [])

            with patch("app.routers.schedule.db.get_setting", return_value=str(main_path)), \
                 patch("app.routers.schedule._get_effective_moq", return_value={}), \
                 patch("app.routers.schedule._prepare_dispatch_context", side_effect=[context_a, context_b]), \
                 patch("app.routers.schedule.build_order_supplement_allocations", return_value={
                     1: {"PART-1": 3000},
                     2: {},
                 }) as mock_allocations, \
                 patch("app.routers.schedule._get_effective_moq", return_value={"PART-1": 500}), \
                 patch("app.routers.schedule.preview_order_batches", return_value={
                     "merged_parts": 4,
                     "shortages": [{"part_number": "PART-1", "moq": 500, "shortage_amount": 12}],
                 }) as mock_preview:
                response = self.client.post("/api/schedule/main-write-preview", json={
                    "order_ids": [1, 2],
                    "decisions": {"part-1": "CreateRequirement"},
                    "supplements": {"part-1": 3000},
                })

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["count"], 2)
        self.assertEqual(data["merged_parts"], 4)
        self.assertEqual(data["shortages"], [{"part_number": "PART-1", "moq": 500, "shortage_amount": 12}])
        mock_allocations.assert_called_once_with([1, 2], {"PART-1": 3000.0})
        self.assertEqual(mock_preview.call_args.kwargs["moq_map"], {"PART-1": 500})
        self.assertEqual(
            mock_preview.call_args.args[1],
            [
                {
                    "order_id": 1,
                    "model": "MODEL-A",
                    "groups": [{"components": []}],
                    "supplements": {"PART-1": 3000},
                    "decisions": {"PART-1": "CreateRequirement"},
                },
                {
                    "order_id": 2,
                    "model": "MODEL-B",
                    "groups": [{"components": []}],
                    "supplements": {},
                    "decisions": {},
                },
            ],
        )
        self.assertEqual(mock_preview.call_args.args[2], {})

    def test_main_write_preview_uses_active_draft_supplements_when_available(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.write_bytes(b"main")
            context = {
                "draft": {
                    "id": 11,
                    "order_id": 1,
                    "shortages": [{
                        "part_number": "EC-10264A",
                        "shortage_amount": 18,
                        "supplement_qty": 3000,
                        "decision": "CreateRequirement",
                    }],
                },
                "order": {"id": 1, "model": "T356789IU+LCD", "code": "1-3"},
                "groups": [{"components": []}],
                "all_components": [{"part_number": "EC-10264A"}],
                "decisions": {"EC-10264A": "CreateRequirement"},
                "supplements": {"EC-10264A": 3000},
            }

            with patch("app.routers.schedule.db.get_setting", return_value=str(main_path)), \
                 patch("app.routers.schedule.db.get_active_merge_draft_ids_by_order_ids", return_value={1: 11}), \
                 patch("app.routers.schedule.rebuild_merge_drafts", return_value=[{"id": 11}]) as mock_rebuild, \
                 patch("app.routers.schedule._get_effective_moq", return_value={"EC-10264A": 500}), \
                 patch("app.routers.schedule._load_active_merge_draft_context", return_value=context), \
                 patch("app.routers.schedule.preview_order_batches", return_value={
                     "merged_parts": 1,
                     "shortages": [{
                         "part_number": "EC-10264A",
                         "shortage_amount": 18,
                         "supplement_qty": 3000,
                         "decision": "CreateRequirement",
                         "model": "T356789IU+LCD",
                         "batch_code": "1-3",
                     }],
                 }):
                response = self.client.post("/api/schedule/main-write-preview", json={
                    "order_ids": [1],
                    "decisions": {},
                    "supplements": {},
                })

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["count"], 1)
        self.assertEqual(data["shortages"][0]["part_number"], "EC-10264A")
        self.assertEqual(data["shortages"][0]["supplement_qty"], 3000)
        self.assertEqual(data["shortages"][0]["model"], "T356789IU+LCD")
        self.assertEqual(data["shortages"][0]["batch_code"], "1-3")
        mock_rebuild.assert_called_once_with([1])

    def test_batch_dispatch_passes_allocated_supplements_to_each_order(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.write_bytes(b"main")

            context_a = ({"id": 1, "po_number": "4500059234", "model": "MODEL-A", "status": "merged"}, [{"components": []}], [])
            context_b = ({"id": 2, "po_number": "4500059235", "model": "MODEL-B", "status": "pending"}, [{"components": []}], [])

            with patch("app.routers.schedule.db.get_setting", return_value=str(main_path)), \
                 patch("app.routers.schedule._get_effective_moq", return_value={}), \
                 patch("app.routers.schedule._prepare_dispatch_context", side_effect=[context_a, context_b]), \
                 patch("app.routers.schedule.build_order_supplement_allocations", return_value={
                     1: {"PART-1": 3000},
                     2: {},
                 }), \
                 patch("app.routers.schedule.preview_order_batches", return_value={
                     "merged_parts": 5,
                     "shortages": [],
                 }), \
                 patch("app.routers.schedule._execute_dispatch", side_effect=[
                     {"order_id": 1, "merged_parts": 3, "backup_path": "C:/b1.xlsx", "session": {"id": 11}},
                     {"order_id": 2, "merged_parts": 2, "backup_path": "C:/b2.xlsx", "session": {"id": 12}},
                 ]) as mock_execute, \
                 patch("app.routers.schedule.db.replace_order_supplements") as mock_replace, \
                 patch("app.routers.schedule.db.log_activity"):
                response = self.client.post("/api/schedule/batch-dispatch", json={
                    "order_ids": [1, 2],
                    "decisions": {"part-1": "CreateRequirement"},
                    "supplements": {"part-1": 3000},
                })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(mock_execute.call_args_list[0].args[5], {"PART-1": 3000})
        self.assertEqual(mock_execute.call_args_list[1].args[5], {})
        mock_replace.assert_called_once_with([1, 2], {1: {"PART-1": 3000}, 2: {}})

    def test_main_write_preview_and_batch_dispatch_share_same_scoped_plan(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.write_bytes(b"main")

            groups_a = [{
                "batch_code": "1-1",
                "po_number": "4500059234",
                "bom_model": "MODEL-A",
                "components": [{"part_number": "PART-A", "needed_qty": 5, "prev_qty_cs": 0}],
            }]
            groups_b = [{
                "batch_code": "2-1",
                "po_number": "4500059235",
                "bom_model": "MODEL-B",
                "components": [{"part_number": "PART-B", "needed_qty": 8, "prev_qty_cs": 0}],
            }]
            context_a = ({"id": 1, "po_number": "4500059234", "model": "MODEL-A", "status": "merged"}, groups_a, [{"part_number": "PART-A"}])
            context_b = ({"id": 2, "po_number": "4500059235", "model": "MODEL-B", "status": "pending"}, groups_b, [{"part_number": "PART-B"}])
            expected_batches = [
                {
                    "order_id": 1,
                    "model": "MODEL-A",
                    "groups": groups_a,
                    "supplements": {"PART-A": 800},
                    "decisions": {"PART-A": "Shortage"},
                },
                {
                    "order_id": 2,
                    "model": "MODEL-B",
                    "groups": groups_b,
                    "supplements": {"PART-B": 1200},
                    "decisions": {"PART-B": "CreateRequirement"},
                },
            ]

            with patch("app.routers.schedule.db.get_setting", return_value=str(main_path)), \
                 patch("app.routers.schedule.db.get_active_merge_draft_ids_by_order_ids", return_value={}), \
                 patch("app.routers.schedule._get_effective_moq", return_value={"PART-A": 500, "PART-B": 1000}), \
                 patch("app.routers.schedule._prepare_dispatch_context", side_effect=[context_a, context_b, context_a, context_b]), \
                 patch("app.routers.schedule._merge_order_decision_allocations", return_value={
                     1: {"PART-A": "Shortage"},
                     2: {"PART-B": "CreateRequirement"},
                 }), \
                 patch("app.routers.schedule._merge_order_supplement_allocations", return_value={
                     1: {"PART-A": 800},
                     2: {"PART-B": 1200},
                 }), \
                 patch("app.routers.schedule.preview_order_batches", side_effect=[
                     {"merged_parts": 2, "shortages": []},
                     {"merged_parts": 2, "shortages": []},
                 ]) as mock_preview, \
                 patch("app.routers.schedule._execute_dispatch", side_effect=[
                     {"order_id": 1, "merged_parts": 1, "session": {"id": 11}},
                     {"order_id": 2, "merged_parts": 1, "session": {"id": 12}},
                 ]) as mock_execute, \
                 patch("app.routers.schedule.db.replace_order_supplements"), \
                 patch("app.routers.schedule.refresh_snapshot_from_main"), \
                 patch("app.routers.schedule.db.log_activity"):
                payload = {
                    "order_ids": [1, 2],
                    "decisions": {"part-global": "CreateRequirement"},
                    "supplements": {"part-global": 5000},
                    "order_decisions": {
                        "1": {"part-a": "Shortage"},
                        "2": {"part-b": "CreateRequirement"},
                    },
                    "order_supplements": {
                        "1": {"part-a": 800},
                        "2": {"part-b": 1200},
                    },
                }
                preview_response = self.client.post("/api/schedule/main-write-preview", json=payload)
                dispatch_response = self.client.post("/api/schedule/batch-dispatch", json=payload)

        self.assertEqual(preview_response.status_code, 200)
        self.assertEqual(dispatch_response.status_code, 200)
        self.assertEqual(mock_preview.call_args_list[0].args[1], expected_batches)
        self.assertEqual(mock_preview.call_args_list[1].args[1], expected_batches)
        self.assertEqual(mock_preview.call_args_list[0].args[2], {})
        self.assertEqual(mock_preview.call_args_list[1].args[2], {})
        self.assertEqual(mock_execute.call_args_list[0].args[4], {"PART-A": "Shortage"})
        self.assertEqual(mock_execute.call_args_list[0].args[5], {"PART-A": 800})
        self.assertEqual(mock_execute.call_args_list[1].args[4], {"PART-B": "CreateRequirement"})
        self.assertEqual(mock_execute.call_args_list[1].args[5], {"PART-B": 1200})

    def test_single_dispatch_and_batch_dispatch_share_same_single_order_plan(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.write_bytes(b"main")

            groups = [{
                "batch_code": "1-1",
                "po_number": "4500059234",
                "bom_model": "MODEL-A",
                "components": [{"part_number": "PART-1", "needed_qty": 5, "prev_qty_cs": 0}],
            }]
            context = ({"id": 1, "po_number": "4500059234", "model": "MODEL-A", "status": "merged"}, groups, [{"part_number": "PART-1"}])

            with patch("app.routers.schedule.db.get_setting", return_value=str(main_path)), \
                 patch("app.routers.schedule.db.get_active_merge_draft_ids_by_order_ids", return_value={}), \
                 patch("app.routers.schedule._get_effective_moq", return_value={"PART-1": 500}), \
                 patch("app.routers.schedule._prepare_dispatch_context", side_effect=[context, context]), \
                 patch("app.routers.schedule.build_order_supplement_allocations", return_value={1: {"PART-1": 3000}}), \
                 patch("app.routers.schedule._merge_order_supplement_allocations", return_value={1: {"PART-1": 3000}}), \
                 patch("app.routers.schedule.preview_order_batches", side_effect=[
                     {"merged_parts": 1, "shortages": []},
                     {"merged_parts": 1, "shortages": []},
                 ]) as mock_preview, \
                 patch("app.routers.schedule._execute_dispatch", side_effect=[
                     {"order_id": 1, "merged_parts": 1},
                     {"order_id": 1, "merged_parts": 1},
                 ]) as mock_execute, \
                 patch("app.routers.schedule.db.replace_order_supplements") as mock_replace, \
                 patch("app.routers.schedule.refresh_snapshot_from_main"), \
                 patch("app.routers.schedule.db.log_activity"):
                direct_response = self.client.post("/api/schedule/orders/1/dispatch", json={
                    "decisions": {"part-1": "CreateRequirement"},
                    "supplements": {"part-1": 3000},
                })
                batch_response = self.client.post("/api/schedule/batch-dispatch", json={
                    "order_ids": [1],
                    "decisions": {"part-1": "CreateRequirement"},
                    "supplements": {"part-1": 3000},
                })

        self.assertEqual(direct_response.status_code, 200)
        self.assertEqual(batch_response.status_code, 200)
        self.assertEqual(mock_preview.call_args_list[0].args[1], mock_preview.call_args_list[1].args[1])
        self.assertEqual(mock_preview.call_args_list[0].args[2], {})
        self.assertEqual(mock_preview.call_args_list[1].args[2], {})
        self.assertEqual(mock_execute.call_args_list[0].args[4], mock_execute.call_args_list[1].args[4])
        self.assertEqual(mock_execute.call_args_list[0].args[5], mock_execute.call_args_list[1].args[5])
        self.assertEqual(mock_replace.call_args_list[0].args, ([1], {1: {"PART-1": 3000}}))
        self.assertEqual(mock_replace.call_args_list[1].args, ([1], {1: {"PART-1": 3000}}))

    def test_batch_dispatch_allows_main_write_when_preview_has_non_ec_shortage(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.write_bytes(b"main")

            context_a = ({"id": 1, "po_number": "4500059234", "model": "MODEL-A", "status": "merged"}, [{"components": []}], [])

            with patch("app.routers.schedule.db.get_setting", return_value=str(main_path)), \
                 patch("app.routers.schedule.db.get_active_merge_draft_ids_by_order_ids", return_value={}), \
                 patch("app.routers.schedule._get_effective_moq", return_value={}), \
                 patch("app.routers.schedule._prepare_dispatch_context", return_value=context_a), \
                 patch("app.routers.schedule.build_order_supplement_allocations", return_value={1: {}}), \
                 patch("app.routers.schedule.preview_order_batches", return_value={
                     "merged_parts": 0,
                     "shortages": [{"part_number": "PART-1", "shortage_amount": 5, "resulting_stock": -5}],
                 }), \
                 patch("app.routers.schedule._execute_dispatch", return_value={
                     "order_id": 1,
                     "merged_parts": 0,
                     "backup_path": "C:/b1.xlsx",
                     "session": {"id": 11},
                 }) as mock_execute, \
                 patch("app.routers.schedule.db.replace_order_supplements") as mock_replace, \
                 patch("app.routers.schedule.refresh_snapshot_from_main"), \
                 patch("app.routers.schedule.db.log_activity"):
                response = self.client.post("/api/schedule/batch-dispatch", json={
                    "order_ids": [1],
                    "decisions": {},
                    "supplements": {},
                })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["shortages"], [{"part_number": "PART-1", "shortage_amount": 5, "resulting_stock": -5}])
        mock_execute.assert_called_once()
        mock_replace.assert_called_once_with([1], {1: {}})

    def test_batch_dispatch_active_draft_uses_modal_request_overrides(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.write_bytes(b"main")

            context = {
                "draft": {"id": 11, "order_id": 1, "shortages": []},
                "order": {"id": 1, "po_number": "4500059234", "model": "MODEL-A", "status": "merged"},
                "groups": [{"components": []}],
                "all_components": [{"part_number": "PART-1", "needed_qty": 5}],
                "decisions": {"PART-1": "CreateRequirement"},
                "supplements": {"PART-1": 1000},
            }

            with patch("app.routers.schedule.db.get_setting", return_value=str(main_path)), \
                 patch("app.routers.schedule.db.get_active_merge_draft_ids_by_order_ids", return_value={1: 11}), \
                 patch("app.routers.schedule.rebuild_merge_drafts", return_value=[{"id": 11}]), \
                 patch("app.routers.schedule._load_active_merge_draft_context", return_value=context), \
                 patch("app.routers.schedule._get_effective_moq", return_value={}), \
                 patch("app.routers.schedule.preview_order_batches", return_value={
                     "merged_parts": 1,
                     "shortages": [],
                 }), \
                 patch("app.routers.schedule._execute_dispatch", return_value={
                     "order_id": 1,
                     "merged_parts": 1,
                     "backup_path": "C:/b1.xlsx",
                     "session": {"id": 11},
                 }) as mock_execute, \
                 patch("app.routers.schedule.db.replace_order_supplements") as mock_replace, \
                 patch("app.routers.schedule.db.get_active_merge_drafts", return_value=[]), \
                 patch("app.routers.schedule.refresh_snapshot_from_main"), \
                 patch("app.routers.schedule.db.log_activity"):
                response = self.client.post("/api/schedule/batch-dispatch", json={
                    "order_ids": [1],
                    "order_decisions": {
                        "1": {"part-1": "CreateRequirement"},
                    },
                    "order_supplements": {
                        "1": {"part-1": 3000},
                    },
                })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(mock_execute.call_args.args[4], {"PART-1": "CreateRequirement"})
        self.assertEqual(mock_execute.call_args.args[5], {"PART-1": 3000.0})
        mock_replace.assert_called_once_with([1], {1: {"PART-1": 3000.0}})

    def test_batch_dispatch_rolls_back_processed_orders_when_later_order_fails(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.write_bytes(b"main")

            context_a = ({"id": 1, "po_number": "4500059234", "model": "MODEL-A", "status": "merged"}, [{"components": []}], [])
            context_b = ({"id": 2, "po_number": "4500059235", "model": "MODEL-B", "status": "pending"}, [{"components": []}], [])

            with patch("app.routers.schedule.db.get_setting", return_value=str(main_path)), \
                 patch("app.routers.schedule._get_effective_moq", return_value={}), \
                 patch("app.routers.schedule._prepare_dispatch_context", side_effect=[context_a, context_b]), \
                 patch("app.routers.schedule.preview_order_batches", return_value={
                     "merged_parts": 3,
                     "shortages": [],
                 }), \
                 patch("app.routers.schedule._execute_dispatch", side_effect=[
                     {"order_id": 1, "merged_parts": 3, "backup_path": "C:/b1.xlsx", "session": {"id": 11, "order_id": 1}},
                     HTTPException(status_code=400, detail="第二筆發料失敗"),
                 ]), \
                 patch("app.routers.schedule._rollback_dispatch_sessions", return_value={"count": 1}) as mock_rollback, \
                 patch("app.routers.schedule.db.log_activity"):
                response = self.client.post("/api/schedule/batch-dispatch", json={
                    "order_ids": [1, 2],
                    "decisions": {},
                })

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["detail"], "第二筆發料失敗")
        mock_rollback.assert_called_once_with([{"id": 11, "order_id": 1}])

    def test_update_schedule_draft_rebuilds_from_saved_payload(self):
        call_order = []

        def _record(name):
            def _inner(*args, **kwargs):
                call_order.append(name)
            return _inner

        with patch("app.routers.schedule.db.get_merge_draft", return_value={"id": 5, "order_id": 1, "status": "active"}), \
             patch("app.routers.schedule.rebuild_merge_drafts", side_effect=lambda order_ids: (call_order.append("rebuild"), [{"id": 8}])[1]) as mock_rebuild, \
             patch("app.routers.schedule.db.replace_order_decisions", side_effect=_record("decisions")) as mock_replace_decisions, \
             patch("app.routers.schedule.db.replace_order_supplements", side_effect=_record("supplements")) as mock_replace_supplements, \
             patch("app.routers.schedule.db.get_active_merge_draft_for_order", return_value={"id": 8, "order_id": 1, "status": "active"}), \
             patch("app.routers.schedule.get_draft_detail", return_value={"draft": {"id": 8, "supplements": {"PART-1": 3000}}}), \
             patch("app.routers.schedule.db.log_activity"):
            response = self.client.put("/api/schedule/drafts/5", json={
                "decisions": {"part-1": "Shortage"},
                "supplements": {"part-1": 3000},
            })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["draft"]["id"], 8)
        self.assertEqual(response.json()["refreshed_count"], 1)
        mock_replace_decisions.assert_called_once_with([1], {1: {"PART-1": "Shortage"}})
        mock_replace_supplements.assert_called_once_with([1], {1: {"PART-1": 3000.0}})
        mock_rebuild.assert_called_once_with([1])
        self.assertEqual(call_order, ["decisions", "supplements", "rebuild"])

    def test_commit_schedule_draft_uses_latest_saved_draft_context(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.write_bytes(b"main")
            context = {
                "draft": {"id": 5, "order_id": 1},
                "order": {"id": 1, "po_number": "4500059234", "model": "MODEL-A"},
                "groups": [{"components": []}],
                "all_components": [{"part_number": "PART-1", "needed_qty": 5}],
                "decisions": {"PART-1": "Shortage"},
                "supplements": {"PART-1": 3000},
            }

            with patch("app.routers.schedule.db.get_setting", return_value=str(main_path)), \
                 patch("app.routers.schedule._get_effective_moq", return_value={}), \
                 patch("app.routers.schedule._load_active_merge_draft_context", return_value=context) as mock_context, \
                 patch("app.routers.schedule.preview_order_batches", return_value={
                     "merged_parts": 4,
                     "shortages": [],
                 }), \
                 patch("app.routers.schedule._execute_dispatch", return_value={"order_id": 1, "merged_parts": 4}) as mock_execute, \
                 patch("app.routers.schedule.db.replace_order_supplements") as mock_replace, \
                 patch("app.routers.schedule.db.mark_merge_draft_committed") as mock_mark, \
                 patch("app.routers.schedule.db.get_active_merge_drafts", return_value=[]), \
                 patch("app.routers.schedule.db.log_activity"):
                response = self.client.post("/api/schedule/drafts/5/commit")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["draft_id"], 5)
        self.assertEqual(response.json()["order_id"], 1)
        self.assertEqual(response.json()["merged_parts"], 4)
        mock_context.assert_called_once_with(5, str(main_path))
        self.assertEqual(mock_execute.call_args.args[4], {"PART-1": "Shortage"})
        self.assertEqual(mock_execute.call_args.args[5], {"PART-1": 3000})
        mock_replace.assert_called_once_with([1], {1: {"PART-1": 3000}})
        mock_mark.assert_called_once_with(5)

    def test_commit_schedule_draft_allows_non_ec_and_ec_shortages(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.write_bytes(b"main")

            blocking_context = {
                "draft": {"id": 5, "order_id": 1, "shortages": [{"part_number": "PART-1", "shortage_amount": 8}]},
                "order": {"id": 1, "po_number": "4500059234", "model": "MODEL-A"},
                "groups": [{"components": []}],
                "all_components": [{"part_number": "PART-1", "needed_qty": 5}],
                "decisions": {"PART-1": "CreateRequirement"},
                "supplements": {},
            }
            allowed_context = {
                "draft": {"id": 6, "order_id": 2, "shortages": [{"part_number": "EC-001", "shortage_amount": 20}]},
                "order": {"id": 2, "po_number": "4500059235", "model": "MODEL-EC"},
                "groups": [{"components": []}],
                "all_components": [{"part_number": "EC-001", "needed_qty": 5}],
                "decisions": {"EC-001": "CreateRequirement"},
                "supplements": {},
            }

            with patch("app.routers.schedule.db.get_setting", return_value=str(main_path)), \
                 patch("app.routers.schedule._get_effective_moq", return_value={}), \
                 patch("app.routers.schedule._load_active_merge_draft_context", side_effect=[blocking_context, allowed_context]), \
                 patch("app.routers.schedule.preview_order_batches", side_effect=[
                     {"merged_parts": 0, "shortages": [{"part_number": "PART-1", "shortage_amount": 8, "resulting_stock": -8}]},
                     {"merged_parts": 0, "shortages": [{"part_number": "EC-001", "shortage_amount": 20, "resulting_stock": 80}]},
                 ]), \
                 patch("app.routers.schedule._execute_dispatch", side_effect=[
                     {"order_id": 1, "merged_parts": 1},
                     {"order_id": 2, "merged_parts": 1},
                 ]) as mock_execute, \
                 patch("app.routers.schedule.db.replace_order_supplements"), \
                 patch("app.routers.schedule.db.mark_merge_draft_committed"), \
                 patch("app.routers.schedule.db.get_active_merge_drafts", return_value=[]), \
                 patch("app.routers.schedule.refresh_snapshot_from_main"), \
                 patch("app.routers.schedule.db.log_activity"):
                blocked = self.client.post("/api/schedule/drafts/5/commit")
                allowed = self.client.post("/api/schedule/drafts/6/commit")

        self.assertEqual(blocked.status_code, 200)
        self.assertEqual(blocked.json()["order_id"], 1)
        self.assertEqual(allowed.status_code, 200)
        self.assertEqual(allowed.json()["order_id"], 2)
        self.assertEqual(mock_execute.call_count, 2)

    def test_commit_schedule_draft_allows_ec_negative_stock(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.write_bytes(b"main")

            context = {
                "draft": {"id": 7, "order_id": 3, "shortages": [{"part_number": "EC-001", "shortage_amount": 101}]},
                "order": {"id": 3, "po_number": "4500059236", "model": "MODEL-EC"},
                "groups": [{"components": []}],
                "all_components": [{"part_number": "EC-001", "needed_qty": 5}],
                "decisions": {"EC-001": "CreateRequirement"},
                "supplements": {},
            }

            with patch("app.routers.schedule.db.get_setting", return_value=str(main_path)), \
                 patch("app.routers.schedule._get_effective_moq", return_value={}), \
                 patch("app.routers.schedule._load_active_merge_draft_context", return_value=context), \
                 patch("app.routers.schedule.preview_order_batches", return_value={
                     "merged_parts": 0,
                     "shortages": [{"part_number": "EC-001", "shortage_amount": 101, "resulting_stock": -1}],
                 }), \
                 patch("app.routers.schedule._execute_dispatch", return_value={"order_id": 3, "merged_parts": 1}) as mock_execute, \
                 patch("app.routers.schedule.db.replace_order_supplements"), \
                 patch("app.routers.schedule.db.mark_merge_draft_committed"), \
                 patch("app.routers.schedule.db.get_active_merge_drafts", return_value=[]), \
                 patch("app.routers.schedule.refresh_snapshot_from_main"), \
                 patch("app.routers.schedule.db.log_activity"):
                response = self.client.post("/api/schedule/drafts/7/commit")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["order_id"], 3)
        mock_execute.assert_called_once()

    def test_active_draft_preview_and_commit_use_same_saved_context(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.write_bytes(b"main")

            expected_shortage = {
                "part_number": "EC-001",
                "shortage_amount": 20,
                "supplement_qty": 3000,
                "decision": "CreateRequirement",
                "resulting_stock": 80,
                "model": "MODEL-EC",
                "batch_code": "1-3",
            }
            context = {
                "draft": {
                    "id": 11,
                    "order_id": 1,
                    "shortages": [{
                        "part_number": "EC-001",
                        "shortage_amount": 20,
                        "supplement_qty": 3000,
                        "decision": "CreateRequirement",
                        "resulting_stock": 80,
                    }],
                },
                "order": {"id": 1, "po_number": "4500059234", "model": "MODEL-EC", "code": "1-3"},
                "groups": [{"components": []}],
                "all_components": [{"part_number": "EC-001", "needed_qty": 5}],
                "decisions": {"EC-001": "CreateRequirement"},
                "supplements": {"EC-001": 3000},
            }

            with patch("app.routers.schedule.db.get_setting", return_value=str(main_path)), \
                 patch("app.routers.schedule.db.get_active_merge_draft_ids_by_order_ids", return_value={1: 11}), \
                 patch("app.routers.schedule.rebuild_merge_drafts", return_value=[{"id": 11}]), \
                 patch("app.routers.schedule._load_active_merge_draft_context", side_effect=[context, context]) as mock_context, \
                 patch("app.routers.schedule._get_effective_moq", return_value={"EC-001": 500}), \
                 patch("app.routers.schedule.preview_order_batches", return_value={
                     "merged_parts": 1,
                     "shortages": [expected_shortage],
                 }), \
                 patch("app.routers.schedule._execute_dispatch", return_value={"order_id": 1, "merged_parts": 1}) as mock_execute, \
                 patch("app.routers.schedule.db.replace_order_supplements"), \
                 patch("app.routers.schedule.db.mark_merge_draft_committed"), \
                 patch("app.routers.schedule.db.get_active_merge_drafts", return_value=[]), \
                 patch("app.routers.schedule.refresh_snapshot_from_main"), \
                 patch("app.routers.schedule.db.log_activity"):
                preview_response = self.client.post("/api/schedule/main-write-preview", json={
                    "order_ids": [1],
                    "decisions": {"ec-001": "Shortage"},
                    "supplements": {"ec-001": 9999},
                })
                commit_response = self.client.post("/api/schedule/drafts/11/commit")

        self.assertEqual(preview_response.status_code, 200)
        self.assertEqual(commit_response.status_code, 200)
        self.assertEqual(preview_response.json()["shortages"], [expected_shortage])
        self.assertEqual(commit_response.json()["shortages"], [expected_shortage])
        self.assertEqual(mock_context.call_count, 2)
        self.assertEqual(mock_execute.call_args.args[4], {"EC-001": "CreateRequirement"})
        self.assertEqual(mock_execute.call_args.args[5], {"EC-001": 3000})

    def test_batch_dispatch_real_write_matches_preview_for_allowed_ec_shortage(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            backup_dir = Path(temp_dir) / "backups"
            self._build_main_workbook(main_path, [("EC-001", 500, 120)])

            groups = [{
                "batch_code": "1-1",
                "po_number": "4500059234",
                "bom_model": "MODEL-EC",
                "components": [{
                    "part_number": "EC-001",
                    "description": "EC part",
                    "is_dash": False,
                    "needed_qty": 30,
                    "prev_qty_cs": 0,
                }],
            }]
            context = (
                {"id": 1, "po_number": "4500059234", "model": "MODEL-EC", "status": "merged"},
                groups,
                [{"part_number": "EC-001", "description": "EC part", "needed_qty": 30, "prev_qty_cs": 0, "is_dash": 0}],
            )

            with patch("app.routers.schedule.db.get_setting", return_value=str(main_path)), \
                 patch("app.routers.schedule.BACKUP_DIR", backup_dir), \
                 patch("app.routers.schedule.db.get_active_merge_draft_ids_by_order_ids", return_value={}), \
                 patch("app.routers.schedule._get_effective_moq", return_value={"EC-001": 500}), \
                 patch("app.routers.schedule.db.get_st_inventory_stock", return_value={}), \
                 patch("app.routers.schedule._prepare_dispatch_context", side_effect=[context, context]), \
                 patch("app.routers.schedule.db.save_dispatch_session", return_value={"id": 11, "order_id": 1}), \
                 patch("app.routers.schedule.db.save_dispatch_records"), \
                 patch("app.routers.schedule.db.update_order"), \
                 patch("app.routers.schedule.db.replace_order_supplements"), \
                 patch("app.routers.schedule.refresh_snapshot_from_main"), \
                 patch("app.routers.schedule.db.log_activity"):
                preview_response = self.client.post("/api/schedule/main-write-preview", json={
                    "order_ids": [1],
                    "decisions": {"ec-001": "CreateRequirement"},
                })
                dispatch_response = self.client.post("/api/schedule/batch-dispatch", json={
                    "order_ids": [1],
                    "decisions": {"ec-001": "CreateRequirement"},
                })

            self.assertEqual(preview_response.status_code, 200)
            self.assertEqual(dispatch_response.status_code, 200)
            self.assertEqual(preview_response.json()["merged_parts"], 1)
            self.assertEqual(dispatch_response.json()["merged_parts"], 1)
            self.assertEqual(preview_response.json()["shortages"], dispatch_response.json()["shortages"])
            self.assertEqual(preview_response.json()["shortages"][0]["part_number"], "EC-001")
            self.assertEqual(preview_response.json()["shortages"][0]["shortage_amount"], 10.0)
            self.assertEqual(preview_response.json()["shortages"][0]["resulting_stock"], 90.0)

            wb = openpyxl.load_workbook(main_path, data_only=True)
            ws = wb.active
            self.assertEqual(ws.cell(row=2, column=9).value, 0)
            self.assertEqual(ws.cell(row=2, column=10).value, 30)
            self.assertEqual(ws.cell(row=2, column=11).value, 90)
            wb.close()

    def test_active_draft_commit_real_write_matches_saved_draft_preview(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            backup_dir = Path(temp_dir) / "backups"
            self._build_main_workbook(main_path, [("EC-001", 500, 120)])

            groups = [{
                "batch_code": "1-3",
                "po_number": "4500059234",
                "bom_model": "MODEL-EC",
                "components": [{
                    "part_number": "EC-001",
                    "description": "EC part",
                    "is_dash": False,
                    "needed_qty": 30,
                    "prev_qty_cs": 0,
                }],
            }]
            expected_preview = schedule_router.preview_order_batches(
                str(main_path),
                [{
                    "order_id": 1,
                    "model": "MODEL-EC",
                    "groups": groups,
                    "supplements": {},
                }],
                {"EC-001": "CreateRequirement"},
                moq_map={"EC-001": 500},
            )
            context = {
                "draft": {
                    "id": 11,
                    "order_id": 1,
                    "shortages": [dict(item) for item in expected_preview["shortages"]],
                },
                "order": {"id": 1, "po_number": "4500059234", "model": "MODEL-EC", "code": "1-3", "status": "merged"},
                "groups": groups,
                "all_components": [{
                    "part_number": "EC-001",
                    "description": "EC part",
                    "needed_qty": 30,
                    "prev_qty_cs": 0,
                    "is_dash": 0,
                }],
                "decisions": {"EC-001": "CreateRequirement"},
                "supplements": {},
            }

            with patch("app.routers.schedule.db.get_setting", return_value=str(main_path)), \
                 patch("app.routers.schedule.BACKUP_DIR", backup_dir), \
                 patch("app.routers.schedule.db.get_active_merge_draft_ids_by_order_ids", return_value={1: 11}), \
                 patch("app.routers.schedule.rebuild_merge_drafts", return_value=[{"id": 11}]), \
                 patch("app.routers.schedule._load_active_merge_draft_context", side_effect=[context, context]), \
                 patch("app.routers.schedule._get_effective_moq", return_value={"EC-001": 500}), \
                 patch("app.routers.schedule.db.save_dispatch_session", return_value={"id": 11, "order_id": 1}), \
                 patch("app.routers.schedule.db.save_dispatch_records"), \
                 patch("app.routers.schedule.db.update_order"), \
                 patch("app.routers.schedule.db.replace_order_supplements"), \
                 patch("app.routers.schedule.db.mark_merge_draft_committed"), \
                 patch("app.routers.schedule.db.get_active_merge_drafts", return_value=[]), \
                 patch("app.routers.schedule.refresh_snapshot_from_main"), \
                 patch("app.routers.schedule.db.log_activity"):
                preview_response = self.client.post("/api/schedule/main-write-preview", json={
                    "order_ids": [1],
                })
                commit_response = self.client.post("/api/schedule/drafts/11/commit")

            self.assertEqual(preview_response.status_code, 200)
            self.assertEqual(commit_response.status_code, 200)
            self.assertEqual(preview_response.json()["shortages"], expected_preview["shortages"])
            self.assertEqual(commit_response.json()["shortages"], expected_preview["shortages"])
            self.assertEqual(commit_response.json()["merged_parts"], expected_preview["merged_parts"])

            wb = openpyxl.load_workbook(main_path, data_only=True)
            ws = wb.active
            self.assertEqual(ws.cell(row=2, column=9).value, 0)
            self.assertEqual(ws.cell(row=2, column=10).value, 30)
            self.assertEqual(ws.cell(row=2, column=11).value, 90)
            wb.close()

    def test_rollback_preview_returns_tail_orders(self):
        orders = {
            5: {"id": 5, "po_number": "4500059234", "model": "MODEL-E", "status": "dispatched"},
            6: {"id": 6, "po_number": "4500059235", "model": "MODEL-F", "status": "completed"},
        }
        session = {
            "id": 9,
            "order_id": 5,
            "backup_path": "C:/backup.xlsx",
            "main_file_path": "C:/main.xlsx",
            "dispatched_at": "2026-03-20T09:00:00",
        }
        tail = [
            {"id": 9, "order_id": 5, "previous_status": "merged"},
            {"id": 10, "order_id": 6, "previous_status": "pending"},
        ]

        with patch("app.routers.schedule.db.get_order", side_effect=lambda order_id: orders.get(order_id)), \
             patch("app.routers.schedule.db.get_active_dispatch_session", return_value=session), \
             patch("app.services.inventory_restore_guard.db.get_defective_batch_summaries_after", return_value=[]), \
             patch("app.services.inventory_restore_guard.db.get_activity_logs_after", return_value=[]), \
             patch("app.routers.schedule.db.get_dispatch_session_tail", return_value=tail):
            response = self.client.get("/api/schedule/orders/5/rollback-preview")

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["count"], 2)
        self.assertEqual(
            data["orders"],
            [
                {"id": 5, "po_number": "4500059234", "model": "MODEL-E", "status": "dispatched", "restore_status": "merged"},
                {"id": 6, "po_number": "4500059235", "model": "MODEL-F", "status": "completed", "restore_status": "pending"},
            ],
        )

    def test_rollback_preview_blocks_when_later_inventory_mutation_exists(self):
        order = {"id": 5, "po_number": "4500059234", "model": "MODEL-E", "status": "dispatched"}
        session = {
            "id": 9,
            "order_id": 5,
            "backup_path": "C:/backup.xlsx",
            "main_file_path": "C:/main.xlsx",
            "dispatched_at": "2026-03-20T09:00:00",
        }

        with patch("app.routers.schedule.db.get_order", return_value=order), \
             patch("app.routers.schedule.db.get_active_dispatch_session", return_value=session), \
             patch("app.services.inventory_restore_guard.db.get_defective_batch_summaries_after", return_value=[{"id": 88}]), \
             patch("app.services.inventory_restore_guard.db.get_activity_logs_after", return_value=[]), \
             patch("app.routers.schedule.db.get_dispatch_session_tail") as mock_tail:
            response = self.client.get("/api/schedule/orders/5/rollback-preview")

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.json()["detail"],
            "後面已有其他庫存異動，不能直接回復。請先下載目前主檔，手動修正後重新上傳主檔。重新上傳後一定要重設快照。",
        )
        mock_tail.assert_not_called()

    def test_rollback_preview_allows_prior_order_rollback_log(self):
        orders = {
            5: {"id": 5, "po_number": "4500059234", "model": "MODEL-E", "status": "dispatched"},
            6: {"id": 6, "po_number": "4500059235", "model": "MODEL-F", "status": "completed"},
        }
        session = {
            "id": 9,
            "order_id": 5,
            "backup_path": "C:/backup.xlsx",
            "main_file_path": "C:/main.xlsx",
            "dispatched_at": "2026-03-20T09:00:00",
        }
        tail = [
            {"id": 9, "order_id": 5, "previous_status": "merged"},
            {"id": 10, "order_id": 6, "previous_status": "pending"},
        ]

        with patch("app.routers.schedule.db.get_order", side_effect=lambda order_id: orders.get(order_id)), \
             patch("app.routers.schedule.db.get_active_dispatch_session", return_value=session), \
             patch("app.services.inventory_restore_guard.db.get_defective_batch_summaries_after", return_value=[]), \
             patch("app.services.inventory_restore_guard.db.get_activity_logs_after", return_value=[]) as mock_logs, \
             patch("app.routers.schedule.db.get_dispatch_session_tail", return_value=tail):
            response = self.client.get("/api/schedule/orders/5/rollback-preview")

        self.assertEqual(response.status_code, 200)
        self.assertNotIn("order_rollback", mock_logs.call_args.kwargs["actions"])

    def test_rollback_preview_force_bypasses_later_inventory_mutation_guard(self):
        orders = {
            5: {"id": 5, "po_number": "4500059234", "model": "MODEL-E", "status": "dispatched"},
            6: {"id": 6, "po_number": "4500059235", "model": "MODEL-F", "status": "completed"},
        }
        session = {
            "id": 9,
            "order_id": 5,
            "backup_path": "C:/backup.xlsx",
            "main_file_path": "C:/main.xlsx",
            "dispatched_at": "2026-03-20T09:00:00",
        }
        tail = [
            {"id": 9, "order_id": 5, "previous_status": "merged"},
            {"id": 10, "order_id": 6, "previous_status": "pending"},
        ]

        with patch("app.routers.schedule.db.get_order", side_effect=lambda order_id: orders.get(order_id)), \
             patch("app.routers.schedule.db.get_active_dispatch_session", return_value=session), \
             patch("app.services.inventory_restore_guard.db.get_defective_batch_summaries_after") as mock_defective, \
             patch("app.services.inventory_restore_guard.db.get_activity_logs_after") as mock_logs, \
             patch("app.routers.schedule.db.get_dispatch_session_tail", return_value=tail):
            response = self.client.get("/api/schedule/orders/5/rollback-preview?force=1")

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["forced"])
        self.assertEqual(response.json()["count"], 2)
        mock_defective.assert_not_called()
        mock_logs.assert_not_called()

    def test_rollback_restores_backup_and_reverts_tail_orders(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            backup_path = Path(temp_dir) / "backup.xlsx"
            main_path.write_text("after-dispatch", encoding="utf-8")
            backup_path.write_text("before-dispatch", encoding="utf-8")

            orders = {
                5: {"id": 5, "po_number": "4500059234", "model": "MODEL-E", "status": "dispatched"},
                6: {"id": 6, "po_number": "4500059235", "model": "MODEL-F", "status": "completed"},
            }
            session = {
                "id": 9,
                "order_id": 5,
                "backup_path": str(backup_path),
                "main_file_path": str(main_path),
                "dispatched_at": "2026-03-20T09:00:00",
            }
            tail = [
                {"id": 9, "order_id": 5, "previous_status": "merged", "backup_path": str(backup_path), "main_file_path": str(main_path)},
                {"id": 10, "order_id": 6, "previous_status": "pending", "backup_path": str(backup_path), "main_file_path": str(main_path)},
            ]

            with patch("app.routers.schedule.db.get_order", side_effect=lambda order_id: orders.get(order_id)), \
                 patch("app.routers.schedule.db.get_active_dispatch_session", return_value=session), \
                 patch("app.services.inventory_restore_guard.db.get_defective_batch_summaries_after", return_value=[]), \
                 patch("app.services.inventory_restore_guard.db.get_activity_logs_after", return_value=[]), \
                 patch("app.routers.schedule.db.get_dispatch_session_tail", return_value=tail), \
                 patch("app.routers.schedule.db.get_setting", return_value=str(main_path)), \
                 patch("app.routers.schedule.db.delete_dispatch_records_for_orders") as mock_delete_records, \
                 patch("app.routers.schedule.db.mark_dispatch_sessions_rolled_back") as mock_mark_rolled_back, \
                 patch("app.routers.schedule.restore_recent_committed_merge_drafts", return_value=[5, 6]) as mock_restore_drafts, \
                 patch("app.routers.schedule.db.update_order") as mock_update_order, \
                 patch("app.routers.schedule.db.log_activity"):
                response = self.client.post("/api/schedule/orders/5/rollback")
                restored_text = main_path.read_text(encoding="utf-8")

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["count"], 2)
        self.assertEqual(restored_text, "before-dispatch")
        mock_delete_records.assert_called_once_with([5, 6])
        mock_mark_rolled_back.assert_called_once_with([9, 10])
        mock_restore_drafts.assert_called_once_with([5, 6])
        self.assertEqual(data["restored_draft_count"], 2)
        self.assertEqual(data["restored_draft_order_ids"], [5, 6])
        mock_update_order.assert_has_calls([
            call(5, status="merged", folder=""),
            call(6, status="pending", folder=""),
        ])

    def test_rollback_force_restores_backup_even_when_guard_would_block(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            backup_path = Path(temp_dir) / "backup.xlsx"
            main_path.write_text("after-dispatch", encoding="utf-8")
            backup_path.write_text("before-dispatch", encoding="utf-8")

            orders = {
                5: {"id": 5, "po_number": "4500059234", "model": "MODEL-E", "status": "dispatched"},
                6: {"id": 6, "po_number": "4500059235", "model": "MODEL-F", "status": "completed"},
            }
            session = {
                "id": 9,
                "order_id": 5,
                "backup_path": str(backup_path),
                "main_file_path": str(main_path),
                "dispatched_at": "2026-03-20T09:00:00",
            }
            tail = [
                {"id": 9, "order_id": 5, "previous_status": "merged", "backup_path": str(backup_path), "main_file_path": str(main_path)},
                {"id": 10, "order_id": 6, "previous_status": "pending", "backup_path": str(backup_path), "main_file_path": str(main_path)},
            ]

            with patch("app.routers.schedule.db.get_order", side_effect=lambda order_id: orders.get(order_id)), \
                 patch("app.routers.schedule.db.get_active_dispatch_session", return_value=session), \
                 patch("app.routers.schedule.db.get_dispatch_session_tail", return_value=tail), \
                 patch("app.routers.schedule.db.get_setting", return_value=str(main_path)), \
                 patch("app.routers.schedule.db.delete_dispatch_records_for_orders") as mock_delete_records, \
                 patch("app.routers.schedule.db.mark_dispatch_sessions_rolled_back") as mock_mark_rolled_back, \
                 patch("app.routers.schedule.restore_recent_committed_merge_drafts", return_value=[5, 6]) as mock_restore_drafts, \
                 patch("app.routers.schedule.db.update_order") as mock_update_order, \
                 patch("app.routers.schedule.db.log_activity") as mock_log_activity, \
                 patch("app.services.inventory_restore_guard.db.get_defective_batch_summaries_after") as mock_defective, \
                 patch("app.services.inventory_restore_guard.db.get_activity_logs_after") as mock_logs:
                response = self.client.post("/api/schedule/orders/5/rollback?force=1")
                restored_text = main_path.read_text(encoding="utf-8")

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["forced"])
        self.assertEqual(data["count"], 2)
        self.assertEqual(restored_text, "before-dispatch")
        mock_delete_records.assert_called_once_with([5, 6])
        mock_mark_rolled_back.assert_called_once_with([9, 10])
        mock_restore_drafts.assert_called_once_with([5, 6])
        mock_update_order.assert_has_calls([
            call(5, status="merged", folder=""),
            call(6, status="pending", folder=""),
        ])
        mock_log_activity.assert_called_once()
        self.assertIn("強制反悔", mock_log_activity.call_args.args[1])
        mock_defective.assert_not_called()
        mock_logs.assert_not_called()

    def test_main_file_data_backfills_missing_moq_from_live_main_file(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.write_bytes(b"placeholder")

            def fake_setting(key, default=""):
                mapping = {
                    "main_file_path": str(main_path),
                    "main_part_count": "1",
                    "main_loaded_at": "2026-03-12T08:00:00",
                    "main_filename": "main.xlsx",
                }
                return mapping.get(key, default)

            snapshot = {"AAA": {"stock_qty": 5, "moq": 8}}

            with patch("app.routers.main_file.db.get_setting", side_effect=fake_setting), \
                 patch("app.routers.main_file.db.get_snapshot", return_value=snapshot), \
                 patch("app.routers.main_file.find_legacy_snapshot_stock_fixes", return_value={}), \
                 patch("app.routers.main_file.db.update_snapshot_stock", return_value=0), \
                 patch("app.routers.main_file.read_moq", return_value={"AAA": 99, "BBB": 12}):
                response = self.client.get("/api/main-file/data")

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["stock"], {"AAA": 5})
        self.assertEqual(data["moq"]["AAA"], 8)
        self.assertEqual(data["moq"]["BBB"], 12)

    def test_main_file_data_repairs_legacy_snapshot_stock_bug(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.write_bytes(b"placeholder")

            def fake_setting(key, default=""):
                mapping = {
                    "main_file_path": str(main_path),
                    "main_part_count": "2",
                    "main_loaded_at": "2026-03-12T08:00:00",
                    "main_filename": "main.xlsx",
                }
                return mapping.get(key, default)

            snapshot = {
                "AAA": {"stock_qty": 8, "moq": 8},
                "BBB": {"stock_qty": 5, "moq": 12},
            }

            with patch("app.routers.main_file.db.get_setting", side_effect=fake_setting), \
                 patch("app.routers.main_file.db.get_snapshot", return_value=snapshot), \
                 patch("app.routers.main_file.find_legacy_snapshot_stock_fixes", return_value={"AAA": 0.0}), \
                 patch("app.routers.main_file.db.update_snapshot_stock", return_value=1), \
                 patch("app.routers.main_file.db.log_activity"), \
                 patch("app.routers.main_file.read_moq", return_value={"AAA": 99, "BBB": 12}):
                response = self.client.get("/api/main-file/data")

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["stock"]["AAA"], 0.0)
        self.assertEqual(data["stock"]["BBB"], 5)

    def test_update_main_vendor_writes_main_file_b_column(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["料號", "廠商", "MOQ", "庫存"])
            sheet.append(["IC-100", "舊廠商", 100, 5])
            workbook.save(main_path)
            workbook.close()

            with patch("app.routers.main_file.db.get_setting", return_value=str(main_path)), \
                 patch("app.routers.main_file.backup_main_file", return_value=Path(temp_dir) / "backup.xlsx"), \
                 patch("app.routers.main_file.db.log_activity"):
                response = self.client.patch(
                    "/api/main-file/vendor",
                    json={"part_number": "ic-100", "vendor": "新廠商"},
                )

            saved = openpyxl.load_workbook(main_path)
            vendor = saved.active.cell(row=2, column=2).value
            saved.close()

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["ok"])
        self.assertEqual(response.json()["old_vendor"], "舊廠商")
        self.assertEqual(response.json()["vendor"], "新廠商")
        self.assertEqual(vendor, "新廠商")

    def test_purchase_reminder_status_endpoint_persists_status(self):
        expected = {
            "part_number": "IC-100",
            "notified": True,
            "notified_at": "2026-04-23T10:00:00",
            "note": "通知完成",
            "ignored": False,
            "ignored_at": "",
            "updated_at": "2026-04-23T10:00:00",
        }

        with patch("app.routers.main_file.db.set_purchase_reminder_status", return_value=expected) as mock_set, \
             patch("app.routers.main_file.db.log_activity"):
            response = self.client.patch(
                "/api/main-file/purchase-reminder-status",
                json={"part_number": " ic-100 ", "notified": True, "note": "通知完成"},
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["status"], expected)
        mock_set.assert_called_once_with("IC-100", True, "通知完成")

    def test_purchase_reminder_status_endpoint_persists_ignored_status(self):
        expected = {
            "part_number": "IC-100",
            "notified": False,
            "notified_at": "",
            "note": "",
            "ignored": True,
            "ignored_at": "2026-04-23T10:00:00",
            "updated_at": "2026-04-23T10:00:00",
        }

        with patch("app.routers.main_file.db.set_purchase_reminder_ignored", return_value=expected) as mock_set, \
             patch("app.routers.main_file.db.log_activity"):
            response = self.client.patch(
                "/api/main-file/purchase-reminder-status",
                json={"part_number": " ic-100 ", "ignored": True},
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["status"], expected)
        mock_set.assert_called_once_with("IC-100", True)

    def test_purchase_reminder_ignored_clears_notified_status(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            with patch.object(database, "DB_PATH", Path(temp_dir) / "system.db"):
                database.init_db()
                database.set_purchase_reminder_status("ic-100", True, "通知完成")
                ignored = database.set_purchase_reminder_ignored("ic-100", True)
                statuses = database.get_purchase_reminder_statuses()

        self.assertFalse(ignored["notified"])
        self.assertEqual(ignored["notified_at"], "")
        self.assertEqual(ignored["note"], "")
        self.assertTrue(ignored["ignored"])
        self.assertFalse(statuses["IC-100"]["notified"])
        self.assertTrue(statuses["IC-100"]["ignored"])

    def test_purchase_reminder_notified_clears_ignored_status(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            with patch.object(database, "DB_PATH", Path(temp_dir) / "system.db"):
                database.init_db()
                database.set_purchase_reminder_ignored("ic-100", True)
                notified = database.set_purchase_reminder_status("ic-100", True, "已寄信")
                statuses = database.get_purchase_reminder_statuses()

        self.assertTrue(notified["notified"])
        self.assertFalse(notified["ignored"])
        self.assertTrue(statuses["IC-100"]["notified"])
        self.assertFalse(statuses["IC-100"]["ignored"])

    def test_purchase_reminder_export_groups_by_vendor(self):
        response = self.client.post(
            "/api/main-file/purchase-reminders/export",
            json={
                "items": [
                    {
                        "vendor": "Vendor-B",
                        "part_number": "OC-200",
                        "description": "Spring",
                        "current_stock": 0,
                        "threshold": 100,
                        "moq": 100,
                        "suggested_qty": 100,
                        "notified": True,
                        "notified_at": "2026-04-23T09:00:00",
                        "note": "已寄信",
                    },
                    {
                        "vendor": "Vendor-A",
                        "part_number": "IC-100",
                        "description": "MCU",
                        "current_stock": 5,
                        "threshold": 100,
                        "moq": 100,
                        "suggested_qty": 100,
                        "notified": False,
                    },
                ],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.headers["content-type"],
        )

        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        detail = workbook["買料提醒"]
        summary = workbook["廠商彙總"]

        self.assertEqual(detail["A5"].value, "Vendor-A")
        self.assertEqual(detail["B5"].value, "待通知")
        self.assertEqual(detail["C5"].value, "IC-100")
        self.assertEqual(detail["A6"].value, "Vendor-B")
        self.assertEqual(detail["B6"].value, "已通知採購")
        self.assertEqual(summary["A2"].value, "Vendor-A")
        self.assertEqual(summary["B2"].value, 1)
        self.assertEqual(summary["A3"].value, "Vendor-B")
        self.assertEqual(summary["C3"].value, 1)
        workbook.close()

    def test_main_file_preview_returns_live_sheet_structure(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "庫存主檔"
            sheet.column_dimensions["A"].width = 18
            sheet.column_dimensions["B"].width = 12
            sheet.merge_cells("A1:C1")
            sheet["A1"] = "主檔即時預覽"
            sheet["A1"].font = Font(name="Calibri", bold=True, size=14)
            sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
            sheet["A2"] = "料號"
            sheet["B2"] = "MOQ"
            sheet["C2"] = "庫存"
            sheet["A3"] = "EC-20023A"
            sheet["B3"] = 10000
            sheet["C3"] = 5625
            workbook.create_sheet("第二頁")
            workbook.save(main_path)
            workbook.close()

            def fake_setting(key, default=""):
                mapping = {
                    "main_file_path": str(main_path),
                    "main_filename": "主檔260306test.xlsx",
                    "main_loaded_at": "2026-03-12T22:45:00",
                }
                return mapping.get(key, default)

            with patch("app.routers.main_file.db.get_setting", side_effect=fake_setting):
                response = self.client.get("/api/main-file/preview")

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["filename"], "主檔260306test.xlsx")
        self.assertEqual(data["selected_sheet"], "庫存主檔")
        self.assertEqual(data["sheet_names"], ["庫存主檔", "第二頁"])
        self.assertTrue(data["style_preserved"])
        self.assertEqual(data["sheet"]["columns"][0]["letter"], "A")
        self.assertGreater(data["sheet"]["columns"][0]["width_px"], 100)
        self.assertEqual(data["sheet"]["rows"][0]["cells"][0]["value"], "主檔即時預覽")
        self.assertEqual(data["sheet"]["rows"][0]["cells"][0]["colspan"], 3)
        self.assertEqual(data["sheet"]["styles"][data["sheet"]["rows"][0]["cells"][0]["style_id"]]["background"], "#D9EAF7")

    def test_main_file_preview_can_switch_sheet(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"

            workbook = openpyxl.Workbook()
            workbook.active.title = "第一頁"
            second = workbook.create_sheet("第二頁")
            second["B2"] = "切換成功"
            workbook.save(main_path)
            workbook.close()

            def fake_setting(key, default=""):
                mapping = {
                    "main_file_path": str(main_path),
                    "main_filename": "main.xlsx",
                    "main_loaded_at": "2026-03-12T22:45:00",
                }
                return mapping.get(key, default)

            with patch("app.routers.main_file.db.get_setting", side_effect=fake_setting):
                response = self.client.get("/api/main-file/preview?sheet=第二頁")

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["selected_sheet"], "第二頁")
        self.assertEqual(data["sheet"]["name"], "第二頁")
        target_cell = next(cell for cell in data["sheet"]["rows"][1]["cells"] if cell["col"] == 2)
        self.assertEqual(target_cell["value"], "切換成功")

    def test_edit_main_cell_returns_affected_balances_and_updates_snapshot_stock(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "庫存主檔"
            sheet.cell(row=1, column=1).value = "料號"
            sheet.cell(row=1, column=8).value = "盤點"
            for col, code in ((9, "1-1"), (12, "1-2"), (15, "1-3")):
                sheet.cell(row=1, column=col).value = code
                sheet.cell(row=1, column=col + 1).value = f"PO-{code}"
                sheet.cell(row=1, column=col + 2).value = f"MODEL-{code}"
            sheet.cell(row=2, column=1).value = "PART-1"
            sheet.cell(row=2, column=8).value = 100
            sheet.cell(row=2, column=9).value = 0
            sheet.cell(row=2, column=10).value = 30
            sheet.cell(row=2, column=11).value = 70
            sheet.cell(row=2, column=12).value = 5
            sheet.cell(row=2, column=13).value = 20
            sheet.cell(row=2, column=14).value = 55
            workbook.save(main_path)
            workbook.close()

            def fake_setting(key, default=""):
                return str(main_path) if key == "main_file_path" else default

            with patch("app.routers.main_file.db.get_setting", side_effect=fake_setting), \
                 patch("app.routers.main_file.backup_main_file"), \
                 patch("app.routers.main_file.db.get_snapshot", return_value={"PART-1": {"stock_qty": 100}}), \
                 patch("app.routers.main_file.db.update_snapshot_stock", return_value=1) as mock_update_stock, \
                 patch("app.routers.main_file.refresh_snapshot_from_main") as mock_refresh, \
                 patch("app.routers.main_file.db.get_active_merge_drafts", return_value=[]), \
                 patch("app.routers.main_file.db.log_activity"):
                response = self.client.patch("/api/main-file/cell", json={
                    "sheet": "庫存主檔",
                    "row": 2,
                    "col": 9,
                    "value": "10",
                })

            self.assertEqual(response.status_code, 200)
            data = response.json()
            self.assertEqual(data["affected_cells"], [
                {"row": 2, "col": 11, "value": 80},
                {"row": 2, "col": 14, "value": 65},
            ])
            self.assertFalse(data["schedule_refresh_required"])
            mock_update_stock.assert_called_once_with({"PART-1": 65.0})
            mock_refresh.assert_not_called()

            saved = openpyxl.load_workbook(main_path)
            try:
                saved_sheet = saved["庫存主檔"]
                self.assertEqual(saved_sheet.cell(row=2, column=9).value, 10)
                self.assertEqual(saved_sheet.cell(row=2, column=11).value, 80)
                self.assertEqual(saved_sheet.cell(row=2, column=14).value, 65)
            finally:
                saved.close()

    def test_edit_main_cell_rebuilds_active_merge_drafts_after_batch_recalc(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "庫存主檔"
            sheet.cell(row=1, column=1).value = "料號"
            sheet.cell(row=1, column=8).value = "盤點"
            for col, code in ((9, "1-1"), (12, "1-2")):
                sheet.cell(row=1, column=col).value = code
                sheet.cell(row=1, column=col + 1).value = f"PO-{code}"
                sheet.cell(row=1, column=col + 2).value = f"MODEL-{code}"
            sheet.cell(row=2, column=1).value = "PART-1"
            sheet.cell(row=2, column=8).value = 100
            sheet.cell(row=2, column=9).value = 0
            sheet.cell(row=2, column=10).value = 30
            sheet.cell(row=2, column=11).value = 70
            workbook.save(main_path)
            workbook.close()

            def fake_setting(key, default=""):
                return str(main_path) if key == "main_file_path" else default

            with patch("app.routers.main_file.db.get_setting", side_effect=fake_setting), \
                 patch("app.routers.main_file.backup_main_file"), \
                 patch("app.routers.main_file.db.get_snapshot", return_value={"PART-1": {"stock_qty": 100}}), \
                 patch("app.routers.main_file.db.update_snapshot_stock", return_value=1), \
                 patch("app.routers.main_file.db.get_active_merge_drafts", return_value=[
                     {"order_id": 5},
                     {"order_id": 6},
                 ]), \
                 patch("app.routers.main_file.rebuild_merge_drafts", return_value=[]) as mock_rebuild, \
                 patch("app.routers.main_file.db.log_activity"):
                response = self.client.patch("/api/main-file/cell", json={
                    "sheet": "庫存主檔",
                    "row": 2,
                    "col": 9,
                    "value": "10",
                })

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["schedule_refresh_required"])
        mock_rebuild.assert_called_once_with([5, 6])

    def test_update_snapshot_moq_endpoint_normalizes_part_number(self):
        with patch("app.routers.main_file.db.upsert_snapshot_moq", return_value="IC-LD39100PUR-TAB") as mock_update, \
             patch("app.routers.main_file.db.log_activity") as mock_log:
            response = self.client.patch("/api/main-file/moq", json={
                "part_number": " ic-ld39100pur-tab ",
                "moq": 2500,
            })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {
            "ok": True,
            "part_number": "IC-LD39100PUR-TAB",
            "moq": 2500,
        })
        mock_update.assert_called_once_with("IC-LD39100PUR-TAB", 2500)
        mock_log.assert_called_once()

    def test_set_snapshot_keeps_manual_moq_overrides_over_excel(self):
        with patch("app.routers.main_file.db.get_setting", return_value="C:/main.xlsx"), \
             patch("app.routers.main_file.Path.exists", return_value=True), \
             patch("app.routers.main_file.db.get_manual_snapshot_moq", return_value={"PART-1": 3000}), \
             patch("app.routers.main_file.read_stock", return_value={"PART-1": 10, "PART-2": 20}), \
             patch("app.routers.main_file.read_moq", return_value={"PART-1": 500, "PART-2": 1200}), \
             patch("app.routers.main_file.db.save_snapshot") as mock_save_snapshot, \
             patch("app.routers.main_file.db.log_activity"):
            response = self.client.post("/api/main-file/snapshot")

        self.assertEqual(response.status_code, 200)
        mock_save_snapshot.assert_called_once_with(
            {"PART-1": 10, "PART-2": 20},
            {"PART-1": 3000, "PART-2": 1200},
            manual_moq_parts={"PART-1"},
        )

    def test_bom_editor_returns_source_metadata(self):
        bom_record = {
            "id": "bom-1",
            "filename": "formal.xlsx",
            "filepath": "C:/formal.xlsx",
            "source_filename": "legacy.xls",
            "source_format": ".xls",
            "is_converted": 1,
            "group_model": "MODEL-A",
            "uploaded_at": "2026-03-12T08:00:00",
        }
        parsed = BomFile(
            id="bom-1",
            filename="formal.xlsx",
            path="C:/formal.xlsx",
            po_number=123,
            model="MODEL-A",
            pcb="PCB-A",
            group_model="MODEL-A",
            order_qty=10,
            uploaded_at="2026-03-12T08:00:00",
            source_filename="legacy.xls",
            source_format=".xls",
            is_converted=True,
            components=[
                BomComponent(part_number="PART-1", qty_per_board=1, needed_qty=300, source_row=5),
            ],
        )

        with patch("app.routers.bom._get_required_bom", return_value=bom_record), \
             patch("app.routers.bom._ensure_editable_bom_record", return_value=bom_record), \
             patch("app.routers.bom.parse_bom_for_storage", return_value=parsed), \
             patch("app.routers.bom.db.save_bom_file") as mock_save_bom:
            response = self.client.get("/api/bom/bom-1/editor")

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["filename"], "formal.xlsx")
        self.assertEqual(data["source_filename"], "legacy.xls")
        self.assertEqual(data["source_format"], ".xls")
        self.assertTrue(data["is_converted"])
        self.assertEqual(data["component_count"], 1)
        mock_save_bom.assert_called_once()
        saved_payload = mock_save_bom.call_args.args[0]
        self.assertEqual(saved_payload["components"][0]["part_number"], "PART-1")
        self.assertEqual(saved_payload["components"][0]["needed_qty"], 300)

    def test_root_and_static_disable_cache(self):
        root_response = self.client.get("/")
        static_response = self.client.get("/static/style.css")

        self.assertEqual(root_response.status_code, 200)
        self.assertEqual(static_response.status_code, 200)
        self.assertEqual(root_response.headers.get("cache-control"), "no-store, no-cache, must-revalidate, max-age=0")
        self.assertEqual(static_response.headers.get("cache-control"), "no-store, no-cache, must-revalidate, max-age=0")

    def test_get_bom_file_normalizes_legacy_xls_before_download(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            bom_path = Path(temp_dir) / "legacy.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["A1"] = "legacy"
            wb.save(bom_path)
            wb.close()

            raw_bom_record = {
                "id": "bom-legacy",
                "filename": "legacy.xls",
                "filepath": str(Path(temp_dir) / "legacy.xls"),
            }
            converted_bom_record = {
                **raw_bom_record,
                "filename": "legacy.xlsx",
                "filepath": str(bom_path),
            }

            with patch("app.routers.bom._get_required_bom", return_value=raw_bom_record), \
                 patch("app.routers.bom._ensure_editable_bom_record", return_value=converted_bom_record) as mock_ensure, \
                 patch("app.routers.bom.append_minute_timestamp", return_value="legacy_20260312_1740.xlsx"):
                response = self.client.get("/api/bom/bom-legacy/file")

        self.assertEqual(response.status_code, 200)
        self.assertIn("legacy_20260312_1740.xlsx", response.headers["content-disposition"])
        mock_ensure.assert_called_once_with(raw_bom_record)

    def test_bom_revision_list_returns_history(self):
        bom_record = {
            "id": "bom-1",
            "filename": "formal.xlsx",
            "filepath": "C:/formal.xlsx",
            "source_filename": "legacy.xls",
        }
        revisions = [
            {
                "id": 11,
                "bom_file_id": "bom-1",
                "revision_number": 2,
                "filename": "formal.xlsx",
                "filepath": "C:/history/v2.xlsx",
                "source_action": "edit",
                "note": "編輯後儲存",
                "created_at": "2026-03-12T17:40:00",
            },
            {
                "id": 7,
                "bom_file_id": "bom-1",
                "revision_number": 1,
                "filename": "formal.xlsx",
                "filepath": "C:/history/v1.xlsx",
                "source_action": "upload",
                "note": "上傳 BOM",
                "created_at": "2026-03-12T16:00:00",
            },
        ]

        with patch("app.routers.bom._get_required_bom", return_value=bom_record), \
             patch("app.routers.bom._ensure_editable_bom_record", return_value=bom_record), \
             patch("app.routers.bom.ensure_bom_revision_history", return_value=revisions) as mock_history:
            response = self.client.get("/api/bom/bom-1/revisions")

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["bom"]["id"], "bom-1")
        self.assertEqual(data["bom"]["source_filename"], "legacy.xls")
        self.assertEqual(len(data["revisions"]), 2)
        self.assertEqual(data["revisions"][0]["revision_number"], 2)
        mock_history.assert_called_once_with(bom_record)

    def test_download_bom_revision_uses_versioned_filename(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            revision_path = Path(temp_dir) / "revision.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["A1"] = "revision"
            wb.save(revision_path)
            wb.close()

            revision = {
                "id": 9,
                "bom_file_id": "bom-1",
                "revision_number": 3,
                "filename": "formal.xlsx",
                "filepath": str(revision_path),
            }

            with patch("app.routers.bom.db.get_bom_revision", return_value=revision), \
                 patch("app.routers.bom._build_revision_download_name", return_value="formal_v003_20260312_1740.xlsx"):
                response = self.client.get("/api/bom/bom-1/revisions/9/file")

        self.assertEqual(response.status_code, 200)
        self.assertIn("formal_v003_20260312_1740.xlsx", response.headers["content-disposition"])

    def test_save_bom_editor_creates_revision_snapshot(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            bom_path = Path(temp_dir) / "formal.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["A1"] = "formal"
            wb.save(bom_path)
            wb.close()

            bom_record = {
                "id": "bom-1",
                "filename": "formal.xlsx",
                "filepath": str(bom_path),
                "uploaded_at": "2026-03-12T08:00:00",
                "source_filename": "formal.xlsx",
                "source_format": ".xlsx",
                "is_converted": 0,
            }
            parsed = BomFile(
                id="bom-1",
                filename="formal.xlsx",
                path=str(bom_path),
                po_number=123,
                model="MODEL-A",
                pcb="PCB-A",
                group_model="MODEL-A",
                order_qty=10,
                uploaded_at="2026-03-12T08:00:00",
                components=[BomComponent(part_number="PART-1", source_row=5)],
            )

            with patch("app.routers.bom._get_required_bom", return_value=bom_record), \
                 patch("app.routers.bom._ensure_editable_bom_record", return_value=bom_record), \
                 patch("app.routers.bom.ensure_bom_revision_history") as mock_ensure_history, \
                 patch("app.routers.bom.backup_bom_file", return_value=str(Path(temp_dir) / "backup.xlsx")), \
                 patch("app.routers.bom.apply_bom_editor_changes"), \
                 patch("app.routers.bom.parse_bom_for_storage", return_value=parsed), \
                 patch("app.routers.bom.db.save_bom_file") as mock_save_bom, \
                 patch("app.routers.bom.snapshot_bom_revision") as mock_snapshot_revision, \
                 patch("app.routers.bom.db.log_activity"):
                response = self.client.put("/api/bom/bom-1/editor", json={
                    "po_number": 123,
                    "order_qty": 10,
                    "model": "MODEL-A",
                    "pcb": "PCB-A",
                    "group_model": "MODEL-A",
                    "components": [
                        {
                            "source_row": 5,
                            "part_number": "PART-1",
                            "description": "",
                            "qty_per_board": 1,
                            "needed_qty": 2,
                            "prev_qty_cs": 3,
                            "is_dash": False,
                        }
                    ],
                })

        self.assertEqual(response.status_code, 200)
        mock_ensure_history.assert_called_once_with(bom_record)
        mock_save_bom.assert_called_once()
        mock_snapshot_revision.assert_called_once()

    def test_bom_download_zip_entries_include_timestamp(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            first_path = Path(temp_dir) / "first.xlsx"
            second_path = Path(temp_dir) / "second.xlsx"

            for path, value in ((first_path, "A"), (second_path, "B")):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws["A1"] = value
                wb.save(path)
                wb.close()

            bom_records = [
                {"id": "bom-1", "filename": "first.xlsx", "filepath": str(first_path)},
                {"id": "bom-2", "filename": "second.xlsx", "filepath": str(second_path)},
            ]

            with patch("app.routers.bom.db.get_bom_files_by_models", return_value=bom_records), \
                 patch("app.routers.bom.append_minute_timestamp", side_effect=lambda filename, now=None: f"{Path(filename).stem}_20260312_1740{Path(filename).suffix}"), \
                 patch("app.routers.bom.build_generated_filename", return_value="BOM_20260312_1740.zip"):
                response = self.client.post("/api/bom/download", json={"models": ["MODEL-A"]})

        self.assertEqual(response.status_code, 200)
        archive = zipfile.ZipFile(io.BytesIO(response.content))
        self.assertEqual(sorted(archive.namelist()), ["first_20260312_1740.xlsx", "second_20260312_1740.xlsx"])
        archive.close()

    def test_dispatch_download_writes_prev_batch_and_supplements_to_separate_columns(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            bom_path = Path(temp_dir) / "dispatch.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.merge_cells("G1:H1")
            ws.cell(row=1, column=7).value = "製單號碼M/O:"
            ws.cell(row=1, column=10).value = "訂單數量:"
            ws.cell(row=2, column=3).value = "MODEL-A"
            ws.cell(row=2, column=4).value = "PCB-A"
            ws.cell(row=3, column=7).value = "上批餘料"
            ws.cell(row=3, column=8).value = "增添料數"
            ws.cell(row=5, column=3).value = "PART-1"
            ws.cell(row=5, column=6).value = 10
            ws.cell(row=5, column=7).value = 3
            ws.cell(row=5, column=8).value = 99
            ws.cell(row=6, column=3).value = "PART-2"
            ws.cell(row=6, column=6).value = 20
            ws.cell(row=6, column=8).value = 8
            ws.cell(row=7, column=3).value = "PART-3"
            ws.cell(row=7, column=6).value = 30
            ws.cell(row=7, column=7).value = 0
            ws.cell(row=7, column=8).value = 0
            ws.cell(row=8, column=3).value = "PART-4"
            ws.cell(row=8, column=6).value = 40
            ws.cell(row=8, column=7).value = 12
            ws.cell(row=8, column=8).value = 5
            wb.save(bom_path)
            wb.close()

            bom_record = {
                "id": "bom-1",
                "filename": "dispatch.xlsx",
                "filepath": str(bom_path),
                "source_filename": "dispatch.xlsx",
                "source_format": ".xlsx",
                "is_converted": 0,
                "po_number": "0",
                "group_model": "MODEL-A",
                "uploaded_at": "2026-03-12T08:00:00",
            }

            with patch("app.routers.bom.db.get_bom_files", return_value=[bom_record]):
                response = self.client.post("/api/bom/dispatch-download", json={
                    "bom_ids": ["bom-1"],
                    "supplements": {"PART-1": 7},
                    "header_overrides": {"bom-1": {"po_number": "4500059234"}},
                    "carry_overs": {"bom-1": {"PART-1": 135, "PART-2": 246}},
                })

        self.assertEqual(response.status_code, 200)
        downloaded = openpyxl.load_workbook(io.BytesIO(response.content), data_only=False)
        ws = downloaded.active
        self.assertEqual(ws.cell(row=1, column=7).value, "製單號碼M/O:4500059234")
        self.assertEqual(ws.cell(row=5, column=7).value, 135)
        self.assertEqual(ws.cell(row=5, column=8).value, 7)
        self.assertEqual(ws.cell(row=6, column=7).value, 246)
        self.assertEqual(ws.cell(row=6, column=8).value, 0)
        self.assertEqual(ws.cell(row=7, column=7).value, 0)
        self.assertEqual(ws.cell(row=7, column=8).value, 0)
        self.assertEqual(ws.cell(row=8, column=7).value, 12)
        self.assertEqual(ws.cell(row=8, column=8).value, 0)
        downloaded.close()

    def test_dispatch_download_uses_recalc_save_helper_for_generated_workbooks(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            bom_path = Path(temp_dir) / "dispatch.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.merge_cells("G1:H1")
            ws.cell(row=1, column=7).value = "製單號碼M/O:"
            ws.cell(row=5, column=3).value = "PART-1"
            ws.cell(row=5, column=6).value = 10
            ws.cell(row=5, column=7).value = 0
            ws.cell(row=5, column=8).value = 0
            wb.save(bom_path)
            wb.close()

            bom_record = {
                "id": "bom-1",
                "filename": "dispatch.xlsx",
                "filepath": str(bom_path),
                "source_filename": "dispatch.xlsx",
                "source_format": ".xlsx",
                "is_converted": 0,
                "po_number": "0",
                "group_model": "MODEL-A",
                "uploaded_at": "2026-03-12T08:00:00",
            }

            def fake_save_bytes(workbook, filename):
                buffer = io.BytesIO()
                workbook.save(buffer)
                buffer.seek(0)
                return buffer

            with patch("app.routers.bom.db.get_bom_files", return_value=[bom_record]), \
                 patch("app.routers.bom.save_workbook_bytes_with_recalc", side_effect=fake_save_bytes) as mock_save:
                response = self.client.post("/api/bom/dispatch-download", json={
                    "bom_ids": ["bom-1"],
                    "supplements": {"PART-1": 7},
                    "header_overrides": {"bom-1": {"po_number": "4500059234"}},
                    "carry_overs": {"bom-1": {"PART-1": 135}},
                })

        self.assertEqual(response.status_code, 200)
        mock_save.assert_called_once()

    def test_dispatch_download_marks_manual_supplement_orange_when_qty_exceeds_st_stock(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            bom_path = Path(temp_dir) / "dispatch.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.merge_cells("G1:H1")
            ws.cell(row=1, column=7).value = "製單號碼M/O:"
            ws.cell(row=5, column=3).value = "PART-1"
            ws.cell(row=5, column=7).value = 0
            ws.cell(row=5, column=8).value = 0
            wb.save(bom_path)
            wb.close()

            bom_record = {
                "id": "bom-1",
                "filename": "dispatch.xlsx",
                "filepath": str(bom_path),
                "source_filename": "dispatch.xlsx",
                "source_format": ".xlsx",
                "is_converted": 0,
                "po_number": "0",
                "group_model": "MODEL-A",
                "uploaded_at": "2026-03-12T08:00:00",
            }

            with patch("app.routers.bom.db.get_bom_files", return_value=[bom_record]), \
                 patch("app.routers.bom.db.get_st_inventory_stock", return_value={"PART-1": 10000}):
                response = self.client.post("/api/bom/dispatch-download", json={
                    "bom_ids": ["bom-1"],
                    "supplements": {"PART-1": 20000},
                    "header_overrides": {"bom-1": {"po_number": "4500059234"}},
                    "carry_overs": {"bom-1": {"PART-1": 135}},
                })

        self.assertEqual(response.status_code, 200)
        downloaded = openpyxl.load_workbook(io.BytesIO(response.content), data_only=False)
        ws = downloaded.active
        self.assertEqual(ws.cell(row=5, column=8).value, 20000)
        self.assertEqual(ws.cell(row=5, column=8).fill.fill_type, "solid")
        self.assertTrue(str(ws.cell(row=5, column=8).fill.start_color.rgb or "").endswith("FFC000"))
        downloaded.close()

    def test_dispatch_download_persists_order_supplements_for_selected_orders(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            bom_path = Path(temp_dir) / "dispatch.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.merge_cells("G1:H1")
            ws.cell(row=1, column=7).value = "製單號碼M/O:"
            ws.cell(row=5, column=3).value = "PART-1"
            wb.save(bom_path)
            wb.close()

            bom_record = {
                "id": "bom-1",
                "filename": "dispatch.xlsx",
                "filepath": str(bom_path),
                "source_filename": "dispatch.xlsx",
                "source_format": ".xlsx",
                "is_converted": 0,
                "po_number": "0",
                "group_model": "MODEL-A",
                "uploaded_at": "2026-03-12T08:00:00",
            }

            with patch("app.routers.bom.db.get_bom_files", return_value=[bom_record]), \
                 patch("app.routers.bom._build_order_based_export_values", return_value=({}, {}, {}, {}, {})), \
                 patch("app.routers.bom.build_order_supplement_allocations", return_value={5: {"PART-1": 3000}}) as mock_allocations, \
                 patch("app.routers.bom.db.replace_order_supplements") as mock_replace:
                response = self.client.post("/api/bom/dispatch-download", json={
                    "bom_ids": ["bom-1"],
                    "order_ids": [5],
                    "supplements": {"PART-1": 3000},
                    "header_overrides": {"bom-1": {"po_number": "4500059234"}},
                    "carry_overs": {"bom-1": {"PART-1": 135}},
                })

        self.assertEqual(response.status_code, 200)
        mock_allocations.assert_called_once_with([5], {"PART-1": 3000})
        mock_replace.assert_called_once_with([5], {5: {"PART-1": 3000}})

    def test_dispatch_download_computes_carry_over_per_bom_in_order(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            bom_a_path = Path(temp_dir) / "board-a.xlsx"
            bom_c_path = Path(temp_dir) / "board-c.xlsx"

            for path, label in ((bom_a_path, "A"), (bom_c_path, "C")):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.merge_cells("G1:H1")
                ws.cell(row=1, column=7).value = "M/O:"
                ws.cell(row=5, column=3).value = "EC-20023A"
                ws.cell(row=5, column=7).value = None
                ws.cell(row=5, column=8).value = None
                ws.cell(row=6, column=3).value = f"PART-{label}"
                ws.cell(row=6, column=7).value = None
                ws.cell(row=6, column=8).value = None
                wb.save(path)
                wb.close()

            bom_a = {
                "id": "bom-a",
                "filename": "board-a.xlsx",
                "filepath": str(bom_a_path),
                "source_filename": "board-a.xlsx",
                "source_format": ".xlsx",
                "is_converted": 1,
                "po_number": "0",
                "model": "T356789IU_MAIN_BOARD_A",
                "group_model": "MODEL-A",
                "sort_order": 0,
                "uploaded_at": "2026-03-12T08:00:00",
            }
            bom_c = {
                "id": "bom-c",
                "filename": "board-c.xlsx",
                "filepath": str(bom_c_path),
                "source_filename": "board-c.xlsx",
                "source_format": ".xlsx",
                "is_converted": 1,
                "po_number": "0",
                "model": "T356789IU_DISPLAY_C",
                "group_model": "MODEL-A",
                "sort_order": 1,
                "uploaded_at": "2026-03-12T08:01:00",
            }

            components_by_bom = {
                "bom-a": [
                    {"part_number": "EC-20023A", "needed_qty": 5505, "prev_qty_cs": 0, "is_dash": 0},
                    {"part_number": "PART-A", "needed_qty": 10, "prev_qty_cs": 0, "is_dash": 0},
                ],
                "bom-c": [
                    {"part_number": "EC-20023A", "needed_qty": 100, "prev_qty_cs": 0, "is_dash": 0},
                    {"part_number": "PART-C", "needed_qty": 20, "prev_qty_cs": 0, "is_dash": 0},
                ],
            }

            with patch("app.routers.bom.db.get_bom_files", return_value=[bom_a, bom_c]), \
                 patch("app.routers.bom.db.get_order", return_value={"id": 1, "model": "MODEL-A"}), \
                 patch("app.routers.bom.db.get_bom_components", side_effect=lambda bom_id: components_by_bom[bom_id]), \
                 patch("app.routers.bom.db.get_snapshot", return_value={
                     "EC-20023A": {"stock_qty": 5625, "moq": 0},
                     "PART-A": {"stock_qty": 30, "moq": 0},
                     "PART-C": {"stock_qty": 40, "moq": 0},
                 }), \
                 patch("app.routers.bom.db.get_setting", return_value=""), \
                 patch("app.routers.bom.db.get_snapshot_taken_at", return_value="2026-03-12T08:00:00"), \
                 patch("app.routers.bom.db.get_all_dispatched_consumption", return_value={}), \
                 patch("app.routers.bom.db.replace_order_supplements"):
                response = self.client.post("/api/bom/dispatch-download", json={
                    "bom_ids": ["bom-a", "bom-c"],
                    "order_ids": [1],
                    "supplements": {},
                    "header_overrides": {
                        "bom-a": {"po_number": "4500059234"},
                        "bom-c": {"po_number": "4500059234"},
                    },
                })

        self.assertEqual(response.status_code, 200)
        archive = zipfile.ZipFile(io.BytesIO(response.content))
        self.assertEqual(len(archive.namelist()), 2)
        bom_a_name = next(n for n in archive.namelist() if "T356789IU_MAIN_BOARD_A" in n)
        bom_c_name = next(n for n in archive.namelist() if "T356789IU_DISPLAY_C" in n)

        wb_a = openpyxl.load_workbook(io.BytesIO(archive.read(bom_a_name)), data_only=False)
        ws_a = wb_a.active
        self.assertEqual(ws_a.cell(row=5, column=7).value, 5625)
        self.assertEqual(ws_a.cell(row=5, column=8).value, 0)
        wb_a.close()

        wb_c = openpyxl.load_workbook(io.BytesIO(archive.read(bom_c_name)), data_only=False)
        ws_c = wb_c.active
        self.assertEqual(ws_c.cell(row=5, column=7).value, 120)
        self.assertEqual(ws_c.cell(row=5, column=8).value, 0)
        wb_c.close()
        archive.close()

    def test_dispatch_download_applies_supplement_once_and_carries_it_forward(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            bom_a_path = Path(temp_dir) / "board-a.xlsx"
            bom_c_path = Path(temp_dir) / "board-c.xlsx"

            for path in (bom_a_path, bom_c_path):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.merge_cells("G1:H1")
                ws.cell(row=1, column=7).value = "M/O:"
                ws.cell(row=5, column=3).value = "EC-20023A"
                ws.cell(row=5, column=7).value = None
                ws.cell(row=5, column=8).value = None
                wb.save(path)
                wb.close()

            bom_a = {
                "id": "bom-a",
                "filename": "board-a.xlsx",
                "filepath": str(bom_a_path),
                "source_filename": "board-a.xlsx",
                "source_format": ".xlsx",
                "is_converted": 1,
                "po_number": "0",
                "model": "T356789IU_MAIN_BOARD_A",
                "group_model": "MODEL-A",
                "sort_order": 0,
                "uploaded_at": "2026-03-12T08:00:00",
            }
            bom_c = {
                "id": "bom-c",
                "filename": "board-c.xlsx",
                "filepath": str(bom_c_path),
                "source_filename": "board-c.xlsx",
                "source_format": ".xlsx",
                "is_converted": 1,
                "po_number": "0",
                "model": "T356789IU_DISPLAY_C",
                "group_model": "MODEL-A",
                "sort_order": 1,
                "uploaded_at": "2026-03-12T08:01:00",
            }

            components_by_bom = {
                "bom-a": [{"part_number": "EC-20023A", "needed_qty": 5505, "prev_qty_cs": 0, "is_dash": 0}],
                "bom-c": [{"part_number": "EC-20023A", "needed_qty": 2000, "prev_qty_cs": 0, "is_dash": 0}],
            }

            with patch("app.routers.bom.db.get_bom_files", return_value=[bom_a, bom_c]), \
                 patch("app.routers.bom.db.get_order", return_value={"id": 1, "model": "MODEL-A"}), \
                 patch("app.routers.bom.db.get_bom_components", side_effect=lambda bom_id: components_by_bom[bom_id]), \
                 patch("app.routers.bom.db.get_snapshot", return_value={
                     "EC-20023A": {"stock_qty": 5625, "moq": 0},
                 }), \
                 patch("app.routers.bom.db.get_setting", return_value=""), \
                 patch("app.routers.bom.db.get_snapshot_taken_at", return_value="2026-03-12T08:00:00"), \
                 patch("app.routers.bom.db.get_all_dispatched_consumption", return_value={}), \
                 patch("app.routers.bom.build_order_supplement_allocations", return_value={1: {"EC-20023A": 3000}}), \
                 patch("app.routers.bom.db.replace_order_supplements"):
                response = self.client.post("/api/bom/dispatch-download", json={
                    "bom_ids": ["bom-a", "bom-c"],
                    "order_ids": [1],
                    "supplements": {"EC-20023A": 3000},
                    "header_overrides": {
                        "bom-a": {"po_number": "4500059234"},
                        "bom-c": {"po_number": "4500059234"},
                    },
                })

        self.assertEqual(response.status_code, 200)
        archive = zipfile.ZipFile(io.BytesIO(response.content))
        self.assertEqual(len(archive.namelist()), 2)
        bom_a_name = next(n for n in archive.namelist() if "T356789IU_MAIN_BOARD_A" in n)
        bom_c_name = next(n for n in archive.namelist() if "T356789IU_DISPLAY_C" in n)

        wb_a = openpyxl.load_workbook(io.BytesIO(archive.read(bom_a_name)), data_only=False)
        ws_a = wb_a.active
        self.assertEqual(ws_a.cell(row=5, column=7).value, 5625)
        self.assertEqual(ws_a.cell(row=5, column=8).value, 0)
        wb_a.close()

        wb_c = openpyxl.load_workbook(io.BytesIO(archive.read(bom_c_name)), data_only=False)
        ws_c = wb_c.active
        self.assertEqual(ws_c.cell(row=5, column=7).value, 120)
        self.assertEqual(ws_c.cell(row=5, column=8).value, 3000)
        wb_c.close()
        archive.close()

    def test_dispatch_download_normalizes_legacy_xls_before_export(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            bom_path = Path(temp_dir) / "dispatch.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.merge_cells("G1:H1")
            ws.cell(row=1, column=7).value = "鋆賢?ⅣM/O:"
            ws.cell(row=3, column=7).value = "銝擗?"
            ws.cell(row=3, column=8).value = "憓溶?"
            ws.cell(row=5, column=3).value = "PART-1"
            ws.cell(row=5, column=7).value = 0
            ws.cell(row=5, column=8).value = 0
            wb.save(bom_path)
            wb.close()

            raw_bom_record = {
                "id": "bom-legacy",
                "filename": "dispatch.xls",
                "filepath": str(Path(temp_dir) / "dispatch.xls"),
                "source_filename": "dispatch.xls",
                "source_format": ".xls",
                "is_converted": 0,
                "po_number": "0",
                "group_model": "MODEL-A",
                "uploaded_at": "2026-03-12T08:00:00",
            }
            converted_bom_record = {
                **raw_bom_record,
                "filename": "dispatch.xlsx",
                "filepath": str(bom_path),
                "is_converted": 1,
            }

            with patch("app.routers.bom.db.get_bom_files", return_value=[raw_bom_record]), \
                 patch("app.routers.bom._ensure_editable_bom_record", return_value=converted_bom_record) as mock_ensure:
                response = self.client.post("/api/bom/dispatch-download", json={
                    "bom_ids": ["bom-legacy"],
                    "supplements": {"PART-1": 7},
                    "header_overrides": {"bom-legacy": {"po_number": "4500059234"}},
                    "carry_overs": {"bom-legacy": {"PART-1": 135}},
                })

        self.assertEqual(response.status_code, 200)
        mock_ensure.assert_called_once_with(raw_bom_record)
        downloaded = openpyxl.load_workbook(io.BytesIO(response.content), data_only=False)
        ws = downloaded.active
        self.assertEqual(ws.cell(row=1, column=7).value, "鋆賢?ⅣM/O:4500059234")
        self.assertEqual(ws.cell(row=5, column=7).value, 135)
        self.assertEqual(ws.cell(row=5, column=8).value, 7)
        downloaded.close()

    def test_dispatch_download_scales_order_qty_and_needed_cells_from_schedule(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            bom_path = Path(temp_dir) / "dispatch.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.cell(row=1, column=11).value = 10
            ws.cell(row=2, column=3).value = "MODEL-A"
            ws.cell(row=2, column=7).value = 10
            ws.cell(row=5, column=2).value = 2
            ws.cell(row=5, column=3).value = "PART-1"
            ws.cell(row=5, column=6).value = 20
            ws.cell(row=5, column=7).value = 0
            ws.cell(row=5, column=8).value = 0
            wb.save(bom_path)
            wb.close()

            bom_record = {
                "id": "bom-scale",
                "filename": "dispatch.xlsx",
                "filepath": str(bom_path),
                "source_filename": "dispatch.xlsx",
                "source_format": ".xlsx",
                "is_converted": 1,
                "po_number": "0",
                "order_qty": 10,
                "group_model": "MODEL-A",
                "uploaded_at": "2026-03-12T08:00:00",
            }

            with patch("app.routers.bom.db.get_bom_files", return_value=[bom_record]), \
                 patch("app.routers.bom.db.get_order", return_value={"id": 1, "model": "MODEL-A", "order_qty": 5}), \
                 patch("app.routers.bom.db.get_bom_components", return_value=[{
                     "part_number": "PART-1",
                     "qty_per_board": 2,
                     "bom_order_qty": 10,
                     "needed_qty": 20,
                     "prev_qty_cs": 0,
                     "is_dash": 0,
                 }]), \
                 patch("app.routers.bom.db.get_snapshot", return_value={"PART-1": {"stock_qty": 50, "moq": 0}}), \
                 patch("app.routers.bom.db.get_setting", return_value=""), \
                 patch("app.routers.bom.db.get_snapshot_taken_at", return_value="2026-03-12T08:00:00"), \
                 patch("app.routers.bom.db.get_all_dispatched_consumption", return_value={}), \
                 patch("app.routers.bom.build_order_supplement_allocations", return_value={1: {}}), \
                 patch("app.routers.bom.db.replace_order_supplements"):
                response = self.client.post("/api/bom/dispatch-download", json={
                    "bom_ids": ["bom-scale"],
                    "order_ids": [1],
                    "supplements": {},
                    "header_overrides": {"bom-scale": {"po_number": "4500059234"}},
                })

        self.assertEqual(response.status_code, 200)
        downloaded = openpyxl.load_workbook(io.BytesIO(response.content), data_only=False)
        ws = downloaded.active
        self.assertEqual(ws.cell(row=1, column=11).value, 5)
        self.assertEqual(ws.cell(row=2, column=7).value, 5)
        self.assertEqual(ws.cell(row=5, column=6).value, 10)
        downloaded.close()

    def test_dispatch_download_keeps_needed_formula_cells(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            bom_path = Path(temp_dir) / "dispatch.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.cell(row=1, column=11).value = 10
            ws.cell(row=2, column=3).value = "MODEL-A"
            ws.cell(row=2, column=7).value = 10
            ws.cell(row=5, column=2).value = 2
            ws.cell(row=5, column=3).value = "PART-1"
            ws.cell(row=5, column=6).value = "=B5*$K$1*1.06"
            ws.cell(row=5, column=7).value = 0
            ws.cell(row=5, column=8).value = 0
            wb.save(bom_path)
            wb.close()

            bom_record = {
                "id": "bom-scale",
                "filename": "dispatch.xlsx",
                "filepath": str(bom_path),
                "source_filename": "dispatch.xlsx",
                "source_format": ".xlsx",
                "is_converted": 1,
                "po_number": "0",
                "order_qty": 10,
                "group_model": "MODEL-A",
                "uploaded_at": "2026-03-12T08:00:00",
            }

            with patch("app.routers.bom.db.get_bom_files", return_value=[bom_record]), \
                 patch("app.routers.bom.db.get_order", return_value={"id": 1, "model": "MODEL-A", "order_qty": 5}), \
                 patch("app.routers.bom.db.get_bom_components", return_value=[{
                     "part_number": "PART-1",
                     "qty_per_board": 2,
                     "bom_order_qty": 10,
                     "needed_qty": 21.2,
                     "prev_qty_cs": 0,
                     "is_dash": 0,
                 }]), \
                 patch("app.routers.bom.db.get_snapshot", return_value={"PART-1": {"stock_qty": 50, "moq": 0}}), \
                 patch("app.routers.bom.db.get_setting", return_value=""), \
                 patch("app.routers.bom.db.get_snapshot_taken_at", return_value="2026-03-12T08:00:00"), \
                 patch("app.routers.bom.db.get_all_dispatched_consumption", return_value={}), \
                 patch("app.routers.bom.build_order_supplement_allocations", return_value={1: {}}), \
                 patch("app.routers.bom.db.replace_order_supplements"):
                response = self.client.post("/api/bom/dispatch-download", json={
                    "bom_ids": ["bom-scale"],
                    "order_ids": [1],
                    "supplements": {},
                    "header_overrides": {"bom-scale": {"po_number": "4500059234"}},
                })

        self.assertEqual(response.status_code, 200)
        downloaded = openpyxl.load_workbook(io.BytesIO(response.content), data_only=False)
        ws = downloaded.active
        self.assertEqual(ws.cell(row=1, column=11).value, 5)
        self.assertEqual(ws.cell(row=2, column=7).value, 5)
        self.assertEqual(ws.cell(row=5, column=6).value, "=B5*$K$1*1.06")
        downloaded.close()

    def test_get_database_backups_returns_overview(self):
        overview = {
            "enabled": True,
            "hour": 2,
            "minute": 0,
            "keep_count": 14,
            "backups": [{"name": "system_backup_20260313_020000.db"}],
        }

        with patch("app.routers.system.db_backup.get_database_backup_overview", return_value=overview):
            response = self.client.get("/api/system/db-backups")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), overview)

    def test_get_system_app_meta_returns_version_payload(self):
        meta = {
            "app_name": "出貨排程系統",
            "version": "v2026.03.16.1",
            "released_at": "2026-03-16",
            "headline": "這版聚焦在發料流程與版本可視性。",
            "sections": [{"title": "發料", "items": ["加入版本號與更新說明"]}],
        }

        with patch("app.routers.system.get_app_meta", return_value=meta):
            response = self.client.get("/api/system/app-meta")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), meta)

    def test_get_st_inventory_data_returns_snapshot_payload(self):
        snapshot = {
            "PART-1": {"stock_qty": 8.0, "description": "Capacitor"},
            "PART-2": {"stock_qty": 3.0, "description": "Resistor"},
        }

        with patch("app.routers.system.db.get_st_inventory_snapshot", return_value=snapshot), \
             patch("app.routers.system.db.get_setting", side_effect=lambda key, default="": {
                 "st_inventory_loaded_at": "2026-03-16T09:30:00",
                 "st_inventory_filename": "st_inventory.xlsx",
             }.get(key, default)):
            response = self.client.get("/api/system/st-inventory/data")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {
            "stock": {"PART-1": 8.0, "PART-2": 3.0},
            "descriptions": {"PART-1": "Capacitor", "PART-2": "Resistor"},
            "part_count": 2,
            "loaded_at": "2026-03-16T09:30:00",
            "filename": "st_inventory.xlsx",
        })

    def test_get_missing_moq_st_packages_returns_rows(self):
        with patch("app.routers.system.build_missing_moq_package_rows", return_value=[
            {
                "part_number": "PART-1",
                "description": "Capacitor",
                "stock_qty": 1000,
                "package_text": "200,300,500",
                "package_values": [200, 300, 500],
                "package_sum": 1000,
                "diff_qty": 0,
                "matches_stock": True,
                "updated_at": "2026-04-07T10:00:00",
            },
        ]), \
             patch("app.routers.system.db.get_setting", side_effect=lambda key, default="": {
                 "main_loaded_at": "2026-04-07T09:00:00",
                 "main_filename": "main.xlsx",
             }.get(key, default)):
            response = self.client.get("/api/system/st-packages/missing-moq")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {
            "rows": [
                {
                    "part_number": "PART-1",
                    "description": "Capacitor",
                    "stock_qty": 1000,
                    "package_text": "200,300,500",
                    "package_values": [200, 300, 500],
                    "package_sum": 1000,
                    "diff_qty": 0,
                    "matches_stock": True,
                    "updated_at": "2026-04-07T10:00:00",
                },
            ],
            "count": 1,
            "loaded_at": "2026-04-07T09:00:00",
            "filename": "main.xlsx",
        })

    def test_update_missing_moq_st_package_requires_edit_auth_and_saves(self):
        with patch.dict(os.environ, {"PYTEST_CURRENT_TEST": ""}, clear=False):
            blocked = self.client.put("/api/system/st-packages/PART-1", json={"package_text": "200,300,500"})
            self.assertEqual(blocked.status_code, 403)

            login = self.client.post("/api/system/edit-auth/login", json={"password": "123"})
            self.assertEqual(login.status_code, 200)

            with patch("app.routers.system.save_missing_moq_package_text", return_value={
                "part_number": "PART-1",
                "description": "Capacitor",
                "stock_qty": 1000,
                "package_text": "200,300,500",
                "package_values": [200, 300, 500],
                "package_sum": 1000,
                "diff_qty": 0,
                "matches_stock": True,
                "updated_at": "2026-04-07T10:00:00",
            }) as mock_save, \
                 patch("app.routers.system.db.log_activity") as mock_log:
                response = self.client.put("/api/system/st-packages/PART-1", json={"package_text": "200,300,500"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["row"]["package_text"], "200,300,500")
        mock_save.assert_called_once_with("PART-1", "200,300,500")
        mock_log.assert_called_once()

    def test_update_database_backup_settings_uses_request_payload(self):
        with patch(
            "app.routers.system.db_backup.update_database_backup_settings",
            return_value={"enabled": False, "hour": 4, "minute": 30, "keep_count": 7},
        ) as mock_update:
            response = self.client.put("/api/system/db-backups/settings", json={
                "enabled": False,
                "hour": 4,
                "minute": 30,
                "keep_count": 7,
            })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {
            "ok": True,
            "enabled": False,
            "hour": 4,
            "minute": 30,
            "keep_count": 7,
        })
        mock_update.assert_called_once_with(enabled=False, hour=4, minute=30, keep_count=7)

    def test_restore_database_backup_endpoint_returns_reload_flag(self):
        with patch(
            "app.routers.system.db_backup.restore_database_backup",
            return_value={
                "restored_backup": {"name": "system_backup_20260313_020000.db"},
                "safety_backup": {"name": "system_backup_20260313_113000.db"},
                "restored_at": "2026-03-13T11:30:00",
                "removed_count": 0,
            },
        ) as mock_restore:
            response = self.client.post("/api/system/db-backups/restore", json={
                "backup_name": "system_backup_20260313_020000.db",
            })

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["requires_reload"])
        mock_restore.assert_called_once_with("system_backup_20260313_020000.db")
