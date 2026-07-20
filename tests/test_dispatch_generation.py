from __future__ import annotations

import io
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl
from fastapi.testclient import TestClient

from main import app
from app.routers import bom as bom_router
from app.routers.dispatch import DispatchRequest, _build_dispatch_result_by_order, _build_order_dispatch_context, _generate_dispatch_response, _get_selected_orders, _load_committed_main_supplements
from app.services.dispatch_form_generator import generate_dispatch_form


class DispatchGenerationTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.client = TestClient(app)

    def test_selected_orders_include_dispatched_and_completed_orders(self):
        orders = {
            1: {"id": 1, "status": "merged"},
            2: {"id": 2, "status": "dispatched"},
            3: {"id": 3, "status": "completed"},
            4: {"id": 4, "status": "cancelled"},
        }

        with patch("app.routers.dispatch.db.get_order", side_effect=lambda order_id: orders.get(order_id)):
            selected = _get_selected_orders([1, 2, 3, 4])

        self.assertEqual([order["id"] for order in selected], [1, 2, 3])

    def test_dispatch_form_writes_explicit_zero_quantity(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "dispatch.xlsx"

            generate_dispatch_form([
                {
                    "batch_code": "1-1",
                    "po_number": "4500059234",
                    "model": "MODEL-A",
                    "date": "2026/03/27",
                    "items": [
                        {
                            "part": "PART-ZERO",
                            "desc": "Zero qty",
                            "qty": 0,
                            "fill_color": None,
                            "is_shortage": False,
                        },
                    ],
                },
            ], str(output_path))

            wb = openpyxl.load_workbook(output_path, data_only=False)
            ws = wb.active
            self.assertEqual(ws.cell(row=3, column=3).value, "PART-ZERO")
            self.assertEqual(ws.cell(row=3, column=5).value, 0)
            wb.close()

    def test_bom_dispatch_download_sample_draft_ignores_ec_minimum_stock(self):
        target_boms = [{"id": "bom-sample", "model": "MODEL-S", "group_model": "MODEL-S"}]
        order = {
            "id": 7,
            "status": "merged",
            "po_number": "PO-S",
            "model": "MODEL-S",
            "order_qty": 1,
        }
        components = [{
            "part_number": "EC-20080A",
            "description": "Sample cap",
            "needed_qty": 1,
            "prev_qty_cs": 0,
            "is_dash": 0,
        }]

        with patch("app.routers.bom._build_dispatch_running_stock", return_value={"EC-20080A": 99.0}), \
             patch("app.routers.bom.db.get_st_inventory_stock", return_value={}), \
             patch("app.routers.bom.db.get_bom_components", return_value=components), \
             patch("app.routers.bom.db.get_order", return_value=order), \
             patch("app.routers.bom.db.get_active_merge_draft_for_order", return_value={"is_sample": 1}):
            _, computed_supplements, _, _, _ = bom_router._build_order_based_export_values(
                target_boms,
                [7],
                {"EC-20080A": 1000},
                order_supplements={7: {"EC-20080A": 1000}},
            )

        self.assertEqual(computed_supplements["bom-sample"].get("EC-20080A", 0), 0)

    def test_dispatch_generation_sample_draft_does_not_fallback_to_ec_suggestion(self):
        order = {"id": 8, "status": "merged", "model": "MODEL-S", "code": "S-1"}
        result_by_order = {
            8: {
                "shortages": [{
                    "part_number": "EC-20080A",
                    "description": "Sample cap",
                    "suggested_qty": 1000,
                    "shortage_amount": 2,
                }],
                "customer_material_shortages": [],
            },
        }
        active_draft = {
            "order_id": 8,
            "is_sample": True,
            "decisions": {},
            "supplements": {},
            "shortages": [],
        }

        with patch("app.routers.dispatch.db.get_decisions_for_order", return_value={}):
            context = _build_order_dispatch_context(
                order,
                result_by_order,
                saved_supplements={8: {}},
                decision_overrides={},
                bom_map={"MODEL-S": [{"part_number": "EC-20080A", "description": "Sample cap"}]},
                active_draft=active_draft,
            )

        self.assertEqual(context["candidate_parts"], [])

    def test_dispatch_generate_uses_selected_orders_saved_supplements_and_sample_layout(self):
        orders = {
            1: {
                "id": 1,
                "status": "merged",
                "po_number": "4500059234",
                "model": "MODEL-A",
                "code": "1-3",
                "delivery_date": "2026-03-27",
            },
            2: {
                "id": 2,
                "status": "pending",
                "po_number": "4500059162",
                "model": "MODEL-B",
                "code": "1-4",
                "delivery_date": "2026-05-08",
            },
        }
        bom_map = {
            "MODEL-A": [
                {"part_number": "PART-SAVED", "description": "Saved desc", "needed_qty": 1, "is_dash": 0},
                {"part_number": "PART-MISSING", "description": "Missing desc", "needed_qty": 1, "is_dash": 0},
            ],
            "MODEL-B": [
                {"part_number": "PART-PENDING", "description": "Pending desc", "needed_qty": 1, "is_dash": 0},
            ],
        }
        calc_results = [
            {
                "order_id": 1,
                "shortages": [
                    {"part_number": "PART-SAVED", "description": "Saved desc", "suggested_qty": 999, "shortage_amount": 999, "purchase_needed_qty": 999},
                    {"part_number": "PART-MISSING", "description": "Missing desc", "suggested_qty": 888, "shortage_amount": 888, "purchase_needed_qty": 0},
                ],
                "customer_material_shortages": [],
            },
            {
                "order_id": 2,
                "shortages": [
                    {"part_number": "PART-PENDING", "description": "Pending desc", "suggested_qty": 2400, "shortage_amount": 2400, "purchase_needed_qty": 2400},
                ],
                "customer_material_shortages": [],
            },
        ]

        with patch("app.routers.dispatch.db.get_all_bom_components_by_model", return_value=bom_map), \
             patch("app.routers.dispatch.db.get_order", side_effect=lambda order_id: orders.get(order_id)), \
             patch("app.routers.dispatch._load_shortage_inputs", return_value=({}, {}, {})), \
             patch("app.routers.dispatch.calc_run", return_value=calc_results), \
             patch("app.routers.dispatch.db.get_order_supplements", return_value={1: {"PART-SAVED": 3000}}), \
             patch("app.routers.dispatch.db.get_decisions_for_order", side_effect=lambda order_id: {
                 1: {"PART-SAVED": "CreateRequirement", "PART-MISSING": "Shortage"},
                 2: {},
             }.get(order_id, {})), \
             patch("app.routers.dispatch.build_generated_filename", return_value="發料單_20260312_1740.xlsx"), \
             patch("app.routers.dispatch.db.log_activity"):
            response = self.client.post("/api/dispatch/generate", json={
                "order_ids": [1, 2],
                "decisions": {"PART-PENDING": "CreateRequirement"},
            })

        self.assertEqual(response.status_code, 200)
        self.assertIn("20260312_1740", response.headers["content-disposition"])

        wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=False)
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).value, "1-3")
        self.assertEqual(ws.cell(row=2, column=1).value, 4500059234)
        self.assertEqual(ws.cell(row=1, column=5).value, "日期")
        self.assertEqual(ws.cell(row=3, column=3).value, "PART-MISSING")
        self.assertEqual(ws.cell(row=3, column=5).value, "缺")
        self.assertEqual(ws.cell(row=4, column=3).value, "PART-SAVED")
        self.assertEqual(ws.cell(row=4, column=5).value, 3000)
        self.assertEqual(ws.cell(row=5, column=1).value, "1-4")
        self.assertEqual(ws.cell(row=7, column=3).value, "PART-PENDING")
        self.assertEqual(ws.cell(row=7, column=5).value, 2400)
        self.assertIn("A1:C1", {str(rng) for rng in ws.merged_cells.ranges})
        self.assertIn("D1:D2", {str(rng) for rng in ws.merged_cells.ranges})
        self.assertEqual(ws.title, "庚霖-TD2U  ")
        self.assertEqual(ws["A1"].font.name, "Calibri")
        self.assertEqual(ws["C3"].font.name, "Calibri")
        self.assertEqual(ws["D3"].font.name, "Calibri")
        self.assertEqual(ws["E3"].font.name, "Calibri")
        self.assertEqual(ws["E4"].font.name, "Calibri")
        self.assertEqual(ws["E3"].fill.fgColor.rgb, "FFFFFFFF")
        self.assertEqual(ws["E4"].fill.fgColor.rgb, "FFFFC000")
        self.assertEqual(ws.column_dimensions["D"].width, 72)
        self.assertTrue(str(ws["D1"].value).startswith("辰尚-庚霖"))
        self.assertFalse(str(ws["D1"].value).startswith(" "))
        self.assertEqual(ws["E2"].font.name, "Calibri")
        self.assertEqual(ws["E2"].font.sz, 9)
        wb.close()

    def test_dispatch_generate_recalculates_dispatched_order_with_its_own_consumption_added_back(self):
        orders = {
            1: {
                "id": 1,
                "status": "dispatched",
                "po_number": "4500059234",
                "model": "MODEL-A",
                "code": "5-1",
                "delivery_date": "2026-05-01",
            },
        }
        bom_map = {
            "MODEL-A": [
                {"part_number": "EC-80004A", "description": "Cap", "needed_qty": 1, "is_dash": 0},
                {"part_number": "EC-LATER", "description": "Later negative", "needed_qty": 1, "is_dash": 0},
                {"part_number": "EC-NEEDS", "description": "Needs supplement", "needed_qty": 1, "is_dash": 0},
            ],
        }
        dispatch_records = [
            {"part_number": "EC-80004A", "needed_qty": 320, "decision": "None"},
            {"part_number": "EC-LATER", "needed_qty": 20, "decision": "None"},
        ]
        calc_results = [
            {
                "order_id": 1,
                "shortages": [
                    {
                        "part_number": "EC-NEEDS",
                        "description": "Needs supplement",
                        "suggested_qty": 100,
                        "shortage_amount": 100,
                    },
                ],
                "customer_material_shortages": [],
            },
        ]

        with patch("app.routers.dispatch.db.get_all_bom_components_by_model", return_value=bom_map), \
             patch("app.routers.dispatch.db.get_order", side_effect=lambda order_id: orders.get(order_id)), \
             patch("app.routers.dispatch._load_shortage_inputs", return_value=({"EC-80004A": 149, "EC-LATER": -10}, {}, {})), \
             patch("app.routers.dispatch.db.get_dispatch_records", return_value=dispatch_records), \
             patch("app.routers.dispatch.calc_run", return_value=calc_results) as mock_calc, \
             patch("app.routers.dispatch.db.get_order_supplements", return_value={}), \
             patch("app.routers.dispatch.db.get_decisions_for_order", return_value={}), \
             patch("app.routers.dispatch.db.get_st_inventory_stock", return_value={}), \
             patch("app.routers.dispatch.build_generated_filename", return_value="發料單測試.xlsx"), \
             patch("app.routers.dispatch.db.log_activity"):
            response = self.client.post("/api/dispatch/generate", json={
                "order_ids": [1],
                "decisions": {},
            })

        self.assertEqual(response.status_code, 200)
        self.assertNotIn("x-dispatch-warning", {key.lower(): value for key, value in response.headers.items()})
        mock_calc.assert_called_once()
        adjusted_consumption = mock_calc.call_args.args[4]
        self.assertEqual(adjusted_consumption["EC-80004A"], -320)
        self.assertEqual(adjusted_consumption["EC-LATER"], -20)

        wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=False)
        ws = wb.active
        parts = [
            str(ws.cell(row=row_idx, column=3).value or "").strip()
            for row_idx in range(1, ws.max_row + 1)
        ]
        self.assertNotIn("EC-80004A", parts)
        self.assertNotIn("EC-LATER", parts)
        self.assertIn("EC-NEEDS", parts)
        self.assertEqual(ws.cell(row=3, column=5).value, 100)
        wb.close()

    def test_dispatch_generate_committed_order_uses_main_batch_supplements_over_stale_db(self):
        class DummyQueryParams:
            def get(self, key, default=None):
                return default

        class DummyRequest:
            query_params = DummyQueryParams()
            headers = {}

        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.cell(row=1, column=4).value = "6-2"
            ws.cell(row=2, column=1).value = "EC-80004A"
            ws.cell(row=3, column=1).value = "EC-OK"
            ws.cell(row=3, column=4).value = 100
            wb.save(main_path)
            wb.close()

            order = {
                "id": 62,
                "status": "dispatched",
                "po_number": "4500059234",
                "model": "MODEL-A",
                "code": "6-2",
                "delivery_date": "2026-05-01",
            }
            bom_map = {
                "MODEL-A": [
                    {"part_number": "EC-80004A", "description": "Stale cap", "needed_qty": 1, "is_dash": 0},
                    {"part_number": "EC-OK", "description": "Main cap", "needed_qty": 1, "is_dash": 0},
                ],
            }
            calc_results = [
                {
                    "order_id": 62,
                    "shortages": [
                        {"part_number": "EC-80004A", "description": "Stale cap", "suggested_qty": 2150, "shortage_amount": 2150},
                        {"part_number": "EC-OK", "description": "Main cap", "suggested_qty": 100, "shortage_amount": 100},
                    ],
                    "customer_material_shortages": [],
                },
            ]
            captured = {}

            def fake_generate(groups, out_path):
                captured["groups"] = groups
                Path(out_path).write_bytes(b"xlsx")
                return out_path

            with patch("app.routers.dispatch.db.get_all_bom_components_by_model", return_value=bom_map), \
                 patch("app.routers.dispatch.db.get_order", return_value=order), \
                 patch("app.routers.dispatch.db.get_setting", return_value=str(main_path)), \
                 patch("app.routers.dispatch._load_shortage_inputs", return_value=({}, {}, {})), \
                 patch("app.routers.dispatch.calc_run", return_value=calc_results), \
                 patch("app.routers.dispatch.db.get_order_supplements", return_value={62: {"EC-80004A": 2150, "EC-OK": 100}}), \
                 patch("app.routers.dispatch.db.get_decisions_for_order", return_value={
                     "EC-80004A": "CreateRequirement",
                     "EC-OK": "CreateRequirement",
                 }), \
                 patch("app.routers.dispatch.db.get_st_inventory_stock", return_value={}), \
                 patch("app.routers.dispatch.generate_dispatch_form", side_effect=fake_generate), \
                 patch("app.routers.dispatch.build_generated_filename", return_value="發料單測試.xlsx"), \
                 patch("app.routers.dispatch.db.log_activity"):
                _generate_dispatch_response(DispatchRequest(order_ids=[62], decisions={}), DummyRequest())

        items = captured["groups"][0]["items"]
        self.assertEqual([item["part"] for item in items], ["EC-OK"])
        self.assertEqual(items[0]["qty"], 100)

    def test_committed_main_supplements_scan_read_only_sheet_once_without_random_cell_reads(self):
        class FakeWorksheet:
            def __init__(self):
                self.data_scan_count = 0

            def iter_rows(self, min_row=None, max_row=None, max_col=None, values_only=False):
                if min_row == 1 and max_row == 1:
                    return iter([("料號", None, None, "6-2", "6-3", "6-2")])
                self.data_scan_count += 1
                self.assert_scan_args = (min_row, max_col, values_only)
                return iter([
                    ("PART-A", None, None, 30, 50, 12),
                    ("PART-B", None, None, 0, 70, 0),
                    # 主檔若有重複料號，維持舊流程只採第一列的行為。
                    ("PART-A", None, None, 999, 999, 999),
                ])

            def cell(self, *args, **kwargs):
                raise AssertionError("read-only 主檔不可逐格隨機讀取")

        class FakeWorkbook:
            def __init__(self, worksheet):
                self.worksheets = [worksheet]
                self.closed = False

            def close(self):
                self.closed = True

        worksheet = FakeWorksheet()
        workbook = FakeWorkbook(worksheet)
        orders = [
            {"id": 62, "status": "dispatched", "model": "MODEL-A", "code": "6-2"},
            {"id": 63, "status": "completed", "model": "MODEL-B", "code": "6-3"},
        ]
        bom_map = {
            "MODEL-A": [{"part_number": "PART-A"}, {"part_number": "PART-MISSING"}],
            "MODEL-B": [{"part_number": "PART-A"}, {"part_number": "PART-B"}],
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.touch()
            with patch("app.routers.dispatch.db.get_setting", return_value=str(main_path)), \
                 patch("app.routers.dispatch.openpyxl.load_workbook", return_value=workbook):
                result = _load_committed_main_supplements(orders, bom_map)

        self.assertEqual(result, {
            62: {"PART-A": 42},
            63: {"PART-A": 50, "PART-B": 70},
        })
        self.assertEqual(worksheet.data_scan_count, 1)
        self.assertEqual(worksheet.assert_scan_args, (2, 6, True))
        self.assertTrue(workbook.closed)

    def test_dispatch_generate_5_1_ec_80004a_current_main_stock_plus_own_dispatch_record_is_enough(self):
        orders = [
            {
                "id": 51,
                "status": "dispatched",
                "po_number": "4500059234",
                "model": "MODEL-A",
                "code": "5-1",
                "delivery_date": "2026-05-01",
                "order_qty": 320,
            },
        ]
        bom_map = {
            "MODEL-A": [
                {"part_number": "EC-80004A", "description": "Cap", "needed_qty": 1, "is_dash": 0},
            ],
        }

        with patch("app.routers.dispatch._load_shortage_inputs", return_value=({"EC-80004A": 149}, {"EC-80004A": 1}, {})), \
             patch("app.routers.dispatch.db.get_dispatch_records", return_value=[
                 {"part_number": "EC-80004A", "needed_qty": 320, "decision": "None"},
             ]), \
             patch("app.routers.dispatch.db.get_st_inventory_stock", return_value={}):
            result_by_order = _build_dispatch_result_by_order(orders, bom_map)

        self.assertEqual(result_by_order[51]["shortages"], [])
        self.assertEqual(result_by_order[51]["status"], "ok")

    def test_dispatch_generate_keeps_st_covered_rows_white_when_no_purchase_is_needed(self):
        orders = {
            1: {
                "id": 1,
                "status": "merged",
                "po_number": "4500059234",
                "model": "MODEL-A",
                "code": "1-3",
                "delivery_date": "2026-03-27",
            },
        }
        bom_map = {
            "MODEL-A": [
                {"part_number": "PART-ST", "description": "ST desc", "needed_qty": 1, "is_dash": 0},
            ],
        }
        calc_results = [
            {
                "order_id": 1,
                "shortages": [
                    {
                        "part_number": "PART-ST",
                        "description": "ST desc",
                        "suggested_qty": 8,
                        "shortage_amount": 8,
                        "st_available_qty": 8,
                        "purchase_needed_qty": 0,
                    },
                ],
                "customer_material_shortages": [],
            },
        ]

        with patch("app.routers.dispatch.db.get_all_bom_components_by_model", return_value=bom_map), \
             patch("app.routers.dispatch.db.get_order", side_effect=lambda order_id: orders.get(order_id)), \
             patch("app.routers.dispatch._load_shortage_inputs", return_value=({}, {}, {})), \
             patch("app.routers.dispatch.calc_run", return_value=calc_results), \
             patch("app.routers.dispatch.db.get_order_supplements", return_value={}), \
             patch("app.routers.dispatch.db.get_decisions_for_order", return_value={}), \
             patch("app.routers.dispatch.build_generated_filename", return_value="發料單測試.xlsx"), \
             patch("app.routers.dispatch.db.log_activity"):
            response = self.client.post("/api/dispatch/generate", json={
                "order_ids": [1],
                "decisions": {"PART-ST": "CreateRequirement"},
            })

        self.assertEqual(response.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=False)
        ws = wb.active
        self.assertEqual(ws.cell(row=3, column=3).value, "PART-ST")
        self.assertEqual(ws.cell(row=3, column=5).value, 8)
        self.assertEqual(ws["E3"].fill.fgColor.rgb, "FFFFFFFF")
        wb.close()

    def test_dispatch_generate_marks_manual_supplement_orange_when_qty_exceeds_st_stock(self):
        orders = {
            1: {
                "id": 1,
                "status": "merged",
                "po_number": "4500059234",
                "model": "MODEL-A",
                "code": "1-3",
                "delivery_date": "2026-03-27",
            },
        }
        bom_map = {
            "MODEL-A": [
                {"part_number": "PART-MANUAL", "description": "Manual desc", "needed_qty": 1, "is_dash": 0},
            ],
        }
        calc_results = [
            {
                "order_id": 1,
                "shortages": [],
                "customer_material_shortages": [],
            },
        ]

        with patch("app.routers.dispatch.db.get_all_bom_components_by_model", return_value=bom_map), \
             patch("app.routers.dispatch.db.get_order", side_effect=lambda order_id: orders.get(order_id)), \
             patch("app.routers.dispatch._load_shortage_inputs", return_value=({}, {}, {})), \
             patch("app.routers.dispatch.calc_run", return_value=calc_results), \
             patch("app.routers.dispatch.db.get_order_supplements", return_value={1: {"PART-MANUAL": 20000}}), \
             patch("app.routers.dispatch.db.get_decisions_for_order", return_value={"PART-MANUAL": "CreateRequirement"}), \
             patch("app.routers.dispatch.db.get_st_inventory_stock", return_value={"PART-MANUAL": 10000}), \
             patch("app.routers.dispatch.build_generated_filename", return_value="發料單測試.xlsx"), \
             patch("app.routers.dispatch.db.log_activity"):
            response = self.client.post("/api/dispatch/generate", json={
                "order_ids": [1],
                "decisions": {},
            })

        self.assertEqual(response.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=False)
        ws = wb.active
        self.assertEqual(ws.cell(row=3, column=3).value, "PART-MANUAL")
        self.assertEqual(ws.cell(row=3, column=5).value, 20000)
        self.assertEqual(ws["E3"].fill.fgColor.rgb, "FFFFC000")
        wb.close()

    def test_dispatch_generate_writes_reviewed_draft_explicit_zero_quantity(self):
        orders = {
            1: {
                "id": 1,
                "status": "merged",
                "po_number": "4500059234",
                "model": "MODEL-A",
                "code": "1-3",
                "delivery_date": "2026-03-27",
            },
        }
        bom_map = {
            "MODEL-A": [
                {"part_number": "PART-ZERO", "description": "Zero desc", "needed_qty": 1, "is_dash": 0},
            ],
        }
        reviewed_draft = {
            "order_id": 1,
            "decisions": {"PART-ZERO": "CreateRequirement"},
            "supplements": {"PART-ZERO": 0},
            "shortages": [
                {"part_number": "PART-ZERO", "description": "Zero desc", "suggested_qty": 1000, "shortage_amount": 1000},
            ],
        }

        with patch("app.routers.dispatch.db.get_all_bom_components_by_model", return_value=bom_map), \
             patch("app.routers.dispatch.db.get_order", side_effect=lambda order_id: orders.get(order_id)), \
             patch("app.routers.dispatch._load_shortage_inputs", return_value=({}, {}, {})), \
             patch("app.routers.dispatch.calc_run", return_value=[]), \
             patch("app.routers.dispatch.db.get_order_supplements", return_value={1: {}}), \
             patch("app.routers.dispatch.db.get_decisions_for_order", return_value=reviewed_draft["decisions"]), \
             patch("app.routers.dispatch.db.get_active_merge_drafts", return_value=[reviewed_draft]), \
             patch("app.routers.dispatch.db.get_st_inventory_stock", return_value={}), \
             patch("app.routers.dispatch.build_generated_filename", return_value="發料單測試.xlsx"), \
             patch("app.routers.dispatch.db.log_activity"):
            response = self.client.post("/api/dispatch/generate", json={
                "order_ids": [1],
                "decisions": {},
            })

        self.assertEqual(response.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=False)
        ws = wb.active
        self.assertEqual(ws.cell(row=3, column=3).value, "PART-ZERO")
        self.assertEqual(ws.cell(row=3, column=5).value, 0)
        wb.close()

    def test_dispatch_generate_addbacks_dispatched_order_st_consumption_before_highlight(self):
        orders = {
            1: {
                "id": 1,
                "status": "dispatched",
                "po_number": "4500059234",
                "model": "MODEL-A",
                "code": "6-4",
                "delivery_date": "2026-05-01",
            },
        }
        bom_map = {
            "MODEL-A": [
                {"part_number": "IC-M24C02-WMN6TP-TAB", "description": "EEPROM", "needed_qty": 1, "is_dash": 0},
                {"part_number": "PART-NEG", "description": "Negative ST desc", "needed_qty": 1, "is_dash": 0},
            ],
        }
        calc_results = [
            {
                "order_id": 1,
                "shortages": [],
                "customer_material_shortages": [],
            },
        ]

        with patch("app.routers.dispatch.db.get_all_bom_components_by_model", return_value=bom_map), \
             patch("app.routers.dispatch.db.get_order", side_effect=lambda order_id: orders.get(order_id)), \
             patch("app.routers.dispatch._load_shortage_inputs", return_value=({}, {}, {})), \
             patch("app.routers.dispatch.calc_run", return_value=calc_results), \
             patch("app.routers.dispatch.db.get_order_supplements", return_value={
                 1: {"IC-M24C02-WMN6TP-TAB": 400, "PART-NEG": 5000},
             }), \
             patch("app.routers.dispatch.db.get_decisions_for_order", return_value={
                 "IC-M24C02-WMN6TP-TAB": "CreateRequirement",
                 "PART-NEG": "CreateRequirement",
             }), \
             patch("app.routers.dispatch.db.get_st_inventory_stock", return_value={
                 "IC-M24C02-WMN6TP-TAB": 0,
                 "PART-NEG": -15000,
             }), \
             patch("app.routers.dispatch.build_generated_filename", return_value="發料單測試.xlsx"), \
             patch("app.routers.dispatch.db.log_activity"):
            response = self.client.post("/api/dispatch/generate", json={
                "order_ids": [1],
                "decisions": {},
            })

        self.assertEqual(response.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=False)
        ws = wb.active
        self.assertEqual(ws.cell(row=3, column=3).value, "IC-M24C02-WMN6TP-TAB")
        self.assertEqual(ws.cell(row=3, column=5).value, 400)
        self.assertNotEqual(ws["E3"].fill.fgColor.rgb, "FFFFC000")
        self.assertEqual(ws.cell(row=4, column=3).value, "PART-NEG")
        self.assertEqual(ws.cell(row=4, column=5).value, 5000)
        self.assertEqual(ws["E4"].fill.fgColor.rgb, "FFFFC000")
        wb.close()

    def test_dispatch_generate_ignores_non_bom_decision_parts_for_selected_order(self):
        orders = {
            1: {
                "id": 1,
                "status": "merged",
                "po_number": "4500059234",
                "model": "TX2",
                "code": "2-1",
                "delivery_date": "2026-03-31",
            },
        }
        bom_map = {
            "TX2": [
                {"part_number": "IC-CSD18531Q5AT-TAB", "description": "Valid TX2 part", "needed_qty": 1, "is_dash": 0},
            ],
        }
        calc_results = [
            {
                "order_id": 1,
                "shortages": [
                    {
                        "part_number": "IC-CSD18531Q5AT-TAB",
                        "description": "Valid TX2 part",
                        "suggested_qty": 1500,
                        "shortage_amount": 1500,
                    },
                ],
                "customer_material_shortages": [],
            },
        ]

        with patch("app.routers.dispatch.db.get_all_bom_components_by_model", return_value=bom_map), \
             patch("app.routers.dispatch.db.get_order", side_effect=lambda order_id: orders.get(order_id)), \
             patch("app.routers.dispatch._load_shortage_inputs", return_value=({}, {}, {})), \
             patch("app.routers.dispatch.calc_run", return_value=calc_results), \
             patch("app.routers.dispatch.db.get_order_supplements", return_value={}), \
             patch("app.routers.dispatch.db.get_decisions_for_order", return_value={
                 "IC-CSD18531Q5AT-TAB": "CreateRequirement",
                 "IC-NB675-TAB": "CreateRequirement",
             }), \
             patch("app.routers.dispatch.build_generated_filename", return_value="發料單測試.xlsx"), \
             patch("app.routers.dispatch.db.log_activity"):
            response = self.client.post("/api/dispatch/generate", json={
                "order_ids": [1],
                "decisions": {},
            })

        self.assertEqual(response.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=False)
        ws = wb.active
        parts = [
            str(ws.cell(row=row_idx, column=3).value or "").strip()
            for row_idx in range(1, ws.max_row + 1)
        ]
        self.assertIn("IC-CSD18531Q5AT-TAB", parts)
        self.assertNotIn("IC-NB675-TAB", parts)
        wb.close()

    def test_dispatch_generate_keeps_order_scoped_ic_parts_separate_per_model(self):
        orders = {
            1: {
                "id": 1,
                "status": "merged",
                "po_number": "4500059234",
                "model": "MODEL-A",
                "code": "1-3",
                "delivery_date": "2026-03-27",
            },
            2: {
                "id": 2,
                "status": "merged",
                "po_number": "4500059235",
                "model": "MODEL-B",
                "code": "1-4",
                "delivery_date": "2026-03-28",
            },
        }
        bom_map = {
            "MODEL-A": [
                {"part_number": "IC-STM32F", "description": "STM part", "needed_qty": 1, "is_dash": 0},
            ],
            "MODEL-B": [
                {"part_number": "IC-STM32F", "description": "STM part", "needed_qty": 1, "is_dash": 0},
            ],
        }
        calc_results = [
            {
                "order_id": 1,
                "shortages": [
                    {"part_number": "IC-STM32F", "description": "STM part", "suggested_qty": 100, "shortage_amount": 100},
                ],
                "customer_material_shortages": [],
            },
            {
                "order_id": 2,
                "shortages": [
                    {"part_number": "IC-STM32F", "description": "STM part", "suggested_qty": 50, "shortage_amount": 50},
                ],
                "customer_material_shortages": [],
            },
        ]

        with patch("app.routers.dispatch.db.get_all_bom_components_by_model", return_value=bom_map), \
             patch("app.routers.dispatch.db.get_order", side_effect=lambda order_id: orders.get(order_id)), \
             patch("app.routers.dispatch._load_shortage_inputs", return_value=({}, {}, {})), \
             patch("app.routers.dispatch.calc_run", return_value=calc_results), \
             patch("app.routers.dispatch.db.get_order_supplements", return_value={}), \
             patch("app.routers.dispatch.db.get_decisions_for_order", return_value={}), \
             patch("app.routers.dispatch.db.get_st_inventory_stock", return_value={}), \
             patch("app.routers.dispatch.build_generated_filename", return_value="發料單測試.xlsx"), \
             patch("app.routers.dispatch.db.log_activity"):
            response = self.client.post("/api/dispatch/generate", json={
                "order_ids": [1, 2],
                "decisions": {"IC-STM32F": "CreateRequirement"},
            })

        self.assertEqual(response.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=False)
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).value, "1-3")
        self.assertEqual(ws.cell(row=3, column=3).value, "IC-STM32F")
        self.assertEqual(ws.cell(row=3, column=5).value, 100)
        self.assertEqual(ws.cell(row=4, column=1).value, "1-4")
        self.assertEqual(ws.cell(row=6, column=3).value, "IC-STM32F")
        self.assertEqual(ws.cell(row=6, column=5).value, 50)
        wb.close()

    def test_dispatch_generate_keeps_same_normal_part_separate_per_order(self):
        orders = {
            1: {
                "id": 1,
                "status": "merged",
                "po_number": "4500059234",
                "model": "MODEL-A",
                "code": "2-2",
                "delivery_date": "2026-03-27",
            },
            2: {
                "id": 2,
                "status": "merged",
                "po_number": "4500059235",
                "model": "MODEL-B",
                "code": "2-3",
                "delivery_date": "2026-03-28",
            },
        }
        bom_map = {
            "MODEL-A": [
                {"part_number": "EC-20080A", "description": "Cap desc", "needed_qty": 1, "is_dash": 0},
            ],
            "MODEL-B": [
                {"part_number": "EC-20080A", "description": "Cap desc", "needed_qty": 1, "is_dash": 0},
            ],
        }
        calc_results = [
            {
                "order_id": 1,
                "shortages": [
                    {"part_number": "EC-20080A", "description": "Cap desc", "suggested_qty": 1000, "shortage_amount": 1000},
                ],
                "customer_material_shortages": [],
            },
            {
                "order_id": 2,
                "shortages": [
                    {"part_number": "EC-20080A", "description": "Cap desc", "suggested_qty": 2000, "shortage_amount": 2000},
                ],
                "customer_material_shortages": [],
            },
        ]

        with patch("app.routers.dispatch.db.get_all_bom_components_by_model", return_value=bom_map), \
             patch("app.routers.dispatch.db.get_order", side_effect=lambda order_id: orders.get(order_id)), \
             patch("app.routers.dispatch._load_shortage_inputs", return_value=({}, {}, {})), \
             patch("app.routers.dispatch.calc_run", return_value=calc_results), \
             patch("app.routers.dispatch.db.get_order_supplements", return_value={1: {}, 2: {"EC-20080A": 3000}}), \
             patch("app.routers.dispatch.db.get_decisions_for_order", return_value={"EC-20080A": "CreateRequirement"}), \
             patch("app.routers.dispatch.db.get_st_inventory_stock", return_value={}), \
             patch("app.routers.dispatch.build_generated_filename", return_value="發料單測試.xlsx"), \
             patch("app.routers.dispatch.db.log_activity"):
            response = self.client.post("/api/dispatch/generate", json={
                "order_ids": [1, 2],
                "decisions": {},
            })

        self.assertEqual(response.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=False)
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).value, "2-2")
        self.assertEqual(ws.cell(row=3, column=3).value, "EC-20080A")
        self.assertEqual(ws.cell(row=3, column=5).value, 1000)
        self.assertEqual(ws.cell(row=4, column=1).value, "2-3")
        self.assertEqual(ws.cell(row=6, column=3).value, "EC-20080A")
        self.assertEqual(ws.cell(row=6, column=5).value, 3000)
        wb.close()

    def test_dispatch_generate_respects_reviewed_draft_parts_cleared_back_to_none(self):
        orders = {
            1: {
                "id": 1,
                "status": "merged",
                "po_number": "4500059319",
                "model": "OT35C",
                "code": "2-8",
                "delivery_date": "2026-05-08",
            },
        }
        bom_map = {
            "OT35C": [
                {"part_number": "EC-50004A", "description": "Diode", "needed_qty": 1, "is_dash": 0},
                {"part_number": "EC-50005A", "description": "Cap", "needed_qty": 1, "is_dash": 0},
            ],
        }
        calc_results = [
            {
                "order_id": 1,
                "shortages": [
                    {"part_number": "EC-50004A", "description": "Diode", "suggested_qty": 1800, "shortage_amount": 6},
                    {"part_number": "EC-50005A", "description": "Cap", "suggested_qty": 2000, "shortage_amount": 2000},
                ],
                "customer_material_shortages": [],
            },
        ]
        reviewed_draft = {
            "order_id": 1,
            "decisions": {"EC-50005A": "CreateRequirement"},
            "supplements": {},
            "shortages": [
                {"part_number": "EC-50004A", "description": "Diode", "suggested_qty": 1800, "shortage_amount": 6, "decision": "None"},
                {"part_number": "EC-50005A", "description": "Cap", "suggested_qty": 2000, "shortage_amount": 2000, "decision": "CreateRequirement"},
            ],
        }

        with patch("app.routers.dispatch.db.get_all_bom_components_by_model", return_value=bom_map), \
             patch("app.routers.dispatch.db.get_order", side_effect=lambda order_id: orders.get(order_id)), \
             patch("app.routers.dispatch._load_shortage_inputs", return_value=({}, {}, {})), \
             patch("app.routers.dispatch.calc_run", return_value=calc_results), \
             patch("app.routers.dispatch.db.get_order_supplements", return_value={1: {}}), \
             patch("app.routers.dispatch.db.get_decisions_for_order", return_value={"EC-50005A": "CreateRequirement"}), \
             patch("app.routers.dispatch.db.get_active_merge_drafts", return_value=[reviewed_draft]), \
             patch("app.routers.dispatch.build_generated_filename", return_value="發料單測試.xlsx"), \
             patch("app.routers.dispatch.db.log_activity"):
            response = self.client.post("/api/dispatch/generate", json={
                "order_ids": [1],
                "decisions": {},
            })

        self.assertEqual(response.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=False)
        ws = wb.active
        parts = [
            str(ws.cell(row=row_idx, column=3).value or "").strip()
            for row_idx in range(1, ws.max_row + 1)
        ]
        self.assertNotIn("EC-50004A", parts)
        self.assertIn("EC-50005A", parts)
        wb.close()

    def test_dispatch_generate_skips_reviewed_draft_create_requirement_without_qty_source(self):
        orders = {
            1: {
                "id": 1,
                "status": "merged",
                "po_number": "4500059291",
                "model": "TA7-2",
                "code": "2-5",
                "delivery_date": "2026-06-05",
            },
        }
        bom_map = {
            "TA7-2": [
                {"part_number": "EC-10032A", "description": "Cap A", "needed_qty": 1, "is_dash": 0},
                {"part_number": "EC-20117A", "description": "Cap B", "needed_qty": 1, "is_dash": 0},
            ],
        }
        calc_results = [
            {
                "order_id": 1,
                "shortages": [
                    {"part_number": "EC-20117A", "description": "Cap B", "suggested_qty": 40000, "shortage_amount": 38829},
                ],
                "customer_material_shortages": [],
            },
        ]
        reviewed_draft = {
            "order_id": 1,
            "decisions": {
                "EC-10032A": "CreateRequirement",
                "EC-20117A": "CreateRequirement",
            },
            "supplements": {},
            "shortages": [
                {"part_number": "EC-20117A", "description": "Cap B", "suggested_qty": 40000, "shortage_amount": 38829},
            ],
        }

        with patch("app.routers.dispatch.db.get_all_bom_components_by_model", return_value=bom_map), \
             patch("app.routers.dispatch.db.get_order", side_effect=lambda order_id: orders.get(order_id)), \
             patch("app.routers.dispatch._load_shortage_inputs", return_value=({}, {}, {})), \
             patch("app.routers.dispatch.calc_run", return_value=calc_results), \
             patch("app.routers.dispatch.db.get_order_supplements", return_value={1: {}}), \
             patch("app.routers.dispatch.db.get_decisions_for_order", return_value=reviewed_draft["decisions"]), \
             patch("app.routers.dispatch.db.get_active_merge_drafts", return_value=[reviewed_draft]), \
             patch("app.routers.dispatch.build_generated_filename", return_value="發料單測試.xlsx"), \
             patch("app.routers.dispatch.db.log_activity"):
            response = self.client.post("/api/dispatch/generate", json={
                "order_ids": [1],
                "decisions": {},
            })

        self.assertEqual(response.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=False)
        ws = wb.active
        parts = [
            str(ws.cell(row=row_idx, column=3).value or "").strip()
            for row_idx in range(1, ws.max_row + 1)
        ]
        self.assertNotIn("EC-10032A", parts)
        self.assertIn("EC-20117A", parts)
        wb.close()
