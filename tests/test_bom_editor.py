from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from app.models import BomEditorComponentUpdate, BomEditorSaveRequest
from app.services.bom_editor import (
    apply_bom_editor_changes,
    build_editable_filename,
    convert_xls_to_xlsx,
    prepare_uploaded_bom_file,
)


class BomEditorTests(unittest.TestCase):
    def _build_bom_workbook(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "BOM"
        ws["H1"] = 1001
        ws["K1"] = 20
        ws["C2"] = "MODEL-A"
        ws["D2"] = "PCB-A"
        ws.append([])
        ws.append([])
        ws["B5"] = 1
        ws["C5"] = "PART-A"
        ws["D5"] = "OLD DESC"
        ws["F5"] = 5
        ws["G5"] = "X"
        ws["H5"] = "-"
        ws["B6"] = 2
        ws["C6"] = "PART-B"
        ws["D6"] = "DESC-B"
        ws["F6"] = 8
        ws["G6"] = None
        ws["H6"] = 1
        wb.save(path)
        wb.close()

    def _build_merged_header_bom_workbook(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "BOM"
        ws.merge_cells("G1:H1")
        ws["G1"] = 1001
        ws.merge_cells("J1:K1")
        ws["J1"] = 20
        ws.merge_cells("B2:C2")
        ws["B2"] = "MODEL-A"
        ws["D2"] = "PCB-A"
        ws["B5"] = 1
        ws["C5"] = "PART-A"
        ws["D5"] = "OLD DESC"
        ws["F5"] = 5
        ws["G5"] = "X"
        ws["H5"] = "-"
        wb.save(path)
        wb.close()

    def test_prepare_uploaded_bom_file_converts_xls_name_to_xlsx(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)

            def fake_convert(src_path: str, dest_path: str):
                wb = Workbook()
                wb.save(dest_path)
                wb.close()

            with patch("app.services.bom_editor.BOM_DIR", temp_path), \
                 patch("app.services.bom_editor.convert_xls_to_xlsx", side_effect=fake_convert):
                result = prepare_uploaded_bom_file("bom-1", "legacy.xls", b"dummy")

        self.assertEqual(Path(result["filepath"]).suffix.lower(), ".xlsx")
        self.assertEqual(result["filename"], "legacy.xlsx")
        self.assertEqual(result["source_filename"], "legacy.xls")
        self.assertEqual(result["source_format"], ".xls")
        self.assertTrue(result["is_converted"])

    def test_build_editable_filename_changes_only_xls(self):
        self.assertEqual(build_editable_filename("legacy.xls"), "legacy.xlsx")
        self.assertEqual(build_editable_filename("modern.xlsx"), "modern.xlsx")
        self.assertEqual(build_editable_filename("macro.xlsm"), "macro.xlsm")

    def test_convert_xls_to_xlsx_uses_libreoffice_when_powershell_unavailable(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            src = temp_path / "legacy.xls"
            dest = temp_path / "legacy.xlsx"
            src.write_bytes(b"legacy")

            def fake_run(cmd, check, capture_output, text):
                exe = Path(str(cmd[0])).name.lower()
                if exe == "powershell":
                    raise RuntimeError("excel com unavailable")
                if exe in {"soffice", "libreoffice"}:
                    wb = Workbook()
                    wb.save(dest)
                    wb.close()
                    return None
                raise AssertionError(f"Unexpected command: {cmd}")

            with patch("app.services.bom_editor.shutil.which", side_effect=lambda name: "soffice" if name == "soffice" else None), \
                 patch("app.services.bom_editor.subprocess.run", side_effect=fake_run), \
                 patch("app.services.bom_editor._copy_xls_to_xlsx_values") as copy_fallback:
                convert_xls_to_xlsx(str(src), str(dest))
                self.assertTrue(dest.exists())
                copy_fallback.assert_not_called()

    def test_convert_xls_to_xlsx_falls_back_to_value_copy_when_all_converters_fail(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            src = temp_path / "legacy.xls"
            dest = temp_path / "legacy.xlsx"
            src.write_bytes(b"legacy")

            with patch("app.services.bom_editor.shutil.which", return_value="soffice"), \
                 patch("app.services.bom_editor.subprocess.run", side_effect=RuntimeError("all converters failed")), \
                 patch("app.services.bom_editor._copy_xls_to_xlsx_values") as copy_fallback:
                convert_xls_to_xlsx(str(src), str(dest))

        copy_fallback.assert_called_once_with(src, dest)

    def test_apply_bom_editor_changes_updates_workbook_cells(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "bom.xlsx"
            self._build_bom_workbook(path)

            req = BomEditorSaveRequest(
                po_number=2002,
                order_qty=88,
                model="MODEL-NEW",
                pcb="PCB-NEW",
                group_model="MODEL-NEW",
                components=[
                    BomEditorComponentUpdate(
                        source_row=5,
                        part_number="PART-A-NEW",
                        description="NEW DESC",
                        qty_per_board=1.5,
                        needed_qty=11,
                        prev_qty_cs=3,
                        is_dash=False,
                    ),
                    BomEditorComponentUpdate(
                        source_row=6,
                        part_number="PART-B",
                        description="DESC-B",
                        qty_per_board=2,
                        needed_qty=8,
                        prev_qty_cs=1,
                        is_dash=True,
                    ),
                ],
            )

            apply_bom_editor_changes(str(path), req)

            wb = load_workbook(path, data_only=False)
            ws = wb.active
            self.assertEqual(ws["H1"].value, 2002)
            self.assertEqual(ws["K1"].value, 88)
            self.assertEqual(ws["C2"].value, "MODEL-NEW")
            self.assertEqual(ws["D2"].value, "PCB-NEW")
            self.assertEqual(ws["C5"].value, "PART-A-NEW")
            self.assertEqual(ws["D5"].value, "NEW DESC")
            self.assertEqual(ws["B5"].value, 1.5)
            self.assertEqual(ws["F5"].value, 11)
            self.assertIsNone(ws["G5"].value)
            self.assertEqual(ws["H5"].value, 3)
            self.assertEqual(ws["G6"].value, "-")
            self.assertEqual(ws["H6"].value, "-")
            wb.close()

    def test_apply_bom_editor_changes_writes_to_merged_header_anchor_cells(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "bom.xlsx"
            self._build_merged_header_bom_workbook(path)

            req = BomEditorSaveRequest(
                po_number=2002,
                order_qty=88,
                model="MODEL-NEW",
                pcb="PCB-NEW",
                group_model="MODEL-NEW",
                components=[
                    BomEditorComponentUpdate(
                        source_row=5,
                        part_number="PART-A-NEW",
                        description="NEW DESC",
                        qty_per_board=1.5,
                        needed_qty=11,
                        prev_qty_cs=3,
                        is_dash=False,
                    ),
                ],
            )

            apply_bom_editor_changes(str(path), req)

            wb = load_workbook(path, data_only=False)
            ws = wb.active
            self.assertEqual(ws["G1"].value, 2002)
            self.assertEqual(ws["J1"].value, 88)
            self.assertEqual(ws["B2"].value, "MODEL-NEW")
            self.assertEqual(ws["D2"].value, "PCB-NEW")
            self.assertEqual(ws["C5"].value, "PART-A-NEW")
            self.assertEqual(ws["D5"].value, "NEW DESC")
            wb.close()
