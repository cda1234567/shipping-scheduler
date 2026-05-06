from __future__ import annotations

import json
from datetime import datetime
import tempfile
import unittest
import zipfile
from pathlib import Path
from types import SimpleNamespace
from openpyxl import Workbook, load_workbook
from unittest.mock import patch

from app.services import merge_drafts


class MergeDraftDetailTests(unittest.TestCase):
    def test_build_draft_display_filename_uses_order_po_and_sanitized_model(self):
        filename = merge_drafts._build_draft_display_filename(
            {
                "po_number": "4500059234",
                "model": "T356789IU-U/A,T356789IU-ALT",
            },
            Path("source.xlsm"),
            now=datetime(2026, 4, 22, 10, 30),
        )

        self.assertEqual(filename, "4500059234_T356789IU-U-A_20260422.xlsm")

    def test_replace_po_in_legacy_filename_preserves_suffix_and_timestamp(self):
        filename = merge_drafts._replace_po_in_filename(
            "BOM_PO#OLD_20260422_1030.xlsx",
            "4500059234",
        )

        self.assertEqual(filename, "BOM_4500059234_20260422_1030.xlsx")

    def test_download_selected_committed_merge_drafts_uses_latest_committed_files(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            draft_file = Path(temp_dir) / "draft.xlsx"
            draft_file.write_bytes(b"xlsx")

            with patch("app.services.merge_drafts.db.get_latest_committed_merge_draft_for_order", return_value={"id": 7}), \
                 patch("app.services.merge_drafts.db.get_order", return_value={"po_number": "4500059999"}), \
                 patch("app.services.merge_drafts.db.get_merge_draft_files", return_value=[{
                     "filename": "01_PO#4500050000 MODEL-A.xlsx",
                     "filepath": str(draft_file),
                 }]), \
                 patch("app.services.merge_drafts._build_download_response", return_value="ok") as mock_response:
                result = merge_drafts.download_selected_committed_merge_drafts([42])

        self.assertEqual(result, "ok")
        entries = mock_response.call_args.args[0]
        self.assertEqual(entries[0]["path"], draft_file)
        self.assertEqual(entries[0]["download_name"], "01_4500059999 MODEL-A.xlsx")
        self.assertEqual(mock_response.call_args.kwargs["archive_label"], "已發料副檔")

    def test_build_download_response_saves_zip_to_server_download_dir_when_requested(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            download_dir = root / "downloads"
            download_dir.mkdir()
            first = root / "first.xlsx"
            second = root / "second.xlsx"
            first.write_bytes(b"first")
            second.write_bytes(b"second")
            request = SimpleNamespace(query_params={"server_save": "1"})

            with patch("app.services.server_downloads.SERVER_DOWNLOAD_DIR", download_dir), \
                 patch("app.services.server_downloads.db.get_setting", side_effect=lambda key, default="": {
                     "server_download_enabled": "1",
                     "server_download_display_path": "D:\\Download\\excel",
                 }.get(key, default)):
                response = merge_drafts._build_download_response(
                    [
                        {"path": first, "download_name": "first.xlsx"},
                        {"path": second, "download_name": "second.xlsx"},
                    ],
                    request=request,
                )

            payload = json.loads(response.body.decode("utf-8"))
            saved_path = download_dir / payload["filename"]

            self.assertTrue(payload["ok"])
            self.assertEqual(payload["directory"], "D:\\Download\\excel")
            self.assertTrue(saved_path.exists())
            with zipfile.ZipFile(saved_path) as zf:
                self.assertEqual(set(zf.namelist()), {"first.xlsx", "second.xlsx"})

    def test_download_selected_merge_drafts_uses_server_download_dir_when_requested(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            download_dir = root / "downloads"
            download_dir.mkdir()
            first = root / "draft-a.xlsx"
            second = root / "draft-b.xlsx"
            first.write_bytes(b"first")
            second.write_bytes(b"second")
            request = SimpleNamespace(query_params={"server_save": "1"})

            with patch("app.services.merge_drafts.db.get_active_merge_draft_ids_by_order_ids", return_value={1: 11, 2: 12}), \
                 patch("app.services.merge_drafts.db.get_order", side_effect=lambda order_id: {"po_number": f"PO{order_id}"}), \
                 patch("app.services.merge_drafts.db.get_merge_draft_files", side_effect=[
                     [{"filename": "draft-a.xlsx", "filepath": str(first)}],
                     [{"filename": "draft-b.xlsx", "filepath": str(second)}],
                 ]), \
                 patch("app.services.server_downloads.SERVER_DOWNLOAD_DIR", download_dir), \
                 patch("app.services.server_downloads.db.get_setting", side_effect=lambda key, default="": {
                     "server_download_enabled": "1",
                     "server_download_display_path": "D:\\Download\\excel",
                 }.get(key, default)):
                response = merge_drafts.download_selected_merge_drafts([1, 2], request=request)

            payload = json.loads(response.body.decode("utf-8"))
            saved_path = download_dir / payload["filename"]

            self.assertTrue(payload["ok"])
            self.assertEqual(payload["directory"], "D:\\Download\\excel")
            self.assertTrue(saved_path.exists())
            with zipfile.ZipFile(saved_path) as zf:
                self.assertEqual(set(zf.namelist()), {"draft-a.xlsx", "draft-b.xlsx"})

    def test_plan_order_draft_keeps_manual_supplement_even_without_shortage(self):
        order = {"id": 12, "code": "1-1", "model": "MODEL-A"}
        draft = {"decisions": {}, "supplements": {"PART-1": 25}}
        bom_files = [{"id": "bom-1", "model": "MODEL-A", "group_model": "MODEL-A"}]
        running_stock = {"PART-1": 100.0}
        moq_map = {"PART-1": 0.0}
        components = [{
            "part_number": "PART-1",
            "description": "Resistor",
            "needed_qty": 10,
            "prev_qty_cs": 0,
            "is_dash": 0,
        }]

        with patch("app.services.merge_drafts.db.get_bom_components", return_value=components):
            plan = merge_drafts._plan_order_draft(order, draft, bom_files, running_stock, moq_map)

        self.assertEqual(plan["file_plans"][0]["supplements"], {"PART-1": 25.0})
        self.assertEqual(plan["running_stock"]["PART-1"], 115.0)

    def test_get_draft_detail_includes_preview_rows_without_shortages(self):
        draft = {
            "id": 5,
            "order_id": 12,
            "status": "active",
            "main_loaded_at": "2026-03-13T15:00:00",
            "updated_at": "2026-03-13T15:05:00",
            "supplements": {},
            "decisions": {},
            "shortages": [],
        }
        file_item = {
            "id": 9,
            "bom_file_id": "bom-1",
            "filename": "draft-1.xlsx",
            "filepath": "C:/draft-1.xlsx",
            "source_filename": "source-1.xlsx",
            "source_format": ".xlsx",
            "model": "MODEL-A",
            "group_model": "MODEL-A",
            "carry_overs": {"PART-1": 120, "PART-2": 45},
            "supplements": {"PART-2": 300},
        }
        components = [
            {
                "part_number": "PART-1",
                "description": "Resistor",
                "needed_qty": 10,
                "prev_qty_cs": 0,
                "is_dash": 0,
                "is_customer_supplied": 0,
            },
            {
                "part_number": "PART-2",
                "description": "Capacitor",
                "needed_qty": 6,
                "prev_qty_cs": 2,
                "is_dash": 0,
                "is_customer_supplied": 0,
            },
        ]

        with patch("app.services.merge_drafts.db.get_merge_draft", return_value=draft), \
             patch("app.services.merge_drafts.db.get_order", return_value={"id": 12, "model": "MODEL-A"}), \
             patch("app.services.merge_drafts.db.get_order_decisions", return_value={12: {"PART-1": "Shortage"}}), \
             patch("app.services.merge_drafts.db.get_order_supplements", return_value={12: {"PART-2": 500}}), \
             patch("app.services.merge_drafts.db.get_merge_draft_files", return_value=[file_item]), \
             patch("app.services.merge_drafts.db.get_bom_components", return_value=components):
            detail = merge_drafts.get_draft_detail(5)

        self.assertEqual(detail["draft"]["decisions"], {"PART-1": "Shortage"})
        self.assertEqual(detail["draft"]["supplements"], {"PART-2": 500.0})
        preview_rows = detail["draft"]["files"][0]["preview_rows"]
        self.assertEqual(len(preview_rows), 2)
        self.assertEqual(preview_rows[0]["part_number"], "PART-1")
        self.assertEqual(preview_rows[0]["carry_over"], 120.0)
        self.assertEqual(preview_rows[0]["supplement_qty"], 0.0)
        self.assertEqual(preview_rows[1]["part_number"], "PART-2")
        self.assertEqual(preview_rows[1]["needed"], 6.0)
        self.assertEqual(preview_rows[1]["prev_qty_cs"], 2.0)
        self.assertEqual(preview_rows[1]["supplement_qty"], 300.0)

    def test_get_schedule_draft_map_uses_persisted_order_settings(self):
        active_draft = {
            "id": 5,
            "order_id": 12,
            "status": "active",
            "model": "MODEL-A",
            "po_number": "4500059234",
            "main_loaded_at": "2026-04-01T09:00:00",
            "updated_at": "2026-04-01T09:05:00",
            "decisions": {"PART-OLD": "IgnoreOnce"},
            "supplements": {"PART-OLD": 1000},
            "shortages": [],
            "files": [{
                "id": 7,
                "bom_file_id": "bom-1",
                "filename": "draft-1.xlsx",
                "filepath": "C:/draft-1.xlsx",
                "source_filename": "source-1.xlsx",
            }],
        }

        with patch("app.services.merge_drafts.db.get_active_merge_drafts", return_value=[active_draft]), \
             patch("app.services.merge_drafts.db.get_order_decisions", return_value={12: {"PART-NEW": "Shortage"}}), \
             patch("app.services.merge_drafts.db.get_order_supplements", return_value={12: {"PART-NEW": 2200}}):
            draft_map = merge_drafts.get_schedule_draft_map()

        self.assertEqual(draft_map[12]["decisions"], {"PART-NEW": "Shortage"})
        self.assertEqual(draft_map[12]["supplements"], {"PART-NEW": 2200.0})

    def test_get_draft_detail_preview_rows_include_resulting_stock_for_shortage(self):
        draft = {
            "id": 6,
            "order_id": 18,
            "status": "committed",
            "main_loaded_at": "2026-04-13T12:00:00",
            "updated_at": "2026-04-13T12:05:00",
            "committed_at": "2026-04-13T12:05:00",
            "supplements": {},
            "decisions": {},
            "shortages": [{
                "part_number": "EC-20121A",
                "shortage_amount": 107.72,
                "resulting_stock": -7.72,
            }],
        }
        file_item = {
            "id": 10,
            "bom_file_id": "bom-ec",
            "filename": "draft-ec.xlsx",
            "filepath": "C:/draft-ec.xlsx",
            "source_filename": "source-ec.xlsx",
            "source_format": ".xlsx",
            "model": "T8U",
            "group_model": "T8U",
            "carry_overs": {"EC-20121A": 1131},
            "supplements": {"EC-20121A": 2000},
        }
        components = [{
            "part_number": "EC-20121A",
            "description": "Capacitor",
            "needed_qty": 3138.72,
            "prev_qty_cs": 0,
            "is_dash": 0,
            "is_customer_supplied": 0,
        }]

        with patch("app.services.merge_drafts.db.get_merge_draft", return_value=draft), \
             patch("app.services.merge_drafts.db.get_order", return_value={"id": 18, "model": "T8U"}), \
             patch("app.services.merge_drafts.db.get_merge_draft_files", return_value=[file_item]), \
             patch("app.services.merge_drafts.db.get_bom_components", return_value=components):
            detail = merge_drafts.get_draft_detail(6)

        preview_row = detail["draft"]["files"][0]["preview_rows"][0]
        self.assertAlmostEqual(preview_row["shortage_amount"], 107.72)
        self.assertAlmostEqual(preview_row["resulting_stock"], -7.72)

    def test_plan_order_draft_treats_ec_below_100_as_shortage(self):
        order = {"id": 18, "code": "2-1", "model": "MODEL-EC"}
        draft = {"decisions": {}, "supplements": {"EC-001": 15}}
        bom_files = [{"id": "bom-ec", "model": "MODEL-EC", "group_model": "MODEL-EC"}]
        running_stock = {"EC-001": 120.0}
        moq_map = {"EC-001": 20.0}
        components = [{
            "part_number": "EC-001",
            "description": "EC Part",
            "needed_qty": 30,
            "prev_qty_cs": 0,
            "is_dash": 0,
        }]

        with patch("app.services.merge_drafts.db.get_bom_components", return_value=components):
            plan = merge_drafts._plan_order_draft(order, draft, bom_files, running_stock, moq_map)

        self.assertEqual(plan["file_plans"][0]["supplements"], {"EC-001": 15.0})
        self.assertEqual(plan["running_stock"]["EC-001"], 105.0)
        self.assertEqual(plan["shortages"], [])

    def test_plan_order_draft_records_resulting_stock_for_remaining_shortage(self):
        order = {"id": 20, "code": "3-1", "model": "MODEL-B"}
        draft = {"decisions": {}, "supplements": {}}
        bom_files = [{"id": "bom-2", "model": "MODEL-B", "group_model": "MODEL-B"}]
        running_stock = {"PART-9": 4.0}
        moq_map = {"PART-9": 10.0}
        components = [{
            "part_number": "PART-9",
            "description": "IC",
            "needed_qty": 7,
            "prev_qty_cs": 2,
            "is_dash": 0,
        }]

        with patch("app.services.merge_drafts.db.get_bom_components", return_value=components):
            plan = merge_drafts._plan_order_draft(order, draft, bom_files, running_stock, moq_map)

        self.assertEqual(len(plan["shortages"]), 1)
        self.assertEqual(plan["shortages"][0]["prev_qty_cs"], 2.0)
        self.assertEqual(plan["shortages"][0]["resulting_stock"], -1.0)

    def test_plan_order_draft_clears_shortage_decision_when_saved_supplement_already_resolves_it(self):
        order = {"id": 21, "code": "3-1", "model": "MODEL-C"}
        draft = {"decisions": {"PART-1": "Shortage"}, "supplements": {"PART-1": 10}}
        bom_files = [{"id": "bom-3", "model": "MODEL-C", "group_model": "MODEL-C"}]
        running_stock = {"PART-1": 5.0}
        moq_map = {"PART-1": 0.0}
        components = [{
            "part_number": "PART-1",
            "description": "Resolved part",
            "needed_qty": 10,
            "prev_qty_cs": 0,
            "is_dash": 0,
        }]

        with patch("app.services.merge_drafts.db.get_bom_components", return_value=components):
            plan = merge_drafts._plan_order_draft(order, draft, bom_files, running_stock, moq_map)

        self.assertEqual(plan["decisions"], {})
        self.assertEqual(plan["shortages"], [])
        self.assertEqual(plan["file_plans"][0]["supplements"], {"PART-1": 10.0})

    def test_plan_order_draft_keeps_total_supply_suggestion_when_st_covers_part_of_shortage(self):
        order = {"id": 22, "code": "3-2", "model": "MODEL-ST"}
        draft = {"decisions": {}, "supplements": {}}
        bom_files = [{"id": "bom-st", "model": "MODEL-ST", "group_model": "MODEL-ST"}]
        running_stock = {"PART-ST": 2.0}
        moq_map = {"PART-ST": 5.0}
        components = [{
            "part_number": "PART-ST",
            "description": "ST assisted part",
            "needed_qty": 10,
            "prev_qty_cs": 0,
            "is_dash": 0,
        }]

        with patch("app.services.merge_drafts.db.get_bom_components", return_value=components):
            plan = merge_drafts._plan_order_draft(
                order,
                draft,
                bom_files,
                running_stock,
                moq_map,
                {"PART-ST": 6.0},
            )

        shortage = plan["shortages"][0]
        self.assertEqual(shortage["shortage_amount"], 8.0)
        self.assertEqual(shortage["st_available_qty"], 6.0)
        self.assertEqual(shortage["purchase_needed_qty"], 2.0)
        self.assertEqual(shortage["purchase_suggested_qty"], 5.0)
        self.assertEqual(shortage["suggested_qty"], 11.0)
        self.assertEqual(plan["file_plans"][0]["purchase_parts"], [])

    def test_plan_order_draft_marks_manual_supplement_orange_when_qty_exceeds_st_stock(self):
        order = {"id": 24, "code": "3-3", "model": "MODEL-MANUAL"}
        draft = {"decisions": {}, "supplements": {"PART-MANUAL": 20000}}
        bom_files = [{"id": "bom-manual", "model": "MODEL-MANUAL", "group_model": "MODEL-MANUAL"}]
        running_stock = {"PART-MANUAL": 50000.0}
        moq_map = {"PART-MANUAL": 0.0}
        components = [{
            "part_number": "PART-MANUAL",
            "description": "Manual part",
            "needed_qty": 10000,
            "prev_qty_cs": 0,
            "is_dash": 0,
        }]

        with patch("app.services.merge_drafts.db.get_bom_components", return_value=components):
            plan = merge_drafts._plan_order_draft(
                order,
                draft,
                bom_files,
                running_stock,
                moq_map,
                {"PART-MANUAL": 10000.0},
            )

        self.assertEqual(plan["file_plans"][0]["supplements"], {"PART-MANUAL": 20000.0})
        self.assertEqual(plan["file_plans"][0]["purchase_parts"], ["PART-MANUAL"])
        self.assertEqual(plan["shortages"], [])

    def test_write_dispatch_values_to_ws_marks_purchase_parts_in_orange(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.cell(row=5, column=3, value="PART-ST")

        merge_drafts._write_dispatch_values_to_ws(
            worksheet,
            supplements={"PART-ST": 11},
            carry_overs={},
            purchase_parts={"PART-ST"},
        )

        h_cell = worksheet.cell(row=5, column=8)
        self.assertEqual(h_cell.value, 11)
        self.assertEqual(h_cell.fill.fill_type, "solid")
        self.assertTrue(str(h_cell.fill.start_color.rgb or "").endswith("FFC000"))

    def test_plan_order_draft_scales_needed_qty_from_schedule_order_qty(self):
        order = {"id": 30, "code": "4-1", "model": "MODEL-S", "order_qty": 5}
        draft = {"decisions": {}, "supplements": {}}
        bom_files = [{"id": "bom-scale", "model": "MODEL-S", "group_model": "MODEL-S", "order_qty": 10}]
        running_stock = {"PART-S": 4.0}
        moq_map = {"PART-S": 0.0}
        components = [{
            "part_number": "PART-S",
            "description": "Scaled",
            "qty_per_board": 2,
            "bom_order_qty": 10,
            "needed_qty": 20,
            "prev_qty_cs": 0,
            "is_dash": 0,
        }]

        with patch("app.services.merge_drafts.db.get_bom_components", return_value=components):
            plan = merge_drafts._plan_order_draft(order, draft, bom_files, running_stock, moq_map)

        self.assertEqual(plan["file_plans"][0]["order_qty"], 5.0)
        self.assertEqual(plan["shortages"][0]["needed"], 10.0)
        self.assertEqual(plan["running_stock"]["PART-S"], -6.0)

    def test_write_dispatch_values_to_ws_updates_order_qty_and_needed_column(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.cell(row=1, column=11, value=10)
        worksheet.cell(row=2, column=7, value=10)
        worksheet.cell(row=5, column=2, value=2)
        worksheet.cell(row=5, column=3, value="PART-S")
        worksheet.cell(row=5, column=6, value=20)
        worksheet.cell(row=5, column=7, value=0)
        worksheet.cell(row=5, column=8, value=0)

        merge_drafts._write_bom_header_values(worksheet, "4500059234", 5)
        merge_drafts._write_dispatch_values_to_ws(
            worksheet,
            supplements={},
            carry_overs={},
            target_order_qty=5,
            source_order_qty=10,
        )

        self.assertEqual(worksheet.cell(row=1, column=11).value, 5)
        self.assertEqual(worksheet.cell(row=2, column=7).value, 5)
        self.assertEqual(worksheet.cell(row=5, column=6).value, 10)

    def test_write_dispatch_values_to_ws_preserves_bom_scrap_rate(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.cell(row=1, column=11, value=10)
        worksheet.cell(row=2, column=7, value=10)
        worksheet.cell(row=5, column=2, value=2)
        worksheet.cell(row=5, column=3, value="PART-S")
        worksheet.cell(row=5, column=5, value=0.06)  # E 欄：拋料率 6%
        worksheet.cell(row=5, column=6, value=21.2)
        worksheet.cell(row=5, column=7, value=0)
        worksheet.cell(row=5, column=8, value=0)

        merge_drafts._write_bom_header_values(worksheet, "4500059234", 5)
        merge_drafts._write_dispatch_values_to_ws(
            worksheet,
            supplements={},
            carry_overs={},
            target_order_qty=5,
            source_order_qty=10,
        )

        self.assertEqual(worksheet.cell(row=1, column=11).value, 5)
        self.assertEqual(worksheet.cell(row=2, column=7).value, 5)
        self.assertAlmostEqual(float(worksheet.cell(row=5, column=6).value), 10.6)

    def test_write_dispatch_values_to_ws_keeps_needed_formula(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.cell(row=1, column=11, value=10)
        worksheet.cell(row=2, column=7, value=10)
        worksheet.cell(row=5, column=2, value=2)
        worksheet.cell(row=5, column=3, value="PART-S")
        worksheet.cell(row=5, column=6, value="=B5*$K$1*1.06")
        worksheet.cell(row=5, column=7, value=0)
        worksheet.cell(row=5, column=8, value=0)

        merge_drafts._write_bom_header_values(worksheet, "4500059234", 5)
        merge_drafts._write_dispatch_values_to_ws(
            worksheet,
            supplements={},
            carry_overs={},
            target_order_qty=5,
            source_order_qty=10,
        )

        self.assertEqual(worksheet.cell(row=1, column=11).value, 5)
        self.assertEqual(worksheet.cell(row=2, column=7).value, 5)
        self.assertEqual(worksheet.cell(row=5, column=6).value, "=B5*$K$1*1.06")

    def test_write_draft_files_uses_first_worksheet_and_sets_it_active(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            source_path = temp_path / "source.xlsx"
            draft_dir = temp_path / "drafts"

            workbook = Workbook()
            first_sheet = workbook.active
            first_sheet.title = "BOM"
            first_sheet.cell(row=1, column=11, value=10)
            first_sheet.cell(row=2, column=7, value=10)
            first_sheet.cell(row=5, column=2, value=2)
            first_sheet.cell(row=5, column=3, value="PART-A")
            first_sheet.cell(row=5, column=6, value=20)
            first_sheet.cell(row=5, column=7, value=0)
            first_sheet.cell(row=5, column=8, value=0)
            other_sheet = workbook.create_sheet("說明頁")
            other_sheet["A1"] = "不要寫到這張"
            workbook.active = 1
            workbook.save(source_path)
            workbook.close()

            with patch("app.services.merge_drafts.MERGE_DRAFT_DIR", draft_dir), \
                 patch("app.services.merge_drafts.db.get_bom_file", return_value={
                     "id": "bom-1",
                     "filename": "source.xlsx",
                     "filepath": str(source_path),
                 }), \
                 patch(
                     "app.services.merge_drafts.save_workbook_with_recalc",
                     side_effect=lambda workbook, output_path: workbook.save(output_path),
                 ) as mock_save:
                written = merge_drafts._write_draft_files(5, [{
                    "bom_file_id": "bom-1",
                    "source_filename": "source.xlsx",
                    "source_format": ".xlsx",
                    "model": "MODEL-A",
                    "group_model": "MODEL-A",
                    "po_number": "4500059234",
                    "order_qty": 5,
                    "source_order_qty": 10,
                    "carry_overs": {"PART-A": 8},
                    "supplements": {"PART-A": 3},
                    "purchase_parts": [],
                }])

            self.assertEqual(len(written), 1)
            self.assertTrue(written[0]["filename"].startswith("4500059234_MODEL-A_"))
            self.assertTrue(written[0]["filename"].endswith(".xlsx"))
            mock_save.assert_called_once()
            output_path = Path(written[0]["filepath"])
            self.assertTrue(output_path.exists())
            self.assertTrue(output_path.name.startswith("01_4500059234_MODEL-A_"))

            saved = load_workbook(output_path)
            try:
                self.assertEqual(saved.sheetnames[0], "BOM")
                self.assertEqual(saved.active.title, "BOM")
                self.assertEqual(saved.worksheets[0].cell(row=1, column=11).value, 5)
                self.assertEqual(saved.worksheets[0].cell(row=5, column=7).value, 8)
                self.assertEqual(saved.worksheets[0].cell(row=5, column=8).value, 3)
                self.assertEqual(saved["說明頁"]["A1"].value, "不要寫到這張")
            finally:
                saved.close()

    def test_ensure_editable_bom_for_draft_converts_old_xls_before_writing(self):
        bom = {
            "id": "bom-legacy",
            "filename": "legacy.xls",
            "filepath": "C:/bom-legacy.xls",
            "uploaded_at": "2026-04-01T12:00:00",
            "group_model": "MODEL-X",
            "source_filename": "legacy.xls",
            "source_format": "",
            "is_converted": 0,
        }

        with patch("app.services.merge_drafts.normalize_bom_record_to_editable", return_value={
            **bom,
            "filename": "legacy.xlsx",
            "filepath": "C:/bom-legacy.xlsx",
            "source_format": ".xls",
            "is_converted": True,
        }) as mock_normalize, \
             patch("app.services.merge_drafts.Path.exists", return_value=True), \
             patch("app.services.merge_drafts.parse_bom_for_storage", return_value=type("ParsedBom", (), {"id": "bom-legacy", "filename": "legacy.xlsx", "path": "C:/bom-legacy.xlsx", "source_filename": "legacy.xls", "source_format": ".xls", "is_converted": True, "po_number": 0, "model": "", "pcb": "", "group_model": "MODEL-X", "order_qty": 0.0, "uploaded_at": "2026-04-01T12:00:00", "components": []})()) as mock_parse, \
             patch("app.services.merge_drafts.build_bom_storage_payload", return_value={"id": "bom-legacy"}) as mock_payload, \
             patch("app.services.merge_drafts.db.save_bom_file") as mock_save, \
             patch("app.services.merge_drafts.db.log_activity") as mock_log, \
             patch("app.services.merge_drafts.db.get_bom_file", return_value={"id": "bom-legacy", "filename": "legacy.xlsx", "filepath": "C:/bom-legacy.xlsx"}):
            result = merge_drafts._ensure_editable_bom_for_draft(bom)

        self.assertEqual(result["filepath"], "C:/bom-legacy.xlsx")
        mock_normalize.assert_called_once_with(bom)
        mock_parse.assert_called_once()
        mock_payload.assert_called_once()
        mock_save.assert_called_once_with({"id": "bom-legacy"})
        mock_log.assert_called_once()

    def test_ensure_editable_bom_for_draft_syncs_existing_xlsx_components(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "bom.xlsx"
            source_path.write_bytes(b"dummy")
            bom = {
                "id": "bom-sync",
                "filename": "bom.xlsx",
                "filepath": str(source_path),
                "uploaded_at": "2026-04-22T10:00:00",
                "group_model": "PB-20111A-TAB",
                "source_filename": "bom.xlsx",
                "source_format": ".xlsx",
                "is_converted": 0,
            }
            parsed = type("ParsedBom", (), {
                "id": "bom-sync",
                "filename": "bom.xlsx",
                "path": str(source_path),
                "source_filename": "bom.xlsx",
                "source_format": ".xlsx",
                "is_converted": False,
                "po_number": 0,
                "model": "",
                "pcb": "",
                "group_model": "PB-20111A-TAB",
                "order_qty": 300.0,
                "uploaded_at": "2026-04-22T10:00:00",
                "components": [],
            })()
            payload = {
                "id": "bom-sync",
                "components": [{"part_number": "PART-1", "qty_per_board": 1, "needed_qty": 300}],
            }

            with patch("app.services.merge_drafts.normalize_bom_record_to_editable", return_value=dict(bom)) as mock_normalize, \
                 patch("app.services.merge_drafts.parse_bom_for_storage", return_value=parsed) as mock_parse, \
                 patch("app.services.merge_drafts.build_bom_storage_payload", return_value=payload), \
                 patch("app.services.merge_drafts.db.save_bom_file") as mock_save, \
                 patch("app.services.merge_drafts.db.log_activity") as mock_log, \
                 patch("app.services.merge_drafts.db.get_bom_file", return_value={**bom, "synced": True}):
                result = merge_drafts._ensure_editable_bom_for_draft(bom)

        self.assertTrue(result["synced"])
        mock_normalize.assert_called_once_with(bom)
        mock_parse.assert_called_once()
        mock_save.assert_called_once_with(payload)
        mock_log.assert_not_called()

    def test_restore_recent_committed_merge_drafts_reactivates_and_rebuilds(self):
        active_drafts = [{"id": 7, "order_id": 21}]

        with patch("app.services.merge_drafts.db.get_active_merge_draft_for_order", return_value=None), \
             patch("app.services.merge_drafts.db.get_latest_committed_merge_draft_for_order", return_value={"id": 7, "order_id": 21}), \
             patch("app.services.merge_drafts.db.reactivate_merge_draft", return_value=1) as mock_reactivate, \
             patch("app.services.merge_drafts.db.get_active_merge_drafts", return_value=active_drafts), \
             patch("app.services.merge_drafts.rebuild_merge_drafts", return_value=active_drafts) as mock_rebuild:
            restored = merge_drafts.restore_recent_committed_merge_drafts([21])

        self.assertEqual(restored, [21])
        mock_reactivate.assert_called_once_with(7)
        mock_rebuild.assert_called_once_with([21])

    def test_rebuild_merge_drafts_uses_latest_persisted_order_settings_when_no_override(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.write_bytes(b"test")
            replace_calls = []

            with patch("app.services.merge_drafts.db.get_setting", side_effect=lambda key, default="": {
                "main_file_path": str(main_path),
                "main_loaded_at": "2026-03-31T10:00:00",
            }.get(key, default)), \
                 patch("app.services.merge_drafts.db.get_order", return_value={"id": 12, "status": "merged", "model": "MODEL-A"}), \
                 patch("app.services.merge_drafts.db.get_active_merge_draft_for_order", return_value={
                     "id": 5,
                     "order_id": 12,
                     "decisions": {"PART-OLD": "IgnoreOnce"},
                     "supplements": {"PART-OLD": 1000},
                     "shortages": [],
                 }), \
                 patch("app.services.merge_drafts.db.get_order_decisions", return_value={12: {"IC-CSD18531Q5AT-TAB": "CreateRequirement"}}), \
                 patch("app.services.merge_drafts.db.get_order_supplements", return_value={12: {"IC-CSD18531Q5AT-TAB": 2000}}), \
                 patch("app.services.merge_drafts.db.replace_merge_draft", side_effect=lambda **kwargs: replace_calls.append(kwargs)), \
                 patch("app.services.merge_drafts.db.get_active_merge_drafts", return_value=[]), \
                 patch("app.services.merge_drafts._build_running_stock", return_value={}), \
                 patch("app.services.merge_drafts._load_effective_moq", return_value={}), \
                 patch("app.services.merge_drafts.db.get_st_inventory_stock", return_value={}):
                merge_drafts.rebuild_merge_drafts([12])

        self.assertEqual(len(replace_calls), 1)
        self.assertEqual(replace_calls[0]["decisions"], {"IC-CSD18531Q5AT-TAB": "CreateRequirement"})
        self.assertEqual(replace_calls[0]["supplements"], {"IC-CSD18531Q5AT-TAB": 2000})

    def test_rebuild_merge_drafts_replans_from_persisted_settings_not_stale_draft_payload(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.write_bytes(b"test")
            plan_calls = []

            with patch("app.services.merge_drafts.db.get_setting", side_effect=lambda key, default="": {
                "main_file_path": str(main_path),
                "main_loaded_at": "2026-04-01T08:00:00",
            }.get(key, default)), \
                 patch("app.services.merge_drafts.db.get_order", return_value={"id": 12, "status": "merged", "model": "MODEL-A"}), \
                 patch("app.services.merge_drafts.db.get_active_merge_draft_for_order", return_value={
                     "id": 5,
                     "order_id": 12,
                     "decisions": {"PART-OLD": "IgnoreOnce"},
                     "supplements": {"PART-OLD": 1000},
                     "shortages": [],
                 }), \
                 patch("app.services.merge_drafts.db.get_order_decisions", return_value={12: {"PART-NEW": "Shortage"}}), \
                 patch("app.services.merge_drafts.db.get_order_supplements", return_value={12: {"PART-NEW": 2200}}), \
                 patch("app.services.merge_drafts.db.replace_merge_draft"), \
                 patch("app.services.merge_drafts.db.get_active_merge_drafts", return_value=[{
                     "id": 5,
                     "order_id": 12,
                     "decisions": {"PART-OLD": "IgnoreOnce"},
                     "supplements": {"PART-OLD": 1000},
                     "shortages": [],
                 }]), \
                 patch("app.services.merge_drafts.db.get_bom_files_by_models", return_value=[]), \
                 patch("app.services.merge_drafts._build_running_stock", return_value={}), \
                 patch("app.services.merge_drafts._load_effective_moq", return_value={}), \
                 patch("app.services.merge_drafts.db.get_st_inventory_stock", return_value={}), \
                 patch("app.services.merge_drafts.db.replace_order_decisions"), \
                 patch("app.services.merge_drafts._cleanup_draft_files"), \
                 patch("app.services.merge_drafts._write_draft_files", return_value=[]), \
                 patch("app.services.merge_drafts.db.replace_merge_draft_files"), \
                 patch("app.services.merge_drafts._plan_order_draft", side_effect=lambda *args, **kwargs: plan_calls.append(kwargs) or {
                     "running_stock": {},
                     "file_plans": [],
                     "shortages": [],
                     "decisions": kwargs["decisions"],
                 }):
                merge_drafts.rebuild_merge_drafts([12])

        self.assertEqual(len(plan_calls), 1)
        self.assertEqual(plan_calls[0]["decisions"], {"PART-NEW": "Shortage"})
        self.assertEqual(plan_calls[0]["supplements"], {"PART-NEW": 2200.0})

    def test_rebuild_merge_drafts_syncs_bom_files_before_planning(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.write_bytes(b"test")
            raw_bom = {"id": "bom-1", "filename": "old.xlsx", "filepath": str(Path(temp_dir) / "old.xlsx")}
            synced_bom = {**raw_bom, "filename": "synced.xlsx"}
            planned_bom_files = []

            def capture_plan(order, draft, bom_files, *args, **kwargs):
                planned_bom_files.extend(bom_files)
                return {
                    "running_stock": {},
                    "file_plans": [],
                    "shortages": [],
                    "decisions": kwargs["decisions"],
                }

            with patch("app.services.merge_drafts.db.get_setting", side_effect=lambda key, default="": {
                "main_file_path": str(main_path),
                "main_loaded_at": "2026-04-22T10:00:00",
            }.get(key, default)), \
                 patch("app.services.merge_drafts.db.get_order", return_value={"id": 12, "status": "merged", "model": "MODEL-A"}), \
                 patch("app.services.merge_drafts.db.get_active_merge_draft_for_order", return_value={
                     "id": 5,
                     "order_id": 12,
                     "decisions": {},
                     "supplements": {},
                     "shortages": [],
                 }), \
                 patch("app.services.merge_drafts.db.get_order_decisions", return_value={12: {}}), \
                 patch("app.services.merge_drafts.db.get_order_supplements", return_value={12: {}}), \
                 patch("app.services.merge_drafts.db.replace_merge_draft"), \
                 patch("app.services.merge_drafts.db.get_active_merge_drafts", return_value=[{
                     "id": 5,
                     "order_id": 12,
                     "decisions": {},
                     "supplements": {},
                     "shortages": [],
                 }]), \
                 patch("app.services.merge_drafts.db.get_bom_files_by_models", return_value=[raw_bom]), \
                 patch("app.services.merge_drafts._ensure_editable_bom_for_draft", return_value=synced_bom) as mock_sync, \
                 patch("app.services.merge_drafts._build_running_stock", return_value={}), \
                 patch("app.services.merge_drafts._load_effective_moq", return_value={}), \
                 patch("app.services.merge_drafts.db.get_st_inventory_stock", return_value={}), \
                 patch("app.services.merge_drafts.db.replace_order_decisions"), \
                 patch("app.services.merge_drafts._cleanup_draft_files"), \
                 patch("app.services.merge_drafts._write_draft_files", return_value=[]), \
                 patch("app.services.merge_drafts.db.replace_merge_draft_files"), \
                 patch("app.services.merge_drafts._plan_order_draft", side_effect=capture_plan):
                merge_drafts.rebuild_merge_drafts([12])

        mock_sync.assert_called_once_with(raw_bom)
        self.assertEqual(planned_bom_files, [synced_bom])

    def test_rebuild_merge_drafts_persists_cleaned_shortage_decisions(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            main_path.write_bytes(b"test")

            with patch("app.services.merge_drafts.db.get_setting", side_effect=lambda key, default="": {
                "main_file_path": str(main_path),
                "main_loaded_at": "2026-04-01T10:00:00",
            }.get(key, default)), \
                 patch("app.services.merge_drafts.db.get_order", return_value={"id": 12, "status": "merged", "model": "MODEL-A"}), \
                 patch("app.services.merge_drafts.db.get_active_merge_draft_for_order", return_value={
                     "id": 5,
                     "order_id": 12,
                     "decisions": {"PART-OLD": "IgnoreOnce"},
                     "supplements": {"PART-OLD": 1000},
                     "shortages": [],
                 }), \
                 patch("app.services.merge_drafts.db.get_order_decisions", return_value={12: {"PART-NEW": "Shortage"}}), \
                 patch("app.services.merge_drafts.db.get_order_supplements", return_value={12: {"PART-NEW": 2200}}), \
                 patch("app.services.merge_drafts.db.replace_merge_draft"), \
                 patch("app.services.merge_drafts.db.get_active_merge_drafts", return_value=[{
                     "id": 5,
                     "order_id": 12,
                     "decisions": {"PART-OLD": "IgnoreOnce"},
                     "supplements": {"PART-OLD": 1000},
                     "shortages": [],
                 }]), \
                 patch("app.services.merge_drafts.db.get_bom_files_by_models", return_value=[]), \
                 patch("app.services.merge_drafts._build_running_stock", return_value={}), \
                 patch("app.services.merge_drafts._load_effective_moq", return_value={}), \
                 patch("app.services.merge_drafts.db.get_st_inventory_stock", return_value={}), \
                 patch("app.services.merge_drafts._cleanup_draft_files"), \
                 patch("app.services.merge_drafts._write_draft_files", return_value=[]), \
                 patch("app.services.merge_drafts.db.replace_merge_draft_files"), \
                 patch("app.services.merge_drafts._plan_order_draft", return_value={
                     "running_stock": {},
                     "file_plans": [],
                     "shortages": [],
                     "decisions": {},
                 }), \
                 patch("app.services.merge_drafts.db.replace_order_decisions") as mock_replace_order_decisions:
                merge_drafts.rebuild_merge_drafts([12])

        mock_replace_order_decisions.assert_called_once_with([12], {12: {}})
