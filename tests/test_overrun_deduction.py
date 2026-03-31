from __future__ import annotations

import sqlite3
import tempfile
import unittest
from contextlib import contextmanager
from pathlib import Path
from unittest.mock import patch

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side

import app.database as db
from app.services.defective_deduction import deduct_defectives_from_main, reverse_defectives_from_main
from app.services.overrun_deduction import (
    apply_overrun_import_confirmations,
    build_overrun_import_preview,
    build_model_overrun_plan,
    parse_overrun_detail_excel,
    preview_deductions_against_main,
)


class InMemoryDbTestCase(unittest.TestCase):
    def setUp(self):
        self.conn = sqlite3.connect(":memory:")
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA foreign_keys=ON")
        self.conn.executescript(db._CREATE_SQL)
        self.conn.execute("ALTER TABLE orders ADD COLUMN folder TEXT NOT NULL DEFAULT ''")

        @contextmanager
        def temp_conn():
            try:
                yield self.conn
                self.conn.commit()
            except Exception:
                self.conn.rollback()
                raise

        self.get_conn_patcher = patch.object(db, "get_conn", temp_conn)
        self.group_patcher = patch.object(db, "_get_bom_model_groups", lambda: {})
        self.get_conn_patcher.start()
        self.group_patcher.start()

    def tearDown(self):
        self.group_patcher.stop()
        self.get_conn_patcher.stop()
        self.conn.close()


class OverrunDeductionTests(InMemoryDbTestCase):
    def test_parse_overrun_detail_excel_reads_vendor_detail_format(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            detail_path = Path(temp_dir) / "detail.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "工作表5"
            ws.cell(row=1, column=1).value = "料號"
            ws.cell(row=1, column=2).value = "12-7  02/02  T356789IU-A (REV3.4) * 4"
            ws.cell(row=2, column=1).value = "M/O"
            ws.cell(row=2, column=2).value = "4500059105  多打  S+D"
            ws.cell(row=3, column=1).value = "EC-10025A"
            ws.cell(row=3, column=2).value = 56
            ws.cell(row=4, column=1).value = "EC-10028A"
            ws.cell(row=4, column=2).value = 24
            wb.save(detail_path)
            wb.close()

            parsed = parse_overrun_detail_excel(str(detail_path))

        self.assertEqual(parsed["title"], "12-7  02/02  T356789IU-A (REV3.4) * 4")
        self.assertEqual(parsed["mo_info"], "4500059105  多打  S+D")
        self.assertEqual(parsed["items"], [
            {"source_row": 3, "part_number": "EC-10025A", "description": "", "defective_qty": 56.0},
            {"source_row": 4, "part_number": "EC-10028A", "description": "", "defective_qty": 24.0},
        ])

    def test_build_model_overrun_plan_aggregates_parts_and_skips_non_stock_rows(self):
        db.save_bom_file({
            "id": "bom-1",
            "filename": "model-a.xlsx",
            "filepath": "C:/bom/model-a.xlsx",
            "po_number": "1001",
            "model": "MODEL-A",
            "pcb": "PCB-A",
            "group_model": "MODEL-A,MODEL-A/A",
            "order_qty": 100,
            "uploaded_at": "2026-03-19T09:00:00",
            "components": [
                {"part_number": "PART-A", "description": "IC-A", "qty_per_board": 0.4, "needed_qty": 40, "prev_qty_cs": 0, "is_dash": False, "is_customer_supplied": False},
                {"part_number": "PART-A", "description": "IC-A", "qty_per_board": 0.4, "needed_qty": 40, "prev_qty_cs": 0, "is_dash": False, "is_customer_supplied": False},
                {"part_number": "PART-B", "description": "CAP-B", "qty_per_board": 2, "needed_qty": 200, "prev_qty_cs": 0, "is_dash": False, "is_customer_supplied": False},
                {"part_number": "PART-C", "description": "SKIP-DASH", "qty_per_board": 1, "needed_qty": 100, "prev_qty_cs": 0, "is_dash": True, "is_customer_supplied": False},
                {"part_number": "PART-D", "description": "客供料", "qty_per_board": 1, "needed_qty": 100, "prev_qty_cs": 0, "is_dash": False, "is_customer_supplied": True},
            ],
        })

        plan = build_model_overrun_plan("model-a/a", 1)

        self.assertEqual(plan["model"], "MODEL-A/A")
        self.assertEqual(plan["matched_models"], ["MODEL-A,MODEL-A/A"])
        self.assertEqual([item["part_number"] for item in plan["items"]], ["PART-A", "PART-B"])
        self.assertEqual(plan["items"][0]["defective_qty"], 1.0)
        self.assertAlmostEqual(plan["items"][0]["qty_per_board_total"], 0.8)
        self.assertEqual(plan["items"][1]["defective_qty"], 2.0)

    def test_preview_deductions_against_main_reports_missing_and_negative_stock(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.cell(row=1, column=1).value = "Part"
            ws.cell(row=1, column=4).value = "Stock"
            ws.cell(row=2, column=1).value = "PART-A"
            ws.cell(row=2, column=4).value = 5
            ws.cell(row=3, column=1).value = "PART-B"
            ws.cell(row=3, column=4).value = 1
            wb.save(main_path)
            wb.close()

            preview = preview_deductions_against_main(str(main_path), [
                {"part_number": "PART-A", "description": "IC-A", "defective_qty": 3},
                {"part_number": "PART-B", "description": "IC-B", "defective_qty": 2},
                {"part_number": "PART-C", "description": "IC-C", "defective_qty": 1},
            ])

        self.assertEqual(preview["deducted_count"], 2)
        self.assertEqual(preview["skipped_parts"], ["PART-C"])
        self.assertEqual(preview["negative_count"], 1)
        result_map = {item["part_number"]: item for item in preview["results"]}
        self.assertEqual(result_map["PART-A"]["stock_before"], 5)
        self.assertEqual(result_map["PART-A"]["stock_after"], 2)
        self.assertEqual(result_map["PART-B"]["stock_after"], -1)

    def test_build_overrun_import_preview_includes_missing_suggestions(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.cell(row=1, column=1).value = "Part"
            ws.cell(row=1, column=4).value = "Stock"
            ws.cell(row=2, column=1).value = "EC-10025A"
            ws.cell(row=2, column=4).value = 100
            ws.cell(row=3, column=1).value = "EC-10028A"
            ws.cell(row=3, column=4).value = 50
            wb.save(main_path)
            wb.close()

            preview = build_overrun_import_preview(str(main_path), {
                "title": "Title",
                "mo_info": "MO",
                "source_filename": "detail.xlsx",
                "items": [
                    {"source_row": 3, "part_number": "EC-10025A", "description": "", "defective_qty": 56},
                    {"source_row": 4, "part_number": "EC-10029A", "description": "", "defective_qty": 24},
                ],
            })

        self.assertEqual(preview["missing_count"], 1)
        self.assertTrue(preview["requires_confirmation"])
        self.assertEqual(preview["missing_items"][0]["part_number"], "EC-10029A")
        self.assertTrue(preview["missing_items"][0]["suggestions"])

    def test_apply_overrun_import_confirmations_supports_replace_and_skip(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.cell(row=1, column=1).value = "Part"
            ws.cell(row=1, column=4).value = "Stock"
            ws.cell(row=2, column=1).value = "EC-10025A"
            ws.cell(row=2, column=4).value = 100
            ws.cell(row=3, column=1).value = "EC-10028A"
            ws.cell(row=3, column=4).value = 50
            wb.save(main_path)
            wb.close()

            applied = apply_overrun_import_confirmations(str(main_path), [
                {"source_row": 3, "part_number": "EC-10025A", "defective_qty": 56, "action": "deduct", "target_part_number": ""},
                {"source_row": 4, "part_number": "EC-10029A", "defective_qty": 24, "action": "replace", "target_part_number": "EC-10028A"},
                {"source_row": 5, "part_number": "EC-10099A", "defective_qty": 10, "action": "skip", "target_part_number": ""},
            ])

        self.assertFalse(applied["unresolved_items"])
        self.assertEqual(len(applied["final_items"]), 2)
        self.assertEqual(len(applied["replaced_items"]), 1)
        self.assertEqual(len(applied["skipped_items"]), 1)

    def test_overrun_headers_are_written_to_main_and_reversal(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.cell(row=1, column=1).value = "Part"
            ws.cell(row=1, column=4).value = "Stock"
            ws.cell(row=2, column=1).value = "PART-A"
            ws.cell(row=2, column=4).value = 100
            wb.save(main_path)
            wb.close()

            deduct_defectives_from_main(
                str(main_path),
                [{"part_number": "PART-A", "description": "IC-A", "defective_qty": 25}],
                entry_header="加工多打扣帳",
            )
            reverse_defectives_from_main(
                str(main_path),
                [{"part_number": "PART-A", "defective_qty": 25}],
                entry_header="加工多打回復",
            )

            result_wb = openpyxl.load_workbook(main_path, data_only=False)
            try:
                result_ws = result_wb.active
                self.assertEqual(result_ws.cell(row=1, column=5).value, "加工多打扣帳")
                self.assertEqual(result_ws.cell(row=1, column=7).value, "加工多打回復")
                self.assertEqual(result_ws.cell(row=2, column=5).value, 25)
                self.assertEqual(result_ws.cell(row=2, column=6).value, 75)
                self.assertEqual(result_ws.cell(row=2, column=7).value, 25)
                self.assertEqual(result_ws.cell(row=2, column=8).value, 100)
            finally:
                result_wb.close()

    def test_deduction_writes_to_first_sheet_even_if_other_sheet_is_active(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            wb = openpyxl.Workbook()
            inventory_ws = wb.active
            inventory_ws.title = "主檔"
            inventory_ws.cell(row=1, column=1).value = "Part"
            inventory_ws.cell(row=1, column=4).value = "Stock"
            inventory_ws.cell(row=2, column=1).value = "PART-A"
            inventory_ws.cell(row=2, column=4).value = 100
            summary_ws = wb.create_sheet("摘要")
            summary_ws.cell(row=1, column=1).value = "Summary"
            wb.active = 1
            wb.save(main_path)
            wb.close()

            deduct_defectives_from_main(
                str(main_path),
                [{"part_number": "PART-A", "description": "IC-A", "defective_qty": 25}],
                entry_header="加工多打扣帳",
            )

            result_wb = openpyxl.load_workbook(main_path, data_only=False)
            try:
                inventory_ws = result_wb["主檔"]
                summary_ws = result_wb["摘要"]
                self.assertEqual(inventory_ws.cell(row=1, column=5).value, "加工多打扣帳")
                self.assertEqual(inventory_ws.cell(row=2, column=6).value, 75)
                self.assertIsNone(summary_ws.cell(row=1, column=2).value)
            finally:
                result_wb.close()

    def test_deduction_copies_recent_column_layout_for_new_columns(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "主檔"
            ws.cell(row=1, column=1).value = "Part"
            ws.cell(row=1, column=4).value = "Stock"
            ws.cell(row=2, column=1).value = "PART-A"
            source_cell = ws.cell(row=2, column=4)
            source_cell.value = 100
            source_cell.number_format = "#,##0.0"
            source_cell.font = Font(name="Calibri", size=11, italic=True)
            source_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            source_cell.border = Border(left=Side(style="thin", color="000000"))
            ws.column_dimensions["D"].width = 24
            wb.save(main_path)
            wb.close()

            deduct_defectives_from_main(
                str(main_path),
                [{"part_number": "PART-A", "description": "IC-A", "defective_qty": 25}],
                entry_header="加工多打扣帳",
            )

            result_wb = openpyxl.load_workbook(main_path, data_only=False)
            try:
                result_ws = result_wb.active
                self.assertEqual(result_ws.column_dimensions["E"].width, 24)
                self.assertEqual(result_ws.column_dimensions["F"].width, 24)
                self.assertEqual(result_ws.cell(row=2, column=5).number_format, "#,##0.0")
                self.assertTrue(result_ws.cell(row=2, column=5).font.italic)
                self.assertEqual(result_ws.cell(row=2, column=5).fill.fill_type, "solid")
                self.assertEqual(result_ws.cell(row=2, column=6).border.left.style, "thin")
            finally:
                result_wb.close()
