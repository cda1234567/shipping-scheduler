from __future__ import annotations

import unittest
import sqlite3
from contextlib import contextmanager
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from openpyxl import Workbook

import app.database as db
from app.services.reconcile_core import theoretical_stock
from app.services.st_reconcile import (
    CATEGORY_GENLIN_BLANK_PHYSICAL,
    CATEGORY_HAVE_OURS_NOT_THEIRS,
    CATEGORY_HAVE_THEIRS_NOT_OURS,
    CATEGORY_MATCHED,
    CATEGORY_STOP_LOSS,
    CATEGORY_UNATTRIBUTED,
    build_st_reconcile_preview,
    commit_st_reconcile_stop_loss,
    parse_st_reconcile_file,
    resolve_cutoff_batch,
)


class StReconcileParserTests(unittest.TestCase):
    def test_parse_real_genlin_file_extracts_book_physical_and_strips_tab_suffix(self):
        sample_path = Path("templates") / "庚霖實際庫存2026Q1_2026-6-29.xlsx"

        parsed = parse_st_reconcile_file(str(sample_path))

        self.assertEqual(parsed["format"], "genlin")
        self.assertEqual(parsed["sheet_name"], "實際庫存-生產結餘")
        rows = {row["part_number"]: row for row in parsed["rows"]}
        self.assertIn("OC-10935B", rows)
        self.assertNotIn("OC-10935B-TAB", rows)
        self.assertEqual(rows["OC-10935B"]["book_qty"], 8100)
        self.assertEqual(rows["OC-10935B"]["physical_qty"], 687)

    def test_parse_real_chenshang_file_extracts_parts_and_forward_filled_groups(self):
        sample_path = Path("templates") / "辰尚庫存狀況20260610_辰尚填寫.xlsx"

        parsed = parse_st_reconcile_file(str(sample_path))

        self.assertEqual(parsed["sheet_name"], "辰尚庫存表20260610")
        self.assertGreater(parsed["part_count"], 0)
        rows = parsed["rows"]
        self.assertTrue(any(row["part_number"] and row["physical"] is not None for row in rows))
        self.assertTrue(all(row["part_number"] == row["part_number"].upper() for row in rows))
        self.assertTrue(any(row.get("customer_code") for row in rows))
        manual_split_rows = [row for row in rows if row.get("needs_manual_split")]
        if manual_split_rows:
            row = manual_split_rows[0]
            self.assertIsNone(row["physical"])
            self.assertGreater(row["group_physical"], 0)
            self.assertGreater(row["group_part_count"], 1)
            self.assertIn(row["part_number"], row["group_parts"])


class StReconcileAttributionTests(unittest.TestCase):
    def test_preview_classifies_four_readonly_attribution_buckets(self):
        parsed = {
            "sheet_name": "盤點",
            "rows": [
                {
                    "part_number": "PART-A",
                    "description": "A",
                    "physical": 12,
                    "customer_code": "C1",
                    "needs_manual_split": False,
                },
                {
                    "part_number": "PART-B",
                    "description": "B",
                    "physical": 5,
                    "customer_code": "C2",
                    "needs_manual_split": False,
                },
                {
                    "part_number": "PART-C",
                    "description": "C",
                    "physical": 7,
                    "customer_code": "C3",
                    "needs_manual_split": False,
                },
                {
                    "part_number": "PART-D",
                    "description": "D",
                    "physical": None,
                    "group_physical": 20,
                    "customer_code": "C4",
                    "group_part_count": 2,
                    "group_parts": ["PART-D", "PART-E"],
                    "needs_manual_split": True,
                },
            ],
        }
        theoretical = {
            "stock": {
                "PART-A": 10,
                "PART-B": 8,
                "PART-C": 7,
                "PART-D": 1,
            },
            "order_details": {
                "PART-A": [{"order_id": 101, "used_qty": 2}],
                "PART-C": [{"order_id": 102, "used_qty": 1}],
            },
        }

        with patch("app.services.st_reconcile.parse_st_reconcile_file", return_value=parsed), \
             patch("app.services.st_reconcile.theoretical_stock_with_details", return_value=theoretical):
            report = build_st_reconcile_preview("ignored.xlsx", "2026-06-10")

        by_part = {row["part_number"]: row for row in report["parts"]}
        self.assertEqual(by_part["PART-A"]["category"], CATEGORY_HAVE_OURS_NOT_THEIRS)
        self.assertEqual(by_part["PART-B"]["category"], CATEGORY_HAVE_THEIRS_NOT_OURS)
        self.assertEqual(by_part["PART-C"]["category"], CATEGORY_MATCHED)
        self.assertEqual(by_part["PART-D"]["category"], CATEGORY_UNATTRIBUTED)
        self.assertIn("群組多料號需人工拆分", by_part["PART-D"]["notes"])
        self.assertEqual(report["summary"][CATEGORY_HAVE_OURS_NOT_THEIRS], 1)
        self.assertEqual(report["summary"][CATEGORY_HAVE_THEIRS_NOT_OURS], 1)
        self.assertEqual(report["summary"][CATEGORY_MATCHED], 1)
        self.assertEqual(report["summary"][CATEGORY_UNATTRIBUTED], 1)


class StReconcileCommitTests(unittest.TestCase):
    def setUp(self):
        self.conn = sqlite3.connect(":memory:")
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA foreign_keys=ON")
        self.conn.executescript(db._CREATE_SQL)
        self.conn.execute("ALTER TABLE orders ADD COLUMN folder TEXT NOT NULL DEFAULT ''")

        @contextmanager
        def temp_conn():
            try:
                yield self.conn
                self.conn.commit()
            except Exception:
                self.conn.rollback()
                raise

        self.get_conn_patcher = patch.object(db, "get_conn", temp_conn)
        self.group_patcher = patch.object(db, "_get_bom_model_groups", lambda: {})
        self.get_conn_patcher.start()
        self.group_patcher.start()

    def tearDown(self):
        self.group_patcher.stop()
        self.get_conn_patcher.stop()
        self.conn.close()

    def _make_genlin_file(self, folder: str, physical_qty: float | None, book_qty: float = 10) -> str:
        return self._make_genlin_file_with_rows(
            folder,
            [{"part": "PART-1-TAB", "description": "測試料", "book_qty": book_qty, "physical_qty": physical_qty}],
        )

    def _make_genlin_file_with_rows(self, folder: str, rows: list[dict]) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "實際庫存-生產結餘"
        ws.cell(row=3, column=3, value="Parts No/Description")
        ws.cell(row=3, column=4, value="consign invoice NO")
        ws.cell(row=3, column=6, value="辰尚庫存")
        ws.cell(row=3, column=7, value="庚霖庫存 當下實際")
        for offset, row in enumerate(rows, start=5):
            ws.cell(row=offset, column=3, value=row.get("description") or "")
            ws.cell(row=offset, column=4, value=row.get("part") or "")
            ws.cell(row=offset, column=6, value=row.get("book_qty"))
            ws.cell(row=offset, column=7, value=row.get("physical_qty"))
        path = Path(folder) / f"genlin_{len(rows)}_rows.xlsx"
        wb.save(path)
        return str(path)

    def _insert_snapshot(self, qty: float, part_number: str = "PART-1") -> None:
        self.conn.execute(
            """
            INSERT INTO st_inventory_snapshot(part_number, stock_qty, description, loaded_at)
            VALUES(?, ?, '', '2026-06-01T08:00:00')
            """,
            (part_number, qty),
        )

    def _insert_audit(self, old_qty: float, new_qty: float, delta: float, reason: str, changed_at: str) -> None:
        self.conn.execute(
            "UPDATE st_inventory_snapshot SET stock_qty=?, loaded_at=? WHERE part_number='PART-1'",
            (new_qty, changed_at),
        )
        self.conn.execute(
            """
            INSERT INTO st_inventory_audit_log(part_number, old_qty, new_qty, delta, reason, actor, changed_at)
            VALUES('PART-1', ?, ?, ?, ?, 'test', ?)
            """,
            (old_qty, new_qty, delta, reason, changed_at),
        )

    def test_commit_sets_stop_loss_anchor_writes_audit_and_second_alignment_reanchors(self):
        self._insert_snapshot(10)
        with TemporaryDirectory() as tmp:
            first_path = self._make_genlin_file(tmp, physical_qty=7, book_qty=10)
            result = commit_st_reconcile_stop_loss(first_path, "2026-06-29", source_filename="first.xlsx")

            self.assertTrue(result["ok"])
            self.assertEqual(result["summary"]["part_count"], 1)
            self.assertEqual(result["preview_summary"][CATEGORY_STOP_LOSS], 1)
            audit_rows = db.get_st_inventory_audit_log("PART-1", limit=10)
            self.assertTrue(any(row["reason"] == "st_reconcile_adjustment" and row["actor"] == "reconcile" for row in audit_rows))

            self._insert_audit(7, 5, -2, "st_consume", "2026-06-30T09:00:00")
            self.assertEqual(theoretical_stock("2026-07-01T00:00:00", part_numbers=["PART-1"])["PART-1"], 5)

            second_path = self._make_genlin_file(tmp, physical_qty=9, book_qty=9)
            second = commit_st_reconcile_stop_loss(second_path, "2026-07-02", source_filename="second.xlsx")
            self.assertEqual(second["preview_summary"][CATEGORY_MATCHED], 1)
            self.assertEqual(db.get_st_inventory_stock()["PART-1"], 9)

            self._insert_audit(9, 8, -1, "st_consume", "2026-07-03T09:00:00")
            self.assertEqual(theoretical_stock("2026-07-04T00:00:00", part_numbers=["PART-1"])["PART-1"], 8)

    def test_blank_genlin_physical_is_previewed_and_skipped_by_commit(self):
        self._insert_snapshot(500)
        with TemporaryDirectory() as tmp:
            path = self._make_genlin_file(tmp, physical_qty=None, book_qty=15)

            preview = build_st_reconcile_preview(path, "2026-06-29")
            by_part = {row["part_number"]: row for row in preview["parts"]}
            self.assertIsNone(by_part["PART-1"]["physical_qty"])
            self.assertEqual(by_part["PART-1"]["category"], CATEGORY_GENLIN_BLANK_PHYSICAL)

            result = commit_st_reconcile_stop_loss(path, "2026-06-29", source_filename="blank.xlsx")

            self.assertEqual(result["summary"]["part_count"], 0)
            self.assertEqual(result["summary"]["updated_count"], 0)
            self.assertEqual(result["adjustments"], [])
            self.assertEqual(db.get_st_inventory_stock()["PART-1"], 500)
            anchor = db.get_latest_st_reconcile_anchor("2026-06-30T00:00:00", ["PART-1"])
            self.assertEqual(anchor["baseline_qty"], {})

    def test_commit_with_part_number_subset_updates_only_selected_parts(self):
        self._insert_snapshot(10, "PART-1")
        self._insert_snapshot(20, "PART-2")
        with TemporaryDirectory() as tmp:
            path = self._make_genlin_file_with_rows(tmp, [
                {"part": "PART-1-TAB", "description": "測試料 1", "book_qty": 10, "physical_qty": 7},
                {"part": "PART-2-TAB", "description": "測試料 2", "book_qty": 20, "physical_qty": 19},
            ])

            result = commit_st_reconcile_stop_loss(
                path,
                "2026-06-29",
                source_filename="subset.xlsx",
                part_numbers=[" part-2 "],
            )

            self.assertEqual(result["summary"]["part_count"], 1)
            self.assertEqual(result["summary"]["adjusted_count"], 1)
            self.assertEqual(db.get_st_inventory_stock()["PART-1"], 10)
            self.assertEqual(db.get_st_inventory_stock()["PART-2"], 19)

            alignment = self.conn.execute(
                "SELECT part_count, diff_count FROM st_reconcile_alignments WHERE id=?",
                (result["summary"]["alignment_id"],),
            ).fetchone()
            self.assertEqual(alignment["part_count"], 1)
            self.assertEqual(alignment["diff_count"], 1)
            parts = self.conn.execute(
                "SELECT part_number FROM st_reconcile_alignment_parts WHERE alignment_id=?",
                (result["summary"]["alignment_id"],),
            ).fetchall()
            self.assertEqual([row["part_number"] for row in parts], ["PART-2"])

    def test_commit_with_empty_part_number_list_returns_plain_error(self):
        self._insert_snapshot(10)
        with TemporaryDirectory() as tmp:
            path = self._make_genlin_file(tmp, physical_qty=7, book_qty=10)

            with self.assertRaisesRegex(ValueError, "請至少勾選 1 支料號再建立停損點"):
                commit_st_reconcile_stop_loss(path, "2026-06-29", part_numbers=[])

    def test_commit_same_genlin_file_twice_records_zero_second_adjustment(self):
        self._insert_snapshot(10)
        with TemporaryDirectory() as tmp:
            path = self._make_genlin_file(tmp, physical_qty=7, book_qty=10)

            first = commit_st_reconcile_stop_loss(path, "2026-06-29", source_filename="same.xlsx")
            second = commit_st_reconcile_stop_loss(path, "2026-06-29", source_filename="same.xlsx")

            self.assertEqual(first["summary"]["adjusted_count"], 1)
            self.assertEqual(second["summary"]["adjusted_count"], 0)
            self.assertEqual(second["adjustments"][0]["adjust_qty"], 0)
            self.assertEqual(db.get_st_inventory_stock()["PART-1"], 7)

    def test_commit_preserves_post_cutoff_consumption_and_second_commit_is_idempotent(self):
        self._insert_snapshot(100)
        self._insert_audit(100, 70, -30, "st_consume", "2026-06-30T09:00:00")
        with TemporaryDirectory() as tmp:
            path = self._make_genlin_file(tmp, physical_qty=80, book_qty=100)

            first = commit_st_reconcile_stop_loss(path, "2026-06-29", source_filename="post-cutoff.xlsx")

            self.assertEqual(first["summary"]["adjusted_count"], 1)
            self.assertEqual(first["adjustments"][0]["adjust_qty"], -20)
            self.assertEqual(db.get_st_inventory_stock()["PART-1"], 50)

            second = commit_st_reconcile_stop_loss(path, "2026-06-29", source_filename="post-cutoff.xlsx")

            self.assertEqual(second["summary"]["adjusted_count"], 0)
            self.assertEqual(second["adjustments"][0]["adjust_qty"], 0)
            self.assertEqual(db.get_st_inventory_stock()["PART-1"], 50)

    def test_preview_lists_defective_parts_not_covered_by_stocktake(self):
        self._insert_snapshot(10)
        for part, qty, created in (
            ("PART-1", 5, "2026-06-21T09:00:00"),
            ("SELF-9", 12, "2026-06-22T09:00:00"),
            ("SELF-9", 3, "2026-06-25T09:00:00"),
            ("SELF-LATE", 4, "2026-07-05T09:00:00"),
        ):
            self.conn.execute(
                "INSERT INTO defective_records(part_number, defective_qty, created_at) VALUES(?, ?, ?)",
                (part, qty, created),
            )

        with TemporaryDirectory() as tmp:
            path = self._make_genlin_file(tmp, physical_qty=7, book_qty=10)
            preview = build_st_reconcile_preview(path, "2026-06-29")

        uncovered = {item["part_number"]: item for item in preview["uncovered_parts"]}
        self.assertNotIn("PART-1", uncovered)          # 檔內有的料不列
        self.assertNotIn("SELF-LATE", uncovered)       # 截止點之後的不算本期
        self.assertIn("SELF-9", uncovered)
        self.assertEqual(uncovered["SELF-9"]["total_qty"], 15)
        self.assertEqual(uncovered["SELF-9"]["record_count"], 2)

    def test_batch_cutoff_absorbs_own_batch_and_replays_only_later_batches(self):
        # 鎖死雙扣回歸：選批次 6-1 時，6-1 自己的消耗（audit 時間晚於 dispatched_at）
        # 必須落在盤點吸收側，只有 6-1 之後的批次被重放。
        self._insert_snapshot(100)
        order_cursor = self.conn.execute(
            "INSERT INTO orders(po_number, model, status, code) VALUES('PO-6-1', 'MODEL-A', 'dispatched', '6-1')"
        )
        session_cursor = self.conn.execute(
            "INSERT INTO dispatch_sessions(order_id, dispatched_at) VALUES(?, '2026-06-28T14:00:00')",
            (order_cursor.lastrowid,),
        )
        self._insert_audit(100, 70, -30, "st_consume", "2026-06-28T14:00:05")
        self.conn.execute(
            """
            INSERT INTO st_dispatch_consumptions(dispatch_session_id, order_id, part_number, used_qty, consumed_at)
            VALUES(?, ?, 'PART-1', 30, '2026-06-28T14:00:06')
            """,
            (session_cursor.lastrowid, order_cursor.lastrowid),
        )
        self._insert_audit(70, 50, -20, "st_consume", "2026-06-30T09:00:05")

        resolved = resolve_cutoff_batch("6-1")
        self.assertEqual(resolved["cutoff_at"], "2026-06-28T14:00:06")

        with TemporaryDirectory() as tmp:
            path = self._make_genlin_file(tmp, physical_qty=68, book_qty=70)
            result = commit_st_reconcile_stop_loss(
                path,
                resolved["cutoff_at"],
                source_filename="batch-cutoff.xlsx",
                cutoff_label=resolved["code"],
            )

            self.assertTrue(result["ok"])
            # 目標 = 實盤 68 + 只有 6-30 那筆 -20 → 48（雙扣錯誤會算出 18）
            self.assertEqual(db.get_st_inventory_stock()["PART-1"], 48)
            self.assertEqual(result["adjustments"][0]["adjust_qty"], -2)


class StReconcileCutoffBatchTests(unittest.TestCase):
    def setUp(self):
        self.conn = sqlite3.connect(":memory:")
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA foreign_keys=ON")
        self.conn.executescript(db._CREATE_SQL)

        @contextmanager
        def temp_conn():
            try:
                yield self.conn
                self.conn.commit()
            except Exception:
                self.conn.rollback()
                raise

        self.get_conn_patcher = patch.object(db, "get_conn", temp_conn)
        self.get_conn_patcher.start()

    def tearDown(self):
        self.get_conn_patcher.stop()
        self.conn.close()

    def _insert_order_with_session(self, code: str, dispatched_at: str, rolled_back_at: str = "") -> int:
        cursor = self.conn.execute(
            "INSERT INTO orders(po_number, model, status, code) VALUES(?, ?, 'dispatched', ?)",
            (f"PO-{code}", f"MODEL-{code}", code),
        )
        order_id = cursor.lastrowid
        session_cursor = self.conn.execute(
            "INSERT INTO dispatch_sessions(order_id, dispatched_at, rolled_back_at) VALUES(?, ?, ?)",
            (order_id, dispatched_at, rolled_back_at),
        )
        return int(session_cursor.lastrowid)

    def _insert_consumption(self, session_id: int, consumed_at: str, part_number: str = "PART-1", used_qty: float = 1) -> None:
        self.conn.execute(
            """
            INSERT INTO st_dispatch_consumptions(dispatch_session_id, order_id, part_number, used_qty, consumed_at)
            SELECT ?, order_id, ?, ?, ? FROM dispatch_sessions WHERE id=?
            """,
            (session_id, part_number, used_qty, consumed_at, session_id),
        )

    def test_cutoff_options_exclude_rolled_back_and_sort_newest_first(self):
        self._insert_order_with_session("1-3", "2026-06-11T09:32:08")
        self._insert_order_with_session("6-3", "2026-06-28T14:05:00")
        self._insert_order_with_session("9-9", "2026-07-01T10:00:00", rolled_back_at="2026-07-02T08:00:00")

        options = db.get_st_reconcile_cutoff_batch_options()

        codes = [option["code"] for option in options]
        self.assertEqual(codes, ["6-3", "1-3"])
        self.assertNotIn("9-9", codes)
        self.assertEqual(options[0]["dispatched_at"], "2026-06-28T14:05:00")

    def test_resolve_cutoff_batch_uses_last_consumption_time_to_absorb_own_batch(self):
        session_id = self._insert_order_with_session("6-1", "2026-06-28T14:32:08")
        self._insert_consumption(session_id, "2026-06-28T14:32:15")

        resolved = resolve_cutoff_batch("6-1")

        self.assertEqual(resolved["dispatched_at"], "2026-06-28T14:32:08")
        self.assertEqual(resolved["cutoff_at"], "2026-06-28T14:32:15")
        with self.assertRaises(ValueError):
            resolve_cutoff_batch("99-99")
        with self.assertRaises(ValueError):
            resolve_cutoff_batch("")

    def test_resolve_cutoff_batch_without_consumption_falls_back_to_dispatched_at(self):
        self._insert_order_with_session("7-1", "2026-07-01T09:00:00")

        resolved = resolve_cutoff_batch("7-1")

        self.assertEqual(resolved["cutoff_at"], "2026-07-01T09:00:00")


if __name__ == "__main__":
    unittest.main()
