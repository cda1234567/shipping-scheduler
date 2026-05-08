from __future__ import annotations

import unittest

import openpyxl

from app.services.main_file_recalc import recalc_batch_balances_for_cell


class MainFileRecalcTests(unittest.TestCase):
    def _build_sheet(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "主檔"

        ws.cell(row=1, column=1).value = "料號"
        ws.cell(row=1, column=8).value = "盤點"

        for col, code in ((9, "1-1"), (12, "1-2"), (15, "1-3"), (18, "1-4")):
            ws.cell(row=1, column=col).value = code
            ws.cell(row=1, column=col + 1).value = f"PO-{code}"
            ws.cell(row=1, column=col + 2).value = f"MODEL-{code}"

        ws.cell(row=2, column=1).value = "PART-1"
        ws.cell(row=2, column=8).value = 100

        ws.cell(row=2, column=9).value = 0
        ws.cell(row=2, column=10).value = 30
        ws.cell(row=2, column=11).value = 70

        ws.cell(row=2, column=12).value = 5
        ws.cell(row=2, column=13).value = 20
        ws.cell(row=2, column=14).value = 55

        # 1-3 這批對此料號完全空白，重算時要跳過且不要寫結餘。

        ws.cell(row=2, column=18).value = 0
        ws.cell(row=2, column=19).value = 5
        ws.cell(row=2, column=20).value = 50
        return wb, ws

    def test_edit_supplement_recalculates_following_balances(self):
        wb, ws = self._build_sheet()
        try:
            ws.cell(row=2, column=9).value = 10

            result = recalc_batch_balances_for_cell(ws, row=2, col=9)

            self.assertTrue(result["recalculated"])
            self.assertEqual(
                result["affected_cells"],
                [
                    {"row": 2, "col": 11, "value": 80},
                    {"row": 2, "col": 14, "value": 65},
                    {"row": 2, "col": 20, "value": 60},
                ],
            )
            self.assertEqual(ws.cell(row=2, column=11).value, 80)
            self.assertEqual(ws.cell(row=2, column=14).value, 65)
            self.assertEqual(ws.cell(row=2, column=20).value, 60)
            self.assertEqual(result["current_stock"], 60)
        finally:
            wb.close()

    def test_blank_supplement_recalculates_as_zero(self):
        wb, ws = self._build_sheet()
        try:
            ws.cell(row=2, column=9).value = None

            result = recalc_batch_balances_for_cell(ws, row=2, col=9)

            self.assertTrue(result["recalculated"])
            self.assertEqual(ws.cell(row=2, column=11).value, 70)
            self.assertEqual(ws.cell(row=2, column=14).value, 55)
            self.assertEqual(ws.cell(row=2, column=20).value, 50)
            self.assertEqual(result["current_stock"], 50)
        finally:
            wb.close()

    def test_edit_usage_recalculates_following_balances(self):
        wb, ws = self._build_sheet()
        try:
            ws.cell(row=2, column=13).value = 50

            result = recalc_batch_balances_for_cell(ws, row=2, col=13)

            self.assertEqual(
                result["affected_cells"],
                [
                    {"row": 2, "col": 14, "value": 25},
                    {"row": 2, "col": 20, "value": 20},
                ],
            )
            self.assertEqual(ws.cell(row=2, column=11).value, 70)
            self.assertEqual(ws.cell(row=2, column=14).value, 25)
            self.assertEqual(ws.cell(row=2, column=20).value, 20)
            self.assertEqual(result["current_stock"], 20)
        finally:
            wb.close()

    def test_edit_balance_itself_does_not_recalculate(self):
        wb, ws = self._build_sheet()
        try:
            ws.cell(row=2, column=11).value = 999

            result = recalc_batch_balances_for_cell(ws, row=2, col=11)

            self.assertFalse(result["recalculated"])
            self.assertEqual(result["affected_cells"], [])
            self.assertEqual(ws.cell(row=2, column=11).value, 999)
            self.assertEqual(ws.cell(row=2, column=14).value, 55)
            self.assertEqual(ws.cell(row=2, column=20).value, 50)
        finally:
            wb.close()

    def test_empty_batch_is_skipped(self):
        wb, ws = self._build_sheet()
        try:
            ws.cell(row=2, column=9).value = 10

            result = recalc_batch_balances_for_cell(ws, row=2, col=9)

            self.assertNotIn({"row": 2, "col": 17, "value": 65}, result["affected_cells"])
            self.assertIsNone(ws.cell(row=2, column=15).value)
            self.assertIsNone(ws.cell(row=2, column=16).value)
            self.assertIsNone(ws.cell(row=2, column=17).value)
        finally:
            wb.close()


if __name__ == "__main__":
    unittest.main()
